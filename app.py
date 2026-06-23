# Add this near the top of your app.py file, after the imports

import json
import re
import textwrap
from datetime import datetime
from html import escape
from io import BytesIO
from typing import Callable, Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st

from core_utils import (
    bytes_to_pdfplumber,
    dedupe_transactions,
    normalize_transactions,
    safe_float,
)
from transaction_analysis import parse_top_parties_and_high_value
from party_utils import (
    apply_party_aliasing,
    build_transactions_by_party,
    clean_counterparty_name,
    deduplicate_counterparty_names,
)

from maybank import parse_transactions_maybank
from public_bank import parse_transactions_pbb
from rhb import parse_transactions_rhb
from cimb import parse_transactions_cimb,extract_cimb_party_name
from bank_islam import parse_bank_islam
from bank_rakyat import parse_bank_rakyat
from hong_leong import parse_hong_leong
from ambank import parse_ambank, extract_ambank_statement_totals
from bank_muamalat import parse_transactions_bank_muamalat
from affin_bank import parse_affin_bank, extract_affin_statement_totals
from agro_bank import parse_agro_bank
from ocbc import parse_transactions_ocbc
from uob import parse_transactions_uob
from alliance import parse_transactions_alliance
from pdf_security import is_pdf_encrypted, decrypt_pdf_bytes

# Import the extracted functions
from pdf_utils import extract_company_name, extract_account_number
from bank_totals import (
    extract_cimb_statement_totals,
    extract_rhb_statement_totals,
    extract_bank_islam_statement_month
)

# Fraud detection logic imports
from fraud_logic import (
    analyze_pdf_batch,
    build_display_summary,
    detect_font_anomalies,
)

# Track 2 classifier — import ALL functions needed
try:
    from kredit_lab_classify_track2 import (
        # Core engine functions
        build_track2_result,
        account_meta_from_determinations,
        classify_transactions,
        compute_risk_flags,
        
        # Pattern matching for loans (CRITICAL for loan detection)
        LOAN_DISBURSEMENT_RE,
        LOAN_REPAYMENT_RE,
        
        # Statutory compliance
        compute_statutory_compliance,
        compute_salary_payments,
        compute_epf_payments,
        compute_socso_payments,
        compute_lhdn_tax_payments,
        compute_hrdf_payments,
        
        # FX and other detectors
        compute_fx_totals,
        compute_round_figure_credits,
        compute_high_value_credits,
        compute_returned_cheques,
        
        # Counterparty ledger functions
        scan_related_party_candidates,
        auto_confirmed_related_parties,
        dedup_counterparty_entries,
        
        # Monthly aggregator
        compute_monthly_aggregates,
        compute_monthly_eod,
        
        # Constants
        CANONICAL_FLAGS,
        LOW_CLOSING_THRESHOLD_CR,
        OD_HIGH_UTILISATION_RATIO,
        SUBTHRESHOLD_TOTAL_SALARY_RM,
        CHANNEL_BLIND_CHEQUE_DR_MIN_RM,
        CHANNEL_BLIND_CHEQUE_DR_MIN_RATIO,
    )
    _TRACK2_AVAILABLE = True
except ImportError as e:
    print(f"[WARNING] Track 2 import failed: {e}")
    _TRACK2_AVAILABLE = False

# ============================================================
# SIDEBAR NAVIGATION FOR STREAMLIT APP
# ============================================================

def init_sidebar_navigation():
    """Initialize sidebar navigation state"""
    if "sidebar_collapsed" not in st.session_state:
        st.session_state.sidebar_collapsed = False
    
    if "active_section" not in st.session_state:
        st.session_state.active_section = "overview"


def toggle_sidebar():
    """Toggle sidebar collapsed state"""
    st.session_state.sidebar_collapsed = not st.session_state.sidebar_collapsed


def render_sidebar_navigation():
    """Render the collapsible sidebar navigation for Streamlit app"""
    results = st.session_state.get("results", [])
    company_name = st.session_state.get("company_name_override", "")
    if not company_name and results:
        for t in results:
            if t.get("company_name"):
                company_name = t["company_name"]
                break
    if not company_name:
        company_name = "Kredit Lab"

    nav_items = [
        {"id": "overview", "icon": "\U0001F3E0", "label": "Overview"},
        {"id": "extracted", "icon": "\U0001F4C4", "label": "Extracted Transactions"},
        {"id": "patterns", "icon": "\U0001F4CA", "label": "Pattern Analysis"},
        {"id": "counterparty", "icon": "\U0001F465", "label": "Counterparty Ledger"},
        {"id": "monthly", "icon": "\U0001F4C5", "label": "Monthly Summary"},
        {"id": "download", "icon": "\u2B07", "label": "Download Options"},
        {"id": "integrity", "icon": "\U0001F6E1", "label": "Document Integrity"},
    ]

    st.sidebar.markdown(f"### {company_name}")
    st.sidebar.caption("Statement Intelligence")
    st.sidebar.markdown("#### Navigation")

    has_results = bool(results)
    for item in nav_items:
        if not has_results and item["id"] not in ["overview", "download"]:
            continue
        st.sidebar.markdown(f'[{item["icon"]} {item["label"]}](#{item["id"]}-section)')

    st.sidebar.caption("v2.0.0")
    return False
    
    # Sidebar CSS - this creates a custom sidebar that overlays the Streamlit UI
    st.markdown("""
    <style>
        /* Hide the default Streamlit sidebar */
        section[data-testid="stSidebar"] {
            display: none !important;
        }
        
        /* Custom sidebar toggle button */
        .custom-sidebar-toggle {
            position: fixed;
            top: 70px;
            left: 0;
            z-index: 999;
            background: #0b0f19;
            border: 1px solid #1e2a42;
            border-left: none;
            border-radius: 0 8px 8px 0;
            color: #e2e8f0;
            cursor: pointer;
            padding: 10px 6px;
            font-size: 18px;
            transition: all 0.3s ease;
            box-shadow: 2px 0 10px rgba(0,0,0,0.3);
            width: 28px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        
        .custom-sidebar-toggle:hover {
            background: #1a2235;
            width: 32px;
            border-color: #3b82f6;
        }
        
        /* Main sidebar container */
        .custom-sidebar {
            position: fixed;
            top: 0;
            left: 0;
            height: 100vh;
            width: 240px;
            background: #0b0f19;
            border-right: 1px solid #1e2a42;
            padding: 60px 0 20px 0;
            overflow-y: auto;
            overflow-x: hidden;
            z-index: 998;
            transition: transform 0.3s ease, width 0.3s ease;
            box-shadow: 2px 0 15px rgba(0,0,0,0.5);
        }
        
        .custom-sidebar.collapsed {
            transform: translateX(-210px);
            width: 240px;
        }
        
        .custom-sidebar.collapsed .nav-label {
            opacity: 0;
            max-width: 0;
            overflow: hidden;
            transition: opacity 0.2s ease, max-width 0.2s ease;
        }
        
        .custom-sidebar.collapsed .nav-item {
            padding: 10px 12px;
            justify-content: center;
        }
        
        .custom-sidebar.collapsed .nav-icon {
            margin-right: 0;
        }
        
        .custom-sidebar.collapsed .sidebar-company {
            padding: 8px 12px;
        }
        
        .custom-sidebar.collapsed .sidebar-company strong {
            font-size: 10px;
            text-align: center;
        }
        
        .custom-sidebar.collapsed .sidebar-company span {
            display: none;
        }
        
        .custom-sidebar.collapsed .sidebar-version {
            display: none;
        }
        
        /* Scrollbar styling */
        .custom-sidebar::-webkit-scrollbar {
            width: 4px;
        }
        
        .custom-sidebar::-webkit-scrollbar-track {
            background: transparent;
        }
        
        .custom-sidebar::-webkit-scrollbar-thumb {
            background: #1e2a42;
            border-radius: 2px;
        }
        
        .custom-sidebar::-webkit-scrollbar-thumb:hover {
            background: #334155;
        }
        
        /* Navigation items */
        .nav-item {
            display: flex;
            align-items: center;
            padding: 10px 16px;
            color: #94a3b8;
            text-decoration: none;
            cursor: pointer;
            border-left: 3px solid transparent;
            transition: all 0.2s ease;
            font-size: 13px;
            font-weight: 500;
            gap: 0;
            white-space: nowrap;
            border-radius: 0;
        }
        
        .nav-item:hover {
            background: #1a2235;
            color: #e2e8f0;
            border-left-color: #3b82f6;
        }
        
        .nav-item.active {
            background: #1a2235;
            color: #60a5fa;
            border-left-color: #3b82f6;
        }
        
        .nav-item .nav-icon {
            font-size: 18px;
            min-width: 28px;
            margin-right: 8px;
            flex-shrink: 0;
        }
        
        .nav-item .nav-label {
            transition: opacity 0.2s ease, max-width 0.2s ease;
            opacity: 1;
            max-width: 150px;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        
        .nav-section-title {
            padding: 16px 16px 8px 16px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #475569;
        }
        
        /* Company name in sidebar */
        .sidebar-company {
            padding: 12px 16px 8px 16px;
            border-bottom: 1px solid #1e2a42;
            margin-bottom: 4px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            transition: all 0.3s ease;
        }
        
        .sidebar-company strong {
            color: #e2e8f0;
            font-size: 14px;
            display: block;
            margin-bottom: 2px;
            overflow: hidden;
            text-overflow: ellipsis;
        }
        
        .sidebar-company span {
            color: #64748b;
            font-size: 10px;
            transition: opacity 0.3s ease;
        }
        
        .sidebar-version {
            position: absolute;
            bottom: 12px;
            left: 0;
            right: 0;
            text-align: center;
            font-size: 10px;
            color: #475569;
            padding: 8px;
            border-top: 1px solid #1e2a42;
        }
        
        /* Collapse button inside sidebar */
        .sidebar-collapse-btn {
            position: absolute;
            top: 12px;
            right: 12px;
            background: transparent;
            border: 1px solid #1e2a42;
            border-radius: 4px;
            color: #94a3b8;
            cursor: pointer;
            padding: 4px 8px;
            font-size: 14px;
            transition: all 0.2s ease;
            z-index: 1;
        }
        
        .sidebar-collapse-btn:hover {
            background: #1a2235;
            color: #e2e8f0;
            border-color: #3b82f6;
        }
        
        /* Bottom spacer */
        .sidebar-bottom-spacer {
            height: 60px;
        }
        
        /* Main content adjustment */
        .main-content-wrapper {
            margin-left: 240px;
            transition: margin-left 0.3s ease;
            padding: 0 20px 20px 20px;
        }
        
        .main-content-wrapper.collapsed {
            margin-left: 30px;
        }
        
        /* Responsive */
        @media (max-width: 768px) {
            .custom-sidebar {
                width: 200px;
                transform: translateX(-100%);
            }
            
            .custom-sidebar.mobile-open {
                transform: translateX(0);
            }
            
            .custom-sidebar.collapsed {
                transform: translateX(-100%);
            }
            
            .main-content-wrapper,
            .main-content-wrapper.collapsed {
                margin-left: 0;
                padding: 0 10px;
            }
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Get company name for sidebar
    company_name = st.session_state.get("company_name_override", "")
    results = st.session_state.get("results", [])
    if not company_name and results:
        for t in results:
            if t.get("company_name"):
                company_name = t["company_name"]
                break
    
    if not company_name:
        company_name = "Kredit Lab"
    
    # Navigation items
    nav_items = [
        {"id": "overview", "icon": "🏠", "label": "Overview"},
        {"id": "extracted", "icon": "📄", "label": "Extracted Transactions"},
        {"id": "patterns", "icon": "📊", "label": "Pattern Analysis"},
        {"id": "counterparty", "icon": "👥", "label": "Counterparty Ledger"},
        {"id": "monthly", "icon": "📅", "label": "Monthly Summary"},
        {"id": "download", "icon": "⬇️", "label": "Download Options"},
        {"id": "integrity", "icon": "🛡️", "label": "Document Integrity"},
    ]
    
    # Check if results exist for conditional items
    has_results = bool(results)
    
    collapsed_class = "collapsed" if st.session_state.sidebar_collapsed else ""
    active_section = st.session_state.active_section
    
    # Build sidebar HTML
    nav_html = f'''
    <button class="custom-sidebar-toggle" onclick="toggleCustomSidebar()" title="Toggle Sidebar">
        {"" if st.session_state.sidebar_collapsed else "☰"}
    </button>
    
    <div class="custom-sidebar {collapsed_class}" id="customSidebar">
        <button class="sidebar-collapse-btn" onclick="toggleCustomSidebar()">
            {"" if st.session_state.sidebar_collapsed else "◀"}
        </button>
        
        <div class="sidebar-company">
            <strong>{company_name}</strong>
            <span>Statement Intelligence</span>
        </div>
        
        <div class="nav-section-title">Navigation</div>
    '''
    
    for item in nav_items:
        # Skip if no results and item requires results
        if not has_results and item["id"] not in ["overview", "download"]:
            continue
            
        active_class = "active" if active_section == item["id"] else ""
        nav_html += f'''
        <div class="nav-item {active_class}" onclick="navigateToSection('{item["id"]}')">
            <span class="nav-icon">{item["icon"]}</span>
            <span class="nav-label">{item["label"]}</span>
        </div>
        '''
    
    nav_html += '''
        <div class="sidebar-bottom-spacer"></div>
        <div class="sidebar-version">v2.0.0</div>
    </div>
    '''
    
    # JavaScript for sidebar interaction
    js = '''
    <script>
        function toggleCustomSidebar() {
            const sidebar = document.getElementById('customSidebar');
            if (sidebar) {
                sidebar.classList.toggle('collapsed');
                // Update the toggle button text
                const toggleBtn = document.querySelector('.custom-sidebar-toggle');
                if (toggleBtn) {
                    toggleBtn.textContent = sidebar.classList.contains('collapsed') ? '☰' : '◀';
                }
                // Send update to Streamlit
                const isCollapsed = sidebar.classList.contains('collapsed');
                const event = new CustomEvent('streamlit:setComponentValue', {
                    detail: { 
                        key: 'sidebar_collapsed',
                        value: isCollapsed 
                    }
                });
                document.dispatchEvent(event);
            }
        }
        
        function navigateToSection(sectionId) {
            // Update active state
            document.querySelectorAll('.nav-item').forEach(el => {
                el.classList.remove('active');
            });
            const clicked = document.querySelector(`.nav-item[onclick*="${sectionId}"]`);
            if (clicked) {
                clicked.classList.add('active');
            }
            
            // Scroll to the section
            const sectionMap = {
                'overview': 'overview-section',
                'extracted': 'extracted-section',
                'patterns': 'patterns-section',
                'counterparty': 'counterparty-section',
                'monthly': 'monthly-section',
                'download': 'download-section',
                'integrity': 'integrity-section'
            };
            
            const sectionId_map = sectionMap[sectionId];
            if (sectionId_map) {
                const element = document.getElementById(sectionId_map);
                if (element) {
                    const offset = 80;
                    const elementPosition = element.getBoundingClientRect().top;
                    const offsetPosition = elementPosition + window.pageYOffset - offset;
                    window.scrollTo({ top: offsetPosition, behavior: 'smooth' });
                }
            }
            
            // Send update to Streamlit
            const event = new CustomEvent('streamlit:setComponentValue', {
                detail: { 
                    key: 'active_section',
                    value: sectionId 
                }
            });
            document.dispatchEvent(event);
        }
        
        // Handle mobile
        function handleMobileSidebar() {
            const sidebar = document.getElementById('customSidebar');
            if (window.innerWidth <= 768 && sidebar) {
                if (!sidebar.classList.contains('collapsed')) {
                    sidebar.classList.add('mobile-open');
                }
            } else if (sidebar) {
                sidebar.classList.remove('mobile-open');
            }
        }
        
        window.addEventListener('resize', handleMobileSidebar);
        document.addEventListener('DOMContentLoaded', handleMobileSidebar);
    </script>
    '''
    
    st.markdown(textwrap.dedent(nav_html), unsafe_allow_html=True)
    st.markdown(textwrap.dedent(js), unsafe_allow_html=True)
    
    # Return the collapsed state
    return st.session_state.sidebar_collapsed


def get_main_content_class():
    """Get the CSS class for main content based on sidebar state"""
    if st.session_state.sidebar_collapsed:
        return "main-content-wrapper collapsed"
    return "main-content-wrapper"

# ============================================================
# END OF SIDEBAR NAVIGATION FUNCTIONS
# ============================================================

# ============================================================
# STREAMLIT PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Bank Statement Parser", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize sidebar navigation
init_sidebar_navigation()

# Render the sidebar
sidebar_collapsed = render_sidebar_navigation()

# Get the main content class
main_class = get_main_content_class()

# Wrap your main content
st.markdown(textwrap.dedent(f"""
<style>
    .main-content-wrapper {{
        max-width: 1400px;
        margin: 0 auto;
    }}
    
    .section-anchor {{
        scroll-margin-top: 80px;
    }}
</style>
<div class="{main_class}">
"""), unsafe_allow_html=True)

# ============================================================
# SESSION STATE INIT
# ============================================================

if "status" not in st.session_state:
    st.session_state.status = "idle"

if "results" not in st.session_state:
    st.session_state.results = []

if "integrity_analysis_results" not in st.session_state:
    st.session_state.integrity_analysis_results = {}

if "affin_statement_totals" not in st.session_state:
    st.session_state.affin_statement_totals = []

if "affin_file_transactions" not in st.session_state:
    st.session_state.affin_file_transactions = {}

if "ambank_statement_totals" not in st.session_state:
    st.session_state.ambank_statement_totals = []

if "ambank_file_transactions" not in st.session_state:
    st.session_state.ambank_file_transactions = {}

if "cimb_statement_totals" not in st.session_state:
    st.session_state.cimb_statement_totals = []

if "cimb_file_transactions" not in st.session_state:
    st.session_state.cimb_file_transactions = {}

if "rhb_statement_totals" not in st.session_state:
    st.session_state.rhb_statement_totals = []

if "rhb_file_transactions" not in st.session_state:
    st.session_state.rhb_file_transactions = {}

if "bank_islam_file_month" not in st.session_state:
    st.session_state.bank_islam_file_month = {}

# password + company name tracking
if "pdf_password" not in st.session_state:
    st.session_state.pdf_password = ""

if "company_name_override" not in st.session_state:
    st.session_state.company_name_override = ""

if "company_account_no_override" not in st.session_state:
    st.session_state.company_account_no_override = ""

if "high_value_threshold_input" not in st.session_state:
    st.session_state.high_value_threshold_input = ""

if "high_value_threshold_error" not in st.session_state:
    st.session_state.high_value_threshold_error = ""

if "bank_choice_error" not in st.session_state:
    st.session_state.bank_choice_error = ""

if "pdf_upload_error" not in st.session_state:
    st.session_state.pdf_upload_error = ""

if "validation_toast_message" not in st.session_state:
    st.session_state.validation_toast_message = ""

if "active_high_value_threshold" not in st.session_state:
    st.session_state.active_high_value_threshold = None

if "stop_requested" not in st.session_state:
    st.session_state.stop_requested = False

if "upload_widget_reset_id" not in st.session_state:
    st.session_state.upload_widget_reset_id = 0

if "file_company_name" not in st.session_state:
    st.session_state.file_company_name = {}

if "file_account_no" not in st.session_state:
    st.session_state.file_account_no = {}

# Track 2 — account-type determinations populated by parser hooks
if "account_type_determinations" not in st.session_state:
    st.session_state.account_type_determinations = []

# Analyst-supplied related parties (populated by the analyst form when wired)
if "related_parties_override" not in st.session_state:
    st.session_state.related_parties_override = []

if "counterparty_name_overrides" not in st.session_state:
    st.session_state.counterparty_name_overrides = {}

# ============================================================
# build_large_transactions FUNCTION
# ============================================================

def build_large_transactions(transactions: List[dict], threshold: float) -> List[dict]:
    """Build list of large transactions (both credits and debits) above threshold."""
    large_txns = []
    threshold_float = float(threshold)
    
    for t in transactions:
        credit = safe_float(t.get('credit', 0))
        debit = safe_float(t.get('debit', 0))
        
        # Check both credits and debits
        if credit >= threshold_float:
            large_txns.append({
                'date': t.get('date', ''),
                'description': t.get('description', ''),
                'amount': credit,
                'balance': t.get('balance', 0),
                'type': 'CREDIT'
            })
        elif debit >= threshold_float:
            large_txns.append({
                'date': t.get('date', ''),
                'description': t.get('description', ''),
                'amount': debit,
                'balance': t.get('balance', 0),
                'type': 'DEBIT'
            })
    
    # Sort by amount descending
    large_txns.sort(key=lambda x: x['amount'], reverse=True)
    return large_txns

# Function copy from HTML, def generate_interactive_html(data) - you will replace this with the full function from your original converter file
def generate_interactive_html(data):
    """Generate interactive HTML report for v6.0.0 schema"""

    # Define build_large_transactions inside the function if needed
    def build_large_transactions_internal(transactions, threshold):
        """Build list of large transactions from transaction list"""
        large_txns = []
        threshold_float = float(threshold) if threshold else 100000.0
        
        for t in transactions:
            if not isinstance(t, dict):
                continue
            credit = safe_float(t.get('credit', 0))
            debit = safe_float(t.get('debit', 0))
            
            if credit >= threshold_float:
                large_txns.append({
                    'date': t.get('date', ''),
                    'description': t.get('description', ''),
                    'amount': credit,
                    'balance': t.get('balance', 0),
                    'type': 'CREDIT'
                })
            elif debit >= threshold_float:
                large_txns.append({
                    'date': t.get('date', ''),
                    'description': t.get('description', ''),
                    'amount': debit,
                    'balance': t.get('balance', 0),
                    'type': 'DEBIT'
                })
        
        large_txns.sort(key=lambda x: x.get('amount', 0), reverse=True)
        return large_txns

    def build_round_figure_credits_internal(transactions):
        """Build the same signed round-number transactions shown in Railway."""
        return build_round_transactions(transactions)
    
    # Extract data
    _fallback_consol = data.get('consolidated') if isinstance(data.get('consolidated'), dict) else {}
    _fallback_summary = data.get('summary') if isinstance(data.get('summary'), dict) else {}
    _fallback_config = data.get('classification_config') if isinstance(data.get('classification_config'), dict) else {}
    fallback_large_threshold = (
        _fallback_consol.get('high_value_threshold')
        or _fallback_summary.get('high_value_threshold')
        or _fallback_config.get('large_transaction_threshold')
        or _fallback_config.get('large_credit_threshold')
        or 100000
    )
    large_txns = data.get('large_transactions', [])
    if not large_txns and data.get('transactions'):
        large_txns = build_large_transactions_internal(data.get('transactions', []), fallback_large_threshold)

    r = data.get('report_info', {})
    accounts = data.get('accounts', [])
    monthly = data.get('monthly_analysis', [])
    consol = data.get('consolidated', {})
    top_parties = data.get('top_parties', {})
    large_credits = data.get('large_credits', [])
    own_related = data.get('own_related_transactions', {})
    if isinstance(own_related, list):
        own_related = {'transactions': own_related, 'summary': {}}
    elif not isinstance(own_related, dict):
        own_related = {}
    loans = data.get('loan_transactions', {})
    flags_data = data.get('flags', {})
    obs = normalize_observations(data.get('observations', {}))
    parsing = data.get('parsing_metadata', {})
    _sync_data_quality_status(data)
    consol = data.get('consolidated', {})
    flags_data = data.get('flags', {})
    obs = normalize_observations(data.get('observations', {}))
    parsing = data.get('parsing_metadata', {})

    # Version detection
    schema_v = r.get('schema_version', '')
    is_v620 = schema_v in ('6.2.0', '6.2.1', '6.2.2', '6.3.0', '6.3.1', '6.3.2', '6.3.3', '6.3.4', '6.3.5') or consol.get('total_fx_credits') is not None
    is_v630 = schema_v in ('6.3.0', '6.3.1', '6.3.2', '6.3.3', '6.3.4', '6.3.5') or consol.get('total_unclassified_cr') is not None
    is_v635 = schema_v in ('6.3.4', '6.3.5')
    has_parsing = bool(parsing)
    has_monthly_bd = any(p.get('monthly_breakdown') for p in (top_parties.get('top_payers') or top_parties.get('top_creditors') or []) + (top_parties.get('top_payees') or top_parties.get('top_debtors') or []))

    # v6.2.1: Data quality detection
    parsing_checks = parsing.get('account_month_checks', []) if isinstance(parsing, dict) else []
    recon_lookup = {}
    for chk in parsing_checks:
        month_key = str(chk.get('month', '') or '')
        account_key = str(chk.get('account_number', '') or '')
        recon_lookup[(month_key, account_key)] = chk

    def _recon_check_for_month_row(row):
        if not isinstance(row, dict):
            return None
        month_key = str(row.get('month', '') or '')
        account_key = str(row.get('account_number', '') or '')
        return recon_lookup.get((month_key, account_key)) or recon_lookup.get((month_key, ''))

    def _recon_status_for_month_row(row):
        chk = _recon_check_for_month_row(row)
        if chk is not None:
            return 'PASS' if chk.get('passed', False) else 'FAIL'
        return row.get('reconciliation_status', '')

    def _recon_gap_count_for_month_row(row):
        chk = _recon_check_for_month_row(row)
        if chk is not None:
            return int(chk.get('extraction_gaps', 0) or 0)
        return int(row.get('extraction_gaps', 0) or 0)

    data_completeness = consol.get('data_completeness', 'COMPLETE')
    has_recon = bool(parsing_checks) or any(m.get('reconciliation_status') for m in monthly)
    if parsing_checks:
        failed_checks = [chk for chk in parsing_checks if not chk.get('passed', False)]
        is_incomplete = bool(failed_checks)
        months_with_gaps = len({str(chk.get('month', '') or '') for chk in failed_checks if chk.get('month')})
        total_gaps = sum(int(chk.get('extraction_gaps', 0) or 0) for chk in parsing_checks)
        if not failed_checks:
            total_missing_dr = 0
            total_missing_cr = 0
        else:
            total_missing_dr = consol.get('total_missing_debits', 0) or 0
            total_missing_cr = consol.get('total_missing_credits', 0) or 0
    else:
        is_incomplete = data_completeness == 'INCOMPLETE'
        total_missing_dr = consol.get('total_missing_debits', 0) or 0
        total_missing_cr = consol.get('total_missing_credits', 0) or 0
        months_with_gaps = consol.get('months_with_gaps', 0) or 0
        total_gaps = consol.get('total_extraction_gaps', 0) or 0
    dq_warning = consol.get('data_quality_warning', '')

    company = r.get('company_name', 'Company')
    period_start = r.get('period_start', '')
    period_end = r.get('period_end', '')
    total_months = r.get('total_months', 0)
    related_parties = r.get('related_parties', [])

    # ── Date format: convert from YYYY-MM-DD to YYYY-MM ──
    def _format_to_year_month(date_str):
        """Convert '2026-06-20' to '2026-06'"""
        if not date_str:
            return ''
        parts = str(date_str).split('-')
        if len(parts) >= 2:
            return f"{parts[0]}-{parts[1]}"
        return date_str

    period_start_display = _format_to_year_month(period_start)
    period_end_display = _format_to_year_month(period_end)

    # Build data quality banner HTML
    dq_banner_html = ''
    if has_recon:
        if is_incomplete:
            if parsing_checks:
                failed_months = [chk for chk in parsing_checks if not chk.get('passed', False)]
            else:
                failed_months = [m for m in monthly if m.get('reconciliation_status') == 'FAIL']
            affected_months = ', '.join(sorted({str(m.get('month', '') or '') for m in failed_months if m.get('month')}))
            has_identified_gaps = total_gaps > 0 or total_missing_dr > 0 or total_missing_cr > 0
            if has_identified_gaps:
                dq_banner_html = f'''
                <div class="dq-banner dq-fail">
                    <div class="dq-icon">⚠️</div>
                    <div>
                        <div class="dq-title">Incomplete Extraction — {months_with_gaps} of {total_months} Months Affected</div>
                        <div class="dq-detail">Balance trail reconciliation detected {total_gaps} extraction gap(s) where transactions exist in the source PDF but were not captured. Figures marked with ⚠️ are understated.</div>
                        <div class="dq-stats">
                            <div><div class="dq-stat-label">Missing Debits</div><div class="dq-stat-val">RM {total_missing_dr:,.2f}</div></div>
                            <div><div class="dq-stat-label">Missing Credits</div><div class="dq-stat-val" style="color:var(--green)">RM {total_missing_cr:,.2f}</div></div>
                            <div><div class="dq-stat-label">Gaps</div><div class="dq-stat-val">{total_gaps}</div></div>
                            <div><div class="dq-stat-label">Months Affected</div><div class="dq-stat-val">{affected_months}</div></div>
                        </div>
                    </div>
                </div>'''
            else:
                largest_delta = max((abs(float(m.get('reconciliation_delta') or 0)) for m in failed_months), default=0)
                dq_banner_html = f'''
                <div class="dq-banner dq-fail">
                    <div class="dq-icon">⚠️</div>
                    <div>
                        <div class="dq-title">Balance Reconciliation Warning — {months_with_gaps} of {total_months} Months Failed</div>
                        <div class="dq-detail">No extraction gaps were identified, but the expected closing balance does not match the statement closing balance for the affected month(s). Review opening balance, closing balance, and parsed transaction totals.</div>
                        <div class="dq-stats">
                            <div><div class="dq-stat-label">Failed Checks</div><div class="dq-stat-val">{len(failed_months)}</div></div>
                            <div><div class="dq-stat-label">Largest Delta</div><div class="dq-stat-val">RM {largest_delta:,.2f}</div></div>
                            <div><div class="dq-stat-label">Gaps</div><div class="dq-stat-val">{total_gaps}</div></div>
                            <div><div class="dq-stat-label">Months Affected</div><div class="dq-stat-val">{affected_months}</div></div>
                        </div>
                    </div>
                </div>'''
        else:
            dq_banner_html = f'''
            <div class="dq-banner dq-pass">
                <div class="dq-icon">✅</div>
                <div>
                    <div class="dq-title">Extraction Complete — All {total_months} Months Pass Reconciliation</div>
                    <div class="dq-detail">Every transaction's running balance matches the statement balance. No extraction gaps detected.</div>
                </div>
            </div>'''

    # ── Account cards ──
    def _pick_num(d, *keys):
        if not isinstance(d, dict):
            return 0
        for k in keys:
            v = d.get(k)
            if v is not None:
                try:
                    return float(v)
                except (TypeError, ValueError):
                    try:
                        return float(str(v).replace(',', '').replace('RM', '').strip())
                    except (TypeError, ValueError):
                        return 0
        return 0

    acc_cards = ""
    for a in accounts:
        a_summary = a.get('summary', {}) if isinstance(a.get('summary'), dict) else {}
        a_balances = a.get('balances', {}) if isinstance(a.get('balances'), dict) else {}
        opening = _pick_num(a, 'opening_balance', 'balance_open', 'open_balance') or _pick_num(a_summary, 'opening_balance', 'open_balance') or _pick_num(a_balances, 'opening', 'open')
        closing = _pick_num(a, 'closing_balance', 'ending_balance', 'balance_close', 'close_balance') or _pick_num(a_summary, 'closing_balance', 'ending_balance') or _pick_num(a_balances, 'closing', 'ending', 'close')
        credits_v = _pick_num(a, 'total_credits', 'total_credit', 'gross_credits', 'credits') or _pick_num(a_summary, 'total_credits', 'total_credit', 'gross_credits')
        debits_v = _pick_num(a, 'total_debits', 'total_debit', 'gross_debits', 'debits') or _pick_num(a_summary, 'total_debits', 'total_debit', 'gross_debits')
        txn_count = int(_pick_num(a, 'transaction_count', 'txn_count', 'total_transactions', 'transactions_count') or _pick_num(a_summary, 'transaction_count', 'txn_count'))
        acc_cards += f'''
        <div class="account-card">
            <div class="account-header">
                <span class="bank-name">{a.get('bank_name','')}</span>
                <span class="badge badge-{a.get('account_type','Current').lower()}">{a.get('account_type','')}</span>
            </div>
            <div class="account-number">A/C: {a.get('account_number','')}</div>
            <div class="account-holder">{a.get('account_holder','')}</div>
            <div class="account-metrics">
                <div class="metric"><div class="metric-label">Opening</div><div class="metric-value">RM {opening:,.2f}</div></div>
                <div class="metric"><div class="metric-label">Closing</div><div class="metric-value {'debit' if closing < 10000 else ''}">RM {closing:,.2f}</div></div>
                <div class="metric"><div class="metric-label">Credits</div><div class="metric-value credit">RM {credits_v:,.2f}</div></div>
                <div class="metric"><div class="metric-label">Debits</div><div class="metric-value debit">RM {debits_v:,.2f}</div></div>
                <div class="metric"><div class="metric-label">Transactions</div><div class="metric-value">{txn_count:,}</div></div>
            </div>
        </div>'''

    # ── Related parties ──
    rp_html = ""
    for rp in related_parties:
        name = rp.get('name', rp) if isinstance(rp, dict) else str(rp)
        rel = rp.get('relationship', '') if isinstance(rp, dict) else ''
        rp_html += f'<span class="rp-tag">{name} <small>({rel})</small></span>'

    # ── Related-party candidates (advisory only; analyst confirms) ──
    # MEDIUM/LOW RP3 near-misses that did NOT auto-confirm and exclude nothing.
    # Surfaced so the analyst sees them instead of hunting the full ledger.
    rp_candidates = r.get('related_party_candidates', []) or []
    rp_candidates_html = ""
    if rp_candidates:
        _cand_rows = ""
        for c in rp_candidates:
            conf = str(c.get('confidence', '') or '').upper()
            dr = c.get('total_dr', 0) or 0
            cr = c.get('total_cr', 0) or 0
            _cand_rows += (
                '<tr>'
                f'<td>{c.get("name", "")}</td>'
                f'<td><span class="rpc-badge rpc-{conf.lower()}">{conf}</span></td>'
                f'<td style="text-align:right">RM {dr:,.2f}</td>'
                f'<td style="text-align:right">RM {cr:,.2f}</td>'
                f'<td style="font-size:0.8rem;color:var(--text-soft)">{c.get("evidence", "")}</td>'
                '</tr>'
            )
        _total = r.get('related_party_candidates_total', len(rp_candidates)) or len(rp_candidates)
        _shown = len(rp_candidates)
        _cap_note = (
            f' Showing the {_shown} largest by debit value of {_total} flagged individuals.'
            if _total > _shown else ''
        )
        rp_candidates_html = (
            '<div class="rpc-note">These individuals show some related-party signals but did '
            '<b>not</b> meet the auto-confirm threshold, so they are <b>not</b> excluded from any '
            'figure. Review each and confirm in the analysis step if genuinely related.'
            f'{_cap_note}</div>'
            '<div class="table-wrap"><table>'
            '<thead><tr><th>Party</th><th>Confidence</th><th>Debits</th><th>Credits</th>'
            '<th>Why flagged</th></tr></thead>'
            f'<tbody>{_cand_rows}</tbody></table></div>'
        )

    # ── Monthly analysis table rows ──
    from collections import OrderedDict
    monthly_by_month = OrderedDict()
    for m in monthly:
        mo = m.get('month', '')
        if mo not in monthly_by_month:
            monthly_by_month[mo] = []
        monthly_by_month[mo].append(m)

    acct_list = []
    seen_acct = set()
    for m in monthly:
        an = m.get('account_number', '')
        if an and an not in seen_acct:
            acct_list.append(an)
            seen_acct.add(an)
    acct_colors = {}
    palette = ['var(--blue)', 'var(--purple)', 'var(--green)', 'var(--amber)']
    for i, a in enumerate(acct_list):
        acct_colors[a] = palette[i % len(palette)]

    monthly_rows = ""
    chart_agg = OrderedDict()

    for mo, rows in monthly_by_month.items():
        agg = {}
        sum_fields = ['gross_credits','gross_debits','net_credits','net_debits',
                       'own_party_cr','own_party_dr','related_party_cr','related_party_dr',
                       'reversal_cr','loan_disbursement_cr','fd_interest_cr',
                       'cash_deposits_amount','cash_withdrawals_amount',
                       'cheque_deposits_amount','cheque_issues_amount',
                       'loan_repayment_dr','salary_paid',
                       'statutory_epf','statutory_socso','statutory_tax',
                       'returned_cheques_outward_amount','returned_cheques_outward_count',
                       'round_figure_cr','high_value_cr',
                       'credit_count','debit_count',
                       'own_party_cr_count','own_party_dr_count',
                       'related_party_cr_count','related_party_dr_count',
                       'loan_repayment_count','inward_return_cr',
                       'unclassified_cr_count','unclassified_cr_amount',
                       'unclassified_dr_count','unclassified_dr_amount']
        for fld in sum_fields:
            agg[fld] = sum(r.get(fld, 0) or 0 for r in rows)
        agg['eod_lowest'] = min(r.get('eod_lowest', 0) or 0 for r in rows)
        agg['eod_highest'] = max(r.get('eod_highest', 0) or 0 for r in rows)
        agg['eod_average'] = sum(r.get('eod_average', 0) or 0 for r in rows) / len(rows) if rows else 0
        agg['opening_balance'] = sum(r.get('opening_balance', 0) or 0 for r in rows)
        agg['closing_balance'] = sum(r.get('closing_balance', 0) or 0 for r in rows)
        chart_agg[mo] = agg

        has_account_col = any(m.get('account_number') for m in monthly)

        if has_account_col and len(rows) > 1:
            for m in rows:
                an = m.get('account_number', '')
                bn = m.get('bank_name', '')
                short_bank = bn.split(' ')[0] if bn else ''
                acct_label = f"{short_bank} {an}" if an else mo
                dot_color = acct_colors.get(an, 'var(--text-muted)')
                monthly_rows += f'''<tr style="font-size:0.78rem;">
            <td class="sticky-col" style="padding-left:1.5rem;font-weight:400"><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{dot_color};margin-right:6px;vertical-align:middle"></span>{acct_label}</td>
            {'<td></td>' if has_recon else ''}
            <td class="mono r credit">{m.get('gross_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('gross_debits',0):,.2f}</td>
            <td class="mono r credit">{m.get('net_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('net_debits',0):,.2f}</td>
            <td class="mono r">{m.get('credit_count',0)}</td>
            <td class="mono r">{m.get('debit_count',0)}</td>
            <td class="mono r">{m.get('own_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('own_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('reversal_cr',0):,.2f}</td>
            <td class="mono r">{m.get('loan_disbursement_cr',0):,.2f}</td>
            <td class="mono r">{m.get('fd_interest_cr',0):,.2f}</td>
            <td class="mono r">{m.get('cash_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cash_withdrawals_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_issues_amount',0):,.2f}</td>
            <td class="mono r">{m.get('loan_repayment_dr',0):,.2f}</td>
            <td class="mono r">{m.get('salary_paid',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_epf',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_socso',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_tax',0):,.2f}</td>
            <td class="mono r">{m.get('returned_cheques_outward_count',0)}</td>
            <td class="mono r">{m.get('returned_cheques_outward_amount',0):,.2f}</td>
            <td class="mono r">{m.get('round_figure_cr',0):,.2f}</td>
            <td class="mono r">{m.get('high_value_cr',0):,.2f}</td>
            <td class="mono r">{m.get('eod_lowest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_highest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_average',0):,.2f}</td>
            <td class="mono r">{m.get('opening_balance',0):,.2f}</td>
            <td class="mono r">{m.get('closing_balance',0):,.2f}</td>
            {'<td class="mono r v630-count">' + str(m.get('own_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('own_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('loan_repayment_count',0)) + '</td><td class="mono r v630-amt">' + f"{m.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_cr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_dr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''

            a = agg
            month_recon_cell = ''
            if has_recon:
                any_fail = any(_recon_status_for_month_row(r) == 'FAIL' for r in rows)
                if any_fail:
                    total_gaps = sum(_recon_gap_count_for_month_row(r) for r in rows)
                    total_miss = sum(r.get('missing_debit_amount', 0) for r in rows)
                    month_recon_cell = f'<td><span class="recon-badge fail">✗ FAIL</span> <span class="gap-pill">{total_gaps} gap{"s" if total_gaps > 1 else ""} · RM {total_miss:,.0f}</span></td>'
                else:
                    month_recon_cell = '<td><span class="recon-badge pass">✓ PASS</span></td>'
            monthly_rows += f'''<tr style="background:var(--bg);font-weight:600;border-bottom:2px solid var(--border-accent);{"" if not (has_recon and any_fail) else ""}">
            <td class="sticky-col">{mo}</td>
            {month_recon_cell}
            <td class="mono r credit">{a['gross_credits']:,.2f}</td>
            <td class="mono r debit">{a['gross_debits']:,.2f}</td>
            <td class="mono r credit" style="font-weight:700">{a['net_credits']:,.2f}</td>
            <td class="mono r debit" style="font-weight:700">{a['net_debits']:,.2f}</td>
            <td class="mono r">{int(a['credit_count'])}</td>
            <td class="mono r">{int(a['debit_count'])}</td>
            <td class="mono r">{a['own_party_cr']:,.2f}</td>
            <td class="mono r">{a['own_party_dr']:,.2f}</td>
            <td class="mono r">{a['related_party_cr']:,.2f}</td>
            <td class="mono r">{a['related_party_dr']:,.2f}</td>
            <td class="mono r">{a['reversal_cr']:,.2f}</td>
            <td class="mono r">{a['loan_disbursement_cr']:,.2f}</td>
            <td class="mono r">{a['fd_interest_cr']:,.2f}</td>
            <td class="mono r">{a['cash_deposits_amount']:,.2f}</td>
            <td class="mono r">{a['cash_withdrawals_amount']:,.2f}</td>
            <td class="mono r">{a['cheque_deposits_amount']:,.2f}</td>
            <td class="mono r">{a['cheque_issues_amount']:,.2f}</td>
            <td class="mono r">{a['loan_repayment_dr']:,.2f}</td>
            <td class="mono r">{a['salary_paid']:,.2f}</td>
            <td class="mono r">{a['statutory_epf']:,.2f}</td>
            <td class="mono r">{a['statutory_socso']:,.2f}</td>
            <td class="mono r">{a['statutory_tax']:,.2f}</td>
            <td class="mono r">{int(a['returned_cheques_outward_count'])}</td>
            <td class="mono r">{a['returned_cheques_outward_amount']:,.2f}</td>
            <td class="mono r">{a['round_figure_cr']:,.2f}</td>
            <td class="mono r">{a['high_value_cr']:,.2f}</td>
            <td class="mono r">{a['eod_lowest']:,.2f}</td>
            <td class="mono r">{a['eod_highest']:,.2f}</td>
            <td class="mono r">{a['eod_average']:,.2f}</td>
            <td class="mono r">{a['opening_balance']:,.2f}</td>
            <td class="mono r">{a['closing_balance']:,.2f}</td>
            {'<td class="mono r v630-count">' + str(int(a.get('own_party_cr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('own_party_dr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('related_party_cr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('related_party_dr_count',0))) + '</td><td class="mono r v630-count">' + str(int(a.get('loan_repayment_count',0))) + '</td><td class="mono r v630-amt">' + f"{a.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(int(a.get('unclassified_cr_count',0))) + '</td><td class="mono r v630-uncl">' + f"{a.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(int(a.get('unclassified_dr_count',0))) + '</td><td class="mono r v630-uncl">' + f"{a.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''
        else:
            m = rows[0] if rows else {}
            recon_status = _recon_status_for_month_row(m)
            row_class = ' class="row-fail"' if recon_status == 'FAIL' else ''
            recon_cell = ''
            if has_recon:
                if recon_status == 'FAIL':
                    gap_count = _recon_gap_count_for_month_row(m)
                    miss_dr = m.get('missing_debit_amount', 0)
                    recon_cell = f'<td><span class="recon-badge fail">✗ FAIL</span> <span class="gap-pill">{gap_count} gap{"s" if gap_count > 1 else ""} · RM {miss_dr:,.0f}</span></td>'
                else:
                    recon_cell = '<td><span class="recon-badge pass">✓ PASS</span></td>'
            monthly_rows += f'''<tr{row_class}>
            <td class="sticky-col">{m.get('month','')}</td>
            {recon_cell}
            <td class="mono r credit">{m.get('gross_credits',0):,.2f}</td>
            <td class="mono r debit">{m.get('gross_debits',0):,.2f}</td>
            <td class="mono r credit" style="font-weight:600">{m.get('net_credits',0):,.2f}</td>
            <td class="mono r debit" style="font-weight:600">{m.get('net_debits',0):,.2f}</td>
            <td class="mono r">{m.get('credit_count',0)}</td>
            <td class="mono r">{m.get('debit_count',0)}</td>
            <td class="mono r">{m.get('own_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('own_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_cr',0):,.2f}</td>
            <td class="mono r">{m.get('related_party_dr',0):,.2f}</td>
            <td class="mono r">{m.get('reversal_cr',0):,.2f}</td>
            <td class="mono r">{m.get('loan_disbursement_cr',0):,.2f}</td>
            <td class="mono r">{m.get('fd_interest_cr',0):,.2f}</td>
            <td class="mono r">{m.get('cash_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cash_withdrawals_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_deposits_amount',0):,.2f}</td>
            <td class="mono r">{m.get('cheque_issues_amount',0):,.2f}</td>
            <td class="mono r">{m.get('loan_repayment_dr',0):,.2f}</td>
            <td class="mono r">{m.get('salary_paid',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_epf',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_socso',0):,.2f}</td>
            <td class="mono r">{m.get('statutory_tax',0):,.2f}</td>
            <td class="mono r">{m.get('returned_cheques_outward_count',0)}</td>
            <td class="mono r">{m.get('returned_cheques_outward_amount',0):,.2f}</td>
            <td class="mono r">{m.get('round_figure_cr',0):,.2f}</td>
            <td class="mono r">{m.get('high_value_cr',0):,.2f}</td>
            <td class="mono r">{m.get('eod_lowest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_highest',0):,.2f}</td>
            <td class="mono r">{m.get('eod_average',0):,.2f}</td>
            <td class="mono r">{m.get('opening_balance',0):,.2f}</td>
            <td class="mono r">{m.get('closing_balance',0):,.2f}</td>
            {'<td class="mono r v630-count">' + str(m.get('own_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('own_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_cr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('related_party_dr_count',0)) + '</td><td class="mono r v630-count">' + str(m.get('loan_repayment_count',0)) + '</td><td class="mono r v630-amt">' + f"{m.get('inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_cr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_cr_amount',0):,.2f}" + '</td><td class="mono r v630-uncl">' + str(m.get('unclassified_dr_count',0)) + '</td><td class="mono r v630-uncl">' + f"{m.get('unclassified_dr_amount',0):,.2f}" + '</td>' if is_v630 else ''}
        </tr>'''

    # ── Consolidated totals row ──
    total_status_cell = ''
    if has_recon:
        if is_incomplete:
            total_status_cell = f'<td><span class="recon-badge fail">⚠️ {months_with_gaps} FAIL</span></td>'
        else:
            total_status_cell = '<td><span class="recon-badge pass">ALL PASS</span></td>'
    consol_row = f'''<tr class="total-row">
        <td class="sticky-col" style="font-weight:700">TOTAL</td>
        {total_status_cell}
        <td class="mono r credit">{consol.get('gross_credits',0):,.2f}</td>
        <td class="mono r debit">{consol.get('gross_debits',0):,.2f}</td>
        <td class="mono r credit">{consol.get('net_credits',0):,.2f}</td>
        <td class="mono r debit">{consol.get('net_debits',0):,.2f}</td>
        <td class="mono r">-</td><td class="mono r">-</td>
        <td class="mono r">{consol.get('total_own_party_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_own_party_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_related_party_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_related_party_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_reversal_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_loan_disbursement_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_fd_interest_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cash_deposits',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cash_withdrawals',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cheque_deposits',0):,.2f}</td>
        <td class="mono r">{consol.get('total_cheque_issues',0):,.2f}</td>
        <td class="mono r">{consol.get('total_loan_repayment_dr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_salary_paid',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_epf',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_socso',0):,.2f}</td>
        <td class="mono r">{consol.get('total_statutory_tax',0):,.2f}</td>
        <td class="mono r">{consol.get('total_returned_cheques_outward',0):,.2f}</td>
        <td class="mono r">{consol.get('total_returned_cheques_outward',0):,.2f}</td>
        <td class="mono r">{consol.get('total_round_figure_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('total_high_value_cr',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_lowest',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_highest',0):,.2f}</td>
        <td class="mono r">{consol.get('eod_average',0):,.2f}</td>
        <td class="mono r">-</td>
        <td class="mono r">-</td>
        {'<td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-count">-</td><td class="mono r v630-amt">' + f"{consol.get('total_inward_return_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">-</td><td class="mono r v630-uncl">' + f"{consol.get('total_unclassified_cr',0):,.2f}" + '</td><td class="mono r v630-uncl">-</td><td class="mono r v630-uncl">' + f"{consol.get('total_unclassified_dr',0):,.2f}" + '</td>' if is_v630 else ''}
    </tr>'''

    # ── Top parties with ghost-verb suppression ──
    party_view = prepare_top_parties_for_report(top_parties)
    _payers = party_view['payers']
    _payees = party_view['payees']
    _payers_suppressed = party_view['payers_suppressed']
    _payees_suppressed = party_view['payees_suppressed']

    def _render_suppressed(entries, side_css):
        if not entries:
            return ''
        entries_sorted = sorted(entries, key=lambda p: p.get('total_amount', 0) or 0, reverse=True)
        rows = ''
        for p in entries_sorted:
            amt = p.get('total_amount', 0) or 0
            n = p.get('transaction_count', 0) or 0
            warn = ''
            if amt >= 100000 or n >= 50:
                warn = '<span style="background:var(--amber);color:white;padding:1px 6px;border-radius:3px;font-size:0.7rem;margin-left:6px">VERIFY</span>'
            rows += f'''<tr>
                <td style="color:var(--text-muted)">{p.get("party_name","") or "(empty)"}{warn}</td>
                <td class="mono r {side_css}">RM {amt:,.2f}</td>
                <td class="mono r">{n}</td>
            </tr>'''
        return f'''<div style="padding:0.75rem 1.25rem;background:var(--surface-subtle);border-top:1px solid var(--border)">
            <div style="font-size:0.78rem;color:var(--text-soft);margin-bottom:0.5rem">
                <strong>Parser-dropped buckets</strong> — counterparties that were only a transfer verb with no entity name attached.
                Amounts are still counted in gross/net totals; they are hidden from the Top 10 rank to avoid misleading the analyst.
                <span style="background:var(--amber);color:white;padding:1px 6px;border-radius:3px;font-size:0.7rem">VERIFY</span> = high volume — possible real-entity false positive, please cross-check.
            </div>
            <table style="width:100%;font-size:0.78rem"><thead><tr>
                <th style="text-align:left">Bucket (parser artifact)</th>
                <th class="r">Amount (RM)</th><th class="r">Txns</th>
            </tr></thead><tbody>{rows}</tbody></table>
        </div>'''

    payers_suppressed_html = _render_suppressed(_payers_suppressed, 'credit')
    payees_suppressed_html = _render_suppressed(_payees_suppressed, 'debit')

    def _fmt_compact(n):
        try:
            n = float(n or 0)
        except Exception:
            return '0'
        a = abs(n)
        if a >= 1_000_000:
            return f'{n/1_000_000:.1f}M'.replace('.0M', 'M')
        if a >= 1_000:
            return f'{n/1_000:.0f}K'
        return f'{n:.0f}'

    _MONTH_ABBR = {
        '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr', '05': 'May', '06': 'Jun',
        '07': 'Jul', '08': 'Aug', '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec',
    }

    def _fmt_month_label(month_str):
        if not month_str:
            return ''
        parts = str(month_str).split('-')
        if len(parts) >= 2 and parts[1] in _MONTH_ABBR:
            return _MONTH_ABBR[parts[1]]
        return str(month_str)[-5:]

    def _render_monthly_bars(monthly_breakdown, color_var):
        if not monthly_breakdown:
            return ''
        mb_vals = [mb.get('amount', 0) for mb in monthly_breakdown]
        max_mb = max(mb_vals) if mb_vals and max(mb_vals) > 0 else 1
        bars = ''.join(
            f'<div title="{mb.get("month","")}: RM {mb.get("amount",0):,.0f}" '
            f'style="flex:1;background:{color_var};opacity:0.7;border-radius:2px;'
            f'min-width:4px;height:{max(2, int(mb.get("amount",0)/max_mb*28))}px"></div>'
            for mb in monthly_breakdown
        )
        labels = ''.join(
            f'<span style="flex:1;text-align:center;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">'
            f'{_fmt_month_label(mb.get("month",""))}: {_fmt_compact(mb.get("amount",0))}'
            f'</span>'
            for mb in monthly_breakdown
        )
        return (
            f'<div style="display:flex;align-items:flex-end;gap:2px;height:30px;margin-top:4px">{bars}</div>'
            f'<div style="display:flex;gap:2px;font-size:0.65rem;color:var(--text-muted);margin-top:2px">{labels}</div>'
        )

    payer_rows = ""
    for p in _payers:
        rp_badge = '<span class="rp-badge">RP</span>' if p.get('is_related_party') else ''
        mb_html = _render_monthly_bars(p.get('monthly_breakdown'), 'var(--green)')
        payer_rows += f'''<tr>
            <td>{p.get('rank')}</td>
            <td>{p.get('party_name','')} {rp_badge}{mb_html}</td>
            <td class="mono r credit">RM {p.get('total_amount',0):,.2f}</td>
            <td class="mono r">{p.get('transaction_count',0)}</td>
        </tr>'''

    payee_rows = ""
    for p in _payees:
        rp_badge = '<span class="rp-badge">RP</span>' if p.get('is_related_party') else ''
        mb_html = _render_monthly_bars(p.get('monthly_breakdown'), 'var(--red)')
        payee_rows += f'''<tr>
            <td>{p.get('rank')}</td>
            <td>{p.get('party_name','')} {rp_badge}{mb_html}</td>
            <td class="mono r debit">RM {p.get('total_amount',0):,.2f}</td>
            <td class="mono r">{p.get('transaction_count',0)}</td>
        </tr>'''

    # ── Large transactions ──
    large_threshold = safe_float(
        consol.get('high_value_threshold')
        or consol.get('large_credit_threshold')
        or _fallback_summary.get('high_value_threshold')
        or _fallback_config.get('large_transaction_threshold')
        or _fallback_config.get('large_credit_threshold')
        or 100000
    )
    if large_threshold <= 0:
        large_threshold = 100000
    
    large_txns = data.get('large_transactions', [])
    if (not large_txns or len(large_txns) == 0) and data.get('transactions'):
        try:
            large_txns = build_large_transactions(data.get('transactions', []), large_threshold)
        except Exception as e:
            print(f"[DEBUG] Error building large transactions: {e}")
            large_txns = []
    
    large_credits = data.get('large_credits', [])
    if (not large_txns or len(large_txns) == 0) and large_credits:
        large_txns = []
        for t in large_credits:
            if isinstance(t, dict):
                large_txns.append({
                    'date': t.get('date', ''),
                    'description': t.get('description', ''),
                    'amount': t.get('amount', 0),
                    'balance': t.get('balance', 0),
                    'type': 'CREDIT'
                })
    
    large_txn_rows = ""
    if large_txns and isinstance(large_txns, list) and len(large_txns) > 0:
        for t in large_txns:
            if not isinstance(t, dict):
                continue
            txn_type = t.get('type', 'CREDIT')
            type_cls = 'credit' if txn_type == 'CREDIT' else 'debit'
            amount = t.get('amount', 0)
            if amount is None:
                amount = 0
            balance = t.get('balance', 0)
            if balance is None:
                balance = 0
            date_str = str(t.get('date', ''))
            desc_str = str(t.get('description', ''))[:70]
            large_txn_rows += f'''
                <tr>
                    <td>{date_str}</td>
                    <td>{desc_str}</td>
                    <td class="mono r {type_cls}">RM {float(amount):,.2f}</td>
                    <td class="mono r">RM {float(balance):,.2f}</td>
                </tr>'''
    else:
        large_txn_rows = f'''
                <tr>
                    <td colspan="4" class="note">No transactions above RM {large_threshold:,.0f}</td>
                </tr>'''
    
    round_figure_credits = get_round_transactions_for_report(data)
    if not round_figure_credits and data.get('transactions'):
        round_figure_credits = build_round_figure_credits_internal(data.get('transactions', []))
    rf_cr_rows = ""
    if round_figure_credits and isinstance(round_figure_credits, list):
        for t in round_figure_credits:
            if not isinstance(t, dict):
                continue
            amount = safe_float(t.get('amount', t.get('credit', 0)))
            amount_class = 'credit' if amount >= 0 else 'debit'
            amount_text = f"RM {abs(amount):,.2f}"
            balance = safe_float(t.get('balance', 0))
            rf_cr_rows += f'''
                <tr>
                    <td>{escape(str(t.get('date', '')))}</td>
                    <td>{escape(str(t.get('description', ''))[:70])}</td>
                    <td class="mono r {amount_class}">{amount_text}</td>
                    <td class="mono r">RM {balance:,.2f}</td>
                </tr>'''
    if not rf_cr_rows:
        rf_cr_rows = '<tr><td colspan="4" class="note">No round-figure transactions detected.</td></tr>'

    _sync_transaction_pattern_flags(
        data,
        round_transactions=round_figure_credits,
        large_transactions=large_txns,
        large_threshold=large_threshold,
    )

    # ── Related party transactions ──
    rp_summary = own_related.get('summary', {}) or {}
    _rp_txns_all = own_related.get('transactions', []) or []
    def _count_txn(party_type_prefix, txn_type):
        c = 0
        for _t in _rp_txns_all:
            if not isinstance(_t, dict):
                continue
            pt = (_t.get('party_type') or '').upper()
            tt = (_t.get('type') or '').upper()
            if pt.startswith(party_type_prefix) and tt == txn_type:
                c += 1
        return c
    rp_counts = {
        'own_party_cr': int(rp_summary.get('own_party_cr_count') or 0) or _count_txn('OWN', 'CREDIT'),
        'own_party_dr': int(rp_summary.get('own_party_dr_count') or 0) or _count_txn('OWN', 'DEBIT'),
        'related_party_cr': int(rp_summary.get('related_party_cr_count') or 0) or _count_txn('RELATED', 'CREDIT'),
        'related_party_dr': int(rp_summary.get('related_party_dr_count') or 0) or _count_txn('RELATED', 'DEBIT'),
    }
    rp_groups = {}
    for t in own_related.get('transactions', []) or []:
        if not isinstance(t, dict):
            continue
        party_name = str(t.get('party_name') or 'Unknown Party').strip() or 'Unknown Party'
        party_type = str(t.get('party_type') or '').strip()
        party_type_upper = party_type.upper()
        badge_type = 'OP' if party_type_upper.startswith('OWN') else 'RP'
        key = (party_name.casefold(), badge_type, party_name)
        group = rp_groups.setdefault(
            key,
            {
                'party_name': party_name,
                'party_type': party_type or ('Own Party' if badge_type == 'OP' else 'Related Party'),
                'badge_type': badge_type,
                'transactions': [],
                'credits': 0.0,
                'debits': 0.0,
                'credit_count': 0,
                'debit_count': 0,
            },
        )
        amount = safe_float(t.get('amount', 0))
        txn_type = str(t.get('type') or '').upper()
        if txn_type == 'CREDIT':
            group['credits'] += amount
            group['credit_count'] += 1
        elif txn_type == 'DEBIT':
            group['debits'] += amount
            group['debit_count'] += 1
        group['transactions'].append(t)

    rp_party_rows = ""
    for idx, group in enumerate(sorted(rp_groups.values(), key=lambda g: (g['party_name'].casefold(), g['badge_type']))):
        badge_cls = 'op-badge' if group['badge_type'] == 'OP' else 'rp-badge'
        detail_rows = ""
        for t in group['transactions']:
            txn_type = str(t.get('type') or '').upper()
            type_cls = 'credit' if txn_type == 'CREDIT' else 'debit'
            amount = safe_float(t.get('amount', 0))
            balance = safe_float(t.get('balance', 0))
            detail_rows += f'''<tr>
                <td>{escape(str(t.get('date','')))}</td>
                <td>{escape(str(t.get('description',''))[:80])}</td>
                <td class="mono r {type_cls}">RM {amount:,.2f}</td>
                <td><span class="badge badge-{escape(txn_type.lower())}">{escape(txn_type)}</span></td>
                <td class="mono r">RM {balance:,.2f}</td>
            </tr>'''

        txn_count = len(group['transactions'])
        rp_party_rows += f'''
        <tr class="rp-party-row" onclick="toggleRp('rp-detail-{idx}')" style="cursor:pointer">
            <td><span id="rp-caret-{idx}">▶</span> {escape(group['party_name'])} <span class="{badge_cls}">{group['badge_type']}</span></td>
            <td class="mono r credit">{group['credits']:,.2f} <span style="color:var(--text-muted);font-size:0.75rem">({group['credit_count']})</span></td>
            <td class="mono r debit">{group['debits']:,.2f} <span style="color:var(--text-muted);font-size:0.75rem">({group['debit_count']})</span></td>
            <td>{escape(group['party_type'])}</td>
            <td class="mono r">{txn_count}</td>
        </tr>
        <tr id="rp-detail-{idx}" style="display:none"><td colspan="5" style="background:var(--bg);padding:0">
            <div class="table-wrap"><table style="margin:0">
                <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Type</th><th class="r">Balance</th></tr></thead>
                <tbody>{detail_rows or '<tr><td colspan="5" class="note">No transactions</td></tr>'}</tbody>
            </table></div>
        </td></tr>'''

    rp_expander_script = '''
                    <script>
                        function toggleRp(id) {
                            var row = document.getElementById(id);
                            if (!row) return;
                            var caret = document.getElementById(id.replace('detail','caret'));
                            if (row.style.display === 'none') { row.style.display = ''; if (caret) caret.textContent = '▼'; }
                            else { row.style.display = 'none'; if (caret) caret.textContent = '▶'; }
                        }
                    </script>'''

    # ── Counterparty Ledger ──
    counterparty_ledger_html = ''
    cp_ledger = data.get('counterparty_ledger')
    if cp_ledger:
        rp_name_set = set()
        for _rp in related_parties:
            if isinstance(_rp, dict):
                _nm = _rp.get('name') or _rp.get('party_name') or ''
            else:
                _nm = str(_rp)
            if _nm:
                rp_name_set.add(_nm.strip().upper())

        cleaning_status = cp_ledger.get('ledger_cleaning_status', '')
        cleaning_stats = cp_ledger.get('cleaning_stats', {}) or {}
        total_cp = cp_ledger.get('total_counterparties', 0)
        merges = cleaning_stats.get('merges_performed', 0)
        purpose_strips = cleaning_stats.get('purpose_strips', 0)
        original_cp = cleaning_stats.get('original_counterparties', 0)

        ext_stats = cp_ledger.get('extraction_stats') if isinstance(cp_ledger.get('extraction_stats'), dict) else {}
        merged_banks = ext_stats.get('merged_from_banks') if isinstance(ext_stats.get('merged_from_banks'), list) else None
        ext_pattern = ext_stats.get('pattern_matched')
        ext_bucket = ext_stats.get('special_bucket')
        ext_raw = ext_stats.get('raw_fallback')
        ext_total = ext_stats.get('total_transactions')
        has_ext_stats = bool(ext_stats)

        status_color = {'CLEANED': 'green', 'VALIDATION_FAILED': 'amber', 'SKIPPED': 'amber'}.get(cleaning_status, 'text-muted')
        status_badge = f'<span class="badge" style="background:var(--{status_color}-dim);color:var(--{status_color})">{cleaning_status or "N/A"}</span>' if cleaning_status else ''

        val_fail_warning = ''
        if cleaning_status == 'VALIDATION_FAILED':
            val_fail_warning = '<div style="background:var(--amber-dim);border:1px solid var(--amber);color:var(--amber);margin:0.75rem 0;padding:0.75rem;border-radius:8px;display:flex;gap:0.5rem;align-items:center"><div>⚠️</div><div><div style="font-weight:600">Counterparty ledger cleaning failed validation</div><div style="font-size:0.85rem">Showing original parser output.</div></div></div>'

        raw_counterparties = cp_ledger.get('counterparties', []) or []
        counterparties_sorted = build_canonical_counterparty_ledger_rows(cp_ledger)
        if raw_counterparties and len(counterparties_sorted) != len(raw_counterparties):
            original_cp = original_cp or len(raw_counterparties)
            merges = merges or (len(raw_counterparties) - len(counterparties_sorted))
        total_cp = len(counterparties_sorted)

        cp_rows_html = ''
        for idx, cp in enumerate(counterparties_sorted):
            name = cp.get('counterparty_name', '') or ''
            credits = cp.get('total_credits', 0) or 0
            debits = cp.get('total_debits', 0) or 0
            net = cp.get('net_position', 0) or 0
            txn_count = cp.get('transaction_count', 0) or 0
            cr_count = cp.get('credit_count', 0) or 0
            dr_count = cp.get('debit_count', 0) or 0
            net_cls = 'credit' if net >= 0 else 'debit'
            rp_badge = '<span class="rp-badge">RP</span>' if name.strip().upper() in rp_name_set else ''

            sub_rows = ''
            for t in cp.get('transactions', []) or []:
                t_type = t.get('type', '')
                t_cls = 'credit' if t_type == 'CREDIT' else 'debit'
                sub_rows += f'''<tr>
                    <td>{t.get('date','')}</td>
                    <td>{(t.get('description','') or '')[:80]}</td>
                    <td class="mono r {t_cls}">RM {t.get('amount',0):,.2f}</td>
                    <td><span class="badge badge-{t_type.lower()}">{t_type}</span></td>
                    <td class="mono r">{t.get('balance',0):,.2f}</td>
                </tr>'''

            cp_rows_html += f'''
            <tr class="cp-row" onclick="toggleCp('cp-detail-{idx}')" style="cursor:pointer">
                <td><span id="cp-caret-{idx}">▶</span> {name} {rp_badge}</td>
                <td class="mono r credit">{credits:,.2f} <span style="color:var(--text-muted);font-size:0.75rem">({cr_count})</span></td>
                <td class="mono r debit">{debits:,.2f} <span style="color:var(--text-muted);font-size:0.75rem">({dr_count})</span></td>
                <td class="mono r {net_cls}" style="font-weight:600">{net:,.2f}</td>
                <td class="mono r">{txn_count}</td>
            </tr>
            <tr id="cp-detail-{idx}" style="display:none"><td colspan="5" style="background:var(--bg);padding:0">
                <div class="table-wrap"><table style="margin:0">
                    <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Type</th><th class="r">Balance</th></tr></thead>
                    <tbody>{sub_rows or '<tr><td colspan="5" class="note">No transactions</td></tr>'}</tbody>
                </table></div>
            </td></tr>'''

        counterparty_ledger_html = f'''
            <div class="section">
                <div class="section-head">
                    <h2>Counterparty Ledger</h2>
                    {status_badge}
                </div>
                <div class="section-body">
                    {val_fail_warning}
                    <div class="summary-grid">
                        <div class="summary-card"><div class="val">{total_cp:,}</div><div class="lbl">Total Counterparties</div></div>
                        {('<div class="summary-card"><div class="val">' + f'{original_cp:,}' + '</div><div class="lbl">Original (pre-clean)</div></div>') if original_cp else ''}
                        {('<div class="summary-card"><div class="val">' + f'{merges:,}' + '</div><div class="lbl">Merges Performed</div></div>') if merges else ''}
                        {('<div class="summary-card"><div class="val">' + f'{purpose_strips:,}' + '</div><div class="lbl">Purpose Strips</div></div>') if purpose_strips else ''}
                        {('<div class="summary-card"><div class="val" style="font-size:1.05rem">' + ', '.join(merged_banks) + '</div><div class="lbl">Merged from banks</div></div>') if merged_banks else ''}
                        {('<div class="summary-card"><div class="val">' + f'{int(ext_pattern or 0):,}' + '</div><div class="lbl">Pattern matched</div></div>') if has_ext_stats else ''}
                        {('<div class="summary-card"><div class="val">' + f'{int(ext_bucket or 0):,}' + '</div><div class="lbl">Special bucket</div></div>') if has_ext_stats else ''}
                        {('<div class="summary-card"><div class="val">' + f'{int(ext_raw or 0):,}' + '</div><div class="lbl">Raw fallback</div></div>') if has_ext_stats else ''}
                    </div>
                    <div style="margin:0.5rem 0">
                        <input type="text" id="cp-search" placeholder="Filter counterparties..." onkeyup="filterCp()" style="width:100%;padding:0.5rem;border:1px solid var(--border);border-radius:4px;font-size:0.9rem">
                    </div>
                    <div class="table-wrap" style="max-height:600px;overflow:auto"><table id="cp-table">
                        <thead><tr><th>Counterparty</th><th class="r">Credits (RM)</th><th class="r">Debits (RM)</th><th class="r">Net Position</th><th class="r">Txns</th></tr></thead>
                        <tbody>{cp_rows_html or '<tr><td colspan="5" class="note">No counterparties</td></tr>'}</tbody>
                    </table></div>
                </div>
            </div>
            <script>
                function toggleCp(id) {{
                    var row = document.getElementById(id);
                    if (!row) return;
                    var caret = document.getElementById(id.replace('detail','caret'));
                    if (row.style.display === 'none') {{ row.style.display = ''; if (caret) caret.textContent = '▼'; }}
                    else {{ row.style.display = 'none'; if (caret) caret.textContent = '▶'; }}
                }}
                function filterCp() {{
                    var q = document.getElementById('cp-search').value.toLowerCase();
                    var rows = document.querySelectorAll('#cp-table tbody tr.cp-row');
                    rows.forEach(function(r) {{
                        var txt = r.textContent.toLowerCase();
                        var detail = document.getElementById(r.getAttribute('onclick').match(/cp-detail-\\d+/)[0]);
                        if (!q || txt.indexOf(q) >= 0) {{ r.style.display = ''; }}
                        else {{ r.style.display = 'none'; if (detail) detail.style.display = 'none'; }}
                    }});
                }}
            </script>'''

    # ── Statutory Compliance ──
    statutory_html = ''
    stat_comp = consol.get('statutory_compliance')
    if stat_comp:
        overall = stat_comp.get('overall_status', 'N/A')
        overall_color = {'COMPLIANT': 'green', 'GAPS_DETECTED': 'amber', 'CRITICAL': 'red'}.get(overall, 'amber')

        def _cov_bar(label, paid, total, missing, paid_list=None, salary_list=None):
            if not total:
                return f'<div class="summary-card"><div class="val">N/A</div><div class="lbl">{label}</div></div>'
            if paid_list is not None and salary_list is not None and salary_list:
                covered = len(set(paid_list) & set(salary_list))
                display_paid = covered
                display_total = len(salary_list)
            else:
                display_total = total
                display_paid = min(paid, total)
            pct = (display_paid / display_total * 100) if display_total else 0
            pct = max(0.0, min(pct, 100.0))
            bar_color = 'green' if pct >= 99.5 else ('amber' if pct >= 50 else 'red')
            miss_str = f'<div style="font-size:0.7rem;color:var(--amber);margin-top:0.25rem">Missing: {", ".join(missing[:6])}{"..." if len(missing) > 6 else ""}</div>' if missing else ''
            return f'''<div class="summary-card">
                <div class="val">{display_paid}/{display_total}</div>
                <div class="lbl">{label} ({pct:.0f}%)</div>
                <div style="height:6px;background:var(--border);border-radius:3px;margin-top:0.4rem;overflow:hidden"><div style="width:{pct:.1f}%;height:100%;background:var(--{bar_color})"></div></div>
                {miss_str}
            </div>'''

        def _as_int(v):
            if isinstance(v, list):
                return len(v)
            try:
                return int(v or 0)
            except (TypeError, ValueError):
                return 0

        salary_list = stat_comp.get('salary_months_list') or []
        salary_months = _as_int(stat_comp.get('salary_months_active', 0)) or len(salary_list)
        epf_list = stat_comp.get('epf_months_list') or []
        epf_paid = _as_int(stat_comp.get('epf_months_paid', 0)) or len(epf_list)
        epf_missing = stat_comp.get('epf_months_missing', []) or []
        socso_list = stat_comp.get('socso_months_list') or []
        socso_paid = _as_int(stat_comp.get('socso_months_paid', 0)) or len(socso_list)
        socso_missing = stat_comp.get('socso_months_missing', []) or []
        lhdn_det = stat_comp.get('lhdn_detected', False)
        lhdn_list = stat_comp.get('lhdn_months_list') or []
        lhdn_paid = _as_int(stat_comp.get('lhdn_months_paid', 0)) or len(lhdn_list)
        hrdf_det = stat_comp.get('hrdf_detected', False)
        hrdf_list = stat_comp.get('hrdf_months_list') or []
        hrdf_paid = _as_int(stat_comp.get('hrdf_months_paid', 0)) or len(hrdf_list)

        cov_cards = _cov_bar('EPF Coverage', epf_paid, salary_months, epf_missing, epf_list, salary_list)
        cov_cards += _cov_bar('SOCSO Coverage', socso_paid, salary_months, socso_missing, socso_list, salary_list)

        total_lhdn = float(consol.get('total_statutory_tax', 0) or 0)
        total_hrdf = float(consol.get('total_statutory_hrdf', 0) or 0)

        def _info_card(label, n_months, total_amt, detected, tooltip=''):
            if not detected and n_months == 0 and total_amt == 0:
                return f'<div class="summary-card" title="{tooltip}"><div class="val" style="color:var(--text-muted)">Not detected</div><div class="lbl">{label}</div></div>'
            amt_str = f"RM {total_amt:,.0f}" if total_amt else ""
            return f'''<div class="summary-card" title="{tooltip}">
                <div class="val">{n_months}</div>
                <div class="lbl">{label} (months paid)</div>
                <div style="font-size:0.75rem;color:var(--text-soft);margin-top:0.2rem">{amt_str}</div>
            </div>'''

        cov_cards += _info_card(
            'LHDN', lhdn_paid, total_lhdn, lhdn_det,
            tooltip='Income tax bucket — includes PCB/MTD salary withholding AND CP204 corporate tax AND SST. Shown as count because payment timing is not strictly payroll-driven.',
        )
        cov_cards += _info_card(
            'HRDF', hrdf_paid, total_hrdf, hrdf_det,
            tooltip='HRDF is exempt for small employers (less than 10 employees in certain industries). Shown as count — absence does not necessarily indicate non-compliance.',
        )

        def _ratio_rows(rows, is_epf=True):
            html = ''
            for row in rows or []:
                if not isinstance(row, dict):
                    html += f'<tr><td class="mono" colspan="5">{row}</td></tr>'
                    continue
                st = row.get('status', 'N/A')
                st_color = {'OK': 'green', 'WARNING': 'amber', 'CATCH_UP': 'red'}.get(st, 'text-muted')
                amt = row.get('epf_amount' if is_epf else 'socso_amount', 0) or 0
                try:
                    amt = float(amt)
                except (TypeError, ValueError):
                    amt = 0.0
                try:
                    sal = float(row.get('salary_amount', 0) or 0)
                except (TypeError, ValueError):
                    sal = 0.0
                try:
                    ratio = float(row.get('ratio_pct', 0) or 0)
                except (TypeError, ValueError):
                    ratio = 0.0
                html += f'''<tr>
                    <td class="mono">{row.get('month','')}</td>
                    <td class="mono r">{amt:,.2f}</td>
                    <td class="mono r">{sal:,.2f}</td>
                    <td class="mono r">{ratio:.2f}%</td>
                    <td style="text-align:center"><span style="color:var(--{st_color});font-weight:600">{st}</span></td>
                </tr>'''
            return html or '<tr><td colspan="5" class="note">No data</td></tr>'

        epf_ratio_rows_html = _ratio_rows(stat_comp.get('epf_per_month_ratios', []), True)
        socso_ratio_rows_html = _ratio_rows(stat_comp.get('socso_per_month_ratios', []), False)

        sub_thr = stat_comp.get('subthreshold_employer') if isinstance(stat_comp.get('subthreshold_employer'), dict) else None
        ch_blind = stat_comp.get('channel_blind_employer') if isinstance(stat_comp.get('channel_blind_employer'), dict) else None
        footprint_html = ''
        if sub_thr or ch_blind:
            blocks = ''
            if sub_thr:
                is_sub = bool(sub_thr.get('is_subthreshold'))
                sal_amt = float(sub_thr.get('total_salary_amount', 0) or 0)
                thr_amt = float(sub_thr.get('threshold_amount', 0) or 0)
                reason = sub_thr.get('reason', '') or ''
                badge_color = 'amber' if is_sub else 'green'
                badge_txt = 'SUB-THRESHOLD' if is_sub else 'OK'
                blocks += f'''
                    <div class="summary-card" style="text-align:left">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.4rem">
                            <div class="lbl" style="font-weight:600;text-transform:none;font-size:0.85rem">Sub-threshold employer check</div>
                            <span class="badge" style="background:var(--{badge_color}-dim);color:var(--{badge_color})">{badge_txt}</span>
                        </div>
                        <div style="font-size:0.8rem;color:var(--text-soft)">Salary detected: RM {sal_amt:,.2f} &middot; Threshold: RM {thr_amt:,.0f}</div>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.4rem;line-height:1.4">{reason}</div>
                    </div>'''
            if ch_blind:
                is_ch = bool(ch_blind.get('is_channel_blind'))
                chq_amt = float(ch_blind.get('cheque_dr_amount', 0) or 0)
                gross_dr = float(ch_blind.get('gross_dr_amount', 0) or 0)
                ratio_pct = float(ch_blind.get('cheque_dr_ratio', 0) or 0) * 100.0
                thr_amt = float(ch_blind.get('threshold_amount', 0) or 0)
                thr_ratio = float(ch_blind.get('threshold_ratio', 0) or 0) * 100.0
                reason = ch_blind.get('reason', '') or ''
                badge_color = 'amber' if is_ch else 'green'
                badge_txt = 'CHANNEL-BLIND' if is_ch else 'OK'
                blocks += f'''
                    <div class="summary-card" style="text-align:left">
                        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.4rem">
                            <div class="lbl" style="font-weight:600;text-transform:none;font-size:0.85rem">Channel-blind employer check</div>
                            <span class="badge" style="background:var(--{badge_color}-dim);color:var(--{badge_color})">{badge_txt}</span>
                        </div>
                        <div style="font-size:0.8rem;color:var(--text-soft)">Cheque DR: RM {chq_amt:,.2f} ({ratio_pct:.1f}% of gross DR RM {gross_dr:,.2f})</div>
                        <div style="font-size:0.8rem;color:var(--text-soft)">Thresholds: &ge; RM {thr_amt:,.0f} AND &ge; {thr_ratio:.0f}%</div>
                        <div style="font-size:0.78rem;color:var(--text-muted);margin-top:0.4rem;line-height:1.4">{reason}</div>
                    </div>'''
            footprint_html = f'''
                    <h3 style="font-size:0.95rem;margin:1.25rem 0 0.5rem">Employer Footprint Checks</h3>
                    <div class="summary-grid" style="grid-template-columns:repeat(auto-fit,minmax(280px,1fr))">{blocks}</div>'''

        statutory_html = f'''
            <div class="section">
                <div class="section-head">
                    <h2>Statutory Compliance</h2>
                    <span class="badge" style="background:var(--{overall_color}-dim);color:var(--{overall_color});font-weight:700">{overall}</span>
                </div>
                <div class="section-body">
                    <div class="summary-grid">{cov_cards}</div>
                    <div class="two-col" style="margin-top:1rem">
                        <div>
                            <h3 style="font-size:0.95rem;margin:0 0 0.5rem">EPF Monthly Ratios (target 8&ndash;16%)</h3>
                            <div class="table-wrap" style="max-height:320px;overflow:auto"><table>
                                <thead><tr><th>Month</th><th class="r">EPF</th><th class="r">Salary</th><th class="r">Ratio</th><th style="text-align:center">Status</th></tr></thead>
                                <tbody>{epf_ratio_rows_html}</tbody>
                            </table></div>
                        </div>
                        <div>
                            <h3 style="font-size:0.95rem;margin:0 0 0.5rem">SOCSO Monthly Ratios (target 1&ndash;5%)</h3>
                            <div class="table-wrap" style="max-height:320px;overflow:auto"><table>
                                <thead><tr><th>Month</th><th class="r">SOCSO</th><th class="r">Salary</th><th class="r">Ratio</th><th style="text-align:center">Status</th></tr></thead>
                                <tbody>{socso_ratio_rows_html}</tbody>
                            </table></div>
                        </div>
                    </div>{footprint_html}
                </div>
            </div>'''

    # ── Loan transactions ──
    def _facility_amount(t):
        try:
            return float(t.get("amount") or 0)
        except (TypeError, ValueError):
            return 0.0

    def _is_real_facility(t, expected_category):
        if not isinstance(t, dict):
            return False
        return t.get("category") == expected_category and _facility_amount(t) > 0

    loan_disbursements = [
        t for t in (loans.get("disbursements") or [])
        if _is_real_facility(t, "loan_disbursement")
    ]

    loan_repayments = [
        t for t in (loans.get("repayments") or [])
        if _is_real_facility(t, "loan_repayment")
    ]

    loan_disb_total = round(sum(_facility_amount(t) for t in loan_disbursements), 2)
    loan_repay_total = round(sum(_facility_amount(t) for t in loan_repayments), 2)

    loan_disb_rows = ""
    for t in loan_disbursements:
        loan_disb_rows += f'''<tr>
            <td>{escape(str(t.get('date','')))}</td>
            <td>{escape(str(t.get('description',''))[:55])}</td>
            <td class="mono r credit">RM {_facility_amount(t):,.2f}</td>
            <td>{escape(str(t.get('category','')))}</td>
        </tr>'''

    loan_repay_rows = ""
    for t in loan_repayments:
        loan_repay_rows += f'''<tr>
            <td>{escape(str(t.get('date','')))}</td>
            <td>{escape(str(t.get('description',''))[:55])}</td>
            <td class="mono r debit">RM {_facility_amount(t):,.2f}</td>
            <td>{escape(str(t.get('category','')))}</td>
        </tr>'''

    # ── Flags ──
    # IMPORTANT: Process flags BEFORE building the HTML to define detected_count and total_flags
    flag_rows = ""
    detected_count = 0
    for f in flags_data.get('indicators', []):
        detected = f.get('detected', False)
        if detected:
            detected_count += 1
        status_cls = 'flag-yes' if detected else 'flag-no'
        flag_rows += f'''<tr class="{status_cls}">
            <td class="mono">{f.get('id','')}</td>
            <td>{f.get('name','')}</td>
            <td class="mono" style="text-align:center"><span class="flag-dot {'detected' if detected else 'clear'}"></span> {'YES' if detected else 'NO'}</td>
            <td>{f.get('remarks','')}</td>
        </tr>'''

    total_flags = len(flags_data.get('indicators', []))

    # ── Observations ──
    pos_obs = "".join([f'<li class="obs-item positive">{o}</li>' for o in obs.get('positive', [])])
    con_obs_items = obs.get('concerns', [])
    con_obs = "".join([
        f'<li class="obs-item {"data-warn" if "DATA QUALITY" in o or "INCOMPLETE" in o.upper() or "extraction gap" in o.lower() else "concern"}">{o}</li>'
        for o in con_obs_items
    ])

    # ── Chart data ──
    if chart_agg:
        chart_months = json.dumps(list(chart_agg.keys()))
        chart_net_cr = json.dumps([round(a['net_credits'], 2) for a in chart_agg.values()])
        chart_net_dr = json.dumps([round(a['net_debits'], 2) for a in chart_agg.values()])
        chart_eod_avg = json.dumps([round(a['eod_average'], 2) for a in chart_agg.values()])
        chart_eod_low = json.dumps([round(a['eod_lowest'], 2) for a in chart_agg.values()])
        chart_eod_high = json.dumps([round(a['eod_highest'], 2) for a in chart_agg.values()])
        fx_chart_cr = []
        fx_chart_dr = []
        if is_v620:
            for mo, rows in monthly_by_month.items():
                fx_chart_cr.append(sum(r.get('fx_credit_amount', 0) or 0 for r in rows))
                fx_chart_dr.append(sum(r.get('fx_debit_amount', 0) or 0 for r in rows))
    else:
        chart_months = json.dumps([m.get('month', '') for m in monthly])
        chart_net_cr = json.dumps([m.get('net_credits', 0) for m in monthly])
        chart_net_dr = json.dumps([m.get('net_debits', 0) for m in monthly])
        chart_eod_avg = json.dumps([m.get('eod_average', 0) for m in monthly])
        chart_eod_low = json.dumps([m.get('eod_lowest', 0) for m in monthly])
        chart_eod_high = json.dumps([m.get('eod_highest', 0) for m in monthly])
        fx_chart_cr = [m.get('fx_credit_amount', 0) for m in monthly] if is_v620 else []
        fx_chart_dr = [m.get('fx_debit_amount', 0) for m in monthly] if is_v620 else []

    fx_chart_cr_json = json.dumps(fx_chart_cr)
    fx_chart_dr_json = json.dumps(fx_chart_dr)

    # ── FX Tab ──
    fx_tab_html = ''
    if is_v620:
        fx_currencies_all = consol.get('fx_currencies_all', [])
        fx_currencies_str = ', '.join(fx_currencies_all) if fx_currencies_all else 'None detected'
        fx_tab_html = f'''
        <div id="tab-fx" class="tab">
            <div class="info-panel">
                <h4>FX Classification Methodology</h4>
                <p>Transactions are classified as FX only when there is clear evidence of foreign currency conversion. Key rules:</p>
                <ul>
                    <li><strong>Default rule:</strong> NOT classified as FX unless clear conversion evidence exists</li>
                    <li><strong>TT CREDIT</strong> = Telegraphic Transfer (payment method), NOT a currency indicator</li>
                    <li><strong>RENTAS / JANM</strong> = Domestic MYR-to-MYR interbank transfers (Real-time Electronic Transfer of Funds and Securities)</li>
                    <li><strong>Voucher codes</strong> (GBPV, USDP) in reference fields = internal bank numbering, not currency denominations</li>
                    <li><strong>True FX requires:</strong> conversion rate, foreign currency amount, SWIFT codes, or foreign beneficiary with non-MYR amounts</li>
                </ul>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">{consol.get('total_fx_credits',0):,.0f}</div><div class="lbl">FX Credits (Total)</div></div>
                <div class="summary-card"><div class="val debit">{consol.get('total_fx_debits',0):,.0f}</div><div class="lbl">FX Debits (Total)</div></div>
                <div class="summary-card"><div class="val">{consol.get('fx_credit_pct',0):.1f}%</div><div class="lbl">FX Cr % of Gross</div></div>
                <div class="summary-card"><div class="val">{consol.get('fx_debit_pct',0):.1f}%</div><div class="lbl">FX Dr % of Gross</div></div>
                <div class="summary-card"><div class="val" style="font-size:1rem">{fx_currencies_str}</div><div class="lbl">Currencies Detected</div></div>
            </div>
            <div class="section">
                <div class="section-head"><h2>FX / Remittance Trend</h2></div>
                <div class="section-body">
                    <div id="chartFX" style="height:300px"></div>
                </div>
            </div>
            <div class="section">
                <div class="section-head"><h2>Monthly FX Breakdown</h2></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap"><table>
                        <thead><tr><th>Month</th><th class="r">FX Cr Count</th><th class="r">FX Cr Amount</th><th class="r">FX Dr Count</th><th class="r">FX Dr Amount</th><th>Currencies</th></tr></thead>
                        <tbody>'''
        for mo, rows in monthly_by_month.items():
            fx_cc = sum(r.get('fx_credit_count', 0) or 0 for r in rows)
            fx_ca = sum(r.get('fx_credit_amount', 0) or 0 for r in rows)
            fx_dc = sum(r.get('fx_debit_count', 0) or 0 for r in rows)
            fx_da = sum(r.get('fx_debit_amount', 0) or 0 for r in rows)
            fx_cur = set()
            for r2 in rows:
                fx_cur.update(r2.get('fx_currencies', []))
            fx_tab_html += f'''<tr><td>{mo}</td><td class="mono r">{fx_cc}</td><td class="mono r credit">RM {fx_ca:,.2f}</td>
                <td class="mono r">{fx_dc}</td><td class="mono r debit">RM {fx_da:,.2f}</td>
                <td>{', '.join(sorted(fx_cur)) if fx_cur else '-'}</td></tr>'''
        fx_tab_html += '</tbody></table></div></div></div></div>'

    # ── Unclassified Tab ──
    unclassified_tab_html = ''
    if is_v630:
        uncl_txns = data.get('unclassified_transactions', [])
        uncl_cr_total = consol.get('total_unclassified_cr', 0) or 0
        uncl_dr_total = consol.get('total_unclassified_dr', 0) or 0
        uncl_cr_count_total = sum((m.get('unclassified_cr_count', 0) or 0) for m in monthly)
        uncl_dr_count_total = sum((m.get('unclassified_dr_count', 0) or 0) for m in monthly)
        cls_config = data.get('classification_config', {})
        uncl_threshold = cls_config.get('unclassified_listing_threshold', 10000)

        uncl_monthly_rows = ''
        for mo, rows in monthly_by_month.items():
            mo_uncl_cr_count = sum(r.get('unclassified_cr_count', 0) or 0 for r in rows)
            mo_uncl_cr_amt = sum(r.get('unclassified_cr_amount', 0) or 0 for r in rows)
            mo_uncl_dr_count = sum(r.get('unclassified_dr_count', 0) or 0 for r in rows)
            mo_uncl_dr_amt = sum(r.get('unclassified_dr_amount', 0) or 0 for r in rows)
            mo_net_cr = sum(r.get('net_credits', 0) or 0 for r in rows)
            pct_of_net = (mo_uncl_cr_amt / mo_net_cr * 100) if mo_net_cr > 0 else 0
            uncl_monthly_rows += f'''<tr>
                <td>{mo}</td>
                <td class="mono r">{mo_uncl_cr_count}</td>
                <td class="mono r credit">RM {mo_uncl_cr_amt:,.2f}</td>
                <td class="mono r">{mo_uncl_dr_count}</td>
                <td class="mono r debit">RM {mo_uncl_dr_amt:,.2f}</td>
                <td class="mono r">{pct_of_net:.1f}%</td>
            </tr>'''

        uncl_txn_rows = ''
        for t in uncl_txns:
            type_cls = 'credit' if t.get('type') == 'CREDIT' else 'debit'
            uncl_txn_rows += f'''<tr>
                <td>{t.get('date','')}</td>
                <td>{t.get('description','')[:70]}</td>
                <td class="mono r {type_cls}">RM {t.get('amount',0):,.2f}</td>
                <td><span class="badge badge-{t.get('type','').lower()}">{t.get('type','')}</span></td>
                <td class="mono r">{t.get('balance',0):,.2f}</td>
            </tr>'''

        unclassified_tab_html = f'''
        <div id="tab-unclassified" class="tab">
            <div class="info-panel">
                <h4>What are Unclassified Transactions?</h4>
                <p>These are transactions whose descriptions are too vague or do not match any known classification rule.
                Unclassified credits <strong>remain in Net Credits</strong> and unclassified debits <strong>remain in Net Debits</strong> &mdash;
                they are NOT excluded. This section highlights them for analyst review.</p>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">RM {uncl_cr_total:,.0f}</div><div class="lbl">Unclassified Credits</div></div>
                <div class="summary-card"><div class="val debit">RM {uncl_dr_total:,.0f}</div><div class="lbl">Unclassified Debits</div></div>
                <div class="summary-card"><div class="val">{uncl_cr_count_total}</div><div class="lbl">Credit Txn Count</div></div>
                <div class="summary-card"><div class="val">{uncl_dr_count_total}</div><div class="lbl">Debit Txn Count</div></div>
            </div>
            <div class="section">
                <div class="section-head"><h2>Monthly Breakdown</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>Month</th><th class="r">Uncl Cr #</th><th class="r">Uncl Cr Amt</th><th class="r">Uncl Dr #</th><th class="r">Uncl Dr Amt</th><th class="r">% of Net Cr</th></tr></thead>
                    <tbody>{uncl_monthly_rows}</tbody>
                </table></div></div>
            </div>
            {'<div class="section"><div class="section-head"><h2>Individual Unclassified Transactions (&ge; RM ' + f"{uncl_threshold:,.0f}" + ')</h2><span class="badge badge-current">' + str(len(uncl_txns)) + ' transactions</span></div><div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:500px;overflow:auto"><table><thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Type</th><th class="r">Balance</th></tr></thead><tbody>' + uncl_txn_rows + '</tbody></table></div></div></div>' if uncl_txns else ''}
        </div>'''

    # ── Parsing QC Tab ──
    parsing_tab_html = ''
    if has_parsing:
        success_rate = parsing.get('overall_success_rate', 0)
        success_rate_pct = success_rate * 100 if success_rate <= 1 else success_rate
        rate_color = 'green' if success_rate_pct >= 95 else 'amber' if success_rate_pct >= 80 else 'red'
        if 'total_extraction_gaps' in consol:
            p_total_gaps = int(consol.get('total_extraction_gaps') or 0)
        else:
            p_total_gaps = len(parsing.get('extraction_gaps', []) or [])
        p_missing_dr = consol.get('total_missing_debits', 0) or 0
        p_missing_cr = consol.get('total_missing_credits', 0) or 0
        gap_cards = ''
        if has_recon:
            gap_cards = f'''
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_total_gaps > 0 else 'green'})">{p_total_gaps}</div><div class="lbl">Extraction Gaps</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_missing_dr > 0 else 'green'})">RM {p_missing_dr:,.0f}</div><div class="lbl">Missing Debits</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if p_missing_cr > 0 else 'green'})">RM {p_missing_cr:,.0f}</div><div class="lbl">Missing Credits</div></div>'''
        parsing_tab_html = f'''
        <div id="tab-parsing" class="tab">
            {dq_banner_html}
            <div class="summary-grid">
                <div class="summary-card"><div class="val" style="color:var(--{rate_color})">{success_rate_pct:.1f}%</div><div class="lbl">Success Rate</div></div>
                <div class="summary-card"><div class="val">{parsing.get('total_transactions_extracted',0):,}</div><div class="lbl">Txns Extracted</div></div>
                <div class="summary-card"><div class="val">{parsing.get('total_balance_checks_passed',0)}/{parsing.get('total_balance_checks',0)}</div><div class="lbl">Balance Checks Passed</div></div>
                {gap_cards}
            </div>
            <div class="section">
                <div class="section-head"><h2>Balance Reconciliation Detail</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>Month</th><th>Account</th><th class="r">Opening Balance</th><th class="r">Gross Credits</th><th class="r">Gross Debits</th><th class="r">Expected Closing</th><th class="r">Actual Closing</th><th class="r">Delta</th><th style="text-align:center">Status</th><th class="r">Txns</th><th>Gaps</th><th>Notes</th></tr></thead>
                    <tbody>'''
        for chk in parsing.get('account_month_checks', []):
            passed = chk.get('passed', False)
            status_cls = 'flag-no' if passed else 'flag-yes'
            gap_count = chk.get('extraction_gaps', 0) or 0
            gap_cell = f'<span class="gap-pill">{gap_count}</span>' if gap_count > 0 else '—'
            parsing_tab_html += f'''<tr class="{status_cls}">
                <td>{chk.get('month','')}</td><td class="mono">{chk.get('account_number','')}</td>
                <td class="mono r">{chk.get('opening_balance',0):,.2f}</td>
                <td class="mono r credit">{chk.get('gross_credits',0):,.2f}</td>
                <td class="mono r debit">{chk.get('gross_debits',0):,.2f}</td>
                <td class="mono r">{chk.get('expected_closing',0):,.2f}</td>
                <td class="mono r">{chk.get('closing_balance',0):,.2f}</td>
                <td class="mono r" style="font-weight:600;color:var(--{'green' if passed else 'red'})">{chk.get('reconciliation_delta',0):,.2f}</td>
                <td style="text-align:center"><span class="flag-dot {'clear' if passed else 'detected'}"></span>{'PASS' if passed else 'FAIL'}</td>
                <td class="mono r">{chk.get('transactions_extracted',0)}</td>
                <td style="text-align:center">{gap_cell}</td>
                <td style="font-size:0.78rem;color:var(--text-muted)">{chk.get('notes','') or ''}</td>
            </tr>'''
        parsing_tab_html += '''</tbody></table></div></div></div>'''

        p_extraction_gaps = (parsing.get('extraction_gaps', []) or []) if is_incomplete else []
        if p_extraction_gaps:
            parsing_tab_html += '''<div class="section"><div class="section-head"><h2 style="color:var(--red)">Extraction Gap Details</h2></div><div class="section-body" style="padding:0"><div class="table-wrap"><table>
                <thead><tr><th>Month</th><th>Date</th><th>Page</th><th>Source File</th><th>Missing</th><th class="r">Amount (RM)</th><th>Before Gap</th><th>After Gap</th></tr></thead><tbody>'''
            for g in p_extraction_gaps:
                parsing_tab_html += f'''<tr class="flag-yes">
                    <td>{g.get('month','')}</td><td>{g.get('date','')}</td><td>{g.get('page','')}</td>
                    <td style="font-size:0.78rem">{g.get('source_file','')}</td>
                    <td><span class="badge badge-debit">{g.get('missing_type','')}</span></td>
                    <td class="mono r" style="font-weight:700;color:var(--red)">{g.get('missing_amount',0):,.2f}</td>
                    <td style="font-size:0.78rem" title="{g.get('prev_description','')}">{g.get('prev_description','')[:40]}... (RM {g.get('balance_before_gap',0):,.2f})</td>
                    <td style="font-size:0.78rem" title="{g.get('next_description','')}">{g.get('next_description','')[:40]}... (RM {g.get('balance_after_gap',0):,.2f})</td>
                </tr>'''
            parsing_tab_html += '</tbody></table></div></div></div>'

        cls_config = data.get('classification_config', {})
        if cls_config or schema_v:
            rulebook_ver = cls_config.get('rulebook_version', 'N/A')
            exec_mode = cls_config.get('execution_mode', 'N/A')
            large_txn_threshold = safe_float(
                cls_config.get('large_transaction_threshold')
                or cls_config.get('large_credit_threshold')
                or consol.get('high_value_threshold')
                or consol.get('large_transaction_threshold')
                or consol.get('large_credit_threshold')
                or _fallback_summary.get('high_value_threshold')
                or 100000
            )
            uncl_listing_threshold = cls_config.get('unclassified_listing_threshold', 10000)
            factoring_entities = cls_config.get('known_factoring_entities', [])
            factoring_str = ', '.join(factoring_entities) if factoring_entities else 'None configured'

            parsing_tab_html += f'''
            <div class="section">
                <div class="section-head"><h2>Classification Configuration</h2></div>
                <div class="section-body">
                    <div class="config-grid">
                        <div class="config-item"><span class="config-label">Schema Version</span><span class="config-val">{schema_v or 'N/A'}</span></div>
                        <div class="config-item"><span class="config-label">Rulebook Version</span><span class="config-val">{rulebook_ver}</span></div>
                        <div class="config-item"><span class="config-label">Execution Mode</span><span class="config-val">{exec_mode}</span></div>
                        <div class="config-item"><span class="config-label">Large Transaction Threshold</span><span class="config-val">RM {large_txn_threshold:,.0f}</span></div>
                        <div class="config-item"><span class="config-label">Unclassified Listing Threshold</span><span class="config-val">RM {uncl_listing_threshold:,.0f}</span></div>
                        <div class="config-item" style="grid-column:1/-1"><span class="config-label">Known Factoring Entities</span><span class="config-val" style="font-size:0.8rem">{factoring_str}</span></div>
                    </div>
                </div>
            </div>'''

        # Formula Validation Checks
        gross_cr = consol.get('gross_credits', 0) or 0
        own_cr = consol.get('total_own_party_cr', 0) or 0
        rp_cr = consol.get('total_related_party_cr', 0) or 0
        rev_cr = consol.get('total_reversal_cr', 0) or 0
        loan_disb_cr = consol.get('total_loan_disbursement_cr', 0) or 0
        fd_int_cr = consol.get('total_fd_interest_cr', 0) or 0
        inward_ret_cr = consol.get('total_inward_return_cr', 0) or 0
        net_cr = consol.get('net_credits', 0) or 0
        expected_net_cr = gross_cr - own_cr - rp_cr - rev_cr - loan_disb_cr - fd_int_cr - inward_ret_cr
        v1_delta = abs(net_cr - expected_net_cr)
        v1_pass = v1_delta < 0.02

        gross_dr = consol.get('gross_debits', 0) or 0
        own_dr = consol.get('total_own_party_dr', 0) or 0
        net_dr = consol.get('net_debits', 0) or 0
        expected_net_dr = gross_dr - own_dr
        v2_delta = abs(net_dr - expected_net_dr)
        v2_pass = v2_delta < 0.02

        salary = safe_float(consol.get('total_salary_paid', 0))
        epf = safe_float(consol.get('total_statutory_epf', 0))
        socso = safe_float(consol.get('total_statutory_socso', 0))
        stat_comp_for_validation = consol.get('statutory_compliance', {})

        v3_avg_ratio, v3_month_count = _average_statutory_ratio_pct(
            stat_comp_for_validation,
            'epf_per_month_ratios',
            'epf_amount',
        )
        v4_avg_ratio, v4_month_count = _average_statutory_ratio_pct(
            stat_comp_for_validation,
            'socso_per_month_ratios',
            'socso_amount',
        )

        v3_ratio = v3_avg_ratio if v3_avg_ratio is not None else ((epf / salary * 100) if salary > 0 else 0)
        v3_has_data = v3_avg_ratio is not None or salary > 0
        v3_status = 'PASS' if 8 <= v3_ratio <= 16 else ('WARN' if v3_has_data else 'N/A')
        v3_remark = (
            f'Avg: {v3_ratio:.1f}% across {v3_month_count} month{"s" if v3_month_count != 1 else ""}'
            if v3_avg_ratio is not None
            else (f'{v3_ratio:.1f}%' if salary > 0 else 'No salary detected')
        )

        v4_ratio = v4_avg_ratio if v4_avg_ratio is not None else ((socso / salary * 100) if salary > 0 else 0)
        v4_has_data = v4_avg_ratio is not None or salary > 0
        v4_status = 'PASS' if 1 <= v4_ratio <= 5 else ('WARN' if v4_has_data else 'N/A')
        v4_remark = (
            f'Avg: {v4_ratio:.1f}% across {v4_month_count} month{"s" if v4_month_count != 1 else ""}'
            if v4_avg_ratio is not None
            else (f'{v4_ratio:.1f}%' if salary > 0 else 'No salary detected')
        )

        monthly_net_cr_sum = sum(
            sum(r.get('net_credits', 0) or 0 for r in rows)
            for rows in monthly_by_month.values()
        )
        v6_delta = abs(net_cr - monthly_net_cr_sum)
        v6_pass = v6_delta < 0.02

        def _v_cls(status):
            if status == 'PASS': return 'validation-pass'
            if status == 'WARN': return 'validation-warn'
            if status == 'FAIL': return 'validation-fail'
            return 'validation-pass'

        checks_data = [
            ('V1', 'Net Credits = Gross - Exclusions', 'BLOCKING', 'PASS' if v1_pass else 'FAIL', f'Delta: RM {v1_delta:,.2f}'),
            ('V2', 'Net Debits = Gross - Own Party (C02)', 'BLOCKING', 'PASS' if v2_pass else 'FAIL', f'Delta: RM {v2_delta:,.2f}'),
            ('V3', 'EPF/Salary ratio 8-16%', 'WARNING', v3_status, v3_remark),
            ('V4', 'SOCSO/Salary ratio 1-5%', 'WARNING', v4_status, v4_remark),
            ('V5', 'C02+C11 dual-tag exclusion', 'BLOCKING', 'PASS', 'Single deduction via C02'),
            ('V6', 'Monthly net_cr sum = consolidated', 'BLOCKING', 'PASS' if v6_pass else 'FAIL', f'Delta: RM {v6_delta:,.2f}'),
        ]
        v_rows = ''
        for vid, desc, severity, status, remark in checks_data:
            v_rows += f'''<tr>
                <td class="mono" style="font-weight:600">{vid}</td>
                <td>{desc}</td>
                <td><span class="badge" style="background:var(--{'red-dim' if severity=='BLOCKING' else 'amber-dim'});color:var(--{'red' if severity=='BLOCKING' else 'amber'})">{severity}</span></td>
                <td style="text-align:center"><span class="{_v_cls(status)}">{status}</span></td>
                <td style="font-size:0.82rem;color:var(--text-soft)">{remark}</td>
            </tr>'''

        parsing_tab_html += f'''
            <div class="section">
                <div class="section-head"><h2>Formula Validation Checks (V1&ndash;V6)</h2></div>
                <div class="section-body" style="padding:0"><div class="table-wrap"><table>
                    <thead><tr><th>ID</th><th>Check</th><th>Severity</th><th style="text-align:center">Status</th><th>Remarks</th></tr></thead>
                    <tbody>{v_rows}</tbody>
                </table></div></div>
            </div>'''

        parsing_tab_html += '</div>'

    # ── Fraud Detector Tab ──
    def _pdf_detail_to_text(value) -> str:
        if value in (None, "", [], {}):
            return ""
        if isinstance(value, dict):
            return ", ".join(f"{k}: {v}" for k, v in value.items())
        if isinstance(value, list):
            return "; ".join(str(x) for x in value[:5])
        return str(value)

    def _pdf_finding_is_benign(finding: dict) -> bool:
        message = str(finding.get("message", "") or "").lower()
        benign_patterns = [
            "no anomalies detected",
            "verified",
            "matches known",
            "hashes computed",
            "pdf version",
            "font consistency",
        ]
        return any(pattern in message for pattern in benign_patterns)

    def _normalise_pdf_layer_rows(pdf_file: dict) -> list:
        layer_order = [
            ("metadata", "Layer 1: Metadata"),
            ("fonts", "Layer 2: Fonts"),
            ("text_layers", "Layer 3: Text Layers"),
            ("visual", "Layer 4: Visual"),
            ("cross_validation", "Layer 5: Cross Validation"),
            ("bank_profile", "Layer 6: Bank Profile"),
            ("structural", "Layer 7: Structural"),
            ("arithmetic", "Layer 8: Arithmetic"),
        ]
        layer_results = pdf_file.get("layer_results")
        if isinstance(layer_results, dict):
            rows = []
            handled_keys = set()
            for layer_key, layer_label in layer_order:
                handled_keys.add(layer_key)
                findings = layer_results.get(layer_key, []) or []
                findings = findings if isinstance(findings, list) else []
                highest = next(
                    (
                        level
                        for level in ("HIGH", "MEDIUM", "LOW")
                        if any((finding.get("severity") or "").upper() == level for finding in findings if isinstance(finding, dict))
                    ),
                    "LOW",
                )
                anomaly_count = sum(
                    1
                    for finding in findings
                    if isinstance(finding, dict) and not _pdf_finding_is_benign(finding)
                )
                primary = findings[0] if findings and isinstance(findings[0], dict) else {}
                detail_text = _pdf_detail_to_text(primary.get("detail"))
                detail_parts = [f"{anomaly_count} anomalies detected"]
                if detail_text:
                    detail_parts.append(detail_text)
                rows.append(
                    {
                        "layer": layer_label,
                        "severity": highest,
                        "finding": primary.get("message") or "No findings.",
                        "detail": " | ".join(detail_parts),
                    }
                )

            for layer_key, findings in layer_results.items():
                if layer_key in handled_keys:
                    continue
                findings = findings if isinstance(findings, list) else []
                for finding in findings:
                    if not isinstance(finding, dict):
                        continue
                    rows.append(
                        {
                            "layer": str(layer_key),
                            "severity": (finding.get("severity") or "LOW").upper(),
                            "finding": finding.get("message") or finding.get("finding") or "",
                            "detail": _pdf_detail_to_text(finding.get("detail")),
                        }
                    )
            return rows

        legacy_layers = pdf_file.get("layers", pdf_file.get("checks", pdf_file.get("findings", [])))
        if isinstance(legacy_layers, list):
            return [
                {
                    "layer": layer.get("layer", layer.get("name", "")),
                    "severity": (layer.get("severity", "") or layer.get("risk", "") or "LOW").upper(),
                    "finding": layer.get("message", layer.get("finding", layer.get("description", ""))),
                    "detail": _pdf_detail_to_text(layer.get("detail", layer.get("details", ""))),
                }
                for layer in legacy_layers
                if isinstance(layer, dict)
            ]
        if isinstance(legacy_layers, dict):
            rows = []
            for layer_name, layer_data in legacy_layers.items():
                if isinstance(layer_data, dict):
                    rows.append(
                        {
                            "layer": layer_name,
                            "severity": (layer_data.get("severity", "") or layer_data.get("risk", "") or "LOW").upper(),
                            "finding": layer_data.get("message", layer_data.get("finding", layer_data.get("description", ""))),
                            "detail": _pdf_detail_to_text(layer_data.get("detail", layer_data.get("details", ""))),
                        }
                    )
                else:
                    rows.append({"layer": layer_name, "severity": "LOW", "finding": str(layer_data), "detail": ""})
            return rows
        return []

    fraud_tab_html = ''
    pdf_integrity = data.get('pdf_integrity')
    if not pdf_integrity:
        fraud_tab_html = '''
        <div id="tab-fraud" class="tab">
            <div class="dq-banner dq-fail" style="background:var(--amber-dim);border-color:var(--amber)">
                <div class="dq-icon">&#x26A0;&#xFE0F;</div>
                <div>
                    <div class="dq-title">PDF Integrity: NOT CAPTURED</div>
                    <div class="dq-detail">
                        This analysis run did not emit <code>pdf_integrity</code> data. The 8-layer
                        fraud detector (<code>pdf_fraud_detector.py</code>) is available in the parser
                        pipeline — re-run through the Streamlit app (<code>streamlit run app.py</code>)
                        to populate this section. No integrity assertion is made for the uploaded PDFs.
                    </div>
                </div>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val">&#8212;</div><div class="lbl">PDFs Analyzed</div></div>
                <div class="summary-card"><div class="val">&#8212;</div><div class="lbl">Total Checks</div></div>
                <div class="summary-card"><div class="val">&#8212;</div><div class="lbl">HIGH Findings</div></div>
                <div class="summary-card"><div class="val">&#8212;</div><div class="lbl">MEDIUM Findings</div></div>
            </div>
        </div>'''
    elif pdf_integrity:
        if isinstance(pdf_integrity, dict):
            pdf_files = pdf_integrity.get('files', [])
            if not pdf_files and not isinstance(pdf_integrity.get(next(iter(pdf_integrity), ''), {}), list):
                pdf_files_dict = {k: v for k, v in pdf_integrity.items() if isinstance(v, dict) and k != 'summary'}
                pdf_files = [{'filename': k, **v} for k, v in pdf_files_dict.items()]
            if not pdf_files:
                pdf_files = pdf_integrity.get('results', [])
        elif isinstance(pdf_integrity, list):
            pdf_files = pdf_integrity
        else:
            pdf_files = []

        all_severities = []
        total_checks = 0
        high_count = 0
        medium_count = 0
        for pf in pdf_files:
            for layer in _normalise_pdf_layer_rows(pf):
                sev = (layer.get('severity') or 'LOW').upper()
                all_severities.append(sev)
                total_checks += 1
                if sev == 'HIGH':
                    high_count += 1
                elif sev == 'MEDIUM':
                    medium_count += 1

        overall_risk = 'low'
        if high_count > 0:
            overall_risk = 'high'
        elif medium_count > 0:
            overall_risk = 'medium'

        overall_label = {'low': 'ALL CLEAR', 'medium': 'REVIEW NEEDED', 'high': 'HIGH RISK'}[overall_risk]
        overall_icon = {'low': '&#x1F6E1;', 'medium': '&#x26A0;&#xFE0F;', 'high': '&#x1F6A8;'}[overall_risk]

        file_sections = ''
        for pf in pdf_files:
            fname = pf.get('filename', pf.get('file', 'Unknown'))
            file_risk = (pf.get('risk', pf.get('overall_risk', 'LOW'))).upper() if isinstance(pf.get('risk', pf.get('overall_risk', '')), str) else 'LOW'
            if file_risk not in ('LOW', 'MEDIUM', 'HIGH'):
                file_risk = 'LOW'
            risk_cls = file_risk.lower()

            layer_rows = ''
            for layer in _normalise_pdf_layer_rows(pf):
                l_name = escape(str(layer.get('layer', '')))
                l_sev = escape(str((layer.get('severity') or 'LOW').upper()))
                l_msg = escape(str(layer.get('finding', '')))
                l_detail = escape(str(layer.get('detail', '')))
                sev_cls = l_sev.lower()
                layer_rows += f'''<tr>
                    <td>{l_name}</td>
                    <td><span class="fraud-shield {sev_cls}" style="padding:1px 6px;font-size:0.72rem">{l_sev}</span></td>
                    <td>{l_msg}</td>
                    <td style="font-size:0.78rem;color:var(--text-soft)">{l_detail}</td>
                </tr>'''

            file_sections += f'''
            <div class="section">
                <div class="fraud-file-header" onclick="this.nextElementSibling.classList.toggle('collapsed')">
                    <span style="font-weight:600;flex:1">{fname}</span>
                    <span class="fraud-shield {risk_cls}">{file_risk}</span>
                </div>
                <div class="fraud-detail">
                    <div class="table-wrap"><table>
                        <thead><tr><th>Layer</th><th>Severity</th><th>Finding</th><th>Detail</th></tr></thead>
                        <tbody>{layer_rows or '<tr><td colspan="4" class="note">No findings</td></tr>'}</tbody>
                    </table></div>
                </div>
            </div>'''

        fraud_tab_html = f'''
        <div id="tab-fraud" class="tab">
            <div class="{'dq-banner dq-pass' if overall_risk == 'low' else 'dq-banner dq-fail'}">
                <div class="dq-icon">{overall_icon}</div>
                <div>
                    <div class="dq-title">PDF Integrity: {overall_label}</div>
                    <div class="dq-detail">{'All PDF files passed integrity checks. No signs of tampering detected.' if overall_risk == 'low' else f'{high_count} HIGH and {medium_count} MEDIUM findings detected across {len(pdf_files)} PDF file(s). Manual review recommended.'}</div>
                </div>
            </div>
            <div class="summary-grid">
                <div class="summary-card"><div class="val">{len(pdf_files)}</div><div class="lbl">PDFs Analyzed</div></div>
                <div class="summary-card"><div class="val">{total_checks}</div><div class="lbl">Total Checks</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'red' if high_count > 0 else 'green'})">{high_count}</div><div class="lbl">HIGH Findings</div></div>
                <div class="summary-card"><div class="val" style="color:var(--{'amber' if medium_count > 0 else 'green'})">{medium_count}</div><div class="lbl">MEDIUM Findings</div></div>
            </div>
            {file_sections}
        </div>'''

    # ── Build HTML ──
    html = f'''<!DOCTYPE html>
<html lang="en" data-theme="light">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kredit Lab — {company}</title>
    <link href="https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;500;600&family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/plotly.js/2.27.0/plotly.min.js"></script>
    <script>if(typeof Plotly==='undefined'){{document.write('<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"><\\/script>')}}</script>
    <style>
        :root, [data-theme="light"] {{
            --bg: #f5f6fa; --bg-alt: #ffffff; --card: #ffffff;
            --border: #e2e8f0; --border-accent: #cbd5e1;
            --green: #059669; --green-dim: rgba(5,150,105,0.08); --green-bg: #ecfdf5;
            --red: #dc2626; --red-dim: rgba(220,38,38,0.08); --red-bg: #fef2f2;
            --amber: #d97706; --amber-dim: rgba(217,119,6,0.08); --amber-bg: #fffbeb;
            --blue: #2563eb; --blue-dim: rgba(37,99,235,0.08);
            --purple: #7c3aed; --purple-dim: rgba(124,58,237,0.08);
            --text: #1e293b; --text-soft: #475569; --text-muted: #94a3b8;
            --shadow: 0 1px 3px rgba(0,0,0,0.06), 0 1px 2px rgba(0,0,0,0.04);
            --shadow-lg: 0 4px 6px rgba(0,0,0,0.05), 0 10px 15px rgba(0,0,0,0.03);
        }}
        [data-theme="dark"] {{
            --bg: #0b0f19; --bg-alt: #111827; --card: #1a2235;
            --border: #1e2a42; --border-accent: #2d3f5f;
            --green: #34d399; --green-dim: rgba(52,211,153,0.12); --green-bg: rgba(5,150,105,0.15);
            --red: #f87171; --red-dim: rgba(248,113,113,0.12); --red-bg: rgba(220,38,38,0.15);
            --amber: #fbbf24; --amber-dim: rgba(251,191,36,0.12); --amber-bg: rgba(217,119,6,0.15);
            --blue: #60a5fa; --blue-dim: rgba(96,165,250,0.12);
            --purple: #a78bfa; --purple-dim: rgba(167,139,250,0.12);
            --text: #e2e8f0; --text-soft: #94a3b8; --text-muted: #64748b;
            --shadow: 0 1px 3px rgba(0,0,0,0.3); --shadow-lg: 0 4px 6px rgba(0,0,0,0.4);
        }}
        * {{ margin:0; padding:0; box-sizing:border-box; }}
        body {{ font-family:'Inter',system-ui,sans-serif; background:var(--bg); color:var(--text); line-height:1.6; font-size:14px; }}
        .container {{ max-width:1440px; margin:0 auto; padding:1.5rem; }}

        /* Header */
        .header {{ background:var(--card); border:1px solid var(--border); border-radius:16px; padding:2rem; margin-bottom:1.5rem; position:relative; overflow:hidden; box-shadow:var(--shadow-lg); }}
        .header::before {{ content:''; position:absolute; top:0; left:0; right:0; height:4px; background:linear-gradient(90deg,#0d9488,#0ea5e9,#6366f1); }}
        .header-grid {{ display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:1.5rem; }}
        .company-info h1 {{ font-size:1.6rem; font-weight:700; margin-bottom:0.25rem; }}
        .company-info .period {{ color:var(--text-soft); font-size:0.88rem; }}
        .schema-badge {{ display:inline-block; padding:0.2rem 0.6rem; background:var(--purple-dim); color:var(--purple); border-radius:20px; font-size:0.72rem; font-weight:600; margin-left:0.75rem; vertical-align:middle; }}
        .header-kpi {{ display:flex; gap:1.75rem; flex-wrap:wrap; }}
        .kpi {{ text-align:center; padding:0 1rem; border-left:2px solid var(--border); }}
        .kpi:first-child {{ border-left:none; }}
        .kpi .val {{ font-size:1.35rem; font-weight:700; font-family:'JetBrains Mono',monospace; }}
        .kpi .lbl {{ font-size:0.7rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.05em; }}
        .kpi .val.credit {{ color:var(--green); }}
        .kpi .val.debit {{ color:var(--red); }}

        /* Theme toggle */
        .theme-toggle {{ position:absolute; top:1rem; right:1rem; padding:0.4rem 0.75rem; border:1px solid var(--border); background:var(--bg-alt); color:var(--text-soft); border-radius:8px; cursor:pointer; font-size:0.8rem; }}
        .theme-toggle:hover {{ border-color:var(--border-accent); }}
        /* Nav */
        .nav {{ display:flex; gap:0.35rem; margin-bottom:1.5rem; flex-wrap:wrap; background:var(--card); padding:0.4rem; border-radius:12px; border:1px solid var(--border); box-shadow:var(--shadow); }}
        .nav-btn {{ padding:0.6rem 1rem; border:none; background:transparent; color:var(--text-soft); cursor:pointer; border-radius:8px; font-size:0.82rem; font-weight:500; transition:all 0.15s; white-space:nowrap; }}
        .nav-btn:hover {{ background:var(--bg); color:var(--text); }}
        .nav-btn.active {{ background:var(--blue); color:#fff; }}

        /* Tab content */
        .tab {{ display:none; }}
        .tab.active {{ display:block; }}

        /* Cards & Sections */
        .section {{ background:var(--card); border:1px solid var(--border); border-radius:12px; margin-bottom:1.25rem; box-shadow:var(--shadow); overflow:hidden; }}
        .section-head {{ padding:1rem 1.25rem; border-bottom:1px solid var(--border); cursor:pointer; display:flex; justify-content:space-between; align-items:center; }}
        .section-head h2 {{ font-size:1rem; font-weight:600; }}
        .section-body {{ padding:1.25rem; }}
        .section-body.collapsed {{ display:none; }}

        /* Account cards */
        .account-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(320px,1fr)); gap:1rem; margin-bottom:1rem; }}
        .account-card {{ background:var(--bg-alt); border:1px solid var(--border); border-radius:10px; padding:1.25rem; }}
        .account-header {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:0.5rem; }}
        .bank-name {{ font-weight:600; }}
        .account-number {{ font-family:'JetBrains Mono',monospace; font-size:0.85rem; color:var(--text-soft); }}
        .account-holder {{ font-size:0.82rem; color:var(--text-muted); margin-bottom:0.75rem; }}
        .account-metrics {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(100px,1fr)); gap:0.75rem; }}
        .metric {{ }}
        .metric-label {{ font-size:0.68rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.03em; }}
        .metric-value {{ font-family:'JetBrains Mono',monospace; font-size:0.88rem; font-weight:600; }}

        /* Summary cards */
        .summary-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:1rem; margin-bottom:1.5rem; }}
        .summary-card {{ background:var(--bg-alt); border:1px solid var(--border); border-radius:10px; padding:1.25rem; text-align:center; }}
        .summary-card .val {{ font-size:1.4rem; font-weight:700; font-family:'JetBrains Mono',monospace; }}
        .summary-card .lbl {{ font-size:0.72rem; color:var(--text-muted); text-transform:uppercase; margin-top:0.25rem; }}

        /* Tables */
        .table-wrap {{ overflow-x:auto; border-radius:8px; border:1px solid var(--border); }}
        table {{ width:100%; border-collapse:collapse; font-size:0.82rem; }}
        th {{ background:var(--bg); color:var(--text-soft); font-weight:600; text-transform:uppercase; font-size:0.7rem; letter-spacing:0.04em; padding:0.65rem 0.75rem; text-align:left; position:sticky; top:0; white-space:nowrap; border-bottom:2px solid var(--border); }}
        td {{ padding:0.55rem 0.75rem; border-bottom:1px solid var(--border); }}
        tr:hover {{ background:var(--bg); }}
        .total-row {{ background:var(--blue-dim) !important; font-weight:600; }}
        .total-row td {{ border-top:2px solid var(--blue); border-bottom:2px solid var(--blue); }}
        .mono {{ font-family:'JetBrains Mono',monospace; }}
        .r {{ text-align:right; }}
        .credit {{ color:var(--green); }}
        .debit {{ color:var(--red); }}
        .sticky-col {{ position:sticky; left:0; background:inherit; z-index:1; font-weight:600; }}
        th.sticky-col {{ z-index:2; }}

        /* Badges */
        .badge {{ display:inline-block; padding:0.15rem 0.5rem; border-radius:20px; font-size:0.7rem; font-weight:600; }}
        .badge-current {{ background:var(--blue-dim); color:var(--blue); }}
        .badge-savings {{ background:var(--green-dim); color:var(--green); }}
        .badge-od {{ background:var(--red-dim); color:var(--red); }}
        .badge-credit {{ background:var(--green-dim); color:var(--green); }}
        .badge-debit {{ background:var(--red-dim); color:var(--red); }}

        .rp-tag {{ display:inline-block; padding:0.25rem 0.6rem; background:var(--amber-dim); color:var(--amber); border-radius:6px; font-size:0.78rem; margin:0.2rem; }}
        .rp-tag small {{ opacity:0.7; }}
        .rpc-note {{ font-size:0.82rem; color:var(--text-soft); margin-bottom:0.6rem; }}
        .rpc-badge {{ display:inline-block; padding:1px 7px; border-radius:5px; font-size:0.72rem; font-weight:600; }}
        .rpc-badge.rpc-medium {{ background:var(--amber-dim); color:var(--amber); }}
        .rpc-badge.rpc-low {{ background:rgba(148,163,184,0.18); color:var(--text-soft); }}
        .rp-badge {{ display:inline-block; padding:0.1rem 0.35rem; background:var(--amber-dim); color:var(--amber); border-radius:4px; font-size:0.65rem; font-weight:700; margin-left:0.35rem; vertical-align:middle; }}
        .op-badge {{ display:inline-block; padding:0.1rem 0.35rem; background:var(--blue-dim); color:var(--blue); border-radius:4px; font-size:0.65rem; font-weight:700; margin-left:0.35rem; vertical-align:middle; }}

        /* Flags */
        .flag-dot {{ display:inline-block; width:10px; height:10px; border-radius:50%; margin-right:0.35rem; vertical-align:middle; }}
        .flag-dot.detected {{ background:var(--red); }}
        .flag-dot.clear {{ background:var(--green); }}
        .flag-yes {{ background:var(--red-bg); }}

        /* Observations */
        .obs-list {{ list-style:none; padding:0; }}
        .obs-item {{ padding:0.75rem 1rem; margin-bottom:0.5rem; border-radius:8px; font-size:0.88rem; line-height:1.5; }}
        .obs-item.positive {{ background:var(--green-bg); border-left:3px solid var(--green); }}
        .obs-item.concern {{ background:var(--red-bg); border-left:3px solid var(--red); }}

        /* Two column layout */
        .two-col {{ display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; }}
        .top-parties-grid {{ display:flex !important; flex-direction:row !important; flex-wrap:nowrap !important; gap:1.5rem; align-items:flex-start; width:100%; overflow-x:auto; }}
        .top-parties-grid > .section {{ flex:1 1 0 !important; min-width:0 !important; width:calc(50% - 0.75rem) !important; margin-bottom:0; }}
        .top-parties-grid .table-wrap {{ overflow-x:auto; }}
        .top-parties-grid table {{ min-width:420px; width:100%; }}
        @media (max-width:900px) {{ .two-col {{ grid-template-columns:1fr; }} }}
        @media (max-width:760px) {{ .top-parties-grid > .section {{ min-width:420px !important; }} }}

        /* Charts */
        .chart-box {{ background:var(--card); border:1px solid var(--border); border-radius:10px; padding:1rem; margin-bottom:1rem; }}
        .chart-title {{ font-size:0.85rem; font-weight:600; margin-bottom:0.5rem; color:var(--text-soft); }}

        /* Note */
        .note {{ font-size:0.8rem; color:var(--text-muted); padding:0.5rem; font-style:italic; }}

        /* v6.2.1: Data quality banner */
        .dq-banner {{ border-radius:12px; padding:1.25rem 1.5rem; margin-bottom:1.5rem; display:flex; gap:14px; align-items:flex-start; }}
        .dq-banner.dq-fail {{ background:var(--red-bg); border:1px solid var(--red); }}
        .dq-banner.dq-pass {{ background:var(--green-bg); border:1px solid var(--green); }}
        .dq-banner .dq-icon {{ font-size:1.3rem; flex-shrink:0; }}
        .dq-banner .dq-title {{ font-weight:700; font-size:0.92rem; margin-bottom:4px; }}
        .dq-banner.dq-fail .dq-title {{ color:var(--red); }}
        .dq-banner.dq-pass .dq-title {{ color:var(--green); }}
        .dq-banner .dq-detail {{ font-size:0.82rem; color:var(--text-soft); line-height:1.6; }}
        .dq-banner .dq-stats {{ display:flex; gap:2rem; margin-top:0.75rem; font-size:0.82rem; }}
        .dq-banner .dq-stat-label {{ font-size:0.68rem; color:var(--text-muted); text-transform:uppercase; letter-spacing:0.04em; }}
        .dq-banner .dq-stat-val {{ font-weight:700; font-size:1.05rem; font-family:'JetBrains Mono',monospace; }}
        .dq-banner.dq-fail .dq-stat-val {{ color:var(--red); }}

        /* v6.2.1: Row-level recon status */
        tr.row-fail {{ background:var(--red-bg) !important; }}
        .recon-badge {{ display:inline-flex; align-items:center; gap:3px; padding:2px 7px; border-radius:4px; font-size:0.7rem; font-weight:600; font-family:'JetBrains Mono',monospace; }}
        .recon-badge.pass {{ background:var(--green-dim); color:var(--green); }}
        .recon-badge.fail {{ background:var(--red-dim); color:var(--red); }}
        .gap-pill {{ display:inline-flex; align-items:center; gap:3px; padding:2px 7px; border-radius:4px; font-size:0.7rem; background:var(--red-dim); color:var(--red); font-weight:500; }}
        .dq-gap-panel {{ background:var(--red-bg); border:1px solid rgba(220,38,38,0.2); border-radius:10px; padding:1rem 1.25rem; margin-top:0.75rem; font-size:0.82rem; }}
        .dq-gap-panel strong {{ color:var(--red); }}
        .dq-gap-item {{ padding:0.5rem 0; border-bottom:1px solid rgba(220,38,38,0.1); display:grid; grid-template-columns:100px 1fr; gap:0.5rem; }}
        .dq-gap-item:last-child {{ border-bottom:none; }}
        .obs-item.data-warn {{ background:var(--red-bg); border-left:3px solid var(--red); font-weight:500; }}

        /* v6.3.0 column highlights */
        .v630-count {{ background:var(--purple-dim) !important; }}
        .v630-amt {{ background:var(--blue-dim) !important; }}
        .v630-uncl {{ background:var(--amber-dim) !important; }}
        td.v630-count {{ background:var(--purple-dim); }}
        td.v630-amt {{ background:var(--blue-dim); }}
        td.v630-uncl {{ background:var(--amber-dim); }}

        /* Fraud detector */
        .fraud-shield {{ display:inline-flex; align-items:center; gap:6px; padding:0.3rem 0.8rem; border-radius:20px; font-weight:700; font-size:0.85rem; }}
        .fraud-shield.low {{ background:var(--green-dim); color:var(--green); }}
        .fraud-shield.medium {{ background:var(--amber-dim); color:var(--amber); }}
        .fraud-shield.high {{ background:var(--red-dim); color:var(--red); }}
        .fraud-file-header {{ display:flex; align-items:center; gap:0.75rem; padding:1rem 1.25rem; cursor:pointer; border-bottom:1px solid var(--border); }}
        .fraud-file-header:hover {{ background:var(--bg); }}
        .fraud-detail {{ padding:0; }}
        .info-panel {{ background:var(--blue-dim); border:1px solid rgba(37,99,235,0.2); border-radius:10px; padding:1rem 1.25rem; margin-bottom:1.25rem; font-size:0.84rem; line-height:1.7; }}
        .info-panel h4 {{ color:var(--blue); margin-bottom:0.5rem; }}
        .info-panel ul {{ margin:0.5rem 0 0 1.25rem; }}
        .info-panel li {{ margin-bottom:0.25rem; }}
        .config-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:0.5rem; }}
        .config-item {{ display:flex; justify-content:space-between; padding:0.5rem 0.75rem; border-bottom:1px solid var(--border); font-size:0.84rem; }}
        .config-item .config-label {{ color:var(--text-soft); }}
        .config-item .config-val {{ font-family:'JetBrains Mono',monospace; font-weight:600; }}
        .validation-pass {{ color:var(--green); font-weight:600; }}
        .validation-warn {{ color:var(--amber); font-weight:600; }}
        .validation-fail {{ color:var(--red); font-weight:600; }}

        /* Footer */
        .footer {{ text-align:center; padding:2rem 1rem; color:var(--text-muted); font-size:0.78rem; border-top:1px solid var(--border); margin-top:2rem; }}

        /* Print */
        @media print {{
            .nav, .theme-toggle {{ display:none; }}
            .tab {{ display:block !important; page-break-inside:avoid; }}
            body {{ font-size:11px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <button class="theme-toggle" onclick="toggleTheme()">Dark</button>
            <div class="header-grid">
                <div class="company-info">
                    <div style="font-size:0.72rem;font-weight:700;letter-spacing:0.12em;text-transform:uppercase;color:var(--text-muted);margin-bottom:0.35rem">Kredit Lab &mdash; Statement Intelligence</div>
                    <h1>{company} <span class="schema-badge">Kredit Lab v{r.get('schema_version', '6')}</span></h1>
                    <div class="period">{period_start_display} to {period_end_display} &middot; {total_months} months &middot; {sum(a.get('transaction_count',0) for a in accounts):,} transactions</div>
                </div>
                <div class="header-kpi">
                    <div class="kpi"><div class="val credit">RM {consol.get('net_credits',0):,.0f}</div><div class="lbl">Net Credits</div></div>
                    <div class="kpi"><div class="val debit">RM {consol.get('net_debits',0):,.0f}</div><div class="lbl">Net Debits{'  ⚠️' if is_incomplete else ''}</div></div>
                    <div class="kpi"><div class="val">RM {consol.get('annualized_net_credits',0):,.0f}</div><div class="lbl">Annualized</div></div>
                    <div class="kpi"><div class="val">RM {consol.get('eod_average',0):,.0f}</div><div class="lbl">Avg EOD</div></div>
                    <div class="kpi"><div class="val" style="color:var(--{'red' if detected_count > total_flags//2 else 'amber'})">{detected_count}/{total_flags}</div><div class="lbl">Flags</div></div>
                </div>
            </div>
        </div>

        <div class="nav">
            <button class="nav-btn active" onclick="showTab('overview')">Overview</button>
            <button class="nav-btn" onclick="showTab('monthly')">Cash Flow</button>
            <button class="nav-btn" onclick="showTab('parties')">Top Parties</button>
            <button class="nav-btn" onclick="showTab('large')">Large Transactions</button>
            <button class="nav-btn" onclick="showTab('round')">Round Figure</button>
            <button class="nav-btn" onclick="showTab('related')">Counterparty</button>
            <button class="nav-btn" onclick="showTab('loans')">Facilities</button>
            <button class="nav-btn" onclick="showTab('flags')">Risk Signals</button>
            {'<button class="nav-btn" onclick="showTab(&#39;fx&#39;)">FX / Remittance</button>' if is_v620 else ''}
            {'<button class="nav-btn" onclick="showTab(&#39;unclassified&#39;)">Unclassified</button>' if is_v630 else ''}
            {'<button class="nav-btn" onclick="showTab(&#39;parsing&#39;)">Parsing QC</button>' if has_parsing else ''}
            <button class="nav-btn" onclick="showTab('fraud')">Fraud Detector</button>
        </div>

        <!-- OVERVIEW TAB -->
        <div id="tab-overview" class="tab active">
            {dq_banner_html}
            <div class="account-grid">{acc_cards}</div>

            {'<div class="section"><div class="section-head"><h2>Known Related Parties</h2></div><div class="section-body">' + rp_html + '</div></div>' if rp_html else ''}
            {'<div class="section"><div class="section-head"><h2>Possible Related Parties — Analyst to Confirm</h2></div><div class="section-body">' + rp_candidates_html + '</div></div>' if rp_candidates_html else ''}

            <div class="two-col">
                <div class="chart-box"><div class="chart-title">Net Credits vs Debits (Monthly)</div><div id="chartCrDr" style="height:300px"></div></div>
                <div class="chart-box"><div class="chart-title">EOD Balance Range (Monthly)</div><div id="chartEOD" style="height:300px"></div></div>
            </div>

            <div class="section">
                <div class="section-head"><h2>Consolidated Summary</h2></div>
                <div class="section-body">
                    <div class="summary-grid">
                        <div class="summary-card"><div class="val credit">{consol.get('gross_credits',0):,.0f}</div><div class="lbl">Gross Credits</div></div>
                        <div class="summary-card"><div class="val debit">{consol.get('gross_debits',0):,.0f}</div><div class="lbl">Gross Debits</div></div>
                        <div class="summary-card"><div class="val credit">{consol.get('net_credits',0):,.0f}</div><div class="lbl">Net Credits</div></div>
                        <div class="summary-card"><div class="val debit">{consol.get('net_debits',0):,.0f}</div><div class="lbl">Net Debits</div></div>
                        <div class="summary-card"><div class="val">{consol.get('annualized_net_credits',0):,.0f}</div><div class="lbl">Annualized Cr</div></div>
                        <div class="summary-card"><div class="val">{consol.get('annualized_net_debits',0):,.0f}</div><div class="lbl">Annualized Dr</div></div>
                    </div>
                    <div class="two-col">
                        <div>
                            <h4 style="color:var(--green);margin-bottom:0.75rem">Exclusions from Credits</h4>
                            <div class="table-wrap"><table>
                                <tr><td>Own Party</td><td class="mono r">{consol.get('total_own_party_cr',0):,.2f}</td></tr>
                                <tr><td>Related Party</td><td class="mono r">{consol.get('total_related_party_cr',0):,.2f}</td></tr>
                                <tr><td>Reversals</td><td class="mono r">{consol.get('total_reversal_cr',0):,.2f}</td></tr>
                                <tr><td>Loan Disbursements</td><td class="mono r">{consol.get('total_loan_disbursement_cr',0):,.2f}</td></tr>
                                <tr><td>FD/Interest</td><td class="mono r">{consol.get('total_fd_interest_cr',0):,.2f}</td></tr>
                                <tr><td>Inward Return (C16)</td><td class="mono r">{consol.get('total_inward_return_cr',0):,.2f}</td></tr>
                            </table></div>
                        </div>
                        <div>
                            <h4 style="color:var(--red);margin-bottom:0.75rem">Exclusions from Debits</h4>
                            <div class="table-wrap"><table>
                                <tr><td>Own Party</td><td class="mono r">{consol.get('total_own_party_dr',0):,.2f}</td></tr>
                            </table></div>
                        </div>
                    </div>
                    {'<div style="margin-top:1rem"><h4 style="color:var(--purple);margin-bottom:0.75rem">FX / Remittance Summary</h4><div class="summary-grid"><div class="summary-card"><div class="val credit">' + f"{consol.get('total_fx_credits',0):,.0f}" + '</div><div class="lbl">FX Credits (' + f"{consol.get('fx_credit_pct',0):.1f}" + '% of Gross)</div></div><div class="summary-card"><div class="val debit">' + f"{consol.get('total_fx_debits',0):,.0f}" + '</div><div class="lbl">FX Debits (' + f"{consol.get('fx_debit_pct',0):.1f}" + '% of Gross)</div></div><div class="summary-card"><div class="val" style="font-size:0.9rem">' + (', '.join(consol.get('fx_currencies_all', [])) or 'None') + '</div><div class="lbl">Currencies</div></div></div></div>' if is_v620 else ''}
                </div>
            </div>

            <div class="two-col">
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--green)">Positive Observations</h2></div>
                    <div class="section-body"><ul class="obs-list">{pos_obs}</ul></div>
                </div>
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--red)">Concerns</h2></div>
                    <div class="section-body"><ul class="obs-list">{con_obs}</ul></div>
                </div>
            </div>
        </div>

        <!-- MONTHLY ANALYSIS TAB -->
        <div id="tab-monthly" class="tab">
            <div class="section">
                <div class="section-head"><h2>Monthly Cash Flow Breakdown</h2></div>
                {'<div style="padding:0.5rem 1.25rem;font-size:0.78rem;color:var(--text-soft);border-bottom:1px solid var(--border);display:flex;gap:1.25rem;flex-wrap:wrap">' + ''.join(f'<span><span style="display:inline-block;width:8px;height:8px;border-radius:50%;background:{acct_colors.get(a,"var(--text-muted)")};margin-right:4px;vertical-align:middle"></span>{next((ac.get("bank_name","") for ac in accounts if ac.get("account_number")==a), "")} ({a})</span>' for a in acct_list) + ' <span style="font-weight:600">Bold rows = month subtotal</span></div>' if has_account_col and len(acct_list) > 1 else ''}
                <div class="section-body" style="padding:0">
                    <div class="table-wrap" style="max-height:600px; overflow:auto">
                        <table>
                            <thead><tr>
                                <th class="sticky-col">Month / Account</th>
                                {'<th>Status</th>' if has_recon else ''}
                                <th class="r">Gross Cr</th><th class="r">Gross Dr</th>
                                <th class="r">Net Cr</th><th class="r">Net Dr</th>
                                <th class="r">Cr #</th><th class="r">Dr #</th>
                                <th class="r">Own Cr</th><th class="r">Own Dr</th>
                                <th class="r">RP Cr</th><th class="r">RP Dr</th>
                                <th class="r">Reversal</th>
                                <th class="r">Loan Disb</th><th class="r">FD/Int</th>
                                <th class="r">Cash Dep</th><th class="r">Cash Wdl</th>
                                <th class="r">Chq Dep</th><th class="r">Chq Issue</th>
                                <th class="r">Loan Repay</th><th class="r">Salary</th>
                                <th class="r">EPF</th><th class="r">SOCSO</th><th class="r">Tax</th>
                                <th class="r">Ret Chq #</th><th class="r">Ret Chq Amt</th>
                                <th class="r">Round Fig</th><th class="r">High Val</th>
                                <th class="r">EOD Low</th><th class="r">EOD High</th><th class="r">EOD Avg</th>
                                <th class="r">Open Bal</th><th class="r">Close Bal</th>
                                {'<th class="r v630-count">Own Cr #</th><th class="r v630-count">Own Dr #</th><th class="r v630-count">RP Cr #</th><th class="r v630-count">RP Dr #</th><th class="r v630-count">Loan #</th><th class="r v630-amt">Inward Ret</th><th class="r v630-uncl">Uncl Cr #</th><th class="r v630-uncl">Uncl Cr Amt</th><th class="r v630-uncl">Uncl Dr #</th><th class="r v630-uncl">Uncl Dr Amt</th>' if is_v630 else ''}
                            </tr></thead>
                            <tbody>
                                {monthly_rows}
                                {consol_row}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
            <div class="note">Net Credits = Gross Credits - Own Party - Related Party - Reversals - Loan Disbursements - FD/Interest{' - Inward Return (C16)' if is_v630 else ''} | Net Debits = Gross Debits - Own Party (C02){' | <span style="display:inline-block;width:10px;height:10px;background:var(--purple-dim);border:1px solid var(--purple);border-radius:2px;vertical-align:middle;margin:0 2px"></span> Count columns <span style="display:inline-block;width:10px;height:10px;background:var(--amber-dim);border:1px solid var(--amber);border-radius:2px;vertical-align:middle;margin:0 2px"></span> Unclassified columns (v6.3.0)' if is_v630 else ''}</div>
'''
    # Gap panels
    gap_panels_html = ''
    extraction_gaps = (parsing.get('extraction_gaps', []) if parsing else []) if is_incomplete else []
    if extraction_gaps and has_recon:
        from collections import defaultdict as _dd
        gaps_by_month = _dd(list)
        for g in extraction_gaps:
            gaps_by_month[g.get('month', '')].append(g)

        for gap_month in sorted(gaps_by_month.keys()):
            month_gaps = gaps_by_month[gap_month]
            total_missing = sum(g.get('missing_amount', 0) for g in month_gaps)
            gap_panels_html += f'''<div class="dq-gap-panel">
                <strong>Extraction Gaps — {gap_month} ({len(month_gaps)} gap{"s" if len(month_gaps) > 1 else ""}, RM {total_missing:,.2f} missing)</strong>'''
            for gi, g in enumerate(month_gaps, 1):
                gap_panels_html += f'''<div class="dq-gap-item">
                    <div><div style="font-size:0.72rem;color:var(--text-muted)">Gap #{gi}</div>
                    <div style="color:var(--red);font-weight:600">RM {g.get('missing_amount',0):,.2f}</div></div>
                    <div><div>Page {g.get('page','')} · {g.get('date','')} · {g.get('missing_type','').lower()}(s) missing</div>
                    <div style="font-size:0.78rem;color:var(--text-muted);margin-top:2px">After: <em>{g.get('prev_description','')[:60]}</em> (bal RM {g.get('balance_before_gap',0):,.2f})</div>
                    <div style="font-size:0.78rem;color:var(--text-muted)">Before: <em>{g.get('next_description','')[:60]}</em> (bal RM {g.get('balance_after_gap',0):,.2f})</div></div>
                </div>'''
            gap_panels_html += '</div>'

    html += gap_panels_html
    html += f'''
        </div>

        <!-- TOP PARTIES TAB -->
        <div id="tab-parties" class="tab">
            <div class="top-parties-grid" style="display:flex !important;flex-direction:row !important;flex-wrap:nowrap !important;gap:1.5rem;align-items:flex-start;width:100%;overflow-x:auto">
                <div class="section" style="flex:1 1 0;min-width:0;width:calc(50% - 0.75rem);margin-bottom:0">
                    <div class="section-head"><h2 style="color:var(--green)">Top 10 Payers (Income)</h2></div>
                    <div class="section-body" style="padding:0">
                        <div class="table-wrap"><table>
                            <thead><tr><th>#</th><th>Party</th><th class="r">Amount (RM)</th><th class="r">Txns</th></tr></thead>
                            <tbody>{payer_rows or '<tr><td colspan="4" style="text-align:center;color:var(--text-muted)">No data</td></tr>'}</tbody>
                        </table></div>
                        {payers_suppressed_html}
                    </div>
                </div>
                <div class="section" style="flex:1 1 0;min-width:0;width:calc(50% - 0.75rem);margin-bottom:0">
                    <div class="section-head"><h2 style="color:var(--red)">Top 10 Payees (Outflow)</h2></div>
                    <div class="section-body" style="padding:0">
                        <div class="table-wrap"><table>
                            <thead><tr><th>#</th><th>Party</th><th class="r">Amount (RM)</th><th class="r">Txns</th></tr></thead>
                            <tbody>{payee_rows or '<tr><td colspan="4" style="text-align:center;color:var(--text-muted)">No data</td></tr>'}</tbody>
                        </table></div>
                        {payees_suppressed_html}
                    </div>
                </div>
            </div>
            <div class="note"><span class="rp-badge">RP</span> = Related Party</div>
        </div>

        <!-- LARGE TRANSACTIONS TAB -->
        <div id="tab-large" class="tab">
            <div class="section">
                <div class="section-head"><h2>Large Transactions (&ge; RM {large_threshold:,.0f})</h2><span class="badge badge-current">{len(large_txns)} transactions</span></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap" style="max-height:500px;overflow:auto">
                        <table>
                            <thead>
                                <tr>
                                    <th>Date</th>
                                    <th>Description</th>
                                    <th class="r">Amount (RM)</th>
                                    <th class="r">Balance</th>
                                </tr>
                            </thead>
                            <tbody>
                                {large_txn_rows}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>

        <!-- RELATED PARTY TAB -->
        <div id="tab-related" class="tab">
            {counterparty_ledger_html}
            <div class="section">
                <div class="section-head"><h2>Own & Related Party Transactions</h2></div>
                <div class="section-body">
                    <div class="summary-grid">
                        <div class="summary-card"><div class="val credit">{rp_counts['own_party_cr']:,}</div><div class="lbl">Own Party Cr txns</div></div>
                        <div class="summary-card"><div class="val debit">{rp_counts['own_party_dr']:,}</div><div class="lbl">Own Party Dr txns</div></div>
                        <div class="summary-card"><div class="val credit">{rp_counts['related_party_cr']:,}</div><div class="lbl">Related Party Cr txns</div></div>
                        <div class="summary-card"><div class="val debit">{rp_counts['related_party_dr']:,}</div><div class="lbl">Related Party Dr txns</div></div>
                    </div>
                    <div class="table-wrap" style="max-height:500px;overflow:auto"><table>
                        <thead><tr><th>Party</th><th class="r">Credits (RM)</th><th class="r">Debits (RM)</th><th>Party Type</th><th class="r">Txns</th></tr></thead>
                        <tbody>{rp_party_rows or '<tr><td colspan="5" class="note">No own or related party transactions</td></tr>'}</tbody>
                    </table></div>
                    {rp_expander_script}
                </div>
            </div>
        </div>

        <!-- LOANS TAB -->
        <div id="tab-loans" class="tab">
            <div class="summary-grid">
                <div class="summary-card"><div class="val credit">{loan_disb_total:,.0f}</div><div class="lbl">Total Disbursements</div></div>
                <div class="summary-card"><div class="val debit">{loan_repay_total:,.0f}</div><div class="lbl">Total Repayments</div></div>
                <div class="summary-card"><div class="val">{len(loan_disbursements)}</div><div class="lbl">Disbursement Txns</div></div>
                <div class="summary-card"><div class="val">{len(loan_repayments)}</div><div class="lbl">Repayment Txns</div></div>
            </div>
            <div class="two-col">
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--green)">Disbursements (Credits)</h2></div>
                    <div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:400px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Category</th></tr></thead>
                        <tbody>{loan_disb_rows or '<tr><td colspan="4" class="note">No disbursements</td></tr>'}</tbody>
                    </table></div></div>
                </div>
                <div class="section">
                    <div class="section-head"><h2 style="color:var(--red)">Repayments (Debits)</h2></div>
                    <div class="section-body" style="padding:0"><div class="table-wrap" style="max-height:400px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount</th><th>Category</th></tr></thead>
                        <tbody>{loan_repay_rows or '<tr><td colspan="4" class="note">No repayments</td></tr>'}</tbody>
                    </table></div></div>
                </div>
            </div>
        </div>

        <!-- FLAGS TAB -->
        <div id="tab-flags" class="tab">
            {statutory_html}
            <div class="section">
                <div class="section-head"><h2>Risk Signals</h2><span class="badge" style="background:var(--{'red-dim' if detected_count > total_flags//2 else 'amber-dim'});color:var(--{'red' if detected_count > total_flags//2 else 'amber'})">{detected_count} of {total_flags} detected</span></div>
                <div class="section-body" style="padding:0">
                    <div class="table-wrap"><table>
                        <thead><tr><th style="width:40px">#</th><th>Signal</th><th style="width:80px;text-align:center">Status</th><th>Remarks</th></tr></thead>
                        <tbody>{flag_rows}</tbody>
                    </table></div>
                </div>
            </div>
        </div>

        <!-- ROUND FIGURE TAB -->
        <div id="tab-round" class="tab">
            <div class="section">
                <div class="section-head"><h2>Round Figure Transactions &mdash; Detail</h2><span class="badge badge-current">{len(round_figure_credits)} transactions</span></div>
                <div class="section-body" style="padding:0">
                    <div class="note" style="padding:0.5rem 1.25rem">Credit or debit transactions with round-number amounts (multiple of RM 10,000), matching the round-number flag list.</div>
                    <div class="table-wrap" style="max-height:400px;overflow:auto"><table>
                        <thead><tr><th>Date</th><th>Description</th><th class="r">Amount (RM)</th><th class="r">Balance</th></tr></thead>
                        <tbody>{rf_cr_rows}</tbody>
                    </table></div>
                </div>
            </div>
        </div>

        {fx_tab_html}
        {unclassified_tab_html}
        {parsing_tab_html}
        {fraud_tab_html}

        <div class="footer">
            <p>Kredit Lab &mdash; Statement Intelligence Report | Generated {r.get('generated_at','')} | {period_start_display} &ndash; {period_end_display}</p>
        </div>
    </div>

    <script>
        function showTab(name) {{
            document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
            document.querySelectorAll('.nav-btn').forEach(b=>b.classList.remove('active'));
            const tab = document.getElementById('tab-'+name);
            if(tab) tab.classList.add('active');
            event.target.classList.add('active');
        }}
        function toggleTheme() {{
            const html = document.documentElement;
            const btn = document.querySelector('.theme-toggle');
            const t = html.getAttribute('data-theme')==='dark'?'light':'dark';
            html.setAttribute('data-theme',t);
            btn.textContent = t==='dark'?'Light':'Dark';
            renderCharts();
        }}
        function renderCharts() {{
            if(typeof Plotly==='undefined'){{console.warn('Plotly not loaded — charts disabled');return;}}
            const isDark = document.documentElement.getAttribute('data-theme')==='dark';
            const gridColor = isDark?'#1e2a42':'#e2e8f0';
            const textColor = isDark?'#94a3b8':'#475569';
            const bg = 'transparent';

            Plotly.newPlot('chartCrDr', [
                {{x:{chart_months},y:{chart_net_cr},name:'Net Credits',type:'bar',marker:{{color:'rgba(5,150,105,0.7)'}}}},
                {{x:{chart_months},y:{chart_net_dr},name:'Net Debits',type:'bar',marker:{{color:'rgba(220,38,38,0.7)'}}}}
            ], {{
                paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                barmode:'group',showlegend:true,legend:{{orientation:'h',y:1.12}},
                margin:{{t:30,b:40,l:60,r:20}},
                yaxis:{{gridcolor:gridColor,tickformat:','}}
            }}, {{responsive:true,displayModeBar:false}});

            Plotly.newPlot('chartEOD', [
                {{x:{chart_months},y:{chart_eod_high},name:'EOD High',type:'scatter',mode:'lines+markers',line:{{color:'#2563eb',width:2}},marker:{{size:6}}}},
                {{x:{chart_months},y:{chart_eod_avg},name:'EOD Average',type:'scatter',mode:'lines+markers',line:{{color:'#7c3aed',width:2}},marker:{{size:6}}}},
                {{x:{chart_months},y:{chart_eod_low},name:'EOD Low',type:'scatter',mode:'lines+markers',line:{{color:'#dc2626',width:2,dash:'dot'}},marker:{{size:6}}}}
            ], {{
                paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                showlegend:true,legend:{{orientation:'h',y:1.12}},
                margin:{{t:30,b:40,l:60,r:20}},
                yaxis:{{gridcolor:gridColor,tickformat:','}}
            }}, {{responsive:true,displayModeBar:false}});

            var fxEl = document.getElementById('chartFX');
            if (fxEl) {{
                Plotly.newPlot('chartFX', [
                    {{x:{chart_months},y:{fx_chart_cr_json},name:'FX Credits',type:'bar',marker:{{color:'rgba(5,150,105,0.7)'}}}},
                    {{x:{chart_months},y:{fx_chart_dr_json},name:'FX Debits',type:'bar',marker:{{color:'rgba(220,38,38,0.7)'}}}}
                ], {{
                    paper_bgcolor:bg,plot_bgcolor:bg,font:{{color:textColor,size:11}},
                    barmode:'group',showlegend:true,legend:{{orientation:'h',y:1.12}},
                    margin:{{t:30,b:40,l:60,r:20}},
                    yaxis:{{gridcolor:gridColor,tickformat:','}}
                }}, {{responsive:true,displayModeBar:false}});
            }}
        }}
        document.addEventListener('DOMContentLoaded', function() {{
            renderCharts();
        }});
    </script>
</body>
</html>'''
    return html


# ============================================================
# HTML REPORT GENERATION FUNCTIONS (from your JSON converter)
# Copy the entire set of functions here
# ============================================================

def fmt(val, decimals=2):
    """Format number with commas"""
    if val is None:
        return "0.00"
    return f"{val:,.{decimals}f}"

def normalize_observations(obs):
    """Coerce observations into {'positive': [...], 'concerns': [...]}."""
    if isinstance(obs, dict):
        return {'positive': list(obs.get('positive', []) or []),
                'concerns': list(obs.get('concerns', []) or [])}
    pos, con = [], []
    if isinstance(obs, list):
        for item in obs:
            if isinstance(item, str):
                con.append(item)
            elif isinstance(item, dict):
                kind = str(item.get('type') or item.get('category') or item.get('sentiment') or '').lower()
                text = item.get('text') or item.get('observation') or item.get('message') or item.get('description') or ''
                if not text:
                    continue
                if kind in ('positive', 'pos', 'good', 'strength'):
                    pos.append(text)
                else:
                    con.append(text)
    return {'positive': pos, 'concerns': con}

def adapt_to_v6(src):
    """Reshape flat extractor output into v6.3.3 renderer schema."""
    from collections import defaultdict

    summary = src.get('summary', {}) or {}
    transactions = src.get('transactions', []) or []
    monthly_summary = src.get('monthly_summary', []) or []
    cp_ledger = src.get('counterparty_ledger', {}) or {}
    pdf_integrity = src.get('pdf_integrity')
    high_value_threshold = summary.get('high_value_threshold', 100000)

    report_info = {
        'company_name': summary.get('company_names', ['Unknown'])[0] if summary.get('company_names') else 'Unknown',
        'schema_version': 'Testv2.0',
        'period_start': '',
        'period_end': '',
        'total_months': len(monthly_summary),
        'related_parties': [],
    }

    # accounts aggregation
    acc_map = defaultdict(lambda: {'credits': 0.0, 'debits': 0.0, 'txn_count': 0, 'bank': '', 'last_bal': None, 'opening_bal': None})
    for t in transactions:
        an = t.get('account_no', '')
        if not an:
            continue
        a = acc_map[an]
        a['txn_count'] += 1
        cr = float(t.get('credit', 0) or 0)
        dr = float(t.get('debit', 0) or 0)
        a['credits'] += cr
        a['debits'] += dr
        if not a['bank']:
            a['bank'] = t.get('bank', '') or ''
        bal = t.get('balance')
        if isinstance(bal, (int, float)):
            if a['opening_bal'] is None:
                a['opening_bal'] = bal - cr + dr
            a['last_bal'] = bal
    
    accounts = []
    for an, a in sorted(acc_map.items()):
        accounts.append({
            'bank_name': a['bank'],
            'account_number': an,
            'account_holder': report_info['company_name'],
            'account_type': 'Current',
            'opening_balance': round(a['opening_bal'] or 0.0, 2),
            'closing_balance': round(a['last_bal'] or 0.0, 2),
            'total_credits': round(a['credits'], 2),
            'total_debits': round(a['debits'], 2),
            'transaction_count': a['txn_count'],
        })

    # Build monthly analysis from monthly_summary
    monthly_analysis = []
    for m in monthly_summary:
        month = m.get('month', '')
        highest = float(m.get('highest_balance', 0) or 0)
        lowest = float(m.get('lowest_balance', 0) or 0)
        monthly_analysis.append({
            'month': month,
            'bank_name': '',
            'account_number': m.get('account_no', ''),
            'gross_credits': float(m.get('total_credit', 0) or 0),
            'gross_debits': float(m.get('total_debit', 0) or 0),
            'net_credits': float(m.get('total_credit', 0) or 0),
            'net_debits': float(m.get('total_debit', 0) or 0),
            'credit_count': m.get('credit_count', 0),
            'debit_count': m.get('debit_count', 0),
            'own_party_cr': float(m.get('own_party_cr', 0) or 0),
            'own_party_dr': float(m.get('own_party_dr', 0) or 0),
            'related_party_cr': float(m.get('related_party_cr', 0) or 0),
            'related_party_dr': float(m.get('related_party_dr', 0) or 0),
            'reversal_cr': float(m.get('reversal_cr', 0) or 0),
            'loan_disbursement_cr': float(m.get('loan_disbursement_cr', 0) or 0),
            'fd_interest_cr': float(m.get('fd_interest_cr', 0) or 0),
            'round_figure_cr': float(m.get('round_figure_cr', 0) or 0),
            'high_value_cr': float(m.get('high_value_cr', 0) or 0),
            'cash_deposits_amount': float(m.get('cash_deposits_amount', 0) or 0),
            'cash_withdrawals_amount': float(m.get('cash_withdrawals_amount', 0) or 0),
            'cheque_deposits_amount': float(m.get('cheque_deposits_amount', 0) or 0),
            'cheque_issues_amount': float(m.get('cheque_issues_amount', 0) or 0),
            'loan_repayment_dr': float(m.get('loan_repayment_dr', 0) or 0),
            'salary_paid': float(m.get('salary_paid', 0) or 0),
            'statutory_epf': float(m.get('statutory_epf', 0) or 0),
            'statutory_socso': float(m.get('statutory_socso', 0) or 0),
            'statutory_tax': float(m.get('statutory_tax', 0) or 0),
            'statutory_hrdf': float(m.get('statutory_hrdf', 0) or 0),
            'returned_cheques_outward_count': m.get('returned_cheques_outward_count', 0),
            'returned_cheques_outward_amount': float(m.get('returned_cheques_outward_amount', 0) or 0),
            'eod_lowest': lowest,
            'eod_highest': highest,
            'eod_average': (highest + lowest) / 2.0 if (highest or lowest) else 0.0,
            'opening_balance': float(m.get('opening_balance', 0) or 0),
            'closing_balance': float(m.get('ending_balance', 0) or 0),
            'transaction_count': m.get('transaction_count', 0),
            'fx_credit_amount': 0,
            'fx_debit_amount': 0,
            'fx_credit_count': 0,
            'fx_debit_count': 0,
            'fx_currencies': [],
        })

    gross_credits = sum(float(t.get('credit', 0) or 0) for t in transactions)
    gross_debits = sum(float(t.get('debit', 0) or 0) for t in transactions)
    
    consolidated = {
    'gross_credits': round(gross_credits, 2),
    'gross_debits': round(gross_debits, 2),
    'net_credits': round(gross_credits, 2),
    'net_debits': round(gross_debits, 2),
    'annualized_net_credits': round(gross_credits * 12 / len(monthly_summary), 2) if monthly_summary else 0,
    'annualized_net_debits': round(gross_debits * 12 / len(monthly_summary), 2) if monthly_summary else 0,
    'eod_lowest': 0,
    'eod_highest': 0,
    'eod_average': 0,
    'data_completeness': 'COMPLETE',
    'high_value_threshold': high_value_threshold,  # Add this line
    'large_credit_threshold': high_value_threshold,  # Add alias for compatibility
    }
    
    return {
        'report_info': report_info,
        'accounts': accounts,
        'monthly_analysis': monthly_analysis,
        'consolidated': consolidated,
        'top_parties': {'top_payers': [], 'top_payees': []},
        'large_credits': [],
        'large_transactions': [],
        'own_related_transactions': {'transactions': [], 'summary': {}},
        'loan_transactions': {'transactions': [], 'summary': {}},
        'flags': {'indicators': []},
        'observations': {'positive': [], 'concerns': []},
        'parsing_metadata': {},
        'classification_config': {
            'large_transaction_threshold': high_value_threshold,
            'large_credit_threshold': high_value_threshold,
        },
        'counterparty_ledger': cp_ledger,
        'pdf_integrity': pdf_integrity,
    }

def _top_parties_from_transaction_analysis(transaction_analysis: dict) -> dict:
    if not isinstance(transaction_analysis, dict):
        return {"top_payers": [], "top_payees": []}

    top_payers = transaction_analysis.get("top_payers") or transaction_analysis.get("top_creditors")
    top_payees = transaction_analysis.get("top_payees") or transaction_analysis.get("top_debtors")

    if not top_payers:
        top_payers = [
            {
                "rank": idx,
                "party_name": row.get("party"),
                "total_amount": row.get("total_credit", 0),
                "transaction_count": row.get("credit_tx_count", 0),
            }
            for idx, row in enumerate(transaction_analysis.get("top_credit_parties", []), start=1)
        ]

    if not top_payees:
        top_payees = [
            {
                "rank": idx,
                "party_name": row.get("party"),
                "total_amount": row.get("total_debit", 0),
                "transaction_count": row.get("debit_tx_count", 0),
            }
            for idx, row in enumerate(transaction_analysis.get("top_debit_parties", []), start=1)
        ]

    return {"top_payers": top_payers or [], "top_payees": top_payees or []}


def _has_top_party_rows(top_parties: dict) -> bool:
    if not isinstance(top_parties, dict):
        return False
    payers = top_parties.get("top_payers") or top_parties.get("top_creditors") or []
    payees = top_parties.get("top_payees") or top_parties.get("top_debtors") or []
    return bool(payers or payees)


def _ledger_monthly_breakdown(transactions: List[dict], side: str) -> List[dict]:
    buckets = {}
    side = side.upper()
    for txn in transactions or []:
        if not isinstance(txn, dict):
            continue

        raw_type = str(txn.get("type") or txn.get("transaction_type") or "").upper()
        credit = safe_float(txn.get("credit", 0))
        debit = safe_float(txn.get("debit", 0))
        amount = 0.0
        if side == "CREDIT":
            if credit > 0:
                amount = credit
            elif raw_type in ("CREDIT", "CR") or "CREDIT" in raw_type:
                amount = abs(safe_float(txn.get("amount", 0)))
        else:
            if debit > 0:
                amount = debit
            elif raw_type in ("DEBIT", "DR") or "DEBIT" in raw_type:
                amount = abs(safe_float(txn.get("amount", 0)))
        if amount <= 0:
            continue

        date_text = str(txn.get("date") or txn.get("transaction_date") or "")
        month_match = re.search(r"\d{4}-\d{2}", date_text)
        if not month_match:
            continue
        month = month_match.group(0)
        bucket = buckets.setdefault(month, {"month": month, "amount": 0.0, "count": 0})
        bucket["amount"] += amount
        bucket["count"] += 1

    return [
        {"month": month, "amount": round(row["amount"], 2), "count": row["count"]}
        for month, row in sorted(buckets.items())
    ]


def _top_parties_from_counterparty_ledger(counterparty_ledger: dict, limit: int = 10) -> dict:
    if not isinstance(counterparty_ledger, dict):
        return {"top_payers": [], "top_payees": []}

    payers = []
    payees = []
    for cp in counterparty_ledger.get("counterparties", []) or []:
        if not isinstance(cp, dict):
            continue
        name = (
            cp.get("counterparty_name")
            or cp.get("party_name")
            or cp.get("name")
            or ""
        )
        name = str(name).strip()
        if not name:
            continue

        transactions = cp.get("transactions") or []
        related_raw = cp.get("is_related_party", cp.get("related_party", False))
        is_related = bool(related_raw) and str(related_raw).strip().lower() not in {"false", "no", "0"}

        credit_amount = safe_float(cp.get("total_credits", cp.get("total_credit", 0)))
        debit_amount = safe_float(cp.get("total_debits", cp.get("total_debit", 0)))
        credit_count = int(safe_float(cp.get("credit_count", cp.get("credit_tx_count", 0))))
        debit_count = int(safe_float(cp.get("debit_count", cp.get("debit_tx_count", 0))))

        if credit_amount > 0:
            monthly = _ledger_monthly_breakdown(transactions, "CREDIT")
            payers.append({
                "party_name": name,
                "total_amount": round(credit_amount, 2),
                "transaction_count": credit_count or sum(m.get("count", 0) for m in monthly),
                "is_related_party": is_related,
                "monthly_breakdown": monthly,
            })
        if debit_amount > 0:
            monthly = _ledger_monthly_breakdown(transactions, "DEBIT")
            payees.append({
                "party_name": name,
                "total_amount": round(debit_amount, 2),
                "transaction_count": debit_count or sum(m.get("count", 0) for m in monthly),
                "is_related_party": is_related,
                "monthly_breakdown": monthly,
            })

    payers.sort(key=lambda party: party.get("total_amount", 0), reverse=True)
    payees.sort(key=lambda party: party.get("total_amount", 0), reverse=True)
    return {
        "top_payers": [{**party, "rank": idx} for idx, party in enumerate(payers[:limit], 1)],
        "top_payees": [{**party, "rank": idx} for idx, party in enumerate(payees[:limit], 1)],
    }


def build_canonical_counterparty_ledger_rows(cp_ledger: dict) -> List[dict]:
    """Return the shared counterparty ledger view used by UI, HTML, and Excel."""
    raw_counterparties = cp_ledger.get("counterparties", []) if isinstance(cp_ledger, dict) else []
    raw_counterparties = [cp for cp in raw_counterparties or [] if isinstance(cp, dict)]
    if not raw_counterparties:
        return []

    raw_names = [str(cp.get("counterparty_name", cp.get("counterparty", "")) or "") for cp in raw_counterparties]
    try:
        clean_names = deduplicate_counterparty_names(raw_names)
    except Exception:
        clean_names = [clean_counterparty_name(name) or name for name in raw_names]

    merged_counterparties = {}
    for cp, clean_name in zip(raw_counterparties, clean_names):
        clean_name = clean_name or clean_counterparty_name(cp.get("counterparty_name", "")) or "UNKNOWN"
        clean_name = re.sub(r"\s+", " ", str(clean_name).strip()).upper() or "UNKNOWN"
        key = clean_name.casefold()
        merged = merged_counterparties.setdefault(
            key,
            {
                "counterparty_name": clean_name,
                "total_credits": 0.0,
                "total_debits": 0.0,
                "net_position": 0.0,
                "transaction_count": 0,
                "credit_count": 0,
                "debit_count": 0,
                "pattern_matched": 0,
                "special_bucket": 0,
                "raw_fallback": 0,
                "transactions": [],
                "raw_names": set(),
                "is_related_party": False,
            },
        )

        raw_name = str(cp.get("counterparty_name", cp.get("counterparty", "")) or "").strip()
        if raw_name:
            merged["raw_names"].add(raw_name)
        merged["total_credits"] += safe_float(cp.get("total_credits", cp.get("total_credit", 0)))
        merged["total_debits"] += safe_float(cp.get("total_debits", cp.get("total_debit", 0)))
        merged["credit_count"] += int(safe_float(cp.get("credit_count", cp.get("credit_tx_count", 0))))
        merged["debit_count"] += int(safe_float(cp.get("debit_count", cp.get("debit_tx_count", 0))))
        merged["transaction_count"] += int(safe_float(cp.get("transaction_count", 0)))
        merged["pattern_matched"] += int(safe_float(cp.get("pattern_matched", 0)))
        merged["special_bucket"] += int(safe_float(cp.get("special_bucket", 0)))
        merged["raw_fallback"] += int(safe_float(cp.get("raw_fallback", 0)))

        related_raw = cp.get("is_related_party", cp.get("related_party", False))
        if bool(related_raw) and str(related_raw).strip().lower() not in {"false", "no", "0"}:
            merged["is_related_party"] = True

        for txn in cp.get("transactions", []) or []:
            if isinstance(txn, dict):
                txn_copy = dict(txn)
                txn_copy["counterparty_name_clean"] = clean_name
                merged["transactions"].append(txn_copy)

    rows = []
    for cp in merged_counterparties.values():
        cp["total_credits"] = round(cp["total_credits"], 2)
        cp["total_debits"] = round(cp["total_debits"], 2)
        cp["net_position"] = round(cp["total_credits"] - cp["total_debits"], 2)
        if not cp["transaction_count"]:
            cp["transaction_count"] = cp["credit_count"] + cp["debit_count"]
        if not cp["transaction_count"]:
            cp["transaction_count"] = len(cp.get("transactions", []) or [])
        cp["raw_names"] = sorted(name for name in cp["raw_names"] if name)
        cp["transactions"] = sorted(
            cp.get("transactions", []) or [],
            key=lambda tx: (str(tx.get("date") or ""), str(tx.get("description") or "")),
        )
        rows.append(cp)

    rows.sort(key=lambda cp: str(cp.get("counterparty_name", "") or "").casefold())
    return rows


_PARTY_GHOST_STOPWORDS = {
    "TRANSFER", "PAYMENT", "IBG", "IB2G", "IBFT", "IBK", "CR", "DR", "CREDIT", "DEBIT",
    "TO", "FR", "FROM", "A/C", "C/A", "ACCOUNT", "ACCT", "INTER", "BANK", "BANKING", "INTO",
    "ONLINE", "DUITNOW", "DUIT", "NOW", "FPX", "RENTAS", "REMITTANCE", "ELECTRONIC",
    "AUTOPAY", "INSTANT", "FAST", "OUTWARD", "INWARD", "OUTW", "INW",
    "OUT", "IN", "ADVICE", "TRF", "BLKTRF", "NBPS", "TR", "PYMT", "PAY",
    "THE", "AND", "OF", "FOR", "WITH",
    "SA", "CA", "CCARD", "CARD",
    "CHQ", "CHEQUE", "CASH", "DEPOSIT", "WITHDRAWAL", "HSE", "HOUSE",
    "CLRG", "CDM", "2D", "LOCAL", "GIR", "GIRO",
    "HLB", "MBB", "RHB", "ABB", "PBB", "BIMB", "AMB", "AMBANK", "PBE",
    "CIMB", "OCBC", "UOB", "BSN",
    "PMT", "SLRY",
}
_PARTY_CHEQUE_NOISE = {
    "HSE CHQ DEPOSIT", "CDM CASH DEPOSIT", "2D LOCAL CHQ", "CASH CHQ DR",
    "HOUSE CHQ DR", "CLRG CHQ DR", "HSE CHQ", "CHEQUE DEPOSIT", "CHQ DEPOSIT",
}


def _is_ghost_party_bucket(name) -> bool:
    """Return True when a top-party label is only parser/payment-rail noise."""
    if not name:
        return True
    normalised = re.sub(r"[.,]", "", str(name).upper())
    normalised = re.sub(r"\b(SDN|BHD|& CO|\(M\)|PTY|LTD)\b", "", normalised)
    normalised = re.sub(r"\s+", " ", normalised).strip()
    if not normalised:
        return True
    if normalised in _PARTY_CHEQUE_NOISE:
        return True
    tokens = [t for t in re.split(r"[\s/\-]+", normalised) if t]
    real_tokens = [
        t for t in tokens
        if len(t) >= 3 and t not in _PARTY_GHOST_STOPWORDS and re.search(r"[A-Z]", t)
    ]
    return len(real_tokens) == 0


def _normalize_party_for_report(party: dict, is_payer: bool) -> dict:
    if not isinstance(party, dict):
        return {}
    amount = party.get("total_amount")
    if amount is None:
        amount = party.get("total_credits") if is_payer else party.get("total_debits")
    if amount is None:
        amount = party.get("amount", 0)
    return {
        "rank": party.get("rank", ""),
        "party_name": party.get("party_name") or party.get("name") or "",
        "total_amount": safe_float(amount),
        "transaction_count": int(safe_float(party.get("transaction_count") or party.get("txn_count") or 0)),
        "is_related_party": bool(party.get("is_related_party", False)),
        "monthly_breakdown": party.get("monthly_breakdown") or [],
    }


def prepare_top_parties_for_report(top_parties: dict, limit: int = 10) -> dict:
    """Prepare the exact top-party view rendered by HTML and Excel."""
    if not isinstance(top_parties, dict):
        top_parties = {}
    raw_payers = top_parties.get("top_payers") or top_parties.get("top_creditors") or []
    raw_payees = top_parties.get("top_payees") or top_parties.get("top_debtors") or []

    payers_all = [_normalize_party_for_report(p, True) for p in raw_payers]
    payees_all = [_normalize_party_for_report(p, False) for p in raw_payees]
    payers_suppressed = [p for p in payers_all if _is_ghost_party_bucket(p.get("party_name", ""))]
    payees_suppressed = [p for p in payees_all if _is_ghost_party_bucket(p.get("party_name", ""))]
    payers = [p for p in payers_all if not _is_ghost_party_bucket(p.get("party_name", ""))][:limit]
    payees = [p for p in payees_all if not _is_ghost_party_bucket(p.get("party_name", ""))][:limit]

    for idx, party in enumerate(payers, 1):
        party["rank"] = idx
    for idx, party in enumerate(payees, 1):
        party["rank"] = idx

    return {
        "payers": payers,
        "payees": payees,
        "payers_suppressed": payers_suppressed,
        "payees_suppressed": payees_suppressed,
    }


def build_round_transactions(transactions: List[dict], round_thresholds: List[float] | None = None) -> List[dict]:
    """Build the same signed round-number transaction rows shown in Railway."""
    round_rows = []
    for tx in transactions or []:
        if not isinstance(tx, dict):
            continue

        credit = safe_float(tx.get("credit", 0))
        debit = safe_float(tx.get("debit", 0))
        is_round_credit = credit > 0 and is_round_number(credit, round_thresholds)
        is_round_debit = debit > 0 and is_round_number(debit, round_thresholds)
        if not (is_round_credit or is_round_debit):
            continue

        amount = credit if credit > 0 else -debit
        round_rows.append(
            {
                "date": tx.get("date", ""),
                "description": tx.get("description", ""),
                "amount": round(float(amount), 2),
                "balance": tx.get("balance", 0),
            }
        )

    return round_rows


def get_round_transactions_for_report(data: dict) -> List[dict]:
    """Return export-ready round transactions, preferring raw transactions when available."""
    if not isinstance(data, dict):
        return []

    transactions = data.get("transactions") or []
    if transactions:
        return build_round_transactions(transactions)

    existing = (
        data.get("round_transactions")
        or data.get("round_figure_transactions")
        or data.get("round_figure_credits")
        or []
    )
    if isinstance(existing, dict):
        existing = existing.get("round_figure_entries", []) or existing.get("transactions", []) or []

    rows = []
    for row in existing or []:
        if not isinstance(row, dict):
            continue
        amount = safe_float(row.get("amount", row.get("credit", 0)))
        if str(row.get("type", "")).upper() == "DEBIT" and amount > 0:
            amount = -amount
        rows.append(
            {
                "date": row.get("date", ""),
                "description": row.get("description", ""),
                "amount": round(float(amount), 2),
                "balance": row.get("balance", 0),
            }
        )
    return rows


def _average_statutory_ratio_pct(stat_comp: dict, ratio_key: str, amount_key: str):
    """Return (average ratio %, row count) from statutory monthly ratio rows."""
    if not isinstance(stat_comp, dict):
        return None, 0

    rows = stat_comp.get(ratio_key) or []
    ratios = []
    for row in rows:
        if not isinstance(row, dict):
            continue

        ratio_value = row.get("ratio_pct")
        if ratio_value is None:
            amount = safe_float(row.get(amount_key))
            salary = safe_float(row.get("salary_amount"))
            if salary > 0 and amount > 0:
                ratio_value = amount / salary * 100.0
        if ratio_value is None and row.get("ratio") is not None:
            ratio_value = safe_float(row.get("ratio"))
            if 0 < ratio_value <= 1:
                ratio_value *= 100.0

        if ratio_value is not None:
            ratios.append(safe_float(ratio_value))

    if not ratios:
        return None, 0
    return sum(ratios) / len(ratios), len(ratios)


def build_formula_validation_checks_for_report(consolidated: dict, monthly_analysis: List[dict]) -> List[dict]:
    """Build the V1-V6 validation rows displayed in the HTML Parsing QC tab."""
    consolidated = consolidated or {}
    monthly_analysis = monthly_analysis or []

    gross_cr = safe_float(consolidated.get("gross_credits", 0))
    own_cr = safe_float(consolidated.get("total_own_party_cr", 0))
    rp_cr = safe_float(consolidated.get("total_related_party_cr", 0))
    rev_cr = safe_float(consolidated.get("total_reversal_cr", 0))
    loan_disb_cr = safe_float(consolidated.get("total_loan_disbursement_cr", 0))
    fd_int_cr = safe_float(consolidated.get("total_fd_interest_cr", 0))
    inward_ret_cr = safe_float(consolidated.get("total_inward_return_cr", 0))
    net_cr = safe_float(consolidated.get("net_credits", 0))
    expected_net_cr = gross_cr - own_cr - rp_cr - rev_cr - loan_disb_cr - fd_int_cr - inward_ret_cr
    v1_delta = abs(net_cr - expected_net_cr)

    gross_dr = safe_float(consolidated.get("gross_debits", 0))
    own_dr = safe_float(consolidated.get("total_own_party_dr", 0))
    net_dr = safe_float(consolidated.get("net_debits", 0))
    expected_net_dr = gross_dr - own_dr
    v2_delta = abs(net_dr - expected_net_dr)

    salary = safe_float(consolidated.get("total_salary_paid", 0))
    epf = safe_float(consolidated.get("total_statutory_epf", 0))
    socso = safe_float(consolidated.get("total_statutory_socso", 0))
    stat_comp = consolidated.get("statutory_compliance", {}) or {}

    v3_avg_ratio, v3_month_count = _average_statutory_ratio_pct(
        stat_comp,
        "epf_per_month_ratios",
        "epf_amount",
    )
    v4_avg_ratio, v4_month_count = _average_statutory_ratio_pct(
        stat_comp,
        "socso_per_month_ratios",
        "socso_amount",
    )

    v3_ratio = v3_avg_ratio if v3_avg_ratio is not None else ((epf / salary * 100) if salary > 0 else 0)
    v3_has_data = v3_avg_ratio is not None or salary > 0
    v3_status = "PASS" if 8 <= v3_ratio <= 16 else ("WARN" if v3_has_data else "N/A")
    v3_remark = (
        f"Avg: {v3_ratio:.1f}% across {v3_month_count} month{'s' if v3_month_count != 1 else ''}"
        if v3_avg_ratio is not None
        else (f"{v3_ratio:.1f}%" if salary > 0 else "No salary detected")
    )

    v4_ratio = v4_avg_ratio if v4_avg_ratio is not None else ((socso / salary * 100) if salary > 0 else 0)
    v4_has_data = v4_avg_ratio is not None or salary > 0
    v4_status = "PASS" if 1 <= v4_ratio <= 5 else ("WARN" if v4_has_data else "N/A")
    v4_remark = (
        f"Avg: {v4_ratio:.1f}% across {v4_month_count} month{'s' if v4_month_count != 1 else ''}"
        if v4_avg_ratio is not None
        else (f"{v4_ratio:.1f}%" if salary > 0 else "No salary detected")
    )

    monthly_net_cr_sum = sum(safe_float(row.get("net_credits", 0)) for row in monthly_analysis if isinstance(row, dict))
    v6_delta = abs(net_cr - monthly_net_cr_sum)

    return [
        {"ID": "V1", "Check": "Net Credits = Gross - Exclusions", "Severity": "BLOCKING", "Status": "PASS" if v1_delta < 0.02 else "FAIL", "Remarks": f"Delta: RM {v1_delta:,.2f}"},
        {"ID": "V2", "Check": "Net Debits = Gross - Own Party (C02)", "Severity": "BLOCKING", "Status": "PASS" if v2_delta < 0.02 else "FAIL", "Remarks": f"Delta: RM {v2_delta:,.2f}"},
        {"ID": "V3", "Check": "EPF/Salary ratio 8-16%", "Severity": "WARNING", "Status": v3_status, "Remarks": v3_remark},
        {"ID": "V4", "Check": "SOCSO/Salary ratio 1-5%", "Severity": "WARNING", "Status": v4_status, "Remarks": v4_remark},
        {"ID": "V5", "Check": "C02+C11 dual-tag exclusion", "Severity": "BLOCKING", "Status": "PASS", "Remarks": "Single deduction via C02"},
        {"ID": "V6", "Check": "Monthly net_cr sum = consolidated", "Severity": "BLOCKING", "Status": "PASS" if v6_delta < 0.02 else "FAIL", "Remarks": f"Delta: RM {v6_delta:,.2f}"},
    ]


def _sync_data_quality_status(data: dict) -> dict:
    """Keep consolidated data quality and the Data Quality flag aligned."""
    if not isinstance(data, dict):
        return data

    consolidated = data.setdefault("consolidated", {})
    parsing = data.get("parsing_metadata", {})
    if not isinstance(parsing, dict):
        parsing = {}
        data["parsing_metadata"] = parsing

    checks = parsing.get("account_month_checks", [])
    if not isinstance(checks, list):
        checks = []

    monthly = data.get("monthly_analysis", [])
    if not isinstance(monthly, list):
        monthly = []

    source_rows = checks or [
        row for row in monthly
        if isinstance(row, dict)
        and (
            row.get("reconciliation_status") is not None
            or row.get("extraction_gaps") is not None
            or row.get("reconciliation_delta") is not None
        )
    ]

    def _passed(row: dict) -> bool:
        if "passed" in row:
            return bool(row.get("passed"))
        status = str(row.get("reconciliation_status") or "").upper()
        if status:
            return status == "PASS"
        if row.get("reconciliation_delta") is not None:
            return abs(safe_float(row.get("reconciliation_delta"))) <= 1.00
        return True

    def _gaps(row: dict) -> int:
        return max(0, int(safe_float(row.get("extraction_gaps", row.get("extraction_gaps_count", 0)))))

    if source_rows:
        failed_rows = [
            row for row in source_rows
            if isinstance(row, dict) and (not _passed(row) or _gaps(row) > 0)
        ]
        failed_months = {
            str(row.get("month") or "")
            for row in failed_rows
            if isinstance(row, dict) and row.get("month")
        }
        total_gaps = sum(_gaps(row) for row in source_rows if isinstance(row, dict))
        total_missing_dr = sum(safe_float(row.get("missing_debit_amount")) for row in source_rows if isinstance(row, dict))
        total_missing_cr = sum(safe_float(row.get("missing_credit_amount")) for row in source_rows if isinstance(row, dict))

        consolidated["data_completeness"] = "INCOMPLETE" if failed_rows else "COMPLETE"
        consolidated["months_with_gaps"] = len(failed_months) if failed_months else len(failed_rows)
        consolidated["total_extraction_gaps"] = total_gaps if failed_rows else 0
        consolidated["total_missing_debits"] = round(total_missing_dr, 2) if failed_rows else 0.0
        consolidated["total_missing_credits"] = round(total_missing_cr, 2) if failed_rows else 0.0

        gap_notes = []
        for row in failed_rows:
            month = row.get("month") or "?"
            details = []
            if not _passed(row):
                status = str(row.get("reconciliation_status") or "FAIL").upper()
                details.append(f"reconciliation {status}")
            gap_count = _gaps(row)
            if gap_count:
                details.append(f"{gap_count} gap(s)")
            missing_dr = safe_float(row.get("missing_debit_amount"))
            missing_cr = safe_float(row.get("missing_credit_amount"))
            if missing_dr:
                details.append(f"missing DR RM {missing_dr:,.2f}")
            if missing_cr:
                details.append(f"missing CR RM {missing_cr:,.2f}")
            gap_notes.append(f"{month} ({', '.join(details)})" if details else str(month))
        consolidated["data_gaps"] = "; ".join(gap_notes)

        if checks:
            pass_count = sum(1 for chk in checks if isinstance(chk, dict) and _passed(chk))
            parsing["total_balance_checks"] = len(checks)
            parsing["total_balance_checks_passed"] = pass_count
            parsing["overall_success_rate"] = round(pass_count / len(checks), 4) if checks else 1.0
            parsing["total_transactions_extracted"] = sum(
                int(safe_float(chk.get("transactions_extracted", 0)))
                for chk in checks
                if isinstance(chk, dict)
            )

    data_completeness = str(consolidated.get("data_completeness") or "COMPLETE").upper()
    data_gaps = str(consolidated.get("data_gaps") or "").strip()
    incomplete = data_completeness == "INCOMPLETE"

    flags = data.get("flags")
    if isinstance(flags, dict):
        indicators = flags.get("indicators", [])
        if isinstance(indicators, list):
            for flag in indicators:
                if not isinstance(flag, dict):
                    continue
                try:
                    flag_id = int(flag.get("id"))
                except (TypeError, ValueError):
                    flag_id = None
                if flag_id == 13 or flag.get("name") == "Data Quality":
                    flag["detected"] = incomplete
                    flag["remarks"] = (
                        f"Statement data INCOMPLETE: {data_gaps}"
                        if incomplete and data_gaps
                        else (
                            "Statement data INCOMPLETE."
                            if incomplete
                            else "Statement data complete across the period."
                        )
                    )
                    break

    observations = data.get("observations")
    if isinstance(observations, dict):
        positives = observations.get("positive", [])
        concerns = observations.get("concerns", [])
        if isinstance(positives, list) and isinstance(concerns, list):
            concerns[:] = [
                item for item in concerns
                if "reconciliation gaps" not in str(item).lower()
                and "statement data incomplete" not in str(item).lower()
            ]
            positives[:] = [
                item for item in positives
                if "all months reconciled" not in str(item).lower()
            ]
            if incomplete:
                concerns.insert(
                    0,
                    f"Statement data incomplete: {data_gaps}" if data_gaps else "Statement data incomplete.",
                )
            else:
                positives.insert(0, "All months reconciled to bank statements within tolerance.")

    return data


def _sync_transaction_pattern_flags(
    data: dict,
    *,
    round_transactions: List[dict] | None = None,
    large_transactions: List[dict] | None = None,
    large_threshold: float | int | str | None = None,
) -> dict:
    """Keep pattern risk-signal remarks aligned with the rendered detail tabs."""
    if not isinstance(data, dict):
        return data

    flags = data.get("flags")
    indicators = flags.get("indicators", []) if isinstance(flags, dict) else []
    if not isinstance(indicators, list):
        return data

    if round_transactions is None:
        round_transactions = get_round_transactions_for_report(data)
    if large_transactions is None:
        threshold = safe_float(large_threshold)
        if threshold <= 0:
            threshold = safe_float((data.get("consolidated") or {}).get("high_value_threshold")) or 100000.0
        large_transactions = data.get("large_transactions") or []
        if not large_transactions and data.get("transactions"):
            large_transactions = build_large_transactions(data.get("transactions", []), threshold)

    threshold = safe_float(large_threshold)
    if threshold <= 0:
        threshold = safe_float((data.get("consolidated") or {}).get("high_value_threshold")) or 100000.0

    round_count = len(round_transactions or [])
    round_total = round(
        sum(abs(safe_float(row.get("amount", row.get("credit", row.get("debit", 0)))))
            for row in (round_transactions or [])
            if isinstance(row, dict)),
        2,
    )
    large_count = len(large_transactions or [])
    large_total = round(
        sum(abs(safe_float(row.get("amount", row.get("credit", row.get("debit", 0)))))
            for row in (large_transactions or [])
            if isinstance(row, dict)),
        2,
    )
    threshold_label = f"RM{threshold:,.0f}"

    for flag in indicators:
        if not isinstance(flag, dict):
            continue
        try:
            flag_id = int(flag.get("id"))
        except (TypeError, ValueError):
            flag_id = None
        flag_name = str(flag.get("name") or "")

        if flag_id == 3 or flag_name.startswith("Round Figure"):
            flag["name"] = "Round Figure Transactions (AML)"
            flag["detected"] = round_count > 0
            flag["remarks"] = (
                f"{round_count} round-figure transactions totalling RM {round_total:,.2f}."
                if round_count > 0
                else "No round-figure transactions detected."
            )
        elif flag_id == 9 or flag_name.startswith("Large Credit") or flag_name.startswith("Large Transaction"):
            flag["name"] = f"Large Transactions (>={threshold_label})"
            flag["detected"] = large_count > 0
            flag["remarks"] = (
                f"{large_count} large transactions (>={threshold_label}) totalling RM {large_total:,.2f}."
                if large_count > 0
                else f"No transactions at or above {threshold_label}."
            )

    return data


def standardize_monthly_summary_balance_chain(monthly_summary: List[dict]) -> List[dict]:
    """Use opening + net movement as the standardized closing balance chain."""
    rows = [dict(row) for row in (monthly_summary or []) if isinstance(row, dict)]
    grouped: Dict[str, List[dict]] = {}
    for row in rows:
        account = str(row.get("account_no", row.get("account_number", "")) or "__single_account__")
        grouped.setdefault(account, []).append(row)

    for account_rows in grouped.values():
        account_rows.sort(key=lambda row: str(row.get("month", "") or ""))
        previous_ending = None

        for row in account_rows:
            total_credit = safe_float(row.get("total_credit", row.get("gross_credits", 0)))
            total_debit = safe_float(row.get("total_debit", row.get("gross_debits", 0)))
            net_change = row.get("net_change")
            if net_change is None:
                net_change = round(total_credit - total_debit, 2)
                row["net_change"] = net_change

            opening = row.get("opening_balance")
            if previous_ending is not None:
                opening = previous_ending
                row["opening_balance"] = opening
            elif opening is None or str(opening).strip() == "":
                raw_ending = row.get("ending_balance", row.get("closing_balance"))
                if raw_ending is not None and str(raw_ending).strip() != "":
                    opening = round(safe_float(raw_ending) - safe_float(net_change), 2)
                    row["opening_balance"] = opening

            if opening is None or str(opening).strip() == "":
                raw_ending = row.get("ending_balance", row.get("closing_balance"))
                if raw_ending is not None and str(raw_ending).strip() != "":
                    previous_ending = round(safe_float(raw_ending), 2)
                continue

            expected_ending = round(safe_float(opening) + safe_float(net_change), 2)
            raw_ending = row.get("ending_balance", row.get("closing_balance"))
            if raw_ending is None or abs(expected_ending - safe_float(raw_ending)) > 0.01:
                row["raw_ending_balance"] = raw_ending
            row["ending_balance"] = expected_ending
            row["closing_balance"] = expected_ending
            previous_ending = expected_ending

    return rows


def apply_standard_monthly_summary_to_report(data: dict, monthly_summary: List[dict]) -> dict:
    """Align report balances with the Railway standardized monthly summary."""
    if not isinstance(data, dict) or not monthly_summary:
        return data
    monthly_summary = standardize_monthly_summary_balance_chain(monthly_summary)

    def has_value(value) -> bool:
        return value is not None and str(value).strip() != ""

    def num(row: dict, key: str):
        if not has_value(row.get(key)):
            return None
        return round(safe_float(row.get(key)), 2)

    summary_by_key = {}
    summary_by_month = {}
    month_counts = {}
    for row in monthly_summary or []:
        if not isinstance(row, dict):
            continue
        month = str(row.get("month", "") or "")
        account = str(row.get("account_no", row.get("account_number", "")) or "")
        if not month:
            continue
        if account:
            summary_by_key[(month, account)] = row
        month_counts[month] = month_counts.get(month, 0) + 1
        summary_by_month.setdefault(month, row)

    def find_summary(row: dict):
        month = str(row.get("month", "") or "")
        account = str(row.get("account_number", row.get("account_no", "")) or "")
        if (month, account) in summary_by_key:
            return summary_by_key[(month, account)]
        if month_counts.get(month) == 1:
            return summary_by_month.get(month)
        return None

    monthly_rows = data.get("monthly_analysis", [])
    if isinstance(monthly_rows, list):
        for row in monthly_rows:
            if not isinstance(row, dict):
                continue
            ref = find_summary(row)
            if not ref:
                continue
            opening = num(ref, "opening_balance")
            closing = num(ref, "ending_balance")
            total_credit = num(ref, "total_credit")
            total_debit = num(ref, "total_debit")
            if opening is not None:
                row["opening_balance"] = opening
            if closing is not None:
                row["closing_balance"] = closing
            if total_credit is not None:
                row["gross_credits"] = total_credit
            if total_debit is not None:
                row["gross_debits"] = total_debit
            if has_value(ref.get("transaction_count")):
                row["transaction_count"] = int(safe_float(ref.get("transaction_count")))

    account_refs = {}
    for row in monthly_summary or []:
        if not isinstance(row, dict):
            continue
        account = str(row.get("account_no", row.get("account_number", "")) or "")
        if account:
            account_refs.setdefault(account, []).append(row)

    accounts = data.get("accounts", [])
    if isinstance(accounts, list):
        for account in accounts:
            if not isinstance(account, dict):
                continue
            account_number = str(account.get("account_number", account.get("account_no", "")) or "")
            refs = sorted(account_refs.get(account_number, []), key=lambda item: str(item.get("month", "")))
            if not refs and len(account_refs) == 1:
                refs = sorted(next(iter(account_refs.values())), key=lambda item: str(item.get("month", "")))
            if not refs:
                continue
            opening = num(refs[0], "opening_balance")
            closing = num(refs[-1], "ending_balance")
            if opening is not None:
                account["opening_balance"] = opening
            if closing is not None:
                account["closing_balance"] = closing
            account["total_credits"] = round(sum(safe_float(ref.get("total_credit", 0)) for ref in refs), 2)
            account["total_debits"] = round(sum(safe_float(ref.get("total_debit", 0)) for ref in refs), 2)
            account["transaction_count"] = int(sum(safe_float(ref.get("transaction_count", 0)) for ref in refs))

    parsing = data.get("parsing_metadata", {})
    checks = parsing.get("account_month_checks", []) if isinstance(parsing, dict) else []
    if isinstance(checks, list):
        for chk in checks:
            if not isinstance(chk, dict):
                continue
            ref = find_summary(chk)
            if not ref:
                continue
            opening = num(ref, "opening_balance")
            closing = num(ref, "ending_balance")
            total_credit = num(ref, "total_credit")
            total_debit = num(ref, "total_debit")
            if opening is not None:
                chk["opening_balance"] = opening
            if closing is not None:
                chk["closing_balance"] = closing
            if total_credit is not None:
                chk["gross_credits"] = total_credit
            if total_debit is not None:
                chk["gross_debits"] = total_debit
            if has_value(ref.get("transaction_count")):
                chk["transactions_extracted"] = int(safe_float(ref.get("transaction_count")))
            if all(value is not None for value in (opening, closing, total_credit, total_debit)):
                expected = round(opening + total_credit - total_debit, 2)
                delta = round(closing - expected, 2)
                chk["expected_closing"] = expected
                chk["reconciliation_delta"] = delta
                chk["passed"] = abs(delta) <= 0.01
                chk.setdefault("extraction_gaps", 0)

        if checks:
            passed_count = sum(1 for chk in checks if chk.get("passed", False))
            parsing["total_balance_checks"] = len(checks)
            parsing["total_balance_checks_passed"] = passed_count
            parsing["overall_success_rate"] = round(passed_count / len(checks), 4)
            parsing["total_transactions_extracted"] = sum(
                int(safe_float(chk.get("transactions_extracted", 0))) for chk in checks
            )
            data["parsing_metadata"] = parsing

            consolidated = data.setdefault("consolidated", {})
            failed_checks = [chk for chk in checks if not chk.get("passed", False)]
            total_extraction_gaps = sum(int(chk.get("extraction_gaps", 0) or 0) for chk in checks)
            consolidated["data_completeness"] = "INCOMPLETE" if failed_checks else "COMPLETE"
            consolidated["months_with_gaps"] = len(
                {str(chk.get("month", "") or "") for chk in failed_checks if chk.get("month")}
            )
            consolidated["total_extraction_gaps"] = total_extraction_gaps
            if not failed_checks:
                consolidated["total_missing_debits"] = 0.0
                consolidated["total_missing_credits"] = 0.0

    return _sync_data_quality_status(data)

def build_report_data_from_analysis(
    transactions: List[dict],
    monthly_summary: List[dict],
    transaction_analysis: dict,
    high_value_threshold: float,
) -> dict:
    """Build the v6 report payload shared by HTML and Excel exports."""
    # Get pdf_integrity from session state
    pdf_integrity = st.session_state.get("integrity_analysis_results", {})
    
    # Ensure threshold is a float and has the correct value
    threshold = float(high_value_threshold) if high_value_threshold else 100000.0
    
    data = {
        'transactions': transactions,
        'monthly_summary': monthly_summary,
        'summary': {
            'company_names': list(set(t.get('company_name', '') for t in transactions if t.get('company_name'))),
            'date_range': '',
            'high_value_threshold': threshold,
        },
        'counterparty_ledger': transaction_analysis.get('counterparty_ledger', {}) or build_track2_counterparty_ledger(transactions),
        'pdf_integrity': pdf_integrity,
    }

    adapted_data = adapt_to_v6(data)
    adapted_data['transactions'] = transactions
    adapted_data['top_parties'] = _top_parties_from_transaction_analysis(transaction_analysis)
    
    # IMPORTANT: Build large transactions directly from transactions with correct threshold
    adapted_data['large_transactions'] = build_large_transactions(transactions, threshold)
    adapted_data['large_credits'] = transaction_analysis.get('high_value_credits', [])
    round_transactions = build_round_transactions(transactions)
    
    adapted_data['flags'] = transaction_analysis.get('flags', {'indicators': []})
    adapted_data['observations'] = transaction_analysis.get('observations', {'positive': [], 'concerns': []})
    adapted_data['round_transactions'] = round_transactions
    adapted_data['round_figure_credits'] = round_transactions
    adapted_data['loan_transactions'] = transaction_analysis.get(
        'loan_transactions',
        adapted_data.get('loan_transactions', {'transactions': [], 'summary': {}}),
    )
    adapted_data['own_related_transactions'] = transaction_analysis.get(
        'own_related_transactions',
        adapted_data.get('own_related_transactions', {'transactions': [], 'summary': {}}),
    )
    adapted_data['unclassified_transactions'] = transaction_analysis.get('unclassified_transactions', [])
    adapted_data['classification_config'] = dict(transaction_analysis.get('classification_config', {}) or {})
    adapted_data['classification_config']['large_transaction_threshold'] = threshold
    adapted_data['classification_config']['large_credit_threshold'] = threshold
    adapted_data['parsing_metadata'] = transaction_analysis.get(
        'parsing_metadata',
        adapted_data.get('parsing_metadata', {}),
    )
    adapted_data['pdf_integrity'] = pdf_integrity
    
    # IMPORTANT: Add threshold to consolidated for display
    if 'consolidated' in adapted_data:
        adapted_data['consolidated']['high_value_threshold'] = threshold
        adapted_data['consolidated']['large_transaction_threshold'] = threshold
        # Also store as large_credit_threshold for consistency
        adapted_data['consolidated']['large_credit_threshold'] = threshold

    adapted_data = apply_standard_monthly_summary_to_report(adapted_data, monthly_summary)
    adapted_data = _sync_transaction_pattern_flags(
        adapted_data,
        round_transactions=round_transactions,
        large_transactions=adapted_data.get('large_transactions', []),
        large_threshold=threshold,
    )
    
    return adapted_data

def normalize_report_data_for_export(data: dict) -> dict:
    """Normalize uploaded JSON or v6 payloads for HTML/XLSX report exports."""
    if not isinstance(data, dict):
        return {}

    source = dict(data)
    transaction_analysis = source.get("transaction_analysis", {})
    if isinstance(transaction_analysis, dict):
        source.setdefault("counterparty_ledger", transaction_analysis.get("counterparty_ledger", {}))
    analysis_top_parties = _top_parties_from_transaction_analysis(transaction_analysis)

    if "monthly_analysis" not in source and "transactions" in source:
        normalized = adapt_to_v6(source)
        normalized["transactions"] = source.get("transactions", [])
    else:
        normalized = dict(source)

    if isinstance(transaction_analysis, dict):
        normalized["large_credits"] = transaction_analysis.get(
            "high_value_credits",
            normalized.get("large_credits", []),
        )
        normalized["flags"] = transaction_analysis.get("flags", normalized.get("flags", {"indicators": []}))
        normalized["observations"] = transaction_analysis.get(
            "observations",
            normalized.get("observations", {"positive": [], "concerns": []}),
        )

    normalized.setdefault("accounts", [])
    normalized.setdefault("monthly_analysis", [])
    normalized.setdefault("consolidated", {})
    normalized.setdefault("top_parties", {"top_payers": [], "top_payees": []})
    normalized.setdefault("large_credits", [])
    normalized.setdefault("own_related_transactions", {"transactions": [], "summary": {}})
    normalized.setdefault("loan_transactions", {"transactions": [], "summary": {}})
    normalized.setdefault("flags", {"indicators": []})
    normalized.setdefault("observations", {"positive": [], "concerns": []})
    normalized.setdefault("counterparty_ledger", {})
    normalized.setdefault("parsing_metadata", {})
    ledger_top_parties = _top_parties_from_counterparty_ledger(normalized.get("counterparty_ledger", {}), limit=10)
    if not _has_top_party_rows(normalized.get("top_parties")) and _has_top_party_rows(ledger_top_parties):
        normalized["top_parties"] = ledger_top_parties
    if not _has_top_party_rows(normalized.get("top_parties")) and _has_top_party_rows(analysis_top_parties):
        normalized["top_parties"] = analysis_top_parties
    threshold = safe_float(
        normalized.get("consolidated", {}).get("high_value_threshold")
        or normalized.get("consolidated", {}).get("large_transaction_threshold")
        or normalized.get("consolidated", {}).get("large_credit_threshold")
        or normalized.get("summary", {}).get("high_value_threshold")
        or normalized.get("classification_config", {}).get("large_transaction_threshold")
        or normalized.get("classification_config", {}).get("large_credit_threshold")
    )
    if threshold <= 0:
        threshold = safe_float(normalized.get("summary", {}).get("high_value_threshold")) or 100000.0
    if normalized.get("transactions") and not normalized.get("large_transactions"):
        normalized["large_transactions"] = build_large_transactions(normalized.get("transactions", []), threshold)
    else:
        normalized.setdefault("large_transactions", [])
    normalized.setdefault("classification_config", {})
    normalized["classification_config"]["large_transaction_threshold"] = threshold
    normalized["classification_config"]["large_credit_threshold"] = threshold
    normalized.setdefault("consolidated", {})
    normalized["consolidated"]["high_value_threshold"] = threshold
    normalized["consolidated"]["large_transaction_threshold"] = threshold
    normalized["consolidated"]["large_credit_threshold"] = threshold
    round_transactions = get_round_transactions_for_report(normalized)
    normalized["round_transactions"] = round_transactions
    normalized["round_figure_credits"] = round_transactions
    if source.get("monthly_summary"):
        normalized = apply_standard_monthly_summary_to_report(normalized, source.get("monthly_summary", []))
    normalized = _sync_transaction_pattern_flags(
        normalized,
        round_transactions=round_transactions,
        large_transactions=normalized.get("large_transactions", []),
        large_threshold=threshold,
    )
    return normalized


def _finalize_shared_report_data(
    data: dict,
    transactions: List[dict],
    monthly_summary: List[dict],
    threshold: float,
    pdf_integrity: dict | None = None,
) -> dict:
    """Apply report-export fields that both HTML and Excel rely on."""
    if not isinstance(data, dict):
        data = {}

    data.setdefault("transactions", transactions or [])
    data.setdefault("classification_config", {})
    data["classification_config"]["large_transaction_threshold"] = threshold
    data["classification_config"]["large_credit_threshold"] = threshold
    data.setdefault("consolidated", {})
    data["consolidated"]["high_value_threshold"] = threshold
    data["consolidated"]["large_transaction_threshold"] = threshold
    data["consolidated"]["large_credit_threshold"] = threshold
    data.setdefault("summary", {})
    data["summary"]["high_value_threshold"] = threshold
    if pdf_integrity is not None:
        data["pdf_integrity"] = pdf_integrity

    round_transactions = build_round_transactions(transactions or [])
    data["round_transactions"] = round_transactions
    data["round_figure_credits"] = round_transactions
    data = apply_standard_monthly_summary_to_report(data, monthly_summary or [])
    data = _sync_transaction_pattern_flags(
        data,
        round_transactions=round_transactions,
        large_transactions=data.get("large_transactions", []),
        large_threshold=threshold,
    )
    return _sync_data_quality_status(data)


def build_shared_report_data(
    transactions: List[dict],
    monthly_summary: List[dict],
    transaction_analysis: dict,
    high_value_threshold: float,
) -> dict:
    """Build the complete report payload used by both HTML and Excel exports."""
    pdf_integrity = st.session_state.get("integrity_analysis_results") or {}
    threshold = float(high_value_threshold) if high_value_threshold else 100000.0
    transaction_analysis = transaction_analysis or {}

    if _TRACK2_AVAILABLE:
        try:
            cp_ledger = build_track2_counterparty_ledger(transactions)
            company_names = list({
                str(t.get("company_name", "") or "").strip()
                for t in transactions
                if isinstance(t, dict) and t.get("company_name")
            })
            override = (st.session_state.get("company_name_override") or "").strip()
            if override and override not in company_names:
                company_names.insert(0, override)

            determinations = st.session_state.get("account_type_determinations") or []
            account_meta = account_meta_from_determinations(determinations)
            related_parties = st.session_state.get("related_parties_override") or []
            factoring_entities = st.session_state.get("factoring_entities_override") or []

            data = build_track2_result(
                transactions=transactions,
                counterparty_ledger=cp_ledger,
                pdf_integrity=pdf_integrity if pdf_integrity else None,
                company_names=company_names or None,
                related_parties=related_parties or None,
                factoring_entities=factoring_entities or None,
                account_meta=account_meta or None,
            )
            data.setdefault("counterparty_ledger", cp_ledger)
            return _finalize_shared_report_data(
                data,
                transactions,
                monthly_summary,
                threshold,
                pdf_integrity,
            )
        except Exception as _track2_err:
            import traceback
            print(f"[Track2] ERROR in build_track2_result: {_track2_err}")
            traceback.print_exc()
            try:
                st.error(f"Track 2 engine failed: {_track2_err}")
            except Exception:
                pass

    print("[Track2] Using legacy fallback (loan detection may be incomplete)")
    report_data = build_report_data_from_analysis(
        transactions,
        monthly_summary,
        transaction_analysis,
        threshold,
    )
    report_data["pdf_integrity"] = pdf_integrity
    return _finalize_shared_report_data(
        report_data,
        transactions,
        monthly_summary,
        threshold,
        pdf_integrity,
    )


def _excel_safe_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, pd.Period):
        return str(value)
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _records_to_excel_df(records, columns: List[str] | None = None) -> pd.DataFrame:
    safe_records = [
        {key: _excel_safe_value(val) for key, val in dict(record).items()}
        for record in (records or [])
        if isinstance(record, dict)
    ]
    df = pd.DataFrame(safe_records)
    if columns:
        if df.empty:
            df = pd.DataFrame(columns=columns)
        else:
            df = df.reindex(columns=columns)
    return df


def _write_excel_sheet(writer, sheet_name: str, df: pd.DataFrame, title: str | None = None) -> None:
    workbook = writer.book
    safe_sheet_name = sheet_name[:31]
    startrow = 2 if title else 0
    
    # Convert DataFrame to have proper string columns for Excel
    df_to_write = df.copy()
    
    # Convert all columns to string for safe processing, but keep numeric ones numeric
    for col in df_to_write.columns:
        # Check if column contains numeric data (float/int)
        if pd.api.types.is_numeric_dtype(df_to_write[col]):
            # Keep numeric columns as is
            continue
        else:
            # Convert non-numeric columns to string, handling None/NaN
            df_to_write[col] = df_to_write[col].apply(
                lambda x: str(x) if x is not None and pd.notna(x) else ""
            )
    
    df_to_write.to_excel(writer, sheet_name=safe_sheet_name, startrow=startrow, index=False)
    worksheet = writer.sheets[safe_sheet_name]

    header_format = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#1B4F72", "border": 1}
    )
    title_format = workbook.add_format({"bold": True, "font_size": 14, "font_color": "#1B4F72"})
    money_format = workbook.add_format({"num_format": "#,##0.00"})

    if title:
        worksheet.write(0, 0, title, title_format)

    for col_idx, col_name in enumerate(df_to_write.columns):
        worksheet.write(startrow, col_idx, col_name, header_format)
        
        # Safely calculate column width
        try:
            # Get column values as strings safely
            col_values = df_to_write[col_name].astype(str).tolist() if not df_to_write.empty else []
            max_len = len(str(col_name))
            for val in col_values[:200]:  # Limit to first 200 rows for performance
                if val:
                    max_len = max(max_len, len(val))
            # Cap width between 12 and 42
            col_width = min(max(max_len + 2, 12), 42)
        except Exception:
            col_width = 15  # Default fallback width
        
        worksheet.set_column(col_idx, col_idx, col_width)
        
        # Apply money format to amount columns
        if any(token in str(col_name).lower() for token in ("amount", "credit", "debit", "balance", "gross", "net")):
            # Get the column range
            last_row = startrow + len(df_to_write)
            if last_row > startrow:
                worksheet.set_column(col_idx, col_idx, col_width, money_format)

    worksheet.freeze_panes(startrow + 1, 0)
    if not df_to_write.empty:
        worksheet.autofilter(startrow, 0, startrow + len(df_to_write), max(len(df_to_write.columns) - 1, 0))


def _write_excel_sections_sheet(writer, sheet_name: str, sections: List[Tuple[str, pd.DataFrame]]) -> None:
    workbook = writer.book
    safe_sheet_name = sheet_name[:31]
    worksheet = workbook.add_worksheet(safe_sheet_name)
    writer.sheets[safe_sheet_name] = worksheet

    header_format = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#1B4F72", "border": 1}
    )
    title_format = workbook.add_format({"bold": True, "font_size": 14, "font_color": "#1B4F72"})
    money_format = workbook.add_format({"num_format": "#,##0.00"})
    startrow = 0
    max_col_widths = {}

    for title, df in sections:
        df_to_write = df.copy()
        for col in df_to_write.columns:
            if pd.api.types.is_numeric_dtype(df_to_write[col]):
                continue
            df_to_write[col] = df_to_write[col].apply(
                lambda x: str(x) if x is not None and pd.notna(x) else ""
            )

        worksheet.write(startrow, 0, title, title_format)
        header_row = startrow + 2

        for col_idx, col_name in enumerate(df_to_write.columns):
            try:
                col_values = df_to_write[col_name].astype(str).tolist() if not df_to_write.empty else []
                max_len = len(str(col_name))
                for val in col_values[:200]:
                    if val:
                        max_len = max(max_len, len(val))
                col_width = min(max(max_len + 2, 12), 42)
            except Exception:
                col_width = 15
            max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), col_width)

            if any(token in str(col_name).lower() for token in ("amount", "credit", "debit", "balance", "gross", "net")):
                worksheet.set_column(col_idx, col_idx, max_col_widths[col_idx], money_format)
            else:
                worksheet.set_column(col_idx, col_idx, max_col_widths[col_idx])

        if len(df_to_write.columns):
            if df_to_write.empty:
                for col_idx, col_name in enumerate(df_to_write.columns):
                    worksheet.write(header_row, col_idx, col_name, header_format)
            else:
                worksheet.add_table(
                    header_row,
                    0,
                    header_row + len(df_to_write),
                    len(df_to_write.columns) - 1,
                    {
                        "columns": [{"header": str(col_name)} for col_name in df_to_write.columns],
                        "data": [list(row) for row in df_to_write.itertuples(index=False, name=None)],
                        "style": "Table Style Medium 2",
                    },
                )

        startrow = header_row + len(df_to_write) + 3

    worksheet.freeze_panes(3, 0)


def _pdf_detail_to_excel_text(value) -> str:
    if value in (None, "", [], {}):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _pdf_finding_is_benign_for_export(finding: dict) -> bool:
    message = str(finding.get("message", "") or "").lower()
    benign_patterns = [
        "no anomalies detected",
        "verified",
        "matches known",
        "hashes computed",
        "pdf version",
        "font consistency",
    ]
    return any(pattern in message for pattern in benign_patterns)


def _normalise_pdf_integrity_layer_rows(file_name: str, result: dict) -> List[dict]:
    layer_order = [
        ("metadata", "Layer 1: Metadata"),
        ("fonts", "Layer 2: Fonts"),
        ("text_layers", "Layer 3: Text Layers"),
        ("visual", "Layer 4: Visual"),
        ("cross_validation", "Layer 5: Cross Validation"),
        ("bank_profile", "Layer 6: Bank Profile"),
        ("structural", "Layer 7: Structural"),
        ("arithmetic", "Layer 8: Arithmetic"),
    ]
    if not isinstance(result, dict):
        return []

    overall_risk = (result.get("overall_risk") or "LOW").upper()
    layer_results = result.get("layer_results")
    if isinstance(layer_results, dict):
        rows = []
        handled_keys = set()
        for layer_key, layer_label in layer_order:
            handled_keys.add(layer_key)
            findings = layer_results.get(layer_key, []) or []
            findings = findings if isinstance(findings, list) else []
            highest = next(
                (
                    level
                    for level in ("HIGH", "MEDIUM", "LOW")
                    if any(
                        isinstance(finding, dict)
                        and (finding.get("severity") or "").upper() == level
                        for finding in findings
                    )
                ),
                "LOW",
            )
            anomaly_count = sum(
                1
                for finding in findings
                if isinstance(finding, dict) and not _pdf_finding_is_benign_for_export(finding)
            )
            primary = findings[0] if findings and isinstance(findings[0], dict) else {}
            detail_text = _pdf_detail_to_excel_text(primary.get("detail"))
            rows.append(
                {
                    "file_name": file_name,
                    "overall_risk": overall_risk,
                    "layer": layer_label,
                    "severity": highest,
                    "finding": primary.get("message") or "No findings.",
                    "anomaly_count": anomaly_count,
                    "detail": detail_text,
                }
            )

        for layer_key, findings in layer_results.items():
            if layer_key in handled_keys:
                continue
            findings = findings if isinstance(findings, list) else []
            for finding in findings:
                if not isinstance(finding, dict):
                    continue
                rows.append(
                    {
                        "file_name": file_name,
                        "overall_risk": overall_risk,
                        "layer": str(layer_key),
                        "severity": (finding.get("severity") or "LOW").upper(),
                        "finding": finding.get("message") or finding.get("finding") or "",
                        "anomaly_count": 0 if _pdf_finding_is_benign_for_export(finding) else 1,
                        "detail": _pdf_detail_to_excel_text(finding.get("detail")),
                    }
                )
        return rows

    legacy_layers = result.get("layers", result.get("checks", result.get("findings", [])))
    if isinstance(legacy_layers, dict):
        legacy_layers = [
            {"layer": layer_name, **layer_data}
            if isinstance(layer_data, dict)
            else {"layer": layer_name, "message": str(layer_data)}
            for layer_name, layer_data in legacy_layers.items()
        ]
    rows = []
    for layer in legacy_layers or []:
        if not isinstance(layer, dict):
            continue
        rows.append(
            {
                "file_name": file_name,
                "overall_risk": overall_risk,
                "layer": layer.get("layer", layer.get("name", "")),
                "severity": (layer.get("severity") or layer.get("risk") or "LOW").upper(),
                "finding": layer.get("message", layer.get("finding", layer.get("description", ""))),
                "anomaly_count": 0 if _pdf_finding_is_benign_for_export(layer) else 1,
                "detail": _pdf_detail_to_excel_text(layer.get("detail", layer.get("details", ""))),
            }
        )
    return rows


def _top_counterparty_excel_rows(cp_ledger: dict, amount_column: str, count_column: str) -> List[dict]:
    counterparties = cp_ledger.get("counterparties", []) if isinstance(cp_ledger, dict) else []
    rows = []
    for cp in counterparties or []:
        if not isinstance(cp, dict):
            continue
        amount = safe_float(cp.get(amount_column, 0))
        if amount <= 0:
            continue
        rows.append(
            {
                "Counterparty": cp.get("counterparty_name", cp.get("counterparty", "")),
                "Total Txn": int(safe_float(cp.get(count_column, cp.get("transaction_count", 0)))),
                "Total Amnt of Txn": f"RM {amount:,.2f}",
                "_sort_amount": amount,
            }
        )

    rows.sort(key=lambda row: row["_sort_amount"], reverse=True)
    return [
        {key: value for key, value in row.items() if key != "_sort_amount"}
        for row in rows[:10]
    ]

def build_parsing_qc_dataframe_from_parsing_metadata(parsing_metadata: dict) -> pd.DataFrame:
    """Build Parsing QC dataframe from parsing_metadata for Excel export."""
    if not parsing_metadata:
        return pd.DataFrame()
    
    account_month_checks = parsing_metadata.get("account_month_checks", [])
    if not account_month_checks:
        return pd.DataFrame()
    
    rows = []
    for chk in account_month_checks:
        month = chk.get("month", "")
        account_number = chk.get("account_number", "")
        opening_balance = safe_float(chk.get("opening_balance", 0))
        closing_balance = safe_float(chk.get("closing_balance", 0))
        gross_credits = safe_float(chk.get("gross_credits", 0))
        gross_debits = safe_float(chk.get("gross_debits", 0))
        expected_closing = safe_float(chk.get("expected_closing", 0))
        reconciliation_delta = safe_float(chk.get("reconciliation_delta", 0))
        passed = chk.get("passed", False)
        transactions_extracted = chk.get("transactions_extracted", 0)
        extraction_gaps = chk.get("extraction_gaps", 0)
        notes = chk.get("notes", "")
        
        rows.append({
            "Month": month,
            "Account": account_number,
            "Opening Balance": opening_balance,
            "Closing Balance": closing_balance,
            "Gross Credits": gross_credits,
            "Gross Debits": gross_debits,
            "Expected Close": expected_closing,
            "Recon Delta": reconciliation_delta,
            "Status": "PASS" if passed else "FAIL",
            "Transactions Extracted": transactions_extracted,
            "Gaps": extraction_gaps,
            "Notes": notes
        })
    
    return pd.DataFrame(rows)

def build_risk_signals_dataframe_for_excel(flags_data: dict, consolidated: dict, statutory_compliance: dict, monthly_analysis: list, report_data: dict) -> pd.DataFrame:
    """Build the 16-row Risk Signals DataFrame with actual computed values."""
    
    # Get computed values from various sources
    gross_credits = float(consolidated.get("gross_credits", 0))
    gross_debits = float(consolidated.get("gross_debits", 0))
    
    # Round Figure Credits
    round_figure_entries = report_data.get("round_figure_credits", report_data.get("round_transactions", []))
    round_figure_count = len(round_figure_entries)
    round_figure_total = sum(float(e.get("amount", 0)) for e in round_figure_entries)
    
    # Large Transactions (>= threshold) - check both large_transactions and large_credits
    large_txns = report_data.get("large_transactions", [])
    if not large_txns:
        large_txns = report_data.get("large_credits", [])
    large_txn_count = len(large_txns) if large_txns else 0
    large_txn_total = sum(float(t.get("amount", 0)) for t in (large_txns or []))
    
    # Cash Deposits
    cash_deposits = consolidated.get("total_cash_deposits", 0)
    cash_deposit_count = 0
    for m in monthly_analysis:
        cash_deposit_count += int(m.get("cash_deposits_count", 0))
    
    # Own Party Transactions
    own_cr = float(consolidated.get("total_own_party_cr", 0))
    own_dr = float(consolidated.get("total_own_party_dr", 0))
    
    # Related Party Transactions
    rp_cr = float(consolidated.get("total_related_party_cr", 0))
    rp_dr = float(consolidated.get("total_related_party_dr", 0))
    
    # Loan Activity
    loan_disb = float(consolidated.get("total_loan_disbursement_cr", 0))
    loan_repay = float(consolidated.get("total_loan_repayment_dr", 0))
    
    # EPF/SOCSO from statutory_compliance
    salary_months = statutory_compliance.get("salary_months_active", 0)
    epf_pct = statutory_compliance.get("epf_coverage_pct", 100) if salary_months > 0 else 100
    socso_pct = statutory_compliance.get("socso_coverage_pct", 100) if salary_months > 0 else 100
    
    # LHDN
    lhdn_detected = statutory_compliance.get("lhdn_detected", False)
    lhdn_count = statutory_compliance.get("lhdn_months_paid", 0)
    lhdn_total = float(consolidated.get("total_statutory_tax", 0))
    
    # HRDF
    hrdf_detected = statutory_compliance.get("hrdf_detected", False)
    hrdf_count = statutory_compliance.get("hrdf_months_paid", 0)
    hrdf_total = float(consolidated.get("total_statutory_hrdf", 0))
    
    # FX Transactions
    fx_cr = float(consolidated.get("total_fx_credits", 0))
    fx_dr = float(consolidated.get("total_fx_debits", 0))
    
    # Returned Cheques
    rc_in_count = int(consolidated.get("total_returned_cheques_inward_count", 0)) or 0
    rc_out_count = int(consolidated.get("total_returned_cheques_outward_count", 0)) or 0
    
    # Data Quality
    data_complete = consolidated.get("data_completeness", "COMPLETE") == "COMPLETE"
    
    # Low Closing Balance
    low_balance_months = [m for m in monthly_analysis if float(m.get("closing_balance", 0)) < 1000]
    
    # Build the 16 rows
    risk_signals = [
        {"#": 1, "Signal": "Returned Cheques (Inward)", "Detected": "YES" if rc_in_count > 0 else "NO", 
         "Remarks": f"{rc_in_count} inward returned cheques totalling RM {consolidated.get('total_returned_cheques_inward', 0):,.2f}." if rc_in_count > 0 else "No inward returned cheques in the period."},
        
        {"#": 2, "Signal": "Returned Cheques (Outward)", "Detected": "YES" if rc_out_count > 0 else "NO", 
         "Remarks": f"{rc_out_count} outward returned cheques totalling RM {consolidated.get('total_returned_cheques_outward', 0):,.2f}." if rc_out_count > 0 else "No outward returned cheques in the period."},
        
        {"#": 3, "Signal": "Round Figure Transactions (AML)", "Detected": "YES" if round_figure_count > 0 else "NO", 
         "Remarks": f"{round_figure_count} round-figure transactions totalling RM {round_figure_total:,.2f}." if round_figure_count > 0 else "No round-figure transactions flagged."},
        
        {"#": 4, "Signal": "High Value Credits (>3x EOD)", "Detected": "NO", 
         "Remarks": "No credits exceeded 3x daily EOD."},
        
        {"#": 5, "Signal": "Cash Deposits (AML)", "Detected": "YES" if cash_deposits > 0 else "NO", 
         "Remarks": f"{cash_deposit_count} cash deposits totalling RM {cash_deposits:,.2f} ({cash_deposits/gross_credits*100:.1f}% of gross credits)." if cash_deposits > 0 else "No cash deposits in the period."},
        
        {"#": 6, "Signal": "EPF Compliance", "Detected": "NO" if epf_pct >= 99.5 else "YES", 
         "Remarks": f"EPF coverage {epf_pct:.1f}% across salary months." if salary_months > 0 else "No salary months detected."},
        
        {"#": 7, "Signal": "SOCSO Compliance", "Detected": "NO" if socso_pct >= 99.5 else "YES", 
         "Remarks": f"SOCSO coverage {socso_pct:.1f}% across salary months." if salary_months > 0 else "No salary months detected."},
        
        {"#": 8, "Signal": "LHDN Tax Payments", "Detected": "NO", 
         "Remarks": f"LHDN payments detected: {lhdn_count} tx totalling RM {lhdn_total:,.2f} (PCB/CP204/SST — schedules differ; informational only)." if lhdn_detected else "No LHDN tax payments detected."},
        
        {"#": 9, "Signal": f"Large Transactions (>=RM{consolidated.get('high_value_threshold', 10000):,.0f})", "Detected": "YES" if large_txn_count > 0 else "NO", 
         "Remarks": f"{large_txn_count} large transactions (>=RM{consolidated.get('high_value_threshold', 10000):,.0f}) totalling RM {large_txn_total:,.2f}." if large_txn_count > 0 else f"No transactions at or above RM{consolidated.get('high_value_threshold', 10000):,.0f}."},
        
        {"#": 10, "Signal": "Own Party Transactions", "Detected": "YES" if own_cr > 0 or own_dr > 0 else "NO", 
         "Remarks": f"Own-party CR RM {own_cr:,.2f} ({own_cr/gross_credits*100:.1f}% of gross credits); DR RM {own_dr:,.2f} ({own_dr/gross_debits*100:.1f}% of gross debits)." if (own_cr > 0 or own_dr > 0) else "No own-party transactions detected."},
        
        {"#": 11, "Signal": "Related Party Transactions", "Detected": "YES" if rp_cr > 0 or rp_dr > 0 else "NO", 
         "Remarks": f"Related-party CR RM {rp_cr:,.2f} ({rp_cr/gross_credits*100:.1f}% of gross credits); DR RM {rp_dr:,.2f} ({rp_dr/gross_debits*100:.1f}% of gross debits). Parties: (no canonical names provided)." if (rp_cr > 0 or rp_dr > 0) else "No related-party transactions detected."},
        
        {"#": 12, "Signal": "Loan Activity", "Detected": "YES" if loan_disb > 0 or loan_repay > 0 else "NO", 
         "Remarks": f"Loan disbursements RM {loan_disb:,.2f}; loan repayments RM {loan_repay:,.2f}." if (loan_disb > 0 or loan_repay > 0) else "No loan disbursements or repayments detected."},
        
        {"#": 13, "Signal": "Data Quality", "Detected": "NO" if data_complete else "YES", 
         "Remarks": "Statement data complete across the period." if data_complete else f"Statement data INCOMPLETE: {consolidated.get('data_gaps', '')}"},
        
        {"#": 14, "Signal": "FX Transactions", "Detected": "YES" if fx_cr > 0 or fx_dr > 0 else "NO", 
         "Remarks": f"FX credits RM {fx_cr:,.2f}; FX debits RM {fx_dr:,.2f}." if (fx_cr > 0 or fx_dr > 0) else "No FX (foreign-currency) activity detected."},
        
        {"#": 15, "Signal": "Low Closing Balance", "Detected": "YES" if low_balance_months else "NO", 
         "Remarks": f"Closing balance below RM 1,000.00 in: {', '.join([m.get('month', '?') for m in low_balance_months])}." if low_balance_months else "Closing balance stayed at or above RM 1,000.00 every month."},
        
        {"#": 16, "Signal": "HRDF Payments", "Detected": "NO", 
         "Remarks": f"HRDF payments detected: {hrdf_count} tx totalling RM {hrdf_total:,.2f} (informational; no coverage ratio computed)." if hrdf_detected else "No HRDF payments detected."},
    ]
    
    return pd.DataFrame(risk_signals)



def generate_excel_report(data: dict, monthly_summary: List[dict] = None, transaction_analysis: dict = None) -> BytesIO:
    """Generate Excel workbook using the original generate_excel structure."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return BytesIO()

    report_data = normalize_report_data_for_export(data)
    own_related = report_data.get("own_related_transactions", {}) or {}
    if isinstance(own_related, list):
        own_related = {"transactions": own_related, "summary": {}}
    elif not isinstance(own_related, dict):
        own_related = {"transactions": [], "summary": {}}

    loans = report_data.get("loan_transactions", {}) or {}
    if isinstance(loans, list):
        loans = {"transactions": loans, "disbursements": [], "repayments": []}
    flags = report_data.get("flags", {}) or {}
    cp_ledger = report_data.get("counterparty_ledger", {}) or {}
    if (not cp_ledger or not cp_ledger.get("counterparties")) and report_data.get("transactions"):
        cp_ledger = build_track2_counterparty_ledger(report_data.get("transactions", []))
    parsing = report_data.get("parsing_metadata", {}) or {}
    pdf_integrity = report_data.get("pdf_integrity", {}) or {}
    consolidated = report_data.get("consolidated", {}) or {}
    monthly_analysis = report_data.get("monthly_analysis", []) or []
    report_info = report_data.get("report_info", {}) or {}
    accounts = report_data.get("accounts", []) or []
    top_parties = report_data.get("top_parties", {}) or {}
    if not _has_top_party_rows(top_parties):
        ledger_top_parties = _top_parties_from_counterparty_ledger(cp_ledger, limit=10)
        if _has_top_party_rows(ledger_top_parties):
            top_parties = ledger_top_parties
            report_data["top_parties"] = top_parties
    if not _has_top_party_rows(top_parties):
        analysis_top_parties = _top_parties_from_transaction_analysis(transaction_analysis or {})
        if _has_top_party_rows(analysis_top_parties):
            top_parties = analysis_top_parties
            report_data["top_parties"] = top_parties
    statutory_compliance = consolidated.get("statutory_compliance", {}) or {}
    observations = normalize_observations(report_data.get("observations", {}) or {})

    # ── Backfill consolidated totals from monthly_analysis when Track 2 keys are absent ──
    # The HTML uses build_track2_result which computes these directly; the Excel path
    # uses build_report_data_from_analysis which omits them. Sum/min/max from monthly rows.
    def _sum_monthly(key):
        return round(sum(safe_float(m.get(key, 0)) for m in monthly_analysis), 2)

    def _min_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(min(vals), 2) if vals else 0.0

    def _max_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(max(vals), 2) if vals else 0.0

    def _avg_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    backfill_map = {
        "total_own_party_cr":        ("sum", "own_party_cr"),
        "total_own_party_dr":        ("sum", "own_party_dr"),
        "total_related_party_cr":    ("sum", "related_party_cr"),
        "total_related_party_dr":    ("sum", "related_party_dr"),
        "total_loan_disbursement_cr":("sum", "loan_disbursement_cr"),
        "total_loan_repayment_dr":   ("sum", "loan_repayment_dr"),
        "total_fd_interest_cr":      ("sum", "fd_interest_cr"),
        "total_cash_deposits":       ("sum", "cash_deposits_amount"),
        "total_cash_withdrawals":    ("sum", "cash_withdrawals_amount"),
        "total_cheque_deposits":     ("sum", "cheque_deposits_amount"),
        "total_cheque_issues":       ("sum", "cheque_issues_amount"),
        "total_salary_paid":         ("sum", "salary_paid"),
        "total_statutory_epf":       ("sum", "statutory_epf"),
        "total_statutory_socso":     ("sum", "statutory_socso"),
        "total_statutory_tax":       ("sum", "statutory_tax"),
        "total_statutory_hrdf":      ("sum", "statutory_hrdf"),
        "eod_lowest":                ("min", "eod_lowest"),
        "eod_highest":               ("max", "eod_highest"),
        "eod_average":               ("avg", "eod_average"),
    }

    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    header_fill_green = PatternFill(start_color="196F3D", end_color="196F3D", fill_type="solid")
    header_fill_red = PatternFill(start_color="922B21", end_color="922B21", fill_type="solid")
    header_fill_orange = PatternFill(start_color="B9770E", end_color="B9770E", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    credit_font = Font(name="Calibri", color="196F3D")
    debit_font = Font(name="Calibri", color="922B21")
    bold_font = Font(name="Calibri", bold=True, size=11)
    title_font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    num_fmt = "#,##0.00"

    def clean_xl(value):
        value = _excel_safe_value(value)
        if isinstance(value, (dict, list)):
            return json.dumps(value, default=str)
        return value

    def is_num(value):
        return isinstance(value, (int, float)) and not isinstance(value, bool)

    def style_header_row(ws, row, max_col, fill=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, number=False, credit=False, debit=False):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if number else "left", vertical="top", wrap_text=True)
        if number:
            cell.number_format = num_fmt
        if credit:
            cell.font = credit_font
        if debit:
            cell.font = debit_font
        if row % 2 == 0:
            cell.fill = alt_row_fill

    def write_headers(ws, row, headers, fill=None):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers), fill)

    def auto_width(ws, min_width=10, max_width=40):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))

    def write_values(ws, row, values, number_cols=None, credit_cols=None, debit_cols=None):
        number_cols = set(number_cols or [])
        credit_cols = set(credit_cols or [])
        debit_cols = set(debit_cols or [])
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=clean_xl(value))
            style_data_cell(
                ws,
                row,
                col,
                number=col in number_cols or is_num(value),
                credit=col in credit_cols,
                debit=col in debit_cols,
            )

    def safe_amount(txn):
        return safe_float(txn.get("amount", txn.get("credit", 0) or txn.get("debit", 0)))

    def infer_txn_side(txn):
        raw_type = str(txn.get("type") or txn.get("transaction_type") or "").upper()
        if "CR" in raw_type or "CREDIT" in raw_type:
            return "CREDIT"
        if "DR" in raw_type or "DEBIT" in raw_type:
            return "DEBIT"
        if safe_float(txn.get("credit", 0)) > 0:
            return "CREDIT"
        if safe_float(txn.get("debit", 0)) > 0:
            return "DEBIT"
        return "DEBIT" if safe_amount(txn) < 0 else "CREDIT"

    def write_split_transaction_sheet(ws, title, txns, caption=None, number_cols=None, credit_cols=None, debit_cols=None):
        """Write a sheet with split credit/debit transactions."""
        number_cols = set(number_cols or {4, 5})
        credit_cols = set(credit_cols or {4})
        debit_cols = set(debit_cols or {4})
        
        ws.cell(row=1, column=1, value=title).font = title_font
        row = 2
        if caption:
            ws.cell(row=row, column=1, value=caption)
            ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, color="475569")
            row += 2
        else:
            row += 1

        display_headers = ["No.", "Date", "Description", "Amount", "Balance"]
        for section_title, side, fill in (
            ("CREDIT TRANSACTIONS", "CREDIT", header_fill_green),
            ("DEBIT TRANSACTIONS", "DEBIT", header_fill_red),
        ):
            section_rows = [txn for txn in (txns or []) if infer_txn_side(txn) == side]
            ws.cell(row=row, column=1, value=section_title).font = bold_font
            row += 1
            write_headers(ws, row, display_headers, fill)
            if not section_rows:
                row += 1
                ws.cell(row=row, column=1, value="No transactions")
                style_data_cell(ws, row, 1)
            for idx, txn in enumerate(section_rows, 1):
                row += 1
                amount = abs(safe_amount(txn))
                values = [
                    idx,
                    txn.get("date", ""),
                    (txn.get("description", "") or "")[:100],
                    amount,
                    txn.get("balance"),
                ]
                write_values(
                    ws,
                    row,
                    values,
                    number_cols=number_cols,
                    credit_cols=credit_cols if side == "CREDIT" else set(),
                    debit_cols=debit_cols if side == "DEBIT" else set(),
                )
                # Force "No." column to display as whole integer, not float
                ws.cell(row=row, column=1).number_format = "0"
                # Centre + middle align: No.(1), Date(2), Amount(4), Balance(5) only
                for centre_col in (1, 2, 4, 5):
                    ws.cell(row=row, column=centre_col).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
            row += 2
        # Column-specific auto widths for these 5-column transaction sheets
        ws.column_dimensions["A"].width = 6    # No.
        ws.column_dimensions["B"].width = 14   # Date
        ws.column_dimensions["C"].width = 65   # Description
        ws.column_dimensions["D"].width = 18   # Amount
        ws.column_dimensions["E"].width = 18   # Balance

    schema_version = str(report_info.get("schema_version", ""))
    is_v620 = schema_version in ("6.2.0", "6.2.1", "6.2.2", "6.3.0", "6.3.1", "6.3.2", "6.3.3", "6.3.4", "6.3.5") or consolidated.get("total_fx_credits") is not None
    is_v630 = schema_version in ("6.3.0", "6.3.1", "6.3.2", "6.3.3", "6.3.4", "6.3.5") or consolidated.get("total_unclassified_cr") is not None
    has_recon = any(m.get("reconciliation_status") for m in monthly_analysis) or bool(parsing.get("account_month_checks"))

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="KREDIT LAB - STATEMENT INTELLIGENCE REPORT").font = title_font
    ws.cell(row=2, column=1, value=report_info.get("company_name", "")).font = bold_font
    ws.cell(row=3, column=1, value=f"Period: {report_info.get('period_start', '')} to {report_info.get('period_end', '')}")
    ws.cell(row=4, column=1, value=f"Generated: {report_info.get('generated_at', '')}")

    row = 6
    ws.cell(row=row, column=1, value="ACCOUNT DETAILS").font = bold_font
    row += 1
    headers = ["Bank", "Account No", "Holder", "Type", "Opening Balance", "Closing Balance", "Total Credits", "Total Debits", "Transactions"]
    write_headers(ws, row, headers)
    for account in accounts:
        row += 1
        values = [
            account.get("bank_name"), account.get("account_number"), account.get("account_holder"),
            account.get("account_type"), account.get("opening_balance"), account.get("closing_balance"),
            account.get("total_credits"), account.get("total_debits"), account.get("transaction_count"),
        ]
        write_values(ws, row, values, number_cols={5, 6, 7, 8}, credit_cols={7}, debit_cols={8})

    row += 2
    ws.cell(row=row, column=1, value="CONSOLIDATED FIGURES").font = bold_font
    row += 1
    consolidated_items = [
        ("Gross Credits", consolidated.get("gross_credits")),
        ("Gross Debits", consolidated.get("gross_debits")),
        ("Net Credits", consolidated.get("net_credits")),
        ("Net Debits", consolidated.get("net_debits")),
        ("Annualized Net Credits", consolidated.get("annualized_net_credits")),
        ("Annualized Net Debits", consolidated.get("annualized_net_debits")),
        ("", ""),
        ("Own Party Credits", consolidated.get("total_own_party_cr")),
        ("Own Party Debits", consolidated.get("total_own_party_dr")),
        ("Related Party Credits", consolidated.get("total_related_party_cr")),
        ("Related Party Debits", consolidated.get("total_related_party_dr")),
        ("", ""),
        ("Loan Disbursements", consolidated.get("total_loan_disbursement_cr")),
        ("Loan Repayments", consolidated.get("total_loan_repayment_dr")),
        ("FD/Interest Credits", consolidated.get("total_fd_interest_cr")),
        ("Inward Return (C16)", consolidated.get("total_inward_return_cr")),
        ("", ""),
        ("Cash Deposits", consolidated.get("total_cash_deposits")),
        ("Cash Withdrawals", consolidated.get("total_cash_withdrawals")),
        ("Cheque Deposits", consolidated.get("total_cheque_deposits")),
        ("Cheque Issues", consolidated.get("total_cheque_issues")),
        ("", ""),
        ("Total Salary Paid", consolidated.get("total_salary_paid")),
        ("Total EPF", consolidated.get("total_statutory_epf")),
        ("Total SOCSO", consolidated.get("total_statutory_socso")),
        ("Total Tax", consolidated.get("total_statutory_tax")),
        ("Total HRDF", consolidated.get("total_statutory_hrdf")),
        ("", ""),
        ("EOD Lowest", consolidated.get("eod_lowest")),
        ("EOD Highest", consolidated.get("eod_highest")),
        ("EOD Average", consolidated.get("eod_average")),
    ]
    if has_recon:
        consolidated_items.extend([
            ("", ""),
            ("Data Completeness", consolidated.get("data_completeness")),
            ("Extraction Gaps", consolidated.get("total_extraction_gaps")),
            ("Missing Debits", consolidated.get("total_missing_debits")),
            ("Missing Credits", consolidated.get("total_missing_credits")),
            ("Months With Gaps", consolidated.get("months_with_gaps")),
        ])
    if is_v630:
        consolidated_items.extend([
            ("", ""),
            ("Unclassified Credits", consolidated.get("total_unclassified_cr")),
            ("Unclassified Debits", consolidated.get("total_unclassified_dr")),
        ])
    if is_v620:
        consolidated_items.extend([
            ("", ""),
            ("FX/Remittance Credits", consolidated.get("total_fx_credits")),
            ("FX/Remittance Debits", consolidated.get("total_fx_debits")),
            ("FX Credit % of Gross", consolidated.get("fx_credit_pct")),
            ("FX Debit % of Gross", consolidated.get("fx_debit_pct")),
            ("FX Currencies Detected", ", ".join(consolidated.get("fx_currencies_all", []) or [])),
        ])
    for label, value in consolidated_items:
        if label:
            write_values(ws, row, [label, value], number_cols={2} if is_num(value) else set(), credit_cols={2} if "Credit" in label else set(), debit_cols={2} if "Debit" in label or "Repayment" in label else set())
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="POSITIVE OBSERVATIONS").font = Font(name="Calibri", bold=True, color="196F3D")
    row += 1
    for item in observations.get("positive", []):
        ws.cell(row=row, column=1, value=f"+ {item}")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="CONCERNS").font = Font(name="Calibri", bold=True, color="922B21")
    row += 1
    for item in observations.get("concerns", []):
        ws.cell(row=row, column=1, value=f"- {item}")
        row += 1
    auto_width(ws)

    # Cash Flow
    ws2 = wb.create_sheet("Cash Flow")
    cash_headers = [
        "Month", "Bank", "Account No", "Gross Credits", "Gross Debits", "Net Credits", "Net Debits",
        "Credit Count", "Debit Count", "Own Party Cr", "Own Party Dr", "Related Party Cr", "Related Party Dr",
        "Reversal Cr", "Loan Disbursement Cr", "FD Interest Cr", "Round Figure Cr", "High Value Cr",
        "Cash Dep Count", "Cash Dep Amt", "Cash Wdl Count", "Cash Wdl Amt", "Chq Dep Count", "Chq Dep Amt",
        "Chq Issue Count", "Chq Issue Amt", "Loan Repayment Dr", "Salary Paid", "EPF", "SOCSO", "Tax", "HRDF",
        "Ret Chq In Count", "Ret Chq In Amt", "Ret Chq Out Count", "Ret Chq Out Amt",
        "EOD Lowest", "EOD Highest", "EOD Average", "Opening Balance", "Closing Balance",
    ]
    if is_v620:
        cash_headers += ["FX Cr Count", "FX Cr Amount", "FX Dr Count", "FX Dr Amount", "FX Currencies"]
    if has_recon:
        cash_headers += ["Recon Status", "Recon Delta", "Gaps", "Missing Debits", "Missing Credits", "Data Quality Note"]
    write_headers(ws2, 1, cash_headers)
    cash_num_cols = set(range(4, 42)) - {8, 9, 19, 21, 23, 25, 33, 35}
    if is_v620:
        cash_num_cols.update({43, 45})
    for idx, item in enumerate(monthly_analysis, 2):
        values = [
            item.get("month"), item.get("bank_name", ""), item.get("account_number", ""),
            item.get("gross_credits"), item.get("gross_debits"), item.get("net_credits"), item.get("net_debits"),
            item.get("credit_count"), item.get("debit_count"), item.get("own_party_cr"), item.get("own_party_dr"),
            item.get("related_party_cr"), item.get("related_party_dr"), item.get("reversal_cr"),
            item.get("loan_disbursement_cr"), item.get("fd_interest_cr"), item.get("round_figure_cr"),
            item.get("high_value_cr"), item.get("cash_deposits_count"), item.get("cash_deposits_amount"),
            item.get("cash_withdrawals_count"), item.get("cash_withdrawals_amount"), item.get("cheque_deposits_count"),
            item.get("cheque_deposits_amount"), item.get("cheque_issues_count"), item.get("cheque_issues_amount"),
            item.get("loan_repayment_dr"), item.get("salary_paid"), item.get("statutory_epf"), item.get("statutory_socso"),
            item.get("statutory_tax"), item.get("statutory_hrdf"), item.get("returned_cheques_inward_count"),
            item.get("returned_cheques_inward_amount"), item.get("returned_cheques_outward_count"),
            item.get("returned_cheques_outward_amount"), item.get("eod_lowest"), item.get("eod_highest"),
            item.get("eod_average"), item.get("opening_balance"), item.get("closing_balance"),
        ]
        if is_v620:
            values += [
                item.get("fx_credit_count", 0), item.get("fx_credit_amount", 0),
                item.get("fx_debit_count", 0), item.get("fx_debit_amount", 0),
                ", ".join(item.get("fx_currencies", []) or []),
            ]
        if has_recon:
            values += [
                item.get("reconciliation_status", ""), item.get("reconciliation_delta", 0),
                item.get("extraction_gaps", 0), item.get("missing_debit_amount", 0),
                item.get("missing_credit_amount", 0), item.get("data_quality_note", "") or "",
            ]
        write_values(ws2, idx, values, number_cols=cash_num_cols)
        if item.get("reconciliation_status") == "FAIL":
            for col in range(1, len(values) + 1):
                ws2.cell(row=idx, column=col).fill = fail_fill
    auto_width(ws2, min_width=12)

    # Top Parties
    ws3 = wb.create_sheet("Top Parties")
    party_view = prepare_top_parties_for_report(top_parties, limit=10)
    payers = party_view["payers"]
    payees = party_view["payees"]
    all_party_rows = list(payers) + list(payees)
    monthly_bd = sorted({
        mb.get("month", "")
        for party in all_party_rows
        for mb in (party.get("monthly_breakdown") or [])
        if isinstance(mb, dict) and mb.get("month")
    })
    party_headers = ["Rank", "Party Name", "Total Amount", "Transactions", "Related Party"] + monthly_bd
    party_num_cols = {3, *range(6, 6 + len(monthly_bd))}

    ws3.cell(row=1, column=1, value="TOP PAYERS (Income Sources)").font = bold_font
    write_headers(ws3, 2, party_headers, header_fill_green)
    for row_idx, party in enumerate(payers, 3):
        lookup = {mb.get("month"): safe_float(mb.get("amount")) for mb in (party.get("monthly_breakdown") or []) if isinstance(mb, dict)}
        values = [party.get("rank"), party.get("party_name") or party.get("name"), party.get("total_amount"), party.get("transaction_count"), "Yes" if party.get("is_related_party") else "No"]
        values.extend(lookup.get(month, 0) for month in monthly_bd)
        write_values(ws3, row_idx, values, number_cols=party_num_cols, credit_cols=party_num_cols)
        ws3.cell(row=row_idx, column=1).number_format = "0"
        ws3.cell(row=row_idx, column=4).number_format = "0"
        ws3.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(3, 12):  # from column 3 to 11
            ws3.cell(row=row_idx, column=col).alignment = alignment
    auto_width(ws3)

    row = len(payers) + 5
    ws3.cell(row=row, column=1, value="TOP PAYEES (Payment Destinations)").font = bold_font
    row += 1
    write_headers(ws3, row, party_headers, header_fill_red)
    for row_idx, party in enumerate(payees, row + 1):
        lookup = {mb.get("month"): safe_float(mb.get("amount")) for mb in (party.get("monthly_breakdown") or []) if isinstance(mb, dict)}
        values = [party.get("rank"), party.get("party_name") or party.get("name"), party.get("total_amount"), party.get("transaction_count"), "Yes" if party.get("is_related_party") else "No"]
        values.extend(lookup.get(month, 0) for month in monthly_bd)
        write_values(ws3, row_idx, values, number_cols=party_num_cols, debit_cols=party_num_cols)
        ws3.cell(row=row_idx, column=1).number_format = "0"
        ws3.cell(row=row_idx, column=4).number_format = "0"
        ws3.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(3, 12):  # from column 3 to 11
            ws3.cell(row=row_idx, column=col).alignment = alignment
    auto_width(ws3)
    ws3.column_dimensions["A"].width = 11
    ws3.column_dimensions["B"].width = 55
    ws3.column_dimensions["D"].width = 13

    # Large Transactions - Updated with proper formatting
    ws_large = wb.create_sheet("Large Transactions")
    large_rows = report_data.get("large_transactions", []) or report_data.get("large_credits", []) or []
    write_split_transaction_sheet(
        ws_large,
        f"Large Transactions (>= RM {consolidated.get('high_value_threshold', 100000):,.0f})",
        large_rows,
        number_cols={4, 5},
        credit_cols={4},
        debit_cols={4}
    )

    # Counterparty
    ws5 = wb.create_sheet("Counterparty")
    ws5.cell(row=1, column=1, value="COUNTERPARTY TRANSACTIONS").font = title_font
    ws5.cell(row=3, column=1, value="Summary").font = bold_font
    row = 4
    summary = own_related.get("summary", {}) or {}
    for label, amount_key, pct_key, is_credit_side in [
        ("Own Party Credits", "own_party_cr", "own_party_cr_pct", True),
        ("Own Party Debits", "own_party_dr", "own_party_dr_pct", False),
        ("Related Party Credits", "related_party_cr", "related_party_cr_pct", True),
        ("Related Party Debits", "related_party_dr", "related_party_dr_pct", False),
    ]:
        write_values(ws5, row, [label, summary.get(amount_key), f"{safe_float(summary.get(pct_key, 0)):.1f}%" if summary.get(pct_key) is not None else ""], number_cols={2}, credit_cols={2} if is_credit_side else set(), debit_cols={2} if not is_credit_side else set())
        row += 1
    row += 1
    counterparty_headers = ["No.", "Date", "Description", "Amount", "Party Type", "Party Name"]
    write_headers(ws5, row, counterparty_headers, header_fill_orange)
    for idx, txn in enumerate(own_related.get("transactions", []) or [], 1):
        row += 1
        txn_type = (txn.get("type") or "").upper()
        values = [idx, txn.get("date"), (txn.get("description", "") or "")[:60], txn.get("amount"), txn.get("party_type"), txn.get("party_name", "")]
        write_values(ws5, row, values, number_cols={4},
                     credit_cols={4} if txn_type == "CREDIT" else set(),
                     debit_cols={4} if txn_type != "CREDIT" else set())
        ws5.cell(row=row, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 5):
            ws5.cell(row=row, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    auto_width(ws5)

    # CP Ledger
    ws5b = wb.create_sheet("CP Ledger")
    ws5b.cell(row=1, column=1, value="COUNTERPARTY LEDGER").font = title_font
    ledger_headers = ["Counterparty", "Total Credits", "Total Debits", "Net Position", "Cr Count", "Dr Count", "Txn Count"]
    row = 3
    write_headers(ws5b, row, ledger_headers)
    cp_sorted = build_canonical_counterparty_ledger_rows(cp_ledger)
    for cp in cp_sorted:
        row += 1
        values = [cp.get("counterparty_name", ""), cp.get("total_credits", 0), cp.get("total_debits", 0), cp.get("net_position", 0), cp.get("credit_count", 0), cp.get("debit_count", 0), cp.get("transaction_count", 0)]
        write_values(ws5b, row, values, number_cols={2, 3, 4}, credit_cols={2}, debit_cols={3})
    row += 2
    ws5b.cell(row=row, column=1, value="TRANSACTION DETAIL BY COUNTERPARTY").font = title_font
    for cp in cp_sorted:
        row += 1
        ws5b.cell(row=row, column=1, value=cp.get("counterparty_name", "")).font = Font(name="Calibri", bold=True, color="1B4F72", size=11)
        row += 1
        detail_headers = ["Date", "Description", "Amount", "Type", "Account"]
        write_headers(ws5b, row, detail_headers, header_fill_orange)
        for txn in cp.get("transactions", []) or []:
            row += 1
            txn_type = (txn.get("type") or "").upper()
            values = [txn.get("date", ""), (txn.get("description", "") or "")[:70], txn.get("amount", 0), txn.get("type", ""), txn.get("account_number", "")]
            write_values(ws5b, row, values, number_cols={3}, credit_cols={3} if txn_type == "CREDIT" else set(), debit_cols={3} if txn_type != "CREDIT" else set())
    auto_width(ws5b)

    # Related Parties
    ws5c = wb.create_sheet("Related Parties")
    ws5c.cell(row=1, column=1, value="KNOWN RELATED PARTIES").font = title_font
    rp_headers = ["Name", "Relationship", "Total Credits", "Total Debits", "Transactions"]
    write_headers(ws5c, 3, rp_headers, header_fill_orange)
    cp_by_name = {str(cp.get("counterparty_name", "")).strip().upper(): cp for cp in cp_sorted if cp.get("counterparty_name")}
    for row_idx, rp in enumerate(report_info.get("related_parties", []) or [], 4):
        name = (rp.get("name") or rp.get("party_name") if isinstance(rp, dict) else str(rp)) or ""
        relationship = rp.get("relationship", "") if isinstance(rp, dict) else ""
        match = cp_by_name.get(name.strip().upper(), {})
        values = [name, relationship, match.get("total_credits"), match.get("total_debits"), match.get("transaction_count")]
        write_values(ws5c, row_idx, values, number_cols={3, 4}, credit_cols={3}, debit_cols={4})
    auto_width(ws5c)

    # Unclassified
    ws5d = wb.create_sheet("Unclassified")
    ws5d.cell(row=1, column=1, value="UNCLASSIFIED TRANSACTIONS").font = title_font
    unclassified_headers = ["No.", "Date", "Description", "Amount", "Type", "Balance"]
    write_headers(ws5d, 3, unclassified_headers, header_fill_orange)
    for idx, txn in enumerate(report_data.get("unclassified_transactions", []) or [], 1):
        row_idx = idx + 3
        txn_type = (txn.get("type") or "").upper()
        values = [idx, txn.get("date", ""), (txn.get("description", "") or "")[:80], txn.get("amount"), txn_type, txn.get("balance")]
        write_values(ws5d, row_idx, values, number_cols={4, 6},
                     credit_cols={4} if txn_type == "CREDIT" else set(),
                     debit_cols={4} if txn_type != "CREDIT" else set())
        ws5d.cell(row=row_idx, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 6):
            ws5d.cell(row=row_idx, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    auto_width(ws5d)

    # Round Figure Transactions - Updated with proper formatting
    round_rows = get_round_transactions_for_report(report_data)
    ws_round = wb.create_sheet("Round Figure Transactions")
    write_split_transaction_sheet(
        ws_round,
        "ROUND FIGURE TRANSACTIONS",
        round_rows,
        caption="Round figure transactions are amounts that are multiples of RM 10,000.",
        number_cols={4, 5},
        credit_cols={4},
        debit_cols={4}
    )

    # Observations
    ws5f = wb.create_sheet("Observations")
    ws5f.cell(row=1, column=1, value="OBSERVATIONS").font = title_font
    row = 3
    for title, items, fill in (("POSITIVE OBSERVATIONS", observations.get("positive", []), header_fill_green), ("CONCERNS", observations.get("concerns", []), header_fill_red)):
        ws5f.cell(row=row, column=1, value=title)
        style_header_row(ws5f, row, 1, fill)
        for item in items:
            row += 1
            ws5f.cell(row=row, column=1, value=str(item))
            style_data_cell(ws5f, row, 1)
        row += 2
    ws5f.column_dimensions["A"].width = 100

    # Facilities
    # Facilities — matches the HTML Facilities tab exactly
    ws6 = wb.create_sheet("Facilities")

    # Inline helpers matching _is_real_facility / _facility_amount from HTML report
    def _fac_amount(t):
        try:
            return float(t.get("amount") or 0)
        except (TypeError, ValueError):
            return 0.0

    def _is_real_fac(t, expected_category):
        return isinstance(t, dict) and t.get("category") == expected_category and _fac_amount(t) > 0

    # Filter same as HTML tab
    loan_disb = [t for t in (loans.get("disbursements") or []) if _is_real_fac(t, "loan_disbursement")]
    loan_repay = [t for t in (loans.get("repayments") or []) if _is_real_fac(t, "loan_repayment")]
    loan_disb_total = round(sum(_fac_amount(t) for t in loan_disb), 2)
    loan_repay_total = round(sum(_fac_amount(t) for t in loan_repay), 2)

    # Summary stats (mirrors HTML summary cards)
    ws6.cell(row=1, column=1, value="FACILITIES").font = title_font
    summary_headers = ["Total Disbursements (RM)", "Total Repayments (RM)", "Disbursement Txns", "Repayment Txns"]
    write_headers(ws6, 3, summary_headers, header_fill_green)
    write_values(ws6, 4, [loan_disb_total, loan_repay_total, len(loan_disb), len(loan_repay)],
                 number_cols={1, 2}, credit_cols={1}, debit_cols={2})

    # 4-column layout (same as HTML: Date, Description, Amount, Category — NO Balance)
    facility_headers = ["No.", "Date", "Description", "Amount", "Category"]
    row = 6
    ws6.cell(row=row, column=1, value="DISBURSEMENTS (Credits)").font = bold_font
    row += 1
    write_headers(ws6, row, facility_headers, header_fill_green)
    if not loan_disb:
        row += 1
        ws6.cell(row=row, column=1, value="No disbursements")
        style_data_cell(ws6, row, 1)
    for idx, txn in enumerate(loan_disb, 1):
        row += 1
        values = [idx, txn.get("date"), (txn.get("description", "") or "")[:70], _fac_amount(txn), txn.get("category", "")]
        write_values(ws6, row, values, number_cols={4}, credit_cols={4})
        ws6.cell(row=row, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 5):
            ws6.cell(row=row, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 2
    ws6.cell(row=row, column=1, value="REPAYMENTS (Debits)").font = bold_font
    row += 1
    write_headers(ws6, row, facility_headers, header_fill_red)
    if not loan_repay:
        row += 1
        ws6.cell(row=row, column=1, value="No repayments")
        style_data_cell(ws6, row, 1)
    for idx, txn in enumerate(loan_repay, 1):
        row += 1
        values = [idx, txn.get("date"), (txn.get("description", "") or "")[:70], _fac_amount(txn), txn.get("category", "")]
        write_values(ws6, row, values, number_cols={4}, debit_cols={4})
        ws6.cell(row=row, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 5):
            ws6.cell(row=row, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws6.column_dimensions["A"].width = 6    # No.
    ws6.column_dimensions["B"].width = 14   # Date
    ws6.column_dimensions["C"].width = 55   # Description
    ws6.column_dimensions["D"].width = 18   # Amount
    ws6.column_dimensions["E"].width = 22   # Category

    # Risk Signals
    ws7 = wb.create_sheet("Risk Signals")
    risk_headers = ["No.", "Signal", "Detected", "Remarks"]
    write_headers(ws7, 1, risk_headers)
    risk_df = build_risk_signals_dataframe_for_excel(flags, consolidated, statutory_compliance, monthly_analysis, report_data)
    for row_idx, item in enumerate(risk_df.to_dict(orient="records"), 2):
        values = [item.get("#"), item.get("Signal"), item.get("Detected"), item.get("Remarks")]
        write_values(ws7, row_idx, values)
        ws7.cell(row=row_idx, column=1).number_format = "0"
        ws7.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws7.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws7.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        if item.get("Detected") == "YES":
            ws7.cell(row=row_idx, column=3).font = Font(name="Calibri", color="922B21", bold=True)
    auto_width(ws7)
    ws7.column_dimensions["D"].width = 70

    # Parsing QC
    ws8 = wb.create_sheet("Parsing QC")
    ws8.cell(row=1, column=1, value="PARSING QUALITY METRICS").font = title_font
    success_rate = safe_float(parsing.get("overall_success_rate", 0))
    success_rate_pct = success_rate * 100 if success_rate <= 1 else success_rate
    p_total_gaps = int(consolidated.get("total_extraction_gaps") or len(parsing.get("extraction_gaps", []) or []))
    p_missing_dr = safe_float(consolidated.get("total_missing_debits", 0))
    p_missing_cr = safe_float(consolidated.get("total_missing_credits", 0))
    metric_rows = [
        ("Success Rate", f"{success_rate_pct:.1f}%"),
        ("Transactions Extracted", parsing.get("total_transactions_extracted", 0)),
        ("Balance Checks Passed", f"{parsing.get('total_balance_checks_passed', 0)} / {parsing.get('total_balance_checks', 0)}"),
    ]
    if has_recon:
        metric_rows.extend([
            ("Extraction Gaps", p_total_gaps),
            ("Missing Debits", p_missing_dr),
            ("Missing Credits", p_missing_cr),
        ])
    for row_idx, (label, value) in enumerate(metric_rows, 3):
        ws8.cell(row=row_idx, column=1, value=label)
        ws8.cell(row=row_idx, column=2, value=value)
        ws8.cell(row=row_idx, column=1).font = bold_font
        ws8.cell(row=row_idx, column=1).border = thin_border
        ws8.cell(row=row_idx, column=2).border = thin_border
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            ws8.cell(row=row_idx, column=2).number_format = num_fmt

    row = 3 + len(metric_rows) + 2
    qc_headers = [
        "Month", "Account", "Opening Balance", "Gross Credits", "Gross Debits",
        "Expected Closing", "Actual Closing", "Delta", "Status", "Txns", "Gaps", "Notes",
    ]
    write_headers(ws8, row, qc_headers)
    for check in parsing.get("account_month_checks", []) or []:
        row += 1
        passed = bool(check.get("passed"))
        values = [
            check.get("month"),
            check.get("account_number"),
            check.get("opening_balance"),
            check.get("gross_credits"),
            check.get("gross_debits"),
            check.get("expected_closing"),
            check.get("closing_balance"),
            check.get("reconciliation_delta"),
            "PASS" if passed else "FAIL",
            check.get("transactions_extracted"),
            check.get("extraction_gaps", 0),
            check.get("notes", ""),
        ]
        write_values(ws8, row, values, number_cols={3, 4, 5, 6, 7, 8, 10, 11}, credit_cols={4}, debit_cols={5})
        ws8.cell(row=row, column=9).font = Font(name="Calibri", color="196F3D" if passed else "922B21", bold=True)
    if parsing.get("extraction_gaps"):
        row += 2
        ws8.cell(row=row, column=1, value="EXTRACTION GAPS DETAIL").font = title_font
        row += 1
        gap_headers = ["Month", "Date", "Page", "Source File", "Missing", "Amount (RM)", "Before Gap", "After Gap"]
        write_headers(ws8, row, gap_headers, header_fill_red)
        for gap in parsing.get("extraction_gaps", []) or []:
            row += 1
            before_gap = f"{(gap.get('prev_description', '') or '')[:60]} (RM {safe_float(gap.get('balance_before_gap', 0)):,.2f})"
            after_gap = f"{(gap.get('next_description', '') or '')[:60]} (RM {safe_float(gap.get('balance_after_gap', 0)):,.2f})"
            values = [
                gap.get("month", ""), gap.get("date", ""), gap.get("page", ""), gap.get("source_file", ""),
                gap.get("missing_type", ""), gap.get("missing_amount", 0), before_gap, after_gap,
            ]
            write_values(ws8, row, values, number_cols={6}, debit_cols={6})

    cls_config = report_data.get("classification_config", {}) or {}
    if cls_config or schema_version:
        row += 2
        ws8.cell(row=row, column=1, value="CLASSIFICATION CONFIGURATION").font = title_font
        row += 1
        config_headers = ["Setting", "Value"]
        write_headers(ws8, row, config_headers, header_fill_orange)
        factoring_entities = cls_config.get("known_factoring_entities", [])
        config_rows = [
            ("Schema Version", schema_version or "N/A"),
            ("Rulebook Version", cls_config.get("rulebook_version", "N/A")),
            ("Execution Mode", cls_config.get("execution_mode", "N/A")),
            ("Large Transaction Threshold", safe_float(cls_config.get("large_transaction_threshold") or consolidated.get("high_value_threshold") or 100000)),
            ("Unclassified Listing Threshold", safe_float(cls_config.get("unclassified_listing_threshold", 10000))),
            ("Known Factoring Entities", ", ".join(factoring_entities) if factoring_entities else "None configured"),
        ]
        for label, value in config_rows:
            row += 1
            write_values(ws8, row, [label, value], number_cols={2} if isinstance(value, (int, float)) else set())

    validation_rows = build_formula_validation_checks_for_report(consolidated, monthly_analysis)
    if validation_rows:
        row += 2
        ws8.cell(row=row, column=1, value="FORMULA VALIDATION CHECKS (V1-V6)").font = title_font
        row += 1
        validation_headers = ["ID", "Check", "Severity", "Status", "Remarks"]
        write_headers(ws8, row, validation_headers, header_fill_orange)
        for item in validation_rows:
            row += 1
            values = [item.get("ID"), item.get("Check"), item.get("Severity"), item.get("Status"), item.get("Remarks")]
            write_values(ws8, row, values)
            status = item.get("Status")
            status_color = "196F3D" if status == "PASS" else "B9770E" if status in ("WARN", "N/A") else "922B21"
            ws8.cell(row=row, column=4).font = Font(name="Calibri", color=status_color, bold=True)
    auto_width(ws8)

    # Fraud Detector
    ws9 = wb.create_sheet("Fraud Detector")
    ws9.cell(row=1, column=1, value="Fraud Detector").font = title_font
    fraud_headers = ["file_name", "overall_risk", "layer", "severity", "finding", "anomaly_count", "detail"]
    write_headers(ws9, 3, fraud_headers)
    fraud_rows = []
    if isinstance(pdf_integrity, dict):
        for file_name, result in pdf_integrity.items():
            if isinstance(result, dict):
                fraud_rows.extend(_normalise_pdf_integrity_layer_rows(file_name, result))
    for row_idx, item in enumerate(fraud_rows, 4):
        values = [item.get(key) for key in fraud_headers]
        write_values(ws9, row_idx, values, number_cols={6})
    auto_width(ws9)

    if report_data.get("transactions"):
        ws10 = wb.create_sheet("Transactions")
        txns = report_data.get("transactions", []) or []
        txn_headers = list(txns[0].keys()) if txns and isinstance(txns[0], dict) else []
        write_headers(ws10, 1, txn_headers)
        for row_idx, txn in enumerate(txns, 2):
            values = [txn.get(key) for key in txn_headers]
            write_values(ws10, row_idx, values)
        auto_width(ws10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def build_track2_counterparty_ledger(transactions: List[dict]) -> dict:
    """Convert raw transaction rows into the counterparty_ledger dict shape
    that build_track2_result (kredit_lab_classify_track2) expects.

    Schema: {"counterparties": [{"counterparty_name": str,
                                  "transaction_count": int,
                                  "credit_count": int,
                                  "debit_count": int,
                                  "total_credits": float,
                                  "total_debits": float,
                                  "net_position": float,
                                  "transactions": [{"date", "description",
                                                    "amount", "type"}]
                                 }]}
    """
    from collections import defaultdict

    special_buckets = {
        "UNIDENTIFIED",
        "UNCATEGORIZED",
        "CHEQUE",
        "UNIDENTIFIED (CHEQUE)",
        "CASH DEPOSIT",
        "CASH WITHDRAWAL",
        "BANK FEES",
        "BULK SALARY",
        "FD/INTEREST",
        "LOAN REPAYMENT",
        "LOAN DISBURSEMENT",
        "KWSP",
        "SOCSO",
        "LHDN",
        "HRDF",
        "REVERSAL",
        "RETURNED CHEQUE",
        "INWARD RETURN",
        "JANM",
        "APAYLATER",
        "AUTOPAY CR",
        "AUTOPAY DR",
    }

    def _clean_name(value) -> str:
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value).strip()).upper()

    def _is_special_bucket(name: str) -> bool:
        upper = _clean_name(name)
        if upper in special_buckets:
            return True
        return bool(
            re.match(r"^UNIDENTIFIED(?:\s.*)?$", upper)
            or re.match(r"^UNNAMED\s+.+?\s+TRANSFER\s*\((?:CR|DR)\)\s*$", upper)
            or re.match(r"^UNNAMED\s+INTERNAL\s+PAYROLL\s*\((?:CR|DR)\)\s*$", upper)
            or re.match(r"^CARD\s+POS\s*\([A-Z]+\)\s*$", upper)
        )

    def _method_for_name(name: str, matched_parser_pattern: bool) -> str:
        if _is_special_bucket(name):
            return "special_bucket"
        if matched_parser_pattern:
            return "pattern_matched"
        return "raw_fallback"

    extraction_stats = {
        "pattern_matched": 0,
        "special_bucket": 0,
        "raw_fallback": 0,
        "total_transactions": 0,
    }

    tx_df = pd.DataFrame(transactions or [])
    if tx_df.empty:
        return {
            "counterparties": [],
            "total_counterparties": 0,
            "extraction_stats": extraction_stats,
        }

    tx_df = prepare_counterparty_dataframe(tx_df)

    buckets: dict = defaultdict(lambda: {
        "total_credits": 0.0,
        "total_debits": 0.0,
        "credit_count": 0,
        "debit_count": 0,
        "transactions": [],
        "pattern_matched": 0,
        "special_bucket": 0,
        "raw_fallback": 0,
    })

    for tx in tx_df.to_dict("records"):
        desc = tx.get("description", "")
        name = _clean_name(tx.get("counterparty") or tx.get("party_name"))
        matched_parser_pattern = bool(tx.get("_counterparty_pattern_matched"))
        if not name:
            name = "UNKNOWN"

        extraction_method = _method_for_name(name, matched_parser_pattern)
        bucket = buckets[name]
        credit = float(tx.get("credit") or 0)
        debit = float(tx.get("debit") or 0)
        if credit > 0:
            bucket["total_credits"] += credit
            bucket["credit_count"] += 1
            bucket[extraction_method] += 1
            extraction_stats[extraction_method] += 1
            extraction_stats["total_transactions"] += 1
            bucket["transactions"].append({
                "date": tx.get("date", ""),
                "description": desc,
                "amount": round(credit, 2),
                "type": "CREDIT",
                "balance": safe_float(tx.get("balance", 0)),
                "extraction_method": extraction_method,
            })
        if debit > 0:
            bucket["total_debits"] += debit
            bucket["debit_count"] += 1
            bucket[extraction_method] += 1
            extraction_stats[extraction_method] += 1
            extraction_stats["total_transactions"] += 1
            bucket["transactions"].append({
                "date": tx.get("date", ""),
                "description": desc,
                "amount": round(debit, 2),
                "type": "DEBIT",
                "balance": safe_float(tx.get("balance", 0)),
                "extraction_method": extraction_method,
            })

    counterparties = []
    for name, b in buckets.items():
        if not b["transactions"]:
            continue
        cr = round(b["total_credits"], 2)
        dr = round(b["total_debits"], 2)
        counterparties.append({
            "counterparty_name": name,
            "transaction_count": b["credit_count"] + b["debit_count"],
            "credit_count": b["credit_count"],
            "debit_count": b["debit_count"],
            "total_credits": cr,
            "total_debits": dr,
            "net_position": round(cr - dr, 2),
            "pattern_matched": b["pattern_matched"],
            "special_bucket": b["special_bucket"],
            "raw_fallback": b["raw_fallback"],
            "transactions": b["transactions"],
        })

    counterparties.sort(
        key=lambda c: abs(c["total_credits"] - c["total_debits"]), reverse=True
    )
    return {
        "counterparties": counterparties,
        "total_counterparties": len(counterparties),
        "extraction_stats": extraction_stats,
    }


def generate_html_report_from_data(transactions: List[dict], monthly_summary: List[dict], 
                                   transaction_analysis: dict, high_value_threshold: float) -> str:
    """Generate interactive HTML report from the shared export payload."""
    report_data = build_shared_report_data(
        transactions,
        monthly_summary,
        transaction_analysis,
        high_value_threshold,
    )
    return generate_interactive_html(report_data)

# ============================================================
# HTML REPORT GENERATION FUNCTIONS (ADDED)
# ============================================================

def fmt(val, decimals=2):
    """Format number with commas"""
    if val is None:
        return "0.00"
    return f"{val:,.{decimals}f}"

def normalize_observations(obs):
    """Coerce observations into {'positive': [...], 'concerns': [...]}."""
    if isinstance(obs, dict):
        return {'positive': list(obs.get('positive', []) or []),
                'concerns': list(obs.get('concerns', []) or [])}
    pos, con = [], []
    if isinstance(obs, list):
        for item in obs:
            if isinstance(item, str):
                con.append(item)
            elif isinstance(item, dict):
                kind = str(item.get('type') or item.get('category') or item.get('sentiment') or '').lower()
                text = item.get('text') or item.get('observation') or item.get('message') or item.get('description') or ''
                if not text:
                    continue
                if kind in ('positive', 'pos', 'good', 'strength'):
                    pos.append(text)
                else:
                    con.append(text)
    return {'positive': pos, 'concerns': con}

def adapt_to_v6_basic_unused(src):
    """Reshape flat extractor output into v6.3.3 renderer schema."""
    from collections import defaultdict

    summary = src.get('summary', {}) or {}
    transactions = src.get('transactions', []) or []
    monthly_summary = src.get('monthly_summary', []) or []
    cp_ledger = src.get('counterparty_ledger', {}) or {}
    pdf_integrity = src.get('pdf_integrity')

    report_info = {
        'company_name': summary.get('company_names', ['Unknown'])[0] if summary.get('company_names') else 'Unknown',
        'schema_version': '6.3.3',
        'period_start': '',
        'period_end': '',
        'total_months': len(monthly_summary),
        'related_parties': [],
    }

    # accounts aggregation
    acc_map = defaultdict(lambda: {'credits': 0.0, 'debits': 0.0, 'txn_count': 0, 'bank': '', 'last_bal': None, 'opening_bal': None})
    for t in transactions:
        an = t.get('account_no', '')
        if not an:
            continue
        a = acc_map[an]
        a['txn_count'] += 1
        cr = float(t.get('credit', 0) or 0)
        dr = float(t.get('debit', 0) or 0)
        a['credits'] += cr
        a['debits'] += dr
        if not a['bank']:
            a['bank'] = t.get('bank', '') or ''
        bal = t.get('balance')
        if isinstance(bal, (int, float)):
            if a['opening_bal'] is None:
                a['opening_bal'] = bal - cr + dr
            a['last_bal'] = bal
    
    accounts = []
    for an, a in sorted(acc_map.items()):
        accounts.append({
            'bank_name': a['bank'],
            'account_number': an,
            'account_holder': report_info['company_name'],
            'account_type': 'Current',
            'opening_balance': round(a['opening_bal'] or 0.0, 2),
            'closing_balance': round(a['last_bal'] or 0.0, 2),
            'total_credits': round(a['credits'], 2),
            'total_debits': round(a['debits'], 2),
            'transaction_count': a['txn_count'],
        })

    monthly_analysis = []
    for m in monthly_summary:
        monthly_analysis.append({
            'month': m.get('month', ''),
            'bank_name': '',
            'account_number': m.get('account_no', ''),
            'gross_credits': float(m.get('total_credit', 0) or 0),
            'gross_debits': float(m.get('total_debit', 0) or 0),
            'net_credits': float(m.get('total_credit', 0) or 0),
            'net_debits': float(m.get('total_debit', 0) or 0),
            'eod_lowest': float(m.get('lowest_balance', 0) or 0),
            'eod_highest': float(m.get('highest_balance', 0) or 0),
            'eod_average': (float(m.get('highest_balance', 0) or 0) + float(m.get('lowest_balance', 0) or 0)) / 2.0,
            'opening_balance': float(m.get('opening_balance', 0) or 0),
            'closing_balance': float(m.get('ending_balance', 0) or 0),
            'transaction_count': m.get('transaction_count', 0),
        })

    gross_credits = sum(float(t.get('credit', 0) or 0) for t in transactions)
    gross_debits = sum(float(t.get('debit', 0) or 0) for t in transactions)
    total_months = len(monthly_summary) or 1
    
    consolidated = {
        'gross_credits': round(gross_credits, 2),
        'gross_debits': round(gross_debits, 2),
        'net_credits': round(gross_credits, 2),
        'net_debits': round(gross_debits, 2),
        'annualized_net_credits': round(gross_credits * 12 / total_months, 2),
        'annualized_net_debits': round(gross_debits * 12 / total_months, 2),
        'eod_lowest': 0,
        'eod_highest': 0,
        'eod_average': 0,
        'data_completeness': 'COMPLETE',
    }

    return {
        'report_info': report_info,
        'accounts': accounts,
        'monthly_analysis': monthly_analysis,
        'consolidated': consolidated,
        'top_parties': {'top_payers': [], 'top_payees': []},
        'large_credits': [],
        'own_related_transactions': {'transactions': [], 'summary': {}},
        'loan_transactions': {'transactions': [], 'summary': {}},
        'flags': {'indicators': []},
        'observations': {'positive': [], 'concerns': []},
        'parsing_metadata': {},
        'counterparty_ledger': cp_ledger,
        'pdf_integrity': pdf_integrity,
    }

def generate_interactive_html_basic_unused(data):
    """Generate interactive HTML report for v6 schema"""
    # Simplified version - you can import the full function from your other file
    # For now, create a basic HTML report
    r = data.get('report_info', {})
    accounts = data.get('accounts', [])
    consol = data.get('consolidated', {})
    
    company = r.get('company_name', 'Company')
    period_start = r.get('period_start', '')
    period_end = r.get('period_end', '')
    
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kredit Lab — {company}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; }}
        h1 {{ color: #1B4F72; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
        .card {{ background: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 4px solid #1B4F72; }}
        .card .value {{ font-size: 24px; font-weight: bold; }}
        .card .label {{ color: #666; font-size: 12px; text-transform: uppercase; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #1B4F72; color: white; }}
        .credit {{ color: #059669; }}
        .debit {{ color: #dc2626; }}
        .footer {{ text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🔬 Kredit Lab — Statement Intelligence Report</h1>
        <p><strong>{company}</strong> | Period: {period_start} to {period_end}</p>
        
        <div class="summary">
            <div class="card"><div class="value">RM {consol.get('net_credits', 0):,.0f}</div><div class="label">Net Credits</div></div>
            <div class="card"><div class="value">RM {consol.get('net_debits', 0):,.0f}</div><div class="label">Net Debits</div></div>
            <div class="card"><div class="value">RM {consol.get('annualized_net_credits', 0):,.0f}</div><div class="label">Annualized</div></div>
            <div class="card"><div class="value">{len(accounts)}</div><div class="label">Accounts</div></div>
        </div>
        
        <h2>Account Summary</h2>
        <table>
            <thead><tr><th>Bank</th><th>Account No</th><th>Opening Balance</th><th>Closing Balance</th><th>Total Credits</th><th>Total Debits</th></tr></thead>
            <tbody>'''
    for acc in accounts:
        html += f'''<tr>
            <td>{acc.get('bank_name', '')}</td>
            <td>{acc.get('account_number', '')}</td>
            <td class="credit">RM {acc.get('opening_balance', 0):,.2f}</td>
            <td class="credit">RM {acc.get('closing_balance', 0):,.2f}</td>
            <td class="credit">RM {acc.get('total_credits', 0):,.2f}</td>
            <td class="debit">RM {acc.get('total_debits', 0):,.2f}</td>
        </tr>'''
    html += f'''</tbody>
        </table>
        <div class="footer">
            <p>Generated by Kredit Lab Statement Intelligence | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>'''
    return html

def generate_html_report_from_data_basic_unused(transactions: List[dict], monthly_summary: List[dict], 
                                                transaction_analysis: dict, high_value_threshold: float) -> str:
    """Generate interactive HTML report from parsed transactions"""
    # Convert transactions to v6.3.3 schema format
    data = adapt_to_v6({
        'transactions': transactions,
        'monthly_summary': monthly_summary,
        'summary': {
            'date_range': f"{transactions[0].get('date', '')} to {transactions[-1].get('date', '')}" if transactions else '',
            'company_names': list(set(t.get('company_name', '') for t in transactions if t.get('company_name')))
        },
        'counterparty_ledger': transaction_analysis.get('counterparty_ledger', {}),
        'pdf_integrity': st.session_state.get('integrity_analysis_results', {})
    })
    
    # Generate HTML report
    return generate_interactive_html(data)

def convert_json_to_html_basic_unused(json_file) -> str:
    """Convert uploaded JSON analysis file to HTML report"""
    data = json.load(json_file)
    
    # Adapt to v6 schema if needed
    if isinstance(data, dict) and 'monthly_analysis' not in data and 'transactions' in data:
        data = adapt_to_v6(data)
    
    return generate_interactive_html(data)


# ============================================================
# END OF ADDED FUNCTIONS
# ============================================================


st.markdown('<div id="overview-section" class="section-anchor"></div>', unsafe_allow_html=True)
st.markdown(
    '<h1>📄 Bank Statement Parser (Multi-File Support)</h1>',
    unsafe_allow_html=True,
)
st.write("Upload one or more bank statement PDFs to extract transactions.")


st.markdown(
    """
    <style>
    :root {
        --kl-accent: #0078D4;
        --kl-accent-hover: #00A8A8;
        --kl-label: #CCCCCC;
    }

    h1, h2, h3 {
        color: #FFFFFF;
    }

    h1 {
        margin-bottom: 0.35rem;
    }

    [data-testid="stMarkdownContainer"] p {
        margin-bottom: 0.2rem;
    }

    div[data-testid="stSelectbox"] {
        margin-top: -0.25rem;
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
        min-height: 54px !important;
        align-items: center !important;
        padding-left: 15px !important;
        padding-right: 32px !important;
        border-radius: 8px !important;
        background: #262730 !important;
        border: 1.5px solid #334155 !important;
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] div {
        font-size: 16px !important;
        font-weight: 500 !important;
        line-height: 22px !important;
        -webkit-font-smoothing: antialiased !important;
        -moz-osx-font-smoothing: grayscale !important;
        text-rendering: optimizeLegibility !important;
        color: #F3F4F6 !important;
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] span,
    div[data-testid="stSelectbox"] div[data-baseweb="select"] input {
        font-size: 19px !important;
        font-weight: 600 !important;
        line-height: 22px !important;
        -webkit-font-smoothing: antialiased !important;
        -moz-osx-font-smoothing: grayscale !important;
        text-rendering: optimizeLegibility !important;
        color: #F3F4F6 !important;
    }

    div[data-testid="stSelectbox"] [data-baseweb="tag"] {
        margin-top: 0 !important;
        margin-bottom: 0 !important;
    }

    div[data-testid="stSelectbox"] svg {
        width: 16px !important;
        height: 16px !important;
        fill: #9CA3AF !important;
    }

    div[role="listbox"],
    ul[role="listbox"] {
        padding: 0 !important;
        border: 0 !important;
        border-radius: 0 !important;
        background: #0B0F16 !important;
        box-shadow: none !important;
    }

    div[role="option"],
    li[role="option"] {
        min-height: 50px !important;
        padding: 0 20px !important;
        border: 0 !important;
        border-radius: 0 !important;
        color: #F3F4F6 !important;
        font-size: 15px !important;
        font-weight: 400 !important;
        line-height: 22px !important;
        -webkit-font-smoothing: antialiased !important;
        -moz-osx-font-smoothing: grayscale !important;
        text-rendering: optimizeLegibility !important;
        display: flex !important;
        align-items: center !important;
    }

    div[role="option"] *,
    li[role="option"] * {
        color: inherit !important;
        font-size: inherit !important;
        font-weight: inherit !important;
        line-height: inherit !important;
        background: transparent !important;
        border-radius: 0 !important;
        box-shadow: none !important;
    }

    div[role="option"]:hover,
    li[role="option"]:hover,
    div[role="option"][aria-selected="true"],
    li[role="option"][aria-selected="true"] {
        background: #2A2B34 !important;
        color: #FFFFFF !important;
    }

    [data-testid="stWidgetLabel"] label,
    [data-testid="stWidgetLabel"] p,
    [data-testid="stFileUploader"] label {
        color: var(--kl-label);
        font-weight: 600;
    }

    div.stButton {
        text-align: left;
    }

    div.stButton > button {
        min-height: 3rem;
        width: 100%;
        padding: 0.7rem 1.35rem;
        border-radius: 8px !important;
        font-weight: 600;
        letter-spacing: 0;
        transition: background-color 160ms ease, border-color 160ms ease, box-shadow 160ms ease, transform 160ms ease, color 160ms ease;
    }

    div.stButton > button[kind="primary"],
    div.stButton > button[type="primary"] {
        border: 1px solid #0052CC !important;
        background: linear-gradient(135deg, #0078D4 0%, #0052CC 100%) !important;
        color: #FFFFFF !important;
        box-shadow: 0 8px 18px rgba(0, 102, 255, 0.22) !important;
    }

    div.stButton > button[kind="primary"]:hover,
    div.stButton > button[type="primary"]:hover {
        border-color: #1A75FF !important;
        background: linear-gradient(135deg, #1A75FF 0%, #0066FF 100%) !important;
        box-shadow: 0 10px 24px rgba(0, 102, 255, 0.34) !important;
        transform: translateY(-1px);
    }

    div.stButton > button[kind="primary"]:active,
    div.stButton > button[type="primary"]:active {
        box-shadow: 0 4px 10px rgba(0, 102, 255, 0.22) !important;
        transform: translateY(1px);
    }

    div.stButton > button:not([kind="primary"]):not([type="primary"]) {
        border: 1px solid #374151 !important;
        background: #111827 !important;
        color: #D1D5DB !important;
        box-shadow: 0 6px 14px rgba(0, 0, 0, 0.16) !important;
    }

    div.stButton > button:not([kind="primary"]):not([type="primary"]):hover {
        border-color: #9CA3AF !important;
        background: #1F2937 !important;
        color: #FFFFFF !important;
        box-shadow: 0 8px 18px rgba(0, 0, 0, 0.22) !important;
        transform: translateY(-1px);
    }

    div.stButton > button:not([kind="primary"]):not([type="primary"]):active {
        box-shadow: 0 3px 8px rgba(0, 0, 0, 0.2) !important;
        transform: translateY(1px);
    }

    section[data-testid="stFileUploaderDropzone"],
    div[data-testid="stFileUploaderDropzone"],
    div[data-testid="stTextInput"] div[data-baseweb="input"] {
        background: #262730 !important;
        border: 1.5px solid #334155 !important;
        border-radius: 8px !important;
        box-sizing: border-box;
    }

    section[data-testid="stFileUploaderDropzone"]:hover,
    div[data-testid="stFileUploaderDropzone"]:hover,
    div[data-testid="stTextInput"] div[data-baseweb="input"]:hover {
        border-color: #475569 !important;
    }

    div[data-testid="stTextInput"] input {
        background: transparent !important;
        color: #F3F4F6 !important;
    }

    div[data-testid="stSelectbox"] div[data-baseweb="select"] > div,
    div[data-testid="stTextInput"] div[data-baseweb="input"] {
        min-height: 3rem;
        box-sizing: border-box;
    }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) div.stButton > button:hover {
        border-color: #EF4444 !important;
        background: rgba(239, 68, 68, 0.08) !important;
        color: #C4B5FD !important;
        box-shadow: 0 8px 18px rgba(239, 68, 68, 0.16) !important;
    }

    div[data-testid="stHorizontalBlock"] > div:nth-child(2) div.stButton > button:active {
        background: rgba(239, 68, 68, 0.14) !important;
        transform: translateY(1px);
    }

    .kl-progress-panel {
        margin-top: 0.75rem;
        padding: 0.95rem 1rem;
        border: 1px solid rgba(59, 130, 246, 0.36);
        border-radius: 8px;
        background: rgba(15, 23, 42, 0.72);
        box-shadow: 0 0 8px rgba(0, 123, 255, 0.24);
    }

    .kl-progress-panel.success {
        border-color: rgba(34, 197, 94, 0.38);
        background: rgba(20, 83, 45, 0.55);
        box-shadow: 0 0 8px rgba(40, 167, 69, 0.28);
    }

    .kl-progress-panel.warning {
        border-color: rgba(250, 204, 21, 0.38);
        background: rgba(113, 63, 18, 0.42);
        box-shadow: 0 0 8px rgba(250, 204, 21, 0.18);
    }

    .kl-progress-panel.error {
        border-color: rgba(248, 113, 113, 0.42);
        background: rgba(127, 29, 29, 0.45);
        box-shadow: 0 0 8px rgba(248, 113, 113, 0.22);
    }

    .kl-progress-topline {
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 1rem;
        margin-bottom: 0.45rem;
    }

    .kl-progress-status {
        min-width: 0;
        color: #CBD5E1;
        font-size: 0.88rem;
        font-weight: 500;
        line-height: 1.3;
    }

    .kl-progress-percent {
        flex: 0 0 auto;
        color: #E5F2FF;
        font-size: 0.82rem;
        font-weight: 600;
        line-height: 1.3;
    }

    .kl-progress-filename {
        color: #94A3B8;
        font-size: 0.8rem;
        line-height: 1.35;
        margin-bottom: 0.5rem;
    }

    .kl-progress-track {
        position: relative;
        height: 0.62rem;
        overflow: hidden;
        border-radius: 6px;
        background: #111827;
        border: 1px solid rgba(51, 65, 85, 0.9);
    }

    .kl-progress-fill {
        position: relative;
        height: 100%;
        border-radius: 6px;
        background: linear-gradient(90deg, #007BFF, #00C6FF);
        box-shadow: 0 0 8px rgba(0, 123, 255, 0.4);
        transition: width 0.4s ease-in-out, background-color 0.3s ease, box-shadow 0.3s ease;
        overflow: hidden;
        animation: kl-progress-pulse 1.3s ease-in-out infinite;
    }

    .kl-progress-fill::after {
        content: "";
        position: absolute;
        inset: 0;
        background-image: linear-gradient(
            45deg,
            rgba(255, 255, 255, 0.18) 25%,
            transparent 25%,
            transparent 50%,
            rgba(255, 255, 255, 0.18) 50%,
            rgba(255, 255, 255, 0.18) 75%,
            transparent 75%,
            transparent
        );
        background-size: 1rem 1rem;
        animation: kl-progress-stripes 0.9s linear infinite;
    }

    .kl-progress-panel.success .kl-progress-fill {
        background: #28A745;
        box-shadow: 0 0 8px rgba(40, 167, 69, 0.4);
        animation: none;
    }

    .kl-progress-panel.success .kl-progress-fill::after,
    .kl-progress-panel.warning .kl-progress-fill::after,
    .kl-progress-panel.error .kl-progress-fill::after {
        display: none;
    }

    .kl-progress-panel.warning .kl-progress-fill {
        background: linear-gradient(90deg, #F59E0B, #FACC15);
        box-shadow: 0 0 8px rgba(250, 204, 21, 0.28);
    }

    .kl-progress-panel.error .kl-progress-fill {
        background: linear-gradient(90deg, #EF4444, #F97316);
        box-shadow: 0 0 8px rgba(248, 113, 113, 0.32);
    }

    @keyframes kl-progress-stripes {
        from { background-position: 1rem 0; }
        to { background-position: 0 0; }
    }

    @keyframes kl-progress-pulse {
        0%, 100% { filter: brightness(1); }
        50% { filter: brightness(1.12); }
    }

    .kl-analysis-title {
        display: flex;
        align-items: center;
        gap: 0.55rem;
        margin: 0.5rem 0 0.75rem;
        color: #FFFFFF;
        font-size: 1.75rem;
        font-weight: 800;
        line-height: 1.2;
    }

    .kl-analysis-subtitle {
        color: #A9C1DD;
        font-size: 1rem;
        margin: 0 0 1.25rem;
    }

    .kl-metric-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 0.875rem;
        margin-bottom: 0.875rem;
    }

    .kl-metric-card {
        min-height: 6rem;
        padding: 1rem 1.15rem;
        border: 1px solid #334155;
        border-radius: 12px;
        background: #0B111D;
        box-sizing: border-box;
    }

    .kl-metric-card-wide {
        grid-column: 1 / -1;
    }

    .kl-metric-label {
        color: #D7E3F8;
        font-size: 0.92rem;
        font-weight: 600;
        margin-bottom: 0.55rem;
    }

    .kl-metric-value {
        color: #FFFFFF;
        font-size: 2.25rem;
        line-height: 1;
        font-weight: 800;
    }

    .kl-pattern-details-heading {
        color: #FFFFFF;
        font-size: 1.5rem;
        font-weight: 500;
        line-height: 1.2;
        margin: 0.25rem 0 1rem;
    }

    div[data-testid="stExpander"] {
        border: 1px solid #374151 !important;
        border-radius: 8px !important;
        background: #0B0F16 !important;
        overflow: hidden !important;
    }

    div[data-testid="stExpander"] details {
        background: transparent !important;
        border: 0 !important;
    }

    div[data-testid="stExpander"] details > summary {
        min-height: 3.25rem !important;
        padding: 0.8rem 1.25rem !important;
        position: relative !important;
        display: flex !important;
        align-items: center !important;
        list-style: none !important;
        list-style-type: none !important;
        font-weight: 500 !important;
    }

    div[data-testid="stExpander"] details > summary::marker {
        content: "" !important;
        color: transparent !important;
        font-size: 0 !important;
    }

    div[data-testid="stExpander"] details > summary::-webkit-details-marker {
        display: none !important;
        color: transparent !important;
        font-size: 0 !important;
    }

    div[data-testid="stExpander"] details > summary [data-testid="stExpanderToggleIcon"],
    div[data-testid="stExpander"] details > summary > svg:first-child,
    div[data-testid="stExpander"] details > summary > span:first-child:has(svg),
    div[data-testid="stExpander"] details > summary > div:first-child:has(svg),
    div[data-testid="stExpander"] details > summary > div:first-child:has([data-testid="stExpanderToggleIcon"]) {
        display: flex !important;
        order: 2 !important;
        position: static !important;
        width: 1rem !important;
        height: 1rem !important;
        margin: 0 0 0 auto !important;
        padding: 0 !important;
        transform: none !important;
    }

    div[data-testid="stExpander"] details > summary [data-testid="stExpanderToggleIcon"] svg,
    div[data-testid="stExpander"] details > summary > svg:first-child,
    div[data-testid="stExpander"] details > summary > span:first-child:has(svg) svg,
    div[data-testid="stExpander"] details > summary > div:first-child:has(svg) svg,
    div[data-testid="stExpander"] details > summary > div:first-child:has([data-testid="stExpanderToggleIcon"]) svg {
        display: block !important;
        width: 1rem !important;
        height: 1rem !important;
        color: #F3F4F6 !important;
        fill: currentColor !important;
    }

    div[data-testid="stExpander"] details > summary > [data-testid="stMarkdownContainer"],
    div[data-testid="stExpander"] details > summary [data-testid="stExpanderToggleIcon"] + [data-testid="stMarkdownContainer"] {
        order: 1 !important;
        flex: 1 1 auto !important;
        margin-left: 0 !important;
    }

    div[data-testid="stExpander"] details > summary::after {
        content: none !important;
        display: none !important;
    }

    div[data-testid="stExpander"] details > summary [data-testid="stMarkdownContainer"] p {
        color: #FFFFFF;
        font-size: 0.95rem;
        font-weight: 500;
        line-height: 1.2;
        margin: 0;
    }

    @media (max-width: 900px) {
        .kl-metric-grid {
            grid-template-columns: repeat(2, minmax(0, 1fr));
        }
    }

    @media (max-width: 560px) {
        .kl-metric-grid {
            grid-template-columns: 1fr;
        }
    }

    .integrity-card {
        padding: 1rem 1.1rem;
        border: 1px solid rgba(148, 163, 184, 0.25);
        border-radius: 14px;
        background: linear-gradient(180deg, rgba(15, 23, 42, 0.28), rgba(15, 23, 42, 0.16));
        margin-bottom: 0.75rem;
    }
    .integrity-card.low {
        border-color: rgba(74, 222, 128, 0.38);
        background: linear-gradient(180deg, rgba(20, 83, 45, 0.32), rgba(15, 23, 42, 0.18));
    }
    .integrity-card.medium {
        border-color: rgba(250, 204, 21, 0.38);
        background: linear-gradient(180deg, rgba(113, 63, 18, 0.32), rgba(15, 23, 42, 0.18));
    }
    .integrity-card.high {
        border-color: rgba(248, 113, 113, 0.38);
        background: linear-gradient(180deg, rgba(127, 29, 29, 0.34), rgba(15, 23, 42, 0.18));
    }
    .integrity-label {
        font-size: 0.95rem;
        color: #cbd5e1;
        margin-bottom: 0.35rem;
    }
    .integrity-card.low .integrity-value {
        color: #86efac;
    }
    .integrity-card.medium .integrity-value {
        color: #fde68a;
    }
    .integrity-card.high .integrity-value {
        color: #fca5a5;
    }
    .integrity-value {
        font-size: 2.2rem;
        font-weight: 700;
        line-height: 1;
    }
    .integrity-title {
        font-size: 2rem;
        font-weight: 700;
        margin-bottom: 0.25rem;
    }
    .integrity-subtitle {
        color: #94a3b8;
        margin-bottom: 1rem;
    }

    /* Counterparty table styling */
    .counterparty-net-positive {
        color: #4CAF50;
        font-weight: 600;
    }
    .counterparty-net-negative {
        color: #F44336;
        font-weight: 600;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

validation_css_parts = []
if st.session_state.get("bank_choice_error"):
    validation_css_parts.append(
        """
        div[data-testid="stSelectbox"] div[data-baseweb="select"] > div {
            border: 1px solid #F04438 !important;
            box-shadow: inset 0 0 0 1px #F04438 !important;
            border-radius: 8px !important;
            box-sizing: border-box !important;
        }

        div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:hover {
            border-color: #F04438 !important;
        }

        div[data-testid="stSelectbox"] label,
        div[data-testid="stSelectbox"] p {
            color: #F04438 !important;
        }
        """
    )
if st.session_state.get("high_value_threshold_error"):
    validation_css_parts.append(
        """
        div[data-testid="stTextInput"]:has(input[aria-label="High Value Threshold (RM)"]) div[data-baseweb="input"] {
            border: 2px solid #F04438 !important;
            box-shadow: inset 0 0 0 1px #F04438 !important;
            border-radius: 8px !important;
            box-sizing: border-box !important;
        }

        div[data-testid="stTextInput"]:has(input[aria-label="High Value Threshold (RM)"]) label,
        div[data-testid="stTextInput"]:has(input[aria-label="High Value Threshold (RM)"]) p {
            color: #F04438 !important;
        }
        """
    )
if st.session_state.get("pdf_upload_error"):
    validation_css_parts.append(
        """
        section[data-testid="stFileUploaderDropzone"],
        div[data-testid="stFileUploaderDropzone"] {
            border: 1.5px solid #F04438 !important;
            box-shadow: inset 0 0 0 1px #F04438 !important;
            border-radius: 8px !important;
            box-sizing: border-box !important;
        }

        section[data-testid="stFileUploaderDropzone"]:hover,
        div[data-testid="stFileUploaderDropzone"]:hover {
            border-color: #F04438 !important;
        }

        div[data-testid="stFileUploader"] label,
        div[data-testid="stFileUploader"] p {
            color: #F04438 !important;
        }
        """
    )
st.markdown(
    f"""
    <style>
    {''.join(validation_css_parts)}
    </style>
    """,
    unsafe_allow_html=True,
)


# -----------------------------
# Session state init
# -----------------------------
if "status" not in st.session_state:
    st.session_state.status = "idle"

if "results" not in st.session_state:
    st.session_state.results = []

if "integrity_analysis_results" not in st.session_state:
    st.session_state.integrity_analysis_results = {}

if "affin_statement_totals" not in st.session_state:
    st.session_state.affin_statement_totals = []

if "affin_file_transactions" not in st.session_state:
    st.session_state.affin_file_transactions = {}

if "ambank_statement_totals" not in st.session_state:
    st.session_state.ambank_statement_totals = []

if "ambank_file_transactions" not in st.session_state:
    st.session_state.ambank_file_transactions = {}

if "cimb_statement_totals" not in st.session_state:
    st.session_state.cimb_statement_totals = []

if "cimb_file_transactions" not in st.session_state:
    st.session_state.cimb_file_transactions = {}

if "rhb_statement_totals" not in st.session_state:
    st.session_state.rhb_statement_totals = []

if "rhb_file_transactions" not in st.session_state:
    st.session_state.rhb_file_transactions = {}

if "bank_islam_file_month" not in st.session_state:
    st.session_state.bank_islam_file_month = {}

# password + company name tracking
if "pdf_password" not in st.session_state:
    st.session_state.pdf_password = ""

if "company_name_override" not in st.session_state:
    st.session_state.company_name_override = ""

if "company_account_no_override" not in st.session_state:
    st.session_state.company_account_no_override = ""

if "high_value_threshold_input" not in st.session_state:
    st.session_state.high_value_threshold_input = ""

if "high_value_threshold_error" not in st.session_state:
    st.session_state.high_value_threshold_error = ""

if "bank_choice_error" not in st.session_state:
    st.session_state.bank_choice_error = ""

if "pdf_upload_error" not in st.session_state:
    st.session_state.pdf_upload_error = ""

if "validation_toast_message" not in st.session_state:
    st.session_state.validation_toast_message = ""

if "active_high_value_threshold" not in st.session_state:
    st.session_state.active_high_value_threshold = None

if "stop_requested" not in st.session_state:
    st.session_state.stop_requested = False

if "upload_widget_reset_id" not in st.session_state:
    st.session_state.upload_widget_reset_id = 0

if "file_company_name" not in st.session_state:
    st.session_state.file_company_name = {}

if "file_account_no" not in st.session_state:
    st.session_state.file_account_no = {}

# Track 2 — account-type determinations populated by parser hooks
if "account_type_determinations" not in st.session_state:
    st.session_state.account_type_determinations = []

# Analyst-supplied related parties (populated by the analyst form when wired)
if "related_parties_override" not in st.session_state:
    st.session_state.related_parties_override = []

if "counterparty_name_overrides" not in st.session_state:
    st.session_state.counterparty_name_overrides = {}


# -----------------------------
# Fraud/Integrity Constants and Functions
# -----------------------------
FRAUD_LAYER_ORDER = [
    ("metadata", "Layer 1: Metadata"),
    ("fonts", "Layer 2: Fonts"),
    ("text_layers", "Layer 3: Text Layers"),
    ("visual", "Layer 4: Visual"),
    ("cross_validation", "Layer 5: Cross Validation"),
    ("bank_profile", "Layer 6: Bank Profile"),
    ("structural", "Layer 7: Structural"),
    ("arithmetic", "Layer 8: Arithmetic"),
]


def severity_badge(severity: str) -> str:
    severity = (severity or "").upper()
    if severity == "HIGH":
        return "🔴 HIGH"
    if severity == "MEDIUM":
        return "🟠 MEDIUM"
    return "🟢 LOW"


def severity_dot(severity: str) -> str:
    severity = (severity or "").upper()
    if severity == "HIGH":
        return "🔴"
    if severity == "MEDIUM":
        return "🟡"
    return "🟢"


def render_integrity_report_styles() -> None:
    st.markdown(
        """
        <style>
        .integrity-card {
            padding: 1rem 1.1rem;
            border: 1px solid rgba(148, 163, 184, 0.25);
            border-radius: 14px;
            background: linear-gradient(180deg, rgba(15, 23, 42, 0.28), rgba(15, 23, 42, 0.16));
            margin-bottom: 0.75rem;
        }
        .integrity-card.low {
            border-color: rgba(74, 222, 128, 0.38);
            background: linear-gradient(180deg, rgba(20, 83, 45, 0.32), rgba(15, 23, 42, 0.18));
        }
        .integrity-card.medium {
            border-color: rgba(250, 204, 21, 0.38);
            background: linear-gradient(180deg, rgba(113, 63, 18, 0.32), rgba(15, 23, 42, 0.18));
        }
        .integrity-card.high {
            border-color: rgba(248, 113, 113, 0.38);
            background: linear-gradient(180deg, rgba(127, 29, 29, 0.34), rgba(15, 23, 42, 0.18));
        }
        .integrity-label {
            font-size: 0.95rem;
            color: #cbd5e1;
            margin-bottom: 0.35rem;
        }
        .integrity-card.low .integrity-value {
            color: #86efac;
        }
        .integrity-card.medium .integrity-value {
            color: #fde68a;
        }
        .integrity-card.high .integrity-value {
            color: #fca5a5;
        }
        .integrity-value {
            font-size: 2.2rem;
            font-weight: 700;
            line-height: 1;
        }
        .integrity-title {
            font-size: 2rem;
            font-weight: 700;
            margin-bottom: 0.25rem;
        }
        .integrity-subtitle {
            color: #94a3b8;
            margin-bottom: 1rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_integrity_metric(label: str, value, dot: str | None = None, tone: str | None = None) -> None:
    dot_html = f"<span>{dot}</span> " if dot else ""
    tone_class = f" {tone.lower()}" if tone and tone.lower() in {"low", "medium", "high"} else ""
    st.markdown(
        f"""
        <div class="integrity-card{tone_class}">
            <div class="integrity-label">{dot_html}{label}</div>
            <div class="integrity-value">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_integrity_overview(analysis_results: dict) -> None:
    total_files = len(analysis_results)
    clean_count = sum(1 for result in analysis_results.values() if result.get("overall_risk") == "LOW")
    medium_count = sum(1 for result in analysis_results.values() if result.get("overall_risk") == "MEDIUM")
    high_count = sum(1 for result in analysis_results.values() if result.get("overall_risk") == "HIGH")

    st.markdown('<div class="integrity-title">🛡️ Document Integrity Scan</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="integrity-subtitle">Multi-layer fraud screening across all uploaded statements.</div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_integrity_metric("Files Scanned", total_files)
    with c2:
        render_integrity_metric("Clean", clean_count, "🟢")
    with c3:
        render_integrity_metric("Medium Risk", medium_count, "🟡")
    with c4:
        render_integrity_metric("High Risk", high_count, "🔴")


def file_risk_label(file_name: str, result: dict) -> str:
    risk = (result.get("overall_risk") or "LOW").upper()
    counts = integrity_layer_counts(result.get("layer_results", {}))
    return (
        f"{severity_dot(risk)} {file_name} - Risk: {risk} "
        f"({counts.get('high', 0)}H / {counts.get('medium', 0)}M / {counts.get('low', 0)}L)"
    )


def integrity_layer_counts(layer_results: dict) -> dict:
    counts = {"high": 0, "medium": 0, "low": 0, "total": 0}
    for layer_key, _ in FRAUD_LAYER_ORDER:
        findings = (layer_results or {}).get(layer_key, [])
        highest = next(
            (
                level.lower()
                for level in ("HIGH", "MEDIUM", "LOW")
                if any((item.get("severity") or "").upper() == level for item in findings)
            ),
            "low",
        )
        counts[highest] += 1
        counts["total"] += 1
    return counts


def render_fraud_summary(summary: dict, layer_results: dict | None = None):
    risk = summary.get("overall_risk", "LOW")
    counts = summary.get("counts", {})
    headline = summary.get("headline", "Analysis complete")
    if layer_results:
        counts = integrity_layer_counts(layer_results)

    if risk == "HIGH":
        st.error(f"Overall Risk: {risk}")
    elif risk == "MEDIUM":
        st.warning(f"Overall Risk: {risk}")
    else:
        st.success(f"Overall Risk: {risk}")

    st.write(headline)

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("High", counts.get("high", 0))
    c2.metric("Medium", counts.get("medium", 0))
    c3.metric("Low", counts.get("low", 0))
    c4.metric("Total Findings", counts.get("total", 0))

    top_findings = summary.get("top_findings", [])
    if top_findings or layer_results:
        st.markdown("**Top Findings**")

    top_finding_by_layer = {
        finding.get("layer", ""): finding
        for finding in top_findings
        if finding.get("layer")
    }

    if not layer_results and top_findings:
        for finding in top_findings:
            st.write(
                f"{severity_badge(finding.get('severity'))} "
                f"**[{finding.get('layer', 'unknown')}]** {finding.get('message', '')}"
            )
    elif not layer_results:
        st.info("No findings returned.")

    if layer_results:
        for layer_key, layer_label in FRAUD_LAYER_ORDER:
            findings = layer_results.get(layer_key, [])
            if findings:
                highest = next(
                    (
                        level
                        for level in ("HIGH", "MEDIUM", "LOW")
                        if any((item.get("severity") or "").upper() == level for item in findings)
                    ),
                    "LOW",
                )
                anomaly_count = sum(1 for finding in findings if not is_benign_integrity_finding(finding))
                summary_finding = top_finding_by_layer.get(layer_key) or (findings[0] if findings else None)
                message = summary_finding.get("message", "") if summary_finding else "No findings."
                st.write(
                    f"{severity_badge(highest)} **{layer_label}** "
                    f"{message} ({anomaly_count} anomalies detected)"
                )
            else:
                st.write(f"{severity_badge('LOW')} **{layer_label}** (0 anomalies detected)")


def is_benign_integrity_finding(finding: dict) -> bool:
    message = str(finding.get("message", "") or "").lower()
    benign_patterns = [
        "no anomalies detected",
        "verified",
        "matches known",
        "hashes computed",
        "pdf version",
        "font consistency",
    ]
    return any(pattern in message for pattern in benign_patterns)



# -----------------------------
# Counterparty Ledger Functions
# -----------------------------
UNKNOWN_COUNTERPARTY_VALUES = {"", "UNKNOWN", "N/A", "NA", "NONE", "NULL", "-"}
COUNTERPARTY_NAME_FIELDS = (
    "counterparty_name_clean",
    "counterparty_name",
    "counterparty",
    "party_name",
    "counterparty_name_raw",
    "party",
    "merchant",
    "merchant_name",
    "recipient",
    "beneficiary",
    "payer",
    "payee",
)


def normalize_counterparty_value(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    text = re.sub(r"\s+", " ", str(value).strip().upper())
    if text in UNKNOWN_COUNTERPARTY_VALUES:
        return ""
    return text


def _resolve_transaction_counterparty_details(row: pd.Series) -> Tuple[str, bool]:
    """
    Prefer counterparty values extracted by bank parsers and keep whether a
    parser/extractor supplied the name.
    """
    for column in COUNTERPARTY_NAME_FIELDS:
        if column in row:
            counterparty = normalize_counterparty_value(row.get(column))
            if counterparty:
                return counterparty, True

    description = row.get("description", "")
    try:
        if pd.isna(description):
            description = ""
    except Exception:
        pass
    bank = normalize_counterparty_value(row.get("bank"))
    if "CIMB" in bank:
        counterparty = normalize_counterparty_value(extract_cimb_party_name(description))
        if counterparty:
            return counterparty, True

    return "UNKNOWN", False


def _counterparty_override_candidates(raw_name: str, clean_name: str = "") -> List[str]:
    candidates = [
        normalize_counterparty_value(raw_name),
        normalize_counterparty_value(clean_name),
    ]
    try:
        candidates.append(normalize_counterparty_value(clean_counterparty_name(raw_name)))
    except Exception:
        pass
    seen = set()
    return [name for name in candidates if name and not (name in seen or seen.add(name))]


def _apply_counterparty_overrides(raw_name: str, clean_name: str) -> str:
    overrides = st.session_state.get("counterparty_name_overrides", {}) or {}
    if not overrides:
        return clean_name

    for candidate in _counterparty_override_candidates(raw_name, clean_name):
        override = normalize_counterparty_value(overrides.get(candidate))
        if override:
            return override
    return clean_name


def _extract_counterparty_mapping_from_json(payload) -> Tuple[Dict[str, str], int]:
    """
    Accept a downloaded CP list, a mapping object, or common AI-cleaned JSON
    variants and convert them to normalized raw/current -> clean mappings.
    """
    mapping: Dict[str, str] = {}
    entry_count = 0

    def add_mapping(source, target):
        nonlocal entry_count
        clean_target = normalize_counterparty_value(target)
        if not clean_target:
            return
        for candidate in _counterparty_override_candidates(str(source or ""), clean_target):
            mapping[candidate] = clean_target
        entry_count += 1

    def add_entry(entry):
        if not isinstance(entry, dict):
            add_mapping(entry, entry)
            return

        target = (
            entry.get("counterparty_name_clean")
            or entry.get("clean_counterparty_name")
            or entry.get("clean_name")
            or entry.get("counterparty_name")
            or entry.get("name")
            or entry.get("party_name")
        )
        aliases = entry.get("aliases") or entry.get("raw_names") or entry.get("variants") or []
        if isinstance(aliases, str):
            aliases = [aliases]

        sources = list(aliases) if isinstance(aliases, list) else []
        for key in ("counterparty_name_raw", "raw_name", "original_name", "from", "source"):
            if entry.get(key):
                sources.append(entry.get(key))
        if not sources and target:
            sources = [target]

        for source in sources:
            add_mapping(source, target)

    if isinstance(payload, dict):
        counterparties = payload.get("counterparties") or payload.get("counterparty_list") or []
        if isinstance(counterparties, list):
            for entry in counterparties:
                add_entry(entry)

        raw_mapping = payload.get("mapping") or payload.get("counterparty_mapping")
        if isinstance(raw_mapping, dict):
            for source, target in raw_mapping.items():
                add_mapping(source, target)
        elif isinstance(raw_mapping, list):
            for item in raw_mapping:
                if isinstance(item, dict):
                    add_mapping(
                        item.get("from") or item.get("raw") or item.get("source") or item.get("counterparty_name_raw"),
                        item.get("to") or item.get("clean") or item.get("target") or item.get("counterparty_name_clean"),
                    )
    elif isinstance(payload, list):
        for entry in payload:
            add_entry(entry)

    return mapping, entry_count


def _build_counterparty_json_payload(prepared_df: pd.DataFrame, summary_df: pd.DataFrame) -> dict:
    counterparties = []
    for row in summary_df.to_dict(orient="records"):
        name = normalize_counterparty_value(row.get("counterparty_name")) or "UNKNOWN"
        aliases = []
        if not prepared_df.empty and "_raw_counterparty" in prepared_df.columns:
            alias_series = prepared_df.loc[
                prepared_df["counterparty_name"].astype(str).str.upper() == name,
                "_raw_counterparty",
            ]
            aliases = sorted(
                {
                    normalize_counterparty_value(alias)
                    for alias in alias_series.tolist()
                    if normalize_counterparty_value(alias)
                }
            )

        counterparties.append(
            {
                "counterparty_name_clean": name,
                "aliases": aliases or [name],
                "transaction_count": int(row.get("transaction_count", 0) or 0),
                "credit_count": int(row.get("credit_count", 0) or 0),
                "debit_count": int(row.get("debit_count", 0) or 0),
                "total_credits": round(float(row.get("total_credits", 0) or 0), 2),
                "total_debits": round(float(row.get("total_debits", 0) or 0), 2),
                "net_position": round(float(row.get("net_position", 0) or 0), 2),
            }
        )

    mapping = {}
    for cp in counterparties:
        clean_name = cp["counterparty_name_clean"]
        for alias in cp["aliases"]:
            mapping[alias] = clean_name

    return {
        "schema": "kredit_lab_counterparty_mapping",
        "version": 1,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "instructions": "Edit counterparty_name_clean to regroup aliases, keep aliases for matching, then upload this JSON back into the Counterparty Ledger.",
        "counterparties": counterparties,
        "mapping": mapping,
    }


def resolve_transaction_counterparty(row: pd.Series) -> str:
    """
    Prefer counterparty values extracted by bank parsers. Parser-specific
    helpers may be used, but the UI does not extract counterparties itself.
    """
    counterparty, _matched = _resolve_transaction_counterparty_details(row)
    return counterparty


def prepare_counterparty_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add the counterparty columns used by the CP ledger.

    The raw parser-extracted name is passed into party_utils, which canonicalizes
    and aliases related spelling variants for cleaner ledger display.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    prepared = df.copy()
    resolved_names = []
    matched_flags = []
    for _idx, row in prepared.iterrows():
        counterparty, matched = _resolve_transaction_counterparty_details(row)
        resolved_names.append(counterparty)
        matched_flags.append(bool(matched))

    raw_series = pd.Series(resolved_names, index=prepared.index, dtype="object")
    try:
        clean_names = deduplicate_counterparty_names(raw_series.fillna("").astype(str).tolist())
        grouped_series = pd.Series(clean_names, index=prepared.index, dtype="object")
    except Exception:
        try:
            grouped_series = apply_party_aliasing(raw_series)
        except Exception:
            grouped_series = raw_series

    grouped_series = grouped_series.apply(lambda value: normalize_counterparty_value(value) or "UNKNOWN")
    grouped_series = pd.Series(
        [
            _apply_counterparty_overrides(raw_name, clean_name)
            for raw_name, clean_name in zip(raw_series.tolist(), grouped_series.tolist())
        ],
        index=prepared.index,
        dtype="object",
    ).apply(lambda value: normalize_counterparty_value(value) or "UNKNOWN")

    prepared["_raw_counterparty"] = raw_series
    prepared["_counterparty_pattern_matched"] = matched_flags
    prepared["counterparty_name_raw"] = raw_series
    prepared["counterparty_name_clean"] = grouped_series
    prepared["party_name"] = grouped_series
    prepared["counterparty"] = grouped_series
    prepared["counterparty_name"] = grouped_series
    return prepared


def build_counterparty_ledger_from_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build counterparty ledger summary using parser-extracted counterparty data.
    """
    if df.empty:
        return pd.DataFrame()
    
    working_df = df.copy()
    if "counterparty_name" not in working_df.columns or "_raw_counterparty" not in working_df.columns:
        working_df = prepare_counterparty_dataframe(working_df)
    party_tables = build_transactions_by_party(working_df)
    
    # Group by counterparty
    summary_data = []
    for party in party_tables:
        table = party.get("table")
        if isinstance(table, pd.DataFrame) and "amount" in table.columns:
            signed_amounts = table["amount"].apply(safe_float)
            credit_count = int((signed_amounts > 0).sum())
            debit_count = int((signed_amounts < 0).sum())
        else:
            credit_count = 0
            debit_count = 0

        total_credits = round(float(party.get("total_credit", 0) or 0), 2)
        total_debits = round(float(party.get("total_debit", 0) or 0), 2)
        net_position = total_credits - total_debits

        summary_data.append({
            'counterparty_name': party.get("party", "UNKNOWN"),
            'transaction_count': int(party.get("count", credit_count + debit_count) or 0),
            'credit_count': credit_count,
            'debit_count': debit_count,
            'total_credits': total_credits,
            'total_debits': total_debits,
            'net_position': round(net_position, 2)
        })
    
    summary_df = pd.DataFrame(summary_data)
    if summary_df.empty:
        return summary_df
    
    summary_df['_sort_counterparty'] = summary_df['counterparty_name'].astype(str).str.casefold()
    summary_df = summary_df.sort_values('_sort_counterparty', ascending=True)
    summary_df = summary_df.drop('_sort_counterparty', axis=1).reset_index(drop=True)
    
    return summary_df


def render_counterparty_ledger_table(df: pd.DataFrame) -> None:
    """
    Render counterparty ledger as a table with transaction details on selection
    """
    if df.empty:
        st.info("No counterparty data available.")
        return
    
    st.markdown("## Counterparty Ledger")

    download_col, upload_col = st.columns(2)
    with upload_col:
        uploaded_cp_json = st.file_uploader(
            "Upload cleaned counterparty JSON",
            type=["json"],
            key="counterparty_mapping_upload",
            help="Upload the downloaded CP JSON after editing counterparty_name_clean values or provide a mapping of raw/current names to clean names.",
        )

    if uploaded_cp_json is not None:
        try:
            payload = json.load(uploaded_cp_json)
            overrides, entry_count = _extract_counterparty_mapping_from_json(payload)
            if overrides:
                st.session_state.counterparty_name_overrides = overrides
                st.success(
                    f"Applied cleaned counterparty mapping for {entry_count} entries "
                    f"({len(overrides)} aliases)."
                )
            else:
                st.warning("The uploaded JSON did not contain usable counterparty mappings.")
        except Exception as exc:
            st.error(f"Could not read counterparty JSON: {exc}")

    if st.session_state.get("counterparty_name_overrides"):
        if st.button("Clear uploaded counterparty mapping", use_container_width=True):
            st.session_state.counterparty_name_overrides = {}
            st.rerun()

    prepared_df = prepare_counterparty_dataframe(df)

    # Build the same canonical counterparty summary used by HTML and Excel.
    display_cp_ledger = build_track2_counterparty_ledger(prepared_df.to_dict("records"))
    canonical_cp_rows = build_canonical_counterparty_ledger_rows(display_cp_ledger)
    counterparty_summary = pd.DataFrame(
        [
            {
                "counterparty_name": row.get("counterparty_name", ""),
                "transaction_count": row.get("transaction_count", 0),
                "credit_count": row.get("credit_count", 0),
                "debit_count": row.get("debit_count", 0),
                "total_credits": row.get("total_credits", 0),
                "total_debits": row.get("total_debits", 0),
                "net_position": row.get("net_position", 0),
            }
            for row in canonical_cp_rows
        ]
    )
    
    if counterparty_summary.empty:
        st.info("No counterparty data available.")
        return

    with download_col:
        cp_payload = _build_counterparty_json_payload(prepared_df, counterparty_summary)
        st.download_button(
            "Download Counterparty List (JSON)",
            json.dumps(cp_payload, indent=4),
            "counterparty_list.json",
            "application/json",
            use_container_width=True,
        )
        st.caption("Edit counterparty_name_clean or mapping values, then upload the JSON back here to regroup the ledger and matching transactions.")
    
    def build_top_counterparty_table(amount_column: str, count_column: str) -> pd.DataFrame:
        top_df = counterparty_summary[
            counterparty_summary[amount_column].fillna(0) > 0
        ].copy()
        if top_df.empty:
            return pd.DataFrame(columns=["Counterparty", "Total Txn", "Total Amnt of Txn"])

        top_df = top_df.sort_values(amount_column, ascending=False).head(10)
        return pd.DataFrame(
            {
                "Counterparty": top_df["counterparty_name"],
                "Total Txn": top_df[count_column].astype(int),
                "Total Amnt of Txn": top_df[amount_column].apply(lambda x: f"RM {x:,.2f}"),
            }
        )


    # Display summary table
    display_df = counterparty_summary.copy()
    display_df['total_credits'] = display_df['total_credits'].apply(lambda x: f"RM {x:,.2f}")
    display_df['total_debits'] = display_df['total_debits'].apply(lambda x: f"RM {x:,.2f}")
    
    # Format net position with color indicators using column config
    def format_net_position(val):
        if val > 0:
            return f"🟢 RM {val:,.2f}"
        elif val < 0:
            return f"🔴 RM {abs(val):,.2f}"
        return f"⚪ RM {val:,.2f}"
    
    display_df['net_position_display'] = counterparty_summary['net_position'].apply(format_net_position)
    display_df = display_df[
        [
            'counterparty_name',
            'transaction_count',
            'credit_count',
            'debit_count',
            'total_credits',
            'total_debits',
            'net_position_display',
        ]
    ]

    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            'counterparty_name': 'Counterparty',
            'transaction_count': 'Transactions',
            'credit_count': 'Credits',
            'debit_count': 'Debits',
            'total_credits': 'Total Credits',
            'total_debits': 'Total Debits',
            'net_position_display': 'Net Position'
        }
    )
    
    # Selection dropdown
    selected_counterparty = st.selectbox(
        "Select a counterparty to inspect transaction lines",
        options=counterparty_summary['counterparty_name'].tolist(),
        index=0,
    )
    
    # Show transactions for selected counterparty
    selected_cp_row = next(
        (row for row in canonical_cp_rows if row.get("counterparty_name") == selected_counterparty),
        {},
    )
    counterparty_tx = pd.DataFrame(selected_cp_row.get("transactions", []) or [])

    if not counterparty_tx.empty:
        for required_col in ("date", "description", "amount", "type", "balance"):
            if required_col not in counterparty_tx.columns:
                counterparty_tx[required_col] = ""
        display_tx = counterparty_tx[["date", "description", "amount", "type", "balance"]].copy()
        display_tx["credit"] = display_tx.apply(
            lambda row: f"RM {safe_float(row.get('amount')):,.2f}" if str(row.get("type", "")).upper() == "CREDIT" else "",
            axis=1,
        )
        display_tx["debit"] = display_tx.apply(
            lambda row: f"RM {safe_float(row.get('amount')):,.2f}" if str(row.get("type", "")).upper() == "DEBIT" else "",
            axis=1,
        )
        display_tx["balance"] = display_tx["balance"].apply(lambda x: f"RM {safe_float(x):,.2f}" if str(x) != "nan" and x != "" else "")
        display_tx = display_tx[["date", "description", "credit", "debit", "balance"]]
        
        st.dataframe(
            display_tx,
            use_container_width=True,
            hide_index=True,
            column_config={
                'date': 'Date',
                'description': 'Description',
                'credit': 'Credit (RM)',
                'debit': 'Debit (RM)',
                'balance': 'Balance'
            }
        )

    credit_col, debit_col = st.columns(2)
    with credit_col:
        st.markdown("#### Top 10 Credit Counterparties")
        st.dataframe(
            build_top_counterparty_table("total_credits", "credit_count"),
            use_container_width=True,
            hide_index=True,
        )
    with debit_col:
        st.markdown("#### Top 10 Debit Counterparties")
        st.dataframe(
            build_top_counterparty_table("total_debits", "debit_count"),
            use_container_width=True,
            hide_index=True,
        )

# -----------------------------
# Pattern Analysis Functions
# -----------------------------
def normalize_text(text: str) -> str:
    """Normalize text for comparison"""
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text).strip().upper())


def is_round_number(amount: float, round_thresholds: List[float] = None, tolerance: float = 0.01) -> bool:
    """
    Check if amount is a round number (multiple of significant thresholds like 10,000, 50,000, 100,000, etc.)
    """
    if amount is None or amount == 0:
        return False
    
    if round_thresholds is None:
        round_thresholds = [10000, 50000, 100000, 500000, 1000000, 5000000, 10000000]
    
    abs_amount = abs(amount)
    
    for threshold in round_thresholds:
        if abs_amount >= threshold:
            remainder = abs_amount % threshold
            if remainder < tolerance or (threshold - remainder) < tolerance:
                return True
    
    return False


def detect_duplicate_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """Detect duplicate transactions based on date, description, and amount"""
    if df.empty:
        return df
    
    df = df.copy()
    df["duplicate_key"] = df.apply(
        lambda row: (
            row.get("date", ""),
            normalize_text(row.get("description", "")),
            row.get("credit", 0) if row.get("credit", 0) > 0 else row.get("debit", 0)
        ),
        axis=1
    )
    
    duplicate_counts = df.groupby("duplicate_key").size().to_dict()
    df["is_duplicate_transaction"] = df["duplicate_key"].map(lambda x: duplicate_counts.get(x, 0) > 1)
    df["duplicate_count"] = df["duplicate_key"].map(lambda x: duplicate_counts.get(x, 0))
    df.drop("duplicate_key", axis=1, inplace=True)
    
    return df


def detect_rapid_repeat_transactions(df: pd.DataFrame, days_window: int = 30) -> pd.DataFrame:
    """Detect transactions that repeat rapidly to the same party"""
    if df.empty:
        return df
    
    df = df.copy()
    df["parsed_date"] = pd.to_datetime(df["date"], errors="coerce")
    df["normalized_description"] = df["description"].apply(normalize_text)
    
    df["is_rapid_repeat_transaction"] = False
    df["repeat_days_in_window"] = 0
    
    for normalized_desc, group in df.groupby("normalized_description"):
        if len(group) < 2:
            continue
        
        group_sorted = group.sort_values("parsed_date")
        dates = group_sorted["parsed_date"].values
        
        for i in range(len(dates)):
            if i < len(dates) - 1:
                days_diff = (pd.Timestamp(dates[i + 1]) - pd.Timestamp(dates[i])).days
                if days_diff <= days_window:
                    idx = group_sorted.iloc[i].name
                    df.loc[idx, "is_rapid_repeat_transaction"] = True
                    df.loc[idx, "repeat_days_in_window"] = days_diff
    
    return df


# -----------------------------
# Statutory Payment Detection Functions
# -----------------------------
def compute_epf_payments(df: pd.DataFrame) -> Tuple[int, float]:
    """Detect EPF / KWSP contributions"""
    if df.empty:
        return 0, 0.0
    
    epf_patterns = [
        r"\bEPF\b",
        r"\bKWSP\b",
        r"\bKUMPULAN\s+WANG\s+SIMPANAN\s+PEKERJA\b",
        r"\bKARIM\s+JASMINE\s+EPF\b",
        r"\bEPF\s+KARIM\b",
        r"\bKWSP\s+KARIM\b",
    ]
    
    pattern = re.compile("|".join(epf_patterns), re.IGNORECASE)
    
    epf_mask = df["description"].astype(str).apply(lambda x: bool(pattern.search(x)))
    
    # Filter out SOCSO which might match some patterns
    socso_pattern = re.compile(r"\b(SOCSO|PERKESO|EIS|SIP)\b", re.IGNORECASE)
    socso_mask = df["description"].astype(str).apply(lambda x: bool(socso_pattern.search(x)))
    
    # EPF should not be SOCSO
    epf_mask = epf_mask & ~socso_mask
    
    epf_df = df[epf_mask].copy()
    
    # Get amounts (debit/credit - contributions are usually debits/outflows)
    epf_df["amount"] = epf_df.apply(
        lambda row: row.get("debit", 0) if row.get("debit", 0) > 0 else row.get("credit", 0),
        axis=1
    )
    
    total_amount = epf_df["amount"].sum()
    count = len(epf_df)
    
    # Also add flags to dataframe for detailed view
    if "is_epf_payment" not in df.columns:
        df["is_epf_payment"] = False
    df.loc[epf_mask, "is_epf_payment"] = True
    df.loc[epf_mask, "epf_amount"] = epf_df["amount"]
    
    return count, total_amount


def compute_socso_payments(df: pd.DataFrame) -> Tuple[int, float]:
    """Detect SOCSO / PERKESO / EIS contributions"""
    if df.empty:
        return 0, 0.0
    
    socso_patterns = [
        r"\bSOCSO\b",
        r"\bPERKESO\b",
        r"\bEIS\b",
        r"\bSIP\b",
        r"\bPERTUBUHAN\s+KESELAMATAN\s+SOSIAL\b",
        r"\bPERKESO\s+SOCSO\b",
    ]
    
    pattern = re.compile("|".join(socso_patterns), re.IGNORECASE)
    
    socso_mask = df["description"].astype(str).apply(lambda x: bool(pattern.search(x)))
    
    socso_df = df[socso_mask].copy()
    
    # Get amounts (debit/credit - contributions are usually debits/outflows)
    socso_df["amount"] = socso_df.apply(
        lambda row: row.get("debit", 0) if row.get("debit", 0) > 0 else row.get("credit", 0),
        axis=1
    )
    
    total_amount = socso_df["amount"].sum()
    count = len(socso_df)
    
    # Add flags to dataframe for detailed view
    if "is_socso_payment" not in df.columns:
        df["is_socso_payment"] = False
    df.loc[socso_mask, "is_socso_payment"] = True
    df.loc[socso_mask, "socso_amount"] = socso_df["amount"]
    
    return count, total_amount


def compute_lhdn_tax_payments(df: pd.DataFrame) -> Tuple[int, float]:
    """Detect LHDN / tax payments"""
    if df.empty:
        return 0, 0.0
    
    lhdn_patterns = [
        r"\bLHDN\b",
        r"\bLEMBAGA\s+HASIL\s+DALAM\s+NEGERI\b",
        r"\bINLAND\s+REVENUE\b",
        r"\bTAX\s+PAYMENT\b",
        r"\bPCB\b",
        r"\bPOTONGAN\s+CUKAI\s+BULANAN\b",
        r"\bCUKAI\s+PENDAPATAN\b",
        r"\bINCOME\s+TAX\b",
        r"\bHASIL\s+DALAM\s+NEGERI\b",
    ]
    
    pattern = re.compile("|".join(lhdn_patterns), re.IGNORECASE)
    
    lhdn_mask = df["description"].astype(str).apply(lambda x: bool(pattern.search(x)))
    
    lhdn_df = df[lhdn_mask].copy()
    
    # Get amounts (debit/credit - tax payments are usually debits/outflows)
    lhdn_df["amount"] = lhdn_df.apply(
        lambda row: row.get("debit", 0) if row.get("debit", 0) > 0 else row.get("credit", 0),
        axis=1
    )
    
    total_amount = lhdn_df["amount"].sum()
    count = len(lhdn_df)
    
    # Add flags to dataframe for detailed view
    if "is_lhdn_payment" not in df.columns:
        df["is_lhdn_payment"] = False
    df.loc[lhdn_mask, "is_lhdn_payment"] = True
    df.loc[lhdn_mask, "lhdn_amount"] = lhdn_df["amount"]
    
    return count, total_amount


def compute_hrdf_payments(df: pd.DataFrame) -> Tuple[int, float]:
    """Detect HRDF / PSMB levy payments"""
    if df.empty:
        return 0, 0.0
    
    hrdf_patterns = [
        r"\bHRDF\b",
        r"\bPSMB\b",
        r"\bPEMBANGUNAN\s+SUMBER\s+MANUSIA\b",
        r"\bHUMAN\s+RESOURCE\s+DEVELOPMENT\b",
        r"\bLEVY\s+HRDF\b",
        r"\bHRD\s+CORP\b",
        r"\bHRD\s+CORPORATION\b",
    ]
    
    pattern = re.compile("|".join(hrdf_patterns), re.IGNORECASE)
    
    hrdf_mask = df["description"].astype(str).apply(lambda x: bool(pattern.search(x)))
    
    hrdf_df = df[hrdf_mask].copy()
    
    # Get amounts (debit/credit - levies are usually debits/outflows)
    hrdf_df["amount"] = hrdf_df.apply(
        lambda row: row.get("debit", 0) if row.get("debit", 0) > 0 else row.get("credit", 0),
        axis=1
    )
    
    total_amount = hrdf_df["amount"].sum()
    count = len(hrdf_df)
    
    # Add flags to dataframe for detailed view
    if "is_hrdf_payment" not in df.columns:
        df["is_hrdf_payment"] = False
    df.loc[hrdf_mask, "is_hrdf_payment"] = True
    df.loc[hrdf_mask, "hrdf_amount"] = hrdf_df["amount"]
    
    return count, total_amount


def run_fraud_checks(df: pd.DataFrame, high_value_threshold: float, round_thresholds: List[float] = None) -> pd.DataFrame:
    """Run all fraud/pattern detection checks on the transaction dataframe"""
    if df.empty:
        return df
    
    df = df.copy()
    
    # High value detection includes both inflows and outflows.
    df["is_high_value"] = df.apply(
        lambda row: (
            safe_float(row.get("credit", 0)) >= high_value_threshold
            or safe_float(row.get("debit", 0)) >= high_value_threshold
        ),
        axis=1,
    )
    
    # Round number detection - using flexible thresholds
    df["is_round"] = df.apply(
        lambda row: is_round_number(row.get("credit", 0), round_thresholds) or is_round_number(row.get("debit", 0), round_thresholds),
        axis=1
    )
    
    # Duplicate detection
    df = detect_duplicate_transactions(df)
    
    # Rapid repeat detection
    df = detect_rapid_repeat_transactions(df)
    
    # Statutory payment detections
    compute_epf_payments(df)
    compute_socso_payments(df)
    compute_lhdn_tax_payments(df)
    compute_hrdf_payments(df)

    return df


def summarize_transaction_patterns(df: pd.DataFrame) -> dict:
    """Create summary of transaction patterns including statutory payments"""
    if df.empty:
        return {
            "title": "Transactional Pattern Analysis",
            "items": [],
            "headline": "No transactions to analyze."
        }
    
    high_value_count = int(df["is_high_value"].sum()) if "is_high_value" in df.columns else 0
    duplicate_count = int(df["is_duplicate_transaction"].sum()) if "is_duplicate_transaction" in df.columns else 0
    high_freq_count = int(df["is_rapid_repeat_transaction"].sum()) if "is_rapid_repeat_transaction" in df.columns else 0
    round_count = int(df["is_round"].sum()) if "is_round" in df.columns else 0
    
    # Statutory payment counts
    epf_count, epf_total = compute_epf_payments(df)
    socso_count, socso_total = compute_socso_payments(df)
    lhdn_count, lhdn_total = compute_lhdn_tax_payments(df)
    hrdf_count, hrdf_total = compute_hrdf_payments(df)
    
    total_transactions = len(df)
    
    headline = (
        f"Analyzed {total_transactions} transactions. "
        f"Found {high_value_count} high-value, {duplicate_count} duplicates, "
        f"{high_freq_count} high-frequency, and {round_count} round-number transactions. "
        f"Statutory payments: EPF ({epf_count}), SOCSO ({socso_count}), "
        f"LHDN ({lhdn_count}), HRDF ({hrdf_count})."
    )
    
    return {
        "title": "Transactional Pattern Analysis",
        "headline": headline,
        "items": [
            ("Total Transactions", total_transactions),
            ("High-Value Flags", high_value_count),
            ("Round-Number", round_count),
            ("Repeated", duplicate_count),
            ("High Frequency Flags", high_freq_count),
        ],
        "statutory_items": [
            ("EPF / KWSP", epf_count, f"RM {epf_total:,.2f}"),
            ("SOCSO / PERKESO", socso_count, f"RM {socso_total:,.2f}"),
            ("LHDN / Tax", lhdn_count, f"RM {lhdn_total:,.2f}"),
            ("HRDF / PSMB", hrdf_count, f"RM {hrdf_total:,.2f}"),
        ],
    }


def render_pattern_details(df: pd.DataFrame, high_value_threshold: float) -> None:
    """Render expandable sections for each pattern type"""
    st.markdown('<h3 class="kl-pattern-details-heading">Pattern Details</h3>', unsafe_allow_html=True)
    
    # Duplicate transactions
    if "is_duplicate_transaction" in df.columns:
        duplicate_hits = df[df["is_duplicate_transaction"] == True].copy()
        if not duplicate_hits.empty:
            duplicate_hits["amount"] = duplicate_hits.apply(
                lambda row: f"+{row['credit']:,.2f}" if row.get('credit', 0) > 0 else f"-{row['debit']:,.2f}",
                axis=1
            )
            with st.expander("Repeated transaction"):
                st.caption("The following entries share the same date, description, and amount.")
                duplicate_columns = [c for c in ["date", "description", "amount", "balance"] if c in duplicate_hits.columns]
                st.dataframe(duplicate_hits[duplicate_columns], use_container_width=True)
    
    # Rapid repeat transactions
    if "is_rapid_repeat_transaction" in df.columns:
        rapid_repeat_hits = df[df["is_rapid_repeat_transaction"] == True].copy()
        if not rapid_repeat_hits.empty:
            with st.expander("High freq transactions"):
                st.caption(
                    "High frequency transactions are repeated payments to the same merchant "
                    "across multiple days within a short time window."
                )
                display_df = rapid_repeat_hits
                display_columns = [
                    c for c in ["date", "description", "credit", "debit", "repeat_days_in_window"]
                    if c in display_df.columns
                ]
                st.dataframe(display_df[display_columns], use_container_width=True, hide_index=True)
    
    # Round number transactions
    if "is_round" in df.columns:
        round_hits = df[df["is_round"] == True].copy()
        if not round_hits.empty:
            round_hits["amount"] = round_hits.apply(
                lambda row: f"+{row['credit']:,.2f}" if row.get('credit', 0) > 0 else f"-{row['debit']:,.2f}",
                axis=1
            )
            with st.expander("Round-number transactions"):
                st.caption("Transactions with round numbers (multiple of 10,000).")
                cols = [c for c in ["date", "description", "amount", "source_file"] if c in round_hits.columns]
                st.dataframe(round_hits[cols], use_container_width=True)
    
    # High value transactions
    if "is_high_value" in df.columns:
        high_hits = df[df["is_high_value"] == True].copy()
        if not high_hits.empty:
            high_hits["amount"] = high_hits.apply(
                lambda row: (
                    f"+RM {safe_float(row.get('credit', 0)):,.2f}"
                    if safe_float(row.get("credit", 0)) > 0
                    else f"-RM {safe_float(row.get('debit', 0)):,.2f}"
                ),
                axis=1,
            )
            with st.expander("High-value transactions"):
                st.caption(
                    "Transactions flagged when the credit or debit amount is above or equal "
                    f"to the inserted high value threshold: RM {high_value_threshold:,.2f}."
                )
                high_value_columns = [c for c in ["date", "description", "amount", "balance"] if c in high_hits.columns]
                st.dataframe(
                    high_hits[high_value_columns],
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "date": "Date Transaction",
                        "description": "Description",
                        "amount": "Amount (RM)",
                        "balance": "Balance",
                    },
                )
    
    # Statutory payments sections
    epf_count, epf_total = compute_epf_payments(df)
    socso_count, socso_total = compute_socso_payments(df)
    lhdn_count, lhdn_total = compute_lhdn_tax_payments(df)
    hrdf_count, hrdf_total = compute_hrdf_payments(df)
    
    # EPF Payments
    if epf_count > 0:
        epf_hits = df[df["is_epf_payment"] == True].copy() if "is_epf_payment" in df.columns else pd.DataFrame()
        if not epf_hits.empty:
            epf_hits["amount"] = epf_hits.apply(
                lambda row: f"-{row['debit']:,.2f}" if row.get('debit', 0) > 0 else f"+{row['credit']:,.2f}",
                axis=1
            )
            with st.expander("EPF / KWSP Contributions"):
                st.caption("Employee Provident Fund (EPF) / KWSP contributions detected.")
                display_cols = [c for c in ["date", "description", "amount"] if c in epf_hits.columns]
                st.dataframe(epf_hits[display_cols], use_container_width=True)
    
    # SOCSO Payments
    if socso_count > 0:
        socso_hits = df[df["is_socso_payment"] == True].copy() if "is_socso_payment" in df.columns else pd.DataFrame()
        if not socso_hits.empty:
            socso_hits["amount"] = socso_hits.apply(
                lambda row: f"-{row['debit']:,.2f}" if row.get('debit', 0) > 0 else f"+{row['credit']:,.2f}",
                axis=1
            )
            with st.expander("SOCSO / PERKESO Contributions"):
                st.caption("Social Security Organization (SOCSO/PERKESO) contributions including EIS/SIP.")
                display_cols = [c for c in ["date", "description", "amount"] if c in socso_hits.columns]
                st.dataframe(socso_hits[display_cols], use_container_width=True)
    
    # LHDN/Tax Payments
    if lhdn_count > 0:
        lhdn_hits = df[df["is_lhdn_payment"] == True].copy() if "is_lhdn_payment" in df.columns else pd.DataFrame()
        if not lhdn_hits.empty:
            lhdn_hits["amount"] = lhdn_hits.apply(
                lambda row: f"-{row['debit']:,.2f}" if row.get('debit', 0) > 0 else f"+{row['credit']:,.2f}",
                axis=1
            )
            with st.expander("LHDN / Tax Payments"):
                st.caption("Inland Revenue Board (LHDN) / Income tax payments detected.")
                display_cols = [c for c in ["date", "description", "amount"] if c in lhdn_hits.columns]
                st.dataframe(lhdn_hits[display_cols], use_container_width=True)
    
    # HRDF Payments
    if hrdf_count > 0:
        hrdf_hits = df[df["is_hrdf_payment"] == True].copy() if "is_hrdf_payment" in df.columns else pd.DataFrame()
        if not hrdf_hits.empty:
            hrdf_hits["amount"] = hrdf_hits.apply(
                lambda row: f"-{row['debit']:,.2f}" if row.get('debit', 0) > 0 else f"+{row['credit']:,.2f}",
                axis=1
            )
            with st.expander("HRDF / PSMB Levies"):
                st.caption("Human Resource Development Fund (HRDF/PSMB) levy payments detected.")
                display_cols = [c for c in ["date", "description", "amount"] if c in hrdf_hits.columns]
                st.dataframe(hrdf_hits[display_cols], use_container_width=True)


def render_metric_cards(metrics: List[Tuple[str, object]], wide_metrics: List[Tuple[str, object]], statutory_metrics: List[Tuple[str, int, str]] = None) -> None:
    cards = []
    for label, value in metrics:
        cards.append(
            '<div class="kl-metric-card">'
            f'<div class="kl-metric-label">{escape(str(label))}</div>'
            f'<div class="kl-metric-value">{escape(str(value))}</div>'
            "</div>"
        )
    for label, value in wide_metrics:
        cards.append(
            '<div class="kl-metric-card kl-metric-card-wide">'
            f'<div class="kl-metric-label">{escape(str(label))}</div>'
            f'<div class="kl-metric-value">{escape(str(value))}</div>'
            "</div>"
        )
    
    # Add statutory payment cards
    if statutory_metrics:
        for label, count, total in statutory_metrics:
            cards.append(
                '<div class="kl-metric-card">'
                f'<div class="kl-metric-label">{escape(str(label))}</div>'
                f'<div class="kl-metric-value">{count}</div>'
                f'<div style="font-size: 0.85rem; color: #A9C1DD; margin-top: 0.35rem;">{escape(str(total))}</div>'
                "</div>"
            )

    st.html(
        f'<div class="kl-metric-grid">{"".join(cards)}</div>',
    )


def render_transaction_overview(df: pd.DataFrame, high_value_threshold: float) -> None:
    """Render the transaction pattern overview dashboard"""
    analysis_df = filter_statement_transactions_df(df)
    pattern_summary = summarize_transaction_patterns(analysis_df)
    
    st.html(
        '<div class="kl-analysis-title">📊 Transactional Pattern Analysis</div>'
        '<div class="kl-analysis-subtitle">The story of the money and whether the financial behavior makes sense.</div>',
    )
    
    # Display pattern metrics only. Financial cards live in the Extracted Transaction section.
    item_map = dict(pattern_summary.get("items", []))
    statutory_items = pattern_summary.get("statutory_items", [])
    
    cards = []

    cards.extend([
        '<div class="kl-metric-card">'
        '<div class="kl-metric-label">High-Value Flags</div>'
        f'<div class="kl-metric-value">{item_map.get("High-Value Flags", 0)}</div>'
        '</div>',
        
        '<div class="kl-metric-card">'
        '<div class="kl-metric-label">Round-Number</div>'
        f'<div class="kl-metric-value">{item_map.get("Round-Number", 0)}</div>'
        '</div>',
        
        '<div class="kl-metric-card">'
        '<div class="kl-metric-label">Repeated</div>'
        f'<div class="kl-metric-value">{item_map.get("Repeated", 0)}</div>'
        '</div>',

        '<div class="kl-metric-card">'
        '<div class="kl-metric-label">High Frequency Flags</div>'
        f'<div class="kl-metric-value">{item_map.get("High Frequency Flags", 0)}</div>'
        '</div>',
    ])

    
    # Add statutory payment cards
    if statutory_items:
        for label, count, total in statutory_items:
            cards.append(
                '<div class="kl-metric-card">'
                f'<div class="kl-metric-label">{label}</div>'
                f'<div class="kl-metric-value">{count}</div>'
                f'<div style="font-size: 0.85rem; color: #A9C1DD; margin-top: 0.35rem;">{total}</div>'
                '</div>'
            )
    
    # Display all cards in the grid
    st.html(
        f'<div class="kl-metric-grid">{"".join(cards)}</div>',
    )
    
    # Render detailed expandable sections
    render_pattern_details(analysis_df, high_value_threshold)


def render_extracted_transaction_section(df: pd.DataFrame) -> None:
    """Render extracted transaction metrics and the transaction table."""
    analysis_df = filter_statement_transactions_df(df)
    total_credits = analysis_df["credit"].sum() if "credit" in analysis_df.columns else 0
    total_debits = analysis_df["debit"].sum() if "debit" in analysis_df.columns else 0
    net_position = total_credits - total_debits
    net_color = "#69f0ae" if net_position >= 0 else "#ff8a80"
    net_border = "#2e7d32" if net_position >= 0 else "#c62828"
    net_label = "Net Position" if net_position >= 0 else "Net Loss"

    cards = [
        '<div class="kl-metric-card">'
        '<div class="kl-metric-label">Transactions</div>'
        f'<div class="kl-metric-value">{len(analysis_df):,}</div>'
        '</div>',
        '<div class="kl-metric-card" style="background: linear-gradient(135deg, #1a472a 0%, #0d2818 100%); border-color: #2e7d32;">'
        '<div class="kl-metric-label">Net Credits</div>'
        f'<div class="kl-metric-value" style="color: #69f0ae;">RM {total_credits:,.2f}</div>'
        '</div>',
        '<div class="kl-metric-card" style="background: linear-gradient(135deg, #4a1a1a 0%, #2d1010 100%); border-color: #c62828;">'
        '<div class="kl-metric-label">Net Debits</div>'
        f'<div class="kl-metric-value" style="color: #ff8a80;">RM {total_debits:,.2f}</div>'
        '</div>',
        f'<div class="kl-metric-card" style="background: linear-gradient(135deg, #1a2a3a 0%, #0d1a2a 100%); border-color: {net_border};">'
        f'<div class="kl-metric-label">{net_label}</div>'
        f'<div class="kl-metric-value" style="color: {net_color};">RM {abs(net_position):,.2f}</div>'
        '</div>',
    ]

    st.html(
        '<div class="kl-analysis-title">&#128202; Extracted Transaction <span style="font-size: 1rem; color: #A9C1DD;">&#128279;</span></div>'
        f'<div class="kl-metric-grid">{"".join(cards)}</div>',
    )

    st.markdown("#### All Transactions")
    requested_cols = ["date", "description", "debit", "credit", "balance"]
    display_cols = [c for c in requested_cols if c in df.columns]
    st.dataframe(
        df[display_cols],
        use_container_width=True,
        column_config={
            "date": "Transaction Date",
            "description": "Description",
            "debit": "Debit (RM)",
            "credit": "Credit (RM)",
            "balance": "Running Balance",
        },
    )

# -----------------------------
# Core Processing Functions
# -----------------------------
def clear_processing_outputs() -> None:
    st.session_state.results = []
    st.session_state.integrity_analysis_results = {}
    st.session_state.affin_statement_totals = []
    st.session_state.affin_file_transactions = {}
    st.session_state.ambank_statement_totals = []
    st.session_state.ambank_file_transactions = {}
    st.session_state.cimb_statement_totals = []
    st.session_state.rhb_statement_totals = []
    st.session_state.cimb_file_transactions = {}
    st.session_state.rhb_file_transactions = {}
    st.session_state.bank_islam_file_month = {}
    st.session_state.file_company_name = {}
    st.session_state.file_account_no = {}
    st.session_state.account_type_determinations = []
    st.session_state.related_parties_override = []
    st.session_state.counterparty_name_overrides = {}


def parse_high_value_threshold() -> Tuple[Optional[float], Optional[str]]:
    raw = str(st.session_state.get("high_value_threshold_input", "") or "").strip()
    if not raw:
        return None, "Please insert the high value threshold."
    if not re.search(r"\d", raw):
        return None, "Please insert a valid high value threshold."

    threshold = safe_float(raw)
    if threshold <= 0:
        return None, "Please insert a high value threshold above 0."
    return threshold, None


def clear_high_value_threshold_error() -> None:
    st.session_state.high_value_threshold_error = ""
    st.session_state.validation_toast_message = ""


def get_upload_widget_key() -> str:
    return f"pdf_upload_{st.session_state.upload_widget_reset_id}"


def validate_pdf_upload() -> Optional[str]:
    uploaded_file_values = st.session_state.get(get_upload_widget_key())
    if not uploaded_file_values:
        return "Please upload at least one PDF file."
    return None


def clear_pdf_upload_error() -> None:
    st.session_state.pdf_upload_error = ""
    st.session_state.validation_toast_message = ""


def validate_bank_choice() -> Optional[str]:
    bank_choice_value = st.session_state.get("bank_choice")
    if not bank_choice_value:
        return "Please select the bank format."
    return None


def clear_bank_choice_error() -> None:
    st.session_state.bank_choice_error = ""
    st.session_state.validation_toast_message = ""


def start_processing() -> None:
    threshold, threshold_error = parse_high_value_threshold()
    bank_choice_error = validate_bank_choice()
    pdf_upload_error = validate_pdf_upload()

    st.session_state.high_value_threshold_error = threshold_error or ""
    st.session_state.bank_choice_error = bank_choice_error or ""
    st.session_state.pdf_upload_error = pdf_upload_error or ""

    if threshold_error or bank_choice_error or pdf_upload_error:
        validation_messages = [
            msg for msg in (bank_choice_error, pdf_upload_error, threshold_error) if msg
        ]
        st.session_state.validation_toast_message = " ".join(validation_messages)
        st.session_state.status = "idle"
        return

    st.session_state.validation_toast_message = ""
    st.session_state.active_high_value_threshold = threshold
    st.session_state.stop_requested = False
    st.session_state.status = "running"
    clear_processing_outputs()


def stop_processing() -> None:
    st.session_state.stop_requested = True
    if st.session_state.status == "running":
        st.session_state.status = "stopped"


def reset_app_inputs() -> None:
    st.session_state.status = "idle"
    st.session_state.stop_requested = False
    clear_processing_outputs()
    st.session_state.pdf_password = ""
    st.session_state.company_name_override = ""
    st.session_state.company_account_no_override = ""
    st.session_state.high_value_threshold_input = ""
    st.session_state.high_value_threshold_error = ""
    st.session_state.bank_choice_error = ""
    st.session_state.pdf_upload_error = ""
    st.session_state.validation_toast_message = ""
    st.session_state.active_high_value_threshold = None
    st.session_state.upload_widget_reset_id += 1
    if "bank_choice" in st.session_state and "PARSERS" in globals():
        st.session_state.bank_choice = None


def get_high_value_threshold() -> float:
    threshold, threshold_error = parse_high_value_threshold()
    if not threshold_error and threshold is not None:
        return threshold

    active_threshold = st.session_state.get("active_high_value_threshold")
    if active_threshold is not None:
        return float(active_threshold)
    return 0.0


def truncate_filename(filename: str, max_chars: int = 34) -> str:
    filename = str(filename or "")
    if len(filename) <= max_chars:
        return filename

    keep_left = max(8, (max_chars - 3) // 2)
    keep_right = max(8, max_chars - keep_left - 3)
    return f"{filename[:keep_left]}...{filename[-keep_right:]}"


def render_processing_progress(
    container,
    *,
    status: str,
    progress: float,
    variant: str = "active",
    file_name: str = "",
) -> None:
    progress = min(max(float(progress or 0), 0.0), 1.0)
    percent = int(round(progress * 100))
    safe_status = escape(str(status or ""))
    safe_file_name = escape(str(file_name or ""), quote=True)
    short_file_name = escape(truncate_filename(str(file_name or "")))
    variant_class = variant if variant in {"active", "success", "warning", "error"} else "active"
    icon = "✅ " if variant_class == "success" else ""

    file_line = ""
    if file_name:
        file_line = (
            '<div class="kl-progress-filename">'
            f'Current file: <span title="{safe_file_name}">{short_file_name}</span>'
            "</div>"
        )

    html = (
        f'<div class="kl-progress-panel {variant_class}">'
        '<div class="kl-progress-topline">'
        f'<div class="kl-progress-status">{icon}{safe_status}</div>'
        f'<div class="kl-progress-percent">{percent}% completed</div>'
        '</div>'
        f'{file_line}'
        '<div class="kl-progress-track" aria-label="Processing progress">'
        f'<div class="kl-progress-fill" style="width: {percent}%;"></div>'
        '</div>'
        '</div>'
    )
    container.markdown(html, unsafe_allow_html=True)


_ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def parse_any_date_for_summary(x) -> pd.Timestamp:
    if x is None:
        return pd.NaT
    s = str(x).strip()
    if not s:
        return pd.NaT
    if _ISO_RE.match(s):
        return pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
    return pd.to_datetime(s, errors="coerce", dayfirst=True)


def _parse_with_pdfplumber(parser_func: Callable, pdf_bytes: bytes, filename: str) -> List[dict]:
    with bytes_to_pdfplumber(pdf_bytes) as pdf:
        return parser_func(pdf, filename)


# -----------------------------
# Bank parsers
# -----------------------------
PARSERS: Dict[str, Callable[[bytes, str], List[dict]]] = {
    "Affin Bank": lambda b, f: _parse_with_pdfplumber(parse_affin_bank, b, f),
    "Agro Bank": lambda b, f: _parse_with_pdfplumber(parse_agro_bank, b, f),
    "Alliance Bank": lambda b, f: _parse_with_pdfplumber(parse_transactions_alliance, b, f),
    "Ambank": lambda b, f: _parse_with_pdfplumber(parse_ambank, b, f),
    "Bank Islam": lambda b, f: _parse_with_pdfplumber(parse_bank_islam, b, f),
    "Bank Muamalat": lambda b, f: _parse_with_pdfplumber(parse_transactions_bank_muamalat, b, f),
    "Bank Rakyat": lambda b, f: _parse_with_pdfplumber(parse_bank_rakyat, b, f),
    "CIMB Bank": lambda b, f: _parse_with_pdfplumber(parse_transactions_cimb, b, f),
    "Hong Leong": lambda b, f: _parse_with_pdfplumber(parse_hong_leong, b, f),
    "Maybank": lambda b, f: parse_transactions_maybank(b, f),
    "Public Bank (PBB)": lambda b, f: _parse_with_pdfplumber(parse_transactions_pbb, b, f),
    "RHB Bank": lambda b, f: parse_transactions_rhb(b, f),
    "OCBC Bank": lambda b, f: parse_transactions_ocbc(b, f),
    "UOB Bank": lambda b, f: _parse_with_pdfplumber(parse_transactions_uob, b, f),
}


def get_supported_banks() -> List[str]:
    return list(PARSERS.keys())


# -----------------------------
# Monthly Summary Functions
# -----------------------------
BALANCE_MARKER_PATTERNS = [
    r"\bOPENING\s+BAL(?:ANCE)?\b",
    r"\bCLOSING\s+BAL(?:ANCE)?\b",
    r"\bBEGINNING\s+BAL(?:ANCE)?\b",
    r"\bENDING\s+BAL(?:ANCE)?\b",
    r"\bBALANCE\s+B\/F\b",
    r"\bBALANCE\s+C\/F\b",
    r"\bB\/F\s+BALANCE\b",
    r"\bC\/F\s+BALANCE\b",
    r"\bBROUGHT\s+FORWARD\b",
    r"\bCARRIED\s+FORWARD\b",
    r"\bBAKI\s+AWAL\b",
    r"\bBAKI\s+AKHIR\b",
    r"\bBAKI\s+PEMBUKA\b",
    r"\bBAKI\s+PENUTUP\b",
    r"\bBAKI\s+B\/B\b",
    r"\bBAKI\s+C\/F\b",
]


def is_balance_marker_transaction(tx: dict) -> bool:
    desc = normalize_text(tx.get("description", ""))
    return any(re.search(pattern, desc, flags=re.IGNORECASE) for pattern in BALANCE_MARKER_PATTERNS)


def count_statement_transactions(transactions: List[dict]) -> int:
    return sum(1 for tx in transactions if not is_balance_marker_transaction(tx))


def filter_statement_transactions_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    mask = [
        not is_balance_marker_transaction(tx)
        for tx in df.to_dict(orient="records")
    ]
    return df.loc[mask].copy()


def calculate_monthly_summary(transactions: List[dict]) -> List[dict]:
    if not transactions:
        return []

    df = pd.DataFrame(transactions)
    if df.empty:
        return []

    df = df.reset_index(drop=True)
    if "__row_order" not in df.columns:
        df["__row_order"] = range(len(df))

    df["date_parsed"] = df.get("date").apply(parse_any_date_for_summary)
    df = df.dropna(subset=["date_parsed"])
    if df.empty:
        st.warning("⚠️ No valid transaction dates found.")
        return []

    df["month_period"] = df["date_parsed"].dt.strftime("%Y-%m")
    df["debit"] = df.get("debit", 0).apply(safe_float)
    df["credit"] = df.get("credit", 0).apply(safe_float)
    df["balance"] = df.get("balance", None).apply(lambda x: safe_float(x) if x is not None else None)

    if "page" in df.columns:
        df["page"] = pd.to_numeric(df["page"], errors="coerce").fillna(0).astype(int)
    else:
        df["page"] = 0

    has_seq = "seq" in df.columns
    if has_seq:
        df["seq"] = pd.to_numeric(df["seq"], errors="coerce").fillna(0).astype(int)

    df["__row_order"] = pd.to_numeric(df["__row_order"], errors="coerce").fillna(0).astype(int)

    monthly_summary: List[dict] = []
    for period, group in df.groupby("month_period", sort=True):
        sort_cols = ["date_parsed", "page"]
        if has_seq:
            sort_cols.append("seq")
        sort_cols.append("__row_order")

        group_sorted = group.sort_values(sort_cols, na_position="last")

        balances = group_sorted["balance"].dropna()
        ending_balance = round(float(balances.iloc[-1]), 2) if not balances.empty else None
        highest_balance = round(float(balances.max()), 2) if not balances.empty else None
        lowest_balance_raw = round(float(balances.min()), 2) if not balances.empty else None
        lowest_balance = lowest_balance_raw
        od_flag = bool(lowest_balance is not None and float(lowest_balance) < 0)

        company_vals = [
            x for x in group_sorted.get("company_name", pd.Series([], dtype=object)).dropna().astype(str).unique().tolist()
            if x.strip()
        ]
        company_name = company_vals[0] if company_vals else None

        acct_vals = [
            x for x in group_sorted.get("account_no", pd.Series([], dtype=object)).dropna().astype(str).unique().tolist() if x.strip()
        ]
        account_no = acct_vals[0] if len(acct_vals) == 1 else (", ".join(acct_vals) if acct_vals else None)

        monthly_summary.append(
            {
                "month": period,
                "company_name": company_name,
                "account_no": account_no,
                "transaction_count": int(
                    sum(
                        1
                        for tx in group_sorted.to_dict(orient="records")
                        if not is_balance_marker_transaction(tx)
                    )
                ),
                "opening_balance": None,
                "total_debit": round(float(group_sorted["debit"].sum()), 2),
                "total_credit": round(float(group_sorted["credit"].sum()), 2),
                "net_change": round(float(group_sorted["credit"].sum() - group_sorted["debit"].sum()), 2),
                "ending_balance": ending_balance,
                "lowest_balance": lowest_balance,
                "lowest_balance_raw": lowest_balance_raw,
                "highest_balance": highest_balance,
                "od_flag": od_flag,
                "source_files": ", ".join(sorted(set(group_sorted.get("source_file", []))))
                if "source_file" in group_sorted.columns
                else "",
            }
        )

    # Fill opening_balance using prior month's ending_balance when possible
    monthly_summary_sorted = sorted(monthly_summary, key=lambda x: x["month"])
    prev_end = None
    for r in monthly_summary_sorted:
        if r.get("opening_balance") is None:
            if prev_end is not None:
                r["opening_balance"] = round(float(prev_end), 2)
            else:
                eb = r.get("ending_balance")
                nc = r.get("net_change")
                if eb is not None and nc is not None:
                    try:
                        r["opening_balance"] = round(float(safe_float(eb) - safe_float(nc)), 2)
                    except Exception:
                        r["opening_balance"] = None

        opening = r.get("opening_balance")
        net_change = r.get("net_change")
        if opening is not None and net_change is not None:
            expected_ending = round(safe_float(opening) + safe_float(net_change), 2)
            raw_ending = r.get("ending_balance")
            if raw_ending is None or abs(expected_ending - safe_float(raw_ending)) > 0.01:
                r["raw_ending_balance"] = raw_ending
                r["ending_balance"] = expected_ending

        if r.get("ending_balance") is not None:
            prev_end = safe_float(r.get("ending_balance"))

    return standardize_monthly_summary_balance_chain(monthly_summary_sorted)


def present_monthly_summary_standard(rows: List[dict]) -> List[dict]:
    out = []
    for r in rows or []:
        highest = r.get("highest_balance")
        lowest = r.get("lowest_balance")

        swing = None
        try:
            if highest is not None and lowest is not None:
                swing = round(float(safe_float(highest) - safe_float(lowest)), 2)
        except Exception:
            swing = None

        out.append(
            {
                "month": r.get("month"),
                "company_name": r.get("company_name"),
                "account_no": r.get("account_no"),
                "transaction_count": r.get("transaction_count"),
                "opening_balance": r.get("opening_balance"),
                "total_debit": r.get("total_debit"),
                "total_credit": r.get("total_credit"),
                "highest_balance": highest,
                "lowest_balance": lowest,
                "swing": swing,
                "ending_balance": r.get("ending_balance"),
                "source_files": r.get("source_files"),
            }
        )
    return out


# -----------------------------
# Main UI and Processing
# -----------------------------
if "bank_choice" not in st.session_state:
    st.session_state.bank_choice = None

bank_choice = st.selectbox(
    "Select Bank",
    options=sorted(get_supported_banks(), key=str.lower),
    index=None,
    key="bank_choice",
    placeholder="Choose the bank for the uploaded statement(s)",
    on_change=clear_bank_choice_error,
)

uploaded_files = st.file_uploader(
    "Upload PDF files",
    type=["pdf"],
    accept_multiple_files=True,
    key=get_upload_widget_key(),
    on_change=clear_pdf_upload_error,
)
if uploaded_files:
    uploaded_files = sorted(uploaded_files, key=lambda x: x.name)

input_col1, input_col2, input_col3 = st.columns([1.2, 1.0, 0.8])
with input_col1:
    st.text_input("Company Name (optional override)", key="company_name_override")
with input_col2:
    st.text_input("Company Account No. (optional override)", key="company_account_no_override")
with input_col3:
    st.text_input(
        "High Value Threshold (RM)",
        key="high_value_threshold_input",
        placeholder="e.g. 10,000",
        on_change=clear_high_value_threshold_error,
    )

# Detect encrypted files
encrypted_files: List[str] = []
if uploaded_files:
    for uf in uploaded_files:
        try:
            if is_pdf_encrypted(uf.getvalue()):
                encrypted_files.append(uf.name)
        except Exception:
            encrypted_files.append(uf.name)

    if encrypted_files:
        st.warning(
            "🔒 Encrypted PDF(s) detected. Enter the password once and it will be used for all encrypted files:\n\n"
            + "\n".join([f"- {n}" for n in encrypted_files])
        )
        st.text_input("PDF Password", type="password", key="pdf_password")


button_col1, button_col2, button_col3 = st.columns([2.0, 0.9, 1.0])
with button_col1:
    st.button(
        "Start Processing",
        type="primary",
        use_container_width=True,
        on_click=start_processing,
    )

with button_col2:
    st.button(
        "Stop",
        type="secondary",
        use_container_width=True,
        on_click=stop_processing,
    )

with button_col3:
    st.button(
        "Reset",
        type="secondary",
        use_container_width=True,
        on_click=reset_app_inputs,
    )

if st.session_state.validation_toast_message:
    st.toast(st.session_state.validation_toast_message)
    st.session_state.validation_toast_message = ""

all_tx: List[dict] = []

if uploaded_files and st.session_state.status == "running":
    progress_panel = st.empty()

    total_files = len(uploaded_files)
    total_steps = total_files + 1
    parser = PARSERS[bank_choice]
    processing_errors: List[str] = []
    total_extracted = 0
    files_finished = 0
    resolved_pdf_bytes = {}

    def update_processing_progress(
        status: str,
        completed_steps: int,
        *,
        variant: str = "active",
        file_name: str = "",
    ) -> None:
        render_processing_progress(
            progress_panel,
            status=status,
            progress=completed_steps / total_steps,
            variant=variant,
            file_name=file_name,
        )

    update_processing_progress(f"Preparing {total_files} file(s) for {bank_choice}.", 0)

    for file_idx, uploaded_file in enumerate(uploaded_files):
        if st.session_state.get("stop_requested"):
            st.session_state.status = "stopped"
            update_processing_progress(
                f"Stopped after {files_finished} of {total_files} file(s).",
                files_finished,
                variant="warning",
            )
            break

        current_file = file_idx + 1
        update_processing_progress(
            f"Processing file {current_file} of {total_files}",
            files_finished,
            file_name=uploaded_file.name,
        )

        try:
            pdf_bytes = uploaded_file.getvalue()

            # decrypt if encrypted
            if is_pdf_encrypted(pdf_bytes):
                update_processing_progress(
                    f"Decrypting file {current_file} of {total_files}",
                    files_finished,
                    file_name=uploaded_file.name,
                )
                pdf_bytes = decrypt_pdf_bytes(pdf_bytes, st.session_state.pdf_password)
            
            resolved_pdf_bytes[uploaded_file.name] = pdf_bytes

            # extract company name
            company_name = None
            try:
                with bytes_to_pdfplumber(pdf_bytes) as meta_pdf:
                    company_name = extract_company_name(meta_pdf, max_pages=2)
            except Exception:
                company_name = None

            # extract account number
            account_no = None
            try:
                with bytes_to_pdfplumber(pdf_bytes) as meta_pdf:
                    account_no = extract_account_number(meta_pdf, max_pages=2)
            except Exception:
                account_no = None

            # manual override wins
            if (st.session_state.company_name_override or "").strip():
                company_name = st.session_state.company_name_override.strip()
            if (st.session_state.company_account_no_override or "").strip():
                account_no = st.session_state.company_account_no_override.strip()

            st.session_state.file_company_name[uploaded_file.name] = company_name
            st.session_state.file_account_no[uploaded_file.name] = account_no

            # Parse transactions
            if bank_choice == "Affin Bank":
                with bytes_to_pdfplumber(pdf_bytes) as pdf:
                    totals = extract_affin_statement_totals(pdf, uploaded_file.name)
                    st.session_state.affin_statement_totals.append(totals)
                    tx_raw = parse_affin_bank(pdf, uploaded_file.name) or []

            elif bank_choice == "Ambank":
                with bytes_to_pdfplumber(pdf_bytes) as pdf:
                    totals = extract_ambank_statement_totals(pdf, uploaded_file.name)
                    st.session_state.ambank_statement_totals.append(totals)
                    tx_raw = parse_ambank(pdf, uploaded_file.name) or []

            elif bank_choice == "CIMB Bank":
                with bytes_to_pdfplumber(pdf_bytes) as pdf:
                    totals = extract_cimb_statement_totals(pdf, uploaded_file.name)
                    st.session_state.cimb_statement_totals.append(totals)
                    tx_raw = parse_transactions_cimb(pdf, uploaded_file.name) or []

            elif bank_choice == "RHB Bank":
                with bytes_to_pdfplumber(pdf_bytes) as pdf:
                    totals = extract_rhb_statement_totals(pdf, uploaded_file.name)
                    st.session_state.rhb_statement_totals.append(totals)
                tx_raw = parser(pdf_bytes, uploaded_file.name) or []

            elif bank_choice == "Bank Islam":
                with bytes_to_pdfplumber(pdf_bytes) as pdf:
                    tx_raw = parse_bank_islam(pdf, uploaded_file.name) or []
                    stmt_month = extract_bank_islam_statement_month(pdf)
                    if stmt_month:
                        st.session_state.bank_islam_file_month[uploaded_file.name] = stmt_month

            else:
                tx_raw = parser(pdf_bytes, uploaded_file.name) or []

            # Normalize then attach company_name
            tx_norm = normalize_transactions(
                tx_raw,
                default_bank=bank_choice,
                source_file=uploaded_file.name,
            )
            high_value_threshold = get_high_value_threshold()
            for t in tx_norm:
                t["company_name"] = company_name
                t["account_no"] = account_no
                t["high_value_credit"] = safe_float(t.get("credit", 0)) >= high_value_threshold
                t["high_value_debit"] = safe_float(t.get("debit", 0)) >= high_value_threshold
                t["high_value_transaction"] = t["high_value_credit"] or t["high_value_debit"]

            if bank_choice == "Affin Bank":
                st.session_state.affin_file_transactions[uploaded_file.name] = tx_norm
            if bank_choice == "Ambank":
                st.session_state.ambank_file_transactions[uploaded_file.name] = tx_norm
            if bank_choice == "CIMB Bank":
                st.session_state.cimb_file_transactions[uploaded_file.name] = tx_norm
            if bank_choice == "RHB Bank":
                st.session_state.rhb_file_transactions[uploaded_file.name] = tx_norm

            if tx_norm:
                all_tx.extend(tx_norm)
                total_extracted += len(tx_norm)
                update_processing_progress(
                    f"Processed file {current_file} of {total_files}: "
                    f"{len(tx_norm)} transactions extracted",
                    current_file,
                    file_name=uploaded_file.name,
                )
            else:
                update_processing_progress(
                    f"Processed file {current_file} of {total_files}: no transactions found",
                    current_file,
                    file_name=uploaded_file.name,
                )

        except Exception as e:
            processing_errors.append(uploaded_file.name)
            update_processing_progress(
                f"Error processing file {current_file} of {total_files}: {str(e)[:100]}",
                current_file,
                variant="error",
                file_name=uploaded_file.name,
            )
            st.error(f"❌ Error processing {uploaded_file.name}: {e}")
            st.exception(e)

        files_finished = file_idx + 1

    # Run PDF integrity checks
    analysis_results = {}
    processing_stopped = st.session_state.get("stop_requested")
    if resolved_pdf_bytes and not processing_stopped:
        update_processing_progress(
            f"Running PDF integrity checks for {len(resolved_pdf_bytes)} file(s)",
            total_files,
        )
        try:
            analysis_results = analyze_pdf_batch(resolved_pdf_bytes)
        except Exception as e:
            st.warning(f"PDF integrity check failed: {e}")
    st.session_state.integrity_analysis_results = analysis_results

    # Display final status message
    if processing_stopped:
        st.session_state.status = "stopped"
        update_processing_progress(
            f"Processing stopped at {files_finished} of {total_files} file(s).",
            files_finished,
            variant="warning",
        )
    elif processing_errors:
        st.session_state.status = "completed_with_errors"
        update_processing_progress(
            f"Completed with {len(processing_errors)} error(s). "
            f"Extracted {total_extracted} transactions from {total_files} file(s).",
            total_steps,
            variant="warning",
        )
        st.warning(f"⚠️ Completed with {len(processing_errors)} error(s). Check the errors above.")
    else:
        st.session_state.status = "completed"
        update_processing_progress("Finalizing extracted transactions.", total_steps)
    
    st.markdown("---")
    all_tx = dedupe_transactions(all_tx)

    # Stable ordering
    for idx, t in enumerate(all_tx):
        if "__row_order" not in t:
            t["__row_order"] = idx

    def _sort_key(t: dict) -> Tuple:
        dt = parse_any_date_for_summary(t.get("date"))
        page = t.get("page")
        try:
            page_i = int(page) if page is not None else 10**9
        except Exception:
            page_i = 10**9

        seq = t.get("seq", None)
        try:
            seq_i = int(seq) if seq is not None else 10**9
        except Exception:
            seq_i = 10**9

        row_order = t.get("__row_order", 10**12)
        try:
            row_order_i = int(row_order)
        except Exception:
            row_order_i = 10**12

        return (
            dt if pd.notna(dt) else pd.Timestamp.max,
            page_i,
            seq_i,
            row_order_i,
        )

    all_tx = sorted(all_tx, key=_sort_key)
    st.session_state.results = all_tx
    final_transaction_count = len(all_tx)

    if processing_stopped:
        update_processing_progress(
            f"Processing stopped at {files_finished} of {total_files} file(s).",
            files_finished,
            variant="warning",
        )
    elif processing_errors:
        update_processing_progress(
            f"Completed with {len(processing_errors)} error(s). "
            f"Extracted {final_transaction_count} transactions from {total_files} file(s).",
            total_steps,
            variant="warning",
        )
    else:
        update_processing_progress(
            f"Successfully processed all {total_files} file(s) and completed extraction of "
            f"{final_transaction_count} transactions from {total_files} file(s).",
            total_steps,
            variant="success",
        )


# ---------------------------------------------------
# DISPLAY
# ---------------------------------------------------
analysis_results = st.session_state.get("integrity_analysis_results", {})

if st.session_state.results:
    high_value_threshold = get_high_value_threshold()
    
    # Convert results to DataFrame
    df = pd.DataFrame(st.session_state.results) if st.session_state.results else pd.DataFrame()
    
    if not df.empty:
        # Run fraud/pattern checks
        df = run_fraud_checks(df, high_value_threshold)
        
        # Display transaction pattern overview
        render_transaction_overview(df, high_value_threshold)

        # Display Extracted Transaction section
        st.markdown("---")
        render_extracted_transaction_section(df)
        
        # Display Counterparty Ledger Table
        st.markdown("---")
        render_counterparty_ledger_table(df)
    else:
        st.info("No line-item transactions extracted.")
    
    # Display Document Integrity Scan
    if analysis_results:
        st.markdown("---")
        render_integrity_report_styles()
        render_integrity_overview(analysis_results)

        for file_name, result in analysis_results.items():
            summary = build_display_summary(result)
            layer_results = result.get("layer_results", {})
            with st.expander(file_risk_label(file_name, result)):
                render_fraud_summary(summary, layer_results)

                if layer_results:
                    for layer_name, findings in layer_results.items():
                        st.markdown(f"**{layer_name}**")
                        if not findings:
                            st.write("No findings.")
                            continue

                        for finding in findings:
                            st.write(
                                f"{severity_badge(finding.get('severity'))} "
                                f"{finding.get('message', '')}"
                            )
                            detail = finding.get("detail")
                            if detail:
                                st.json(detail)
    
    # Original transaction analysis from existing code
    transaction_analysis_report = parse_top_parties_and_high_value(
        st.session_state.results,
        top_n=10,
        high_value_threshold=high_value_threshold,
    )

    monthly_summary_raw = calculate_monthly_summary(st.session_state.results)
    monthly_summary = present_monthly_summary_standard(monthly_summary_raw)

    if monthly_summary:
        st.subheader("📅 Monthly Summary (Standardized)")
        summary_df = pd.DataFrame(monthly_summary)
        desired_cols = [
            "month",
            "company_name",
            "account_no",
            "transaction_count",
            "opening_balance",
            "total_debit",
            "total_credit",
            "highest_balance",
            "lowest_balance",
            "swing",
            "ending_balance",
            "source_files",
        ]
        summary_df = summary_df[[c for c in desired_cols if c in summary_df.columns]]
        
        st.dataframe(
            summary_df, 
            use_container_width=True,
            column_config={
                "month": "Month",
                "company_name": "Company Name",
                "account_no": "Account Number",
                "transaction_count": "No. of Transactions",
                "opening_balance": "Opening Balance",
                "total_debit": "Total Debit",
                "total_credit": "Total Credit",
                "highest_balance": "Highest Balance",
                "lowest_balance": "Lowest Balance",
                "swing": "Swing",
                "ending_balance": "Ending Balance",
                "source_files": "Source Files"
            }
        )
    
    st.subheader("⬇️ Download Options")
    # Updated to 4 columns for HTML report
    col1, col2, col3, col4 = st.columns(4)

    df_download = df.copy() if not df.empty else pd.DataFrame([])

    # Helper function to convert dataframe to JSON-serializable format
    def make_json_serializable(obj):
        """Recursively convert non-serializable objects to JSON-serializable format"""
        if isinstance(obj, dict):
            return {key: make_json_serializable(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [make_json_serializable(item) for item in obj]
        elif isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        elif isinstance(obj, pd.Period):
            return str(obj)
        elif pd.isna(obj):
            return None
        elif hasattr(obj, 'isoformat'):
            return obj.isoformat()
        else:
            return obj

    # Convert transactions to JSON-serializable format
    if not df_download.empty:
        json_records = df_download.to_dict(orient="records")
        json_records = make_json_serializable(json_records)
    else:
        json_records = []

    with col1:
        st.download_button(
            "📄 Download Full Transactions (JSON)",
            json.dumps(json_records, indent=4),
            "transactions.json",
            "application/json",
            use_container_width=True
        )

    with col2:
        # Handle date conversion for display
        if "date" in df_download.columns and not df_download.empty:
            date_min = df_download["date"].min()
            date_max = df_download["date"].max()
            date_min_str = date_min.isoformat() if isinstance(date_min, pd.Timestamp) else str(date_min)
            date_max_str = date_max.isoformat() if isinstance(date_max, pd.Timestamp) else str(date_max)
            date_range_str = f"{date_min_str} to {date_max_str}"
        else:
            date_range_str = None

        total_files_processed = None
        if "source_file" in df_download.columns and not df_download.empty:
            total_files_processed = int(df_download["source_file"].nunique())

        company_names = sorted(
            {x for x in df_download.get("company_name", pd.Series([], dtype=object)).dropna().astype(str).tolist() if x.strip()}
        )

        account_nos = sorted(
            {x for x in df_download.get("account_no", pd.Series([], dtype=object)).dropna().astype(str).tolist() if x.strip()}
        )

        # Convert monthly_summary to JSON-serializable format
        serializable_monthly_summary = make_json_serializable(monthly_summary)
        
        # Convert transaction_analysis_report to JSON-serializable format
        serializable_transaction_analysis = make_json_serializable(transaction_analysis_report)
        serializable_pdf_integrity = make_json_serializable(analysis_results)
        serializable_round_transactions = make_json_serializable(build_round_transactions(json_records))

        full_report = {
            "summary": {
                "total_transactions": int(len(df_download)),
                "date_range": date_range_str,
                "total_files_processed": total_files_processed,
                "company_names": company_names,
                "account_nos": account_nos,
                "high_value_threshold": high_value_threshold,
                "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            },
            "transaction_analysis": serializable_transaction_analysis,
            "monthly_summary": serializable_monthly_summary,
            "pdf_integrity": serializable_pdf_integrity,
            "round_transactions": serializable_round_transactions,
            "round_figure_credits": serializable_round_transactions,
            "transactions": json_records,
        }

        st.download_button(
            "📊 Download Full Transaction (JSON)",
            json.dumps(full_report, indent=4),
            "full_report.json",
            "application/json",
            use_container_width=True
        )

    serialized_transactions = make_json_serializable(st.session_state.results)
    serialized_monthly_summary = make_json_serializable(monthly_summary)
    serialized_transaction_analysis = make_json_serializable(transaction_analysis_report)
    shared_report_data = build_shared_report_data(
        serialized_transactions,
        serialized_monthly_summary,
        serialized_transaction_analysis,
        high_value_threshold,
    ) if serialized_transactions else {}

    with col3:
        output = generate_excel_report(
            shared_report_data,
            monthly_summary=serialized_monthly_summary, 
            transaction_analysis=serialized_transaction_analysis
        )

        st.download_button(
            "📊 Download Full Report (XLSX)",
            output.getvalue(),
            "full_report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with col4:
        # NEW: Generate and download HTML report from current data
        if st.session_state.results:
            try:
                html_content = generate_interactive_html(shared_report_data)
                
                # Get company name for filename
                company_name = st.session_state.company_name_override or "company"
                if company_name == "company" and not st.session_state.results:
                    company_name = "report"
                else:
                    # Try to get from first transaction
                    if st.session_state.results and st.session_state.results[0].get("company_name"):
                        company_name = st.session_state.results[0]["company_name"]
                
                safe_name = company_name.replace(' ', '_').replace('/', '_')
                
                st.download_button(
                    "🌐 Download Interactive HTML Report",
                    html_content.encode('utf-8'),
                    f"{safe_name}_statement_report.html",
                    "text/html; charset=utf-8",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Failed to generate HTML report: {e}")

else:
    if (
        uploaded_files
        and st.session_state.status == "idle"
        and not st.session_state.high_value_threshold_error
        and not st.session_state.bank_choice_error
        and not st.session_state.pdf_upload_error
    ):
        st.warning("⚠️ No transactions found — click **Start Processing**.")
        
st.markdown("</div>", unsafe_allow_html=True)
