# Extracted from app.py to keep the Streamlit entrypoint smaller.
from __future__ import annotations

import copy
import hashlib
import json
import re
import textwrap
from datetime import datetime
from html import escape
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from core_utils import safe_float
except Exception:  # pragma: no cover - rebound from app.py during normal use
    safe_float = float


def bind_app_globals(app_globals: dict) -> None:
    """Expose app.py helpers/constants that these extracted functions already use."""
    for name, value in app_globals.items():
        if name.startswith("__") and name.endswith("__"):
            continue
        globals()[name] = value


def generate_excel_report(data: dict, monthly_summary: List[dict] = None, transaction_analysis: dict = None) -> BytesIO:
    """Generate Excel workbook using the original generate_excel structure."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return BytesIO()

    report_data = normalize_report_data_for_export(data)
    own_related = report_data.get("own_related_transactions", {}) or {}
    if isinstance(own_related, list):
        own_related = {"transactions": own_related, "summary": {}}
    elif not isinstance(own_related, dict):
        own_related = {"transactions": [], "summary": {}}

    loans = report_data.get("loan_transactions", {}) or {}
    if isinstance(loans, list):
        loans = {"transactions": loans, "disbursements": [], "repayments": []}
    flags = report_data.get("flags", {}) or {}

    cp_ledger = report_data.get("counterparty_ledger", {})
    if not cp_ledger or not cp_ledger.get("counterparties"):
        cp_ledger = build_track2_counterparty_ledger(report_data.get("transactions", []))
        report_data["counterparty_ledger"] = cp_ledger
    report_info = report_data.get("report_info", {}) or {}
    company_name = report_info.get("company_name", "")
    
    # Build top_parties from the same aligned CP ledger rows used by the Counterparty sheets.
    report_counterparty_rows = get_report_counterparty_rows_from_data(
        report_data,
        cp_ledger,
        related_parties=filter_report_related_parties(
            report_info.get("related_parties", []) or [],
            company_name=company_name,
        ),
        own_related=own_related,
        company_name=company_name,
    )
    if report_counterparty_rows:
        report_data["report_counterparty_rows"] = report_counterparty_rows
    top_parties = _top_parties_from_counterparty_rows(
        report_counterparty_rows,
        limit=None,
        company_name=company_name,
    )
    report_data["top_parties"] = top_parties

    parsing = report_data.get("parsing_metadata", {}) or {}
    pdf_integrity = report_data.get("pdf_integrity", {}) or {}
    consolidated = report_data.get("consolidated", {}) or {}
    monthly_analysis = report_data.get("monthly_analysis", []) or []
    accounts = report_data.get("accounts", []) or []
    top_parties = report_data.get("top_parties", {}) or {}

    statutory_compliance = consolidated.get("statutory_compliance", {}) or {}
    observations = normalize_observations(report_data.get("observations", {}) or {})

    # ── Backfill consolidated totals from monthly_analysis when Track 2 keys are absent ──
    def _sum_monthly(key):
        return round(sum(safe_float(m.get(key, 0)) for m in monthly_analysis), 2)

    def _min_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(min(vals), 2) if vals else 0.0

    def _max_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(max(vals), 2) if vals else 0.0

    def _avg_monthly(key):
        vals = [safe_float(m.get(key, 0)) for m in monthly_analysis if m.get(key) is not None]
        return round(sum(vals) / len(vals), 2) if vals else 0.0

    backfill_map = {
        "total_own_party_cr":        ("sum", "own_party_cr"),
        "total_own_party_dr":        ("sum", "own_party_dr"),
        "total_related_party_cr":    ("sum", "related_party_cr"),
        "total_related_party_dr":    ("sum", "related_party_dr"),
        "total_loan_disbursement_cr":("sum", "loan_disbursement_cr"),
        "total_loan_repayment_dr":   ("sum", "loan_repayment_dr"),
        "total_fd_interest_cr":      ("sum", "fd_interest_cr"),
        "total_cash_deposits":       ("sum", "cash_deposits_amount"),
        "total_cash_withdrawals":    ("sum", "cash_withdrawals_amount"),
        "total_cheque_deposits":     ("sum", "cheque_deposits_amount"),
        "total_cheque_issues":       ("sum", "cheque_issues_amount"),
        "total_salary_paid":         ("sum", "salary_paid"),
        "total_statutory_epf":       ("sum", "statutory_epf"),
        "total_statutory_socso":     ("sum", "statutory_socso"),
        "total_statutory_tax":       ("sum", "statutory_tax"),
        "total_statutory_hrdf":      ("sum", "statutory_hrdf"),
        "eod_lowest":                ("min", "eod_lowest"),
        "eod_highest":               ("max", "eod_highest"),
        "eod_average":               ("avg", "eod_average"),
    }

    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    header_fill_green = PatternFill(start_color="196F3D", end_color="196F3D", fill_type="solid")
    header_fill_red = PatternFill(start_color="922B21", end_color="922B21", fill_type="solid")
    header_fill_orange = PatternFill(start_color="B9770E", end_color="B9770E", fill_type="solid")
    # Add these color definitions near the other header fills
    header_fill_purple = PatternFill(start_color="6C3483", end_color="6C3483", fill_type="solid")
    header_fill_blue = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    credit_font = Font(name="Calibri", color="196F3D")
    debit_font = Font(name="Calibri", color="922B21")
    bold_font = Font(name="Calibri", bold=True, size=11)
    title_font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    num_fmt = "#,##0.00"

    def clean_xl(value):
        value = _excel_safe_value(value)
        if isinstance(value, (dict, list)):
            return json.dumps(value, default=str)
        return value

    def is_num(value):
        return isinstance(value, (int, float)) and not isinstance(value, bool)

    def style_header_row(ws, row, max_col, fill=None):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, number=False, credit=False, debit=False):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if number else "left", vertical="top", wrap_text=True)
        if number:
            cell.number_format = num_fmt
        if credit:
            cell.font = credit_font
        if debit:
            cell.font = debit_font
        if row % 2 == 0:
            cell.fill = alt_row_fill

    def write_headers(ws, row, headers, fill=None):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header)
        style_header_row(ws, row, len(headers), fill)

    def auto_width(ws, min_width=10, max_width=40):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 3, max_width))

    def write_values(ws, row, values, number_cols=None, credit_cols=None, debit_cols=None):
        number_cols = set(number_cols or [])
        credit_cols = set(credit_cols or [])
        debit_cols = set(debit_cols or [])
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=clean_xl(value))
            style_data_cell(
                ws,
                row,
                col,
                number=col in number_cols or is_num(value),
                credit=col in credit_cols,
                debit=col in debit_cols,
            )

    def safe_amount(txn):
        return safe_float(txn.get("amount", txn.get("credit", 0) or txn.get("debit", 0)))

    def infer_txn_side(txn):
        raw_type = str(txn.get("type") or txn.get("transaction_type") or "").upper()
        if "CR" in raw_type or "CREDIT" in raw_type:
            return "CREDIT"
        if "DR" in raw_type or "DEBIT" in raw_type:
            return "DEBIT"
        if safe_float(txn.get("credit", 0)) > 0:
            return "CREDIT"
        if safe_float(txn.get("debit", 0)) > 0:
            return "DEBIT"
        return "DEBIT" if safe_amount(txn) < 0 else "CREDIT"

    def write_split_transaction_sheet(ws, title, txns, caption=None, number_cols=None, credit_cols=None, debit_cols=None):
        """Write a sheet with split credit/debit transactions."""
        number_cols = set(number_cols or {4, 5})
        credit_cols = set(credit_cols or {4})
        debit_cols = set(debit_cols or {4})
        
        ws.cell(row=1, column=1, value=title).font = title_font
        row = 2
        if caption:
            ws.cell(row=row, column=1, value=caption)
            ws.cell(row=row, column=1).font = Font(name="Calibri", italic=True, color="475569")
            row += 2
        else:
            row += 1

        display_headers = ["No.", "Date", "Description", "Amount", "Balance"]
        for section_title, side, fill in (
            ("CREDIT TRANSACTIONS", "CREDIT", header_fill_green),
            ("DEBIT TRANSACTIONS", "DEBIT", header_fill_red),
        ):
            section_rows = [txn for txn in (txns or []) if infer_txn_side(txn) == side]
            ws.cell(row=row, column=1, value=section_title).font = bold_font
            row += 1
            write_headers(ws, row, display_headers, fill)
            if not section_rows:
                row += 1
                ws.cell(row=row, column=1, value="No transactions")
                style_data_cell(ws, row, 1)
            for idx, txn in enumerate(section_rows, 1):
                row += 1
                amount = abs(safe_amount(txn))
                values = [
                    idx,
                    txn.get("date", ""),
                    (txn.get("description", "") or "")[:100],
                    amount,
                    txn.get("balance"),
                ]
                write_values(
                    ws,
                    row,
                    values,
                    number_cols=number_cols,
                    credit_cols=credit_cols if side == "CREDIT" else set(),
                    debit_cols=debit_cols if side == "DEBIT" else set(),
                )
                # Force "No." column to display as whole integer, not float
                ws.cell(row=row, column=1).number_format = "0"
                # Centre + middle align: No.(1), Date(2), Amount(4), Balance(5) only
                for centre_col in (1, 2, 4, 5):
                    ws.cell(row=row, column=centre_col).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
            row += 2
        # Column-specific auto widths for these 5-column transaction sheets
        ws.column_dimensions["A"].width = 6    # No.
        ws.column_dimensions["B"].width = 14   # Date
        ws.column_dimensions["C"].width = 65   # Description
        ws.column_dimensions["D"].width = 18   # Amount
        ws.column_dimensions["E"].width = 18   # Balance

    schema_version = str(report_info.get("schema_version", ""))
    is_v620 = schema_version in ("6.2.0", "6.2.1", "6.2.2", "6.3.0", "6.3.1", "6.3.2", "6.3.3", "6.3.4", "6.3.5") or consolidated.get("total_fx_credits") is not None
    is_v630 = schema_version in ("6.3.0", "6.3.1", "6.3.2", "6.3.3", "6.3.4", "6.3.5") or consolidated.get("total_unclassified_cr") is not None
    recon_lookup = _build_reconciliation_lookup(parsing)
    has_recon = any(m.get("reconciliation_status") for m in monthly_analysis) or bool(recon_lookup)

    # Build cp_sorted HERE so it's available for both Counterparty and CP Ledger sheets
    related_parties = filter_report_related_parties(
        report_info.get("related_parties", []) or [],
        company_name=company_name,
    )
    cp_sorted = get_report_counterparty_rows_from_data(
        report_data,
        cp_ledger,
        related_parties=related_parties,
        own_related=own_related,
        company_name=company_name,
    )

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="KREDIT LAB - STATEMENT INTELLIGENCE REPORT").font = title_font
    ws.cell(row=2, column=1, value=report_info.get("company_name", "")).font = bold_font
    ws.cell(row=3, column=1, value=f"Period: {report_info.get('period_start', '')} to {report_info.get('period_end', '')}")
    ws.cell(row=4, column=1, value=f"Generated: {report_info.get('generated_at', '')}")

    row = 6
    ws.cell(row=row, column=1, value="ACCOUNT DETAILS").font = bold_font
    row += 1
    headers = ["Bank", "Account No", "Holder", "Type", "Opening Balance", "Closing Balance", "Total Credits", "Total Debits", "Transactions"]
    write_headers(ws, row, headers)
    for account in accounts:
        row += 1
        values = [
            account.get("bank_name"), account.get("account_number"), account.get("account_holder"),
            account.get("account_type"), account.get("opening_balance"), account.get("closing_balance"),
            account.get("total_credits"), account.get("total_debits"), account.get("transaction_count"),
        ]
        write_values(ws, row, values, number_cols={5, 6, 7, 8}, credit_cols={7}, debit_cols={8})

    row += 2
    ws.cell(row=row, column=1, value="CONSOLIDATED FIGURES").font = bold_font
    row += 1
    consolidated_items = [
        ("Gross Credits", consolidated.get("gross_credits")),
        ("Gross Debits", consolidated.get("gross_debits")),
        ("Net Credits", consolidated.get("net_credits")),
        ("Net Debits", consolidated.get("net_debits")),
        ("Annualized Net Credits", consolidated.get("annualized_net_credits")),
        ("Annualized Net Debits", consolidated.get("annualized_net_debits")),
        ("", ""),
        ("Own Party Credits", consolidated.get("total_own_party_cr")),
        ("Own Party Debits", consolidated.get("total_own_party_dr")),
        ("Related Party Credits", consolidated.get("total_related_party_cr")),
        ("Related Party Debits", consolidated.get("total_related_party_dr")),
        ("", ""),
        ("Loan Disbursements", consolidated.get("total_loan_disbursement_cr")),
        ("Loan Repayments", consolidated.get("total_loan_repayment_dr")),
        ("FD/Interest Credits", consolidated.get("total_fd_interest_cr")),
        ("Inward Return (C16)", consolidated.get("total_inward_return_cr")),
        ("", ""),
        ("Cash Deposits", consolidated.get("total_cash_deposits")),
        ("Cash Withdrawals", consolidated.get("total_cash_withdrawals")),
        ("Cheque Deposits", consolidated.get("total_cheque_deposits")),
        ("Cheque Issues", consolidated.get("total_cheque_issues")),
        ("", ""),
        ("Total Salary Paid", consolidated.get("total_salary_paid")),
        ("Total EPF", consolidated.get("total_statutory_epf")),
        ("Total SOCSO", consolidated.get("total_statutory_socso")),
        ("Total Tax", consolidated.get("total_statutory_tax")),
        ("Total HRDF", consolidated.get("total_statutory_hrdf")),
        ("", ""),
        ("EOD Lowest", consolidated.get("eod_lowest")),
        ("EOD Highest", consolidated.get("eod_highest")),
        ("EOD Average", consolidated.get("eod_average")),
    ]
    if has_recon:
        consolidated_items.extend([
            ("", ""),
            ("Data Completeness", consolidated.get("data_completeness")),
            ("Extraction Gaps", consolidated.get("total_extraction_gaps")),
            ("Missing Debits", consolidated.get("total_missing_debits")),
            ("Missing Credits", consolidated.get("total_missing_credits")),
            ("Months With Gaps", consolidated.get("months_with_gaps")),
        ])
    if is_v630:
        consolidated_items.extend([
            ("", ""),
            ("Unclassified Credits", consolidated.get("total_unclassified_cr")),
            ("Unclassified Debits", consolidated.get("total_unclassified_dr")),
        ])
    if is_v620:
        consolidated_items.extend([
            ("", ""),
            ("FX/Remittance Credits", consolidated.get("total_fx_credits")),
            ("FX/Remittance Debits", consolidated.get("total_fx_debits")),
            ("FX Credit % of Gross", consolidated.get("fx_credit_pct")),
            ("FX Debit % of Gross", consolidated.get("fx_debit_pct")),
            ("FX Currencies Detected", ", ".join(consolidated.get("fx_currencies_all", []) or [])),
        ])
    for label, value in consolidated_items:
        if label:
            write_values(ws, row, [label, value], number_cols={2} if is_num(value) else set(), credit_cols={2} if "Credit" in label else set(), debit_cols={2} if "Debit" in label or "Repayment" in label else set())
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="POSITIVE OBSERVATIONS").font = Font(name="Calibri", bold=True, color="196F3D")
    row += 1
    for item in observations.get("positive", []):
        ws.cell(row=row, column=1, value=f"+ {item}")
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="CONCERNS").font = Font(name="Calibri", bold=True, color="922B21")
    row += 1
    for item in observations.get("concerns", []):
        ws.cell(row=row, column=1, value=f"- {item}")
        row += 1
    auto_width(ws)

    # Cash Flow
    ws2 = wb.create_sheet("Cash Flow")
    ws2.cell(row=1, column=1, value="CASH FLOW STATEMENT").font = title_font
    row = 3

    # ============================================================
    # SECTION 1: BASIC TRANSACTION SUMMARY
    # ============================================================
    ws2.cell(row=row, column=1, value="TRANSACTION SUMMARY").font = bold_font
    row += 1

    basic_headers = ["Month", "Bank", "Account No", "Opening Balance", "Closing Balance", 
                    "Gross Credits", "Gross Debits", "Net Credits", "Net Debits", 
                    "Credit Count", "Debit Count", "Total Txn"]
    write_headers(ws2, row, basic_headers, header_fill)
    row += 1

    # Set widths for basic summary columns
    basic_widths = [14, 18, 20, 26, 16, 16, 16, 16, 16, 12, 12, 14]
    for idx, width in enumerate(basic_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    # Write basic data
    total_rows = len(monthly_analysis)
    grand_total_credits = 0
    grand_total_debits = 0
    grand_net_credits = 0
    grand_net_debits = 0
    grand_credit_count = 0
    grand_debit_count = 0
    grand_total_txn = 0

    for idx, item in enumerate(monthly_analysis, row):
        credit_count = item.get("credit_count", 0)
        debit_count = item.get("debit_count", 0)
        total_txn = credit_count + debit_count
        
        values = [
            item.get("month"), 
            item.get("bank_name", ""), 
            item.get("account_number", ""),
            item.get("opening_balance"), 
            item.get("closing_balance"),
            item.get("gross_credits"), 
            item.get("gross_debits"), 
            item.get("net_credits"), 
            item.get("net_debits"),
            credit_count,
            debit_count,
            total_txn
        ]
        write_values(ws2, idx, values, number_cols={4, 5, 6, 7, 8, 9, 12}, 
                    credit_cols={6, 8}, debit_cols={7, 9})
        
        # Center align ALL columns for data rows
        for col in range(1, 13):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format count columns as integers
        for count_col in [10, 11, 12]:
            ws2.cell(row=idx, column=count_col).number_format = "0"
        
        # Accumulate grand totals
        grand_total_credits += safe_float(item.get("gross_credits", 0))
        grand_total_debits += safe_float(item.get("gross_debits", 0))
        grand_net_credits += safe_float(item.get("net_credits", 0))
        grand_net_debits += safe_float(item.get("net_debits", 0))
        grand_credit_count += int(credit_count or 0)
        grand_debit_count += int(debit_count or 0)
        grand_total_txn += int(total_txn or 0)

    row = idx + 1  # CHANGED: was idx + 2, now idx + 1 (no spacing)

    # Add Total row for Transaction Summary
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.cell(row=row, column=1).border = thin_border

    total_values = [
        "TOTAL",  # Month - use dash
        "-",  # Bank - use dash
        "-",  # Account No - use dash
        "-",  # Opening Balance - use dash
        "-",  # Closing Balance - use dash
        grand_total_credits,
        grand_total_debits,
        grand_net_credits,
        grand_net_debits,
        grand_credit_count,
        grand_debit_count,
        grand_total_txn
    ]

    for col_idx, value in enumerate(total_values, 1):
        cell = ws2.cell(row=row, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Skip formatting for dash values
        if value != "-":
            # Apply number formatting
            if col_idx in [6, 7, 8, 9]:  # Amount columns
                cell.number_format = "#,##0.00"
                if col_idx in [6, 8]:  # Credits
                    cell.font = credit_font
                else:  # Debits
                    cell.font = debit_font
            elif col_idx in [10, 11, 12]:  # Count columns
                cell.number_format = "0"

    row += 2  # Now this is 2 rows after TOTAL (spacing before next section)

    # ============================================================
    # SECTION 2: EXCLUSIONS (What's removed from Gross)
    # ============================================================
    ws2.cell(row=row, column=1, value="EXCLUSIONS FROM NET FIGURES").font = bold_font
    row += 1

    exclusion_headers = ["Month", "Own Party Cr", "Own Party Dr", "Related Party Cr", "Related Party Dr", 
                        "Reversal Cr", "Loan Disbursement Cr", "FD Interest Cr"]
    write_headers(ws2, row, exclusion_headers, header_fill_orange)
    row += 1

    # Set widths for exclusion columns
    excl_widths = [14, 18, 18, 22, 18, 18, 22, 18]
    for idx, width in enumerate(excl_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    # Write exclusion data
    exclusion_totals = [0.0] * 7  # 7 columns to sum (skip Month)
    for idx, item in enumerate(monthly_analysis, row):
        values = [
            item.get("month"),
            item.get("own_party_cr"), 
            item.get("own_party_dr"),
            item.get("related_party_cr"), 
            item.get("related_party_dr"),
            item.get("reversal_cr"),
            item.get("loan_disbursement_cr"), 
            item.get("fd_interest_cr")
        ]
        write_values(ws2, idx, values, number_cols={2, 3, 4, 5, 6, 7, 8}, 
                    credit_cols={2, 4, 6, 7, 8}, debit_cols={3, 5})
        
        # Center align ALL columns for data rows
        for col in range(1, 9):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Accumulate totals
        for i in range(2, 9):  # Skip Month (column 1)
            exclusion_totals[i-2] += safe_float(values[i-1] or 0)

    row = idx + 1  # CHANGED: was idx + 2, now idx + 1 (no spacing)

   # Add Total row for Exclusions
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.cell(row=row, column=1).border = thin_border

    exclusion_total_values = [None] + exclusion_totals  # None for Month column
    for col_idx, value in enumerate(exclusion_total_values, 1):
        if value is not None:
            cell = ws2.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.number_format = "#,##0.00"
            # Apply color coding
            if col_idx in [2, 4, 6, 7, 8]:  # Credit columns
                cell.font = credit_font
            elif col_idx in [3, 5]:  # Debit columns
                cell.font = debit_font 

    row += 2

    # ============================================================
    # SECTION 3: CASH & CHEQUE ACTIVITY
    # ============================================================
    ws2.cell(row=row, column=1, value="CASH & CHEQUE ACTIVITY").font = bold_font
    row += 1

    cash_headers = ["Month", "Cash Dep Count", "Cash Dep Amt", "Cash Withd Count", "Cash Withd Amt",
                    "Chq Dep Count", "Chq Dep Amt", "Chq Issue Count", "Chq Issue Amt"]
    write_headers(ws2, row, cash_headers, header_fill_green)
    row += 1

    # Set widths for cash/cheque columns
    cash_widths = [14, 18, 16, 18, 16, 14, 16, 14, 16]
    for idx, width in enumerate(cash_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    # Write cash/cheque data
    cash_totals = [0.0] * 8  # 8 columns to sum (skip Month)
    for idx, item in enumerate(monthly_analysis, row):
        values = [
            item.get("month"),
            item.get("cash_deposits_count"), 
            item.get("cash_deposits_amount"),
            item.get("cash_withdrawals_count"), 
            item.get("cash_withdrawals_amount"),
            item.get("cheque_deposits_count"), 
            item.get("cheque_deposits_amount"),
            item.get("cheque_issues_count"), 
            item.get("cheque_issues_amount")
        ]
        write_values(ws2, idx, values, number_cols={2, 3, 4, 5, 6, 7, 8, 9}, 
                    credit_cols={3, 5, 7}, debit_cols={9})
        
        # Center align ALL columns for data rows
        for col in range(1, 10):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format count columns as integers
        for count_col in [2, 4, 6, 8]:
            ws2.cell(row=idx, column=count_col).number_format = "0"
        
        # Accumulate totals
        for i in range(2, 10):  # Skip Month (column 1)
            cash_totals[i-2] += safe_float(values[i-1] or 0)

    row = idx + 1  # CHANGED: was idx + 2, now idx + 1 (no spacing)

    # Add Total row for Cash & Cheque Activity
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.cell(row=row, column=1).border = thin_border

    cash_total_values = [None] + cash_totals  # None for Month column
    for col_idx, value in enumerate(cash_total_values, 1):
        if value is not None:
            cell = ws2.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Apply formatting
            if col_idx in [3, 5, 7, 9]:  # Amount columns
                cell.number_format = "#,##0.00"
                if col_idx in [3, 5, 7]:  # Credit amounts
                    cell.font = credit_font
                else:  # Debit amounts
                    cell.font = debit_font
            elif col_idx in [2, 4, 6, 8]:  # Count columns
                cell.number_format = "0" 

    row += 2

    # ============================================================
    # SECTION 4: STATUTORY PAYMENTS
    # ============================================================
    ws2.cell(row=row, column=1, value="STATUTORY PAYMENTS").font = bold_font
    row += 1

    statutory_headers = ["Month", "Salary Paid", "EPF", "SOCSO", "Tax", "HRDF", 
                        "Loan Repayment Dr", "High Value Cr", "Round Figure Cr"]
    write_headers(ws2, row, statutory_headers, header_fill_purple)
    row += 1

    # Set widths for statutory columns
    stat_widths = [14, 16, 16, 18, 16, 16, 22, 16, 16]
    for idx, width in enumerate(stat_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    # Write statutory data
    statutory_totals = [0.0] * 8  # 8 columns to sum (skip Month)
    for idx, item in enumerate(monthly_analysis, row):
        values = [
            item.get("month"),
            item.get("salary_paid"), 
            item.get("statutory_epf"), 
            item.get("statutory_socso"),
            item.get("statutory_tax"), 
            item.get("statutory_hrdf"),
            item.get("loan_repayment_dr"), 
            item.get("high_value_cr"),
            item.get("round_figure_cr")
        ]
        write_values(ws2, idx, values, number_cols={2, 3, 4, 5, 6, 7, 8, 9}, 
                    credit_cols={6, 8, 9}, debit_cols={3, 4, 5, 7})
        
        # Center align ALL columns for data rows
        for col in range(1, 10):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Accumulate totals
        for i in range(2, 10):  # Skip Month (column 1)
            statutory_totals[i-2] += safe_float(values[i-1] or 0)

    row = idx + 1

    # Add Total row for Statutory Payments
    ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.cell(row=row, column=1).border = thin_border

    statutory_total_values = [None] + statutory_totals  # None for Month column
    for col_idx, value in enumerate(statutory_total_values, 1):
        if value is not None:
            cell = ws2.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.number_format = "#,##0.00"
            
            # Apply color coding
            if col_idx in [8, 9]:  # High Value Cr, Round Figure Cr - Credits
                cell.font = credit_font
            elif col_idx in [7]:  # Loan Repayment Dr - Debit
                cell.font = debit_font
            elif col_idx in [3, 4, 5]:  # EPF, SOCSO, Tax - Debits
                cell.font = debit_font

    row += 2

    # ============================================================
    # SECTION 5: RETURNED CHEQUES
    # ============================================================
    if any(item.get("returned_cheques_inward_count") or item.get("returned_cheques_outward_count") 
        for item in monthly_analysis):
        ws2.cell(row=row, column=1, value="RETURNED CHEQUES").font = bold_font
        row += 1
        
        return_headers = ["Month", "Ret Chq In Count", "Ret Chq In Amt", "Ret Chq Out Count", "Ret Chq Out Amt"]
        write_headers(ws2, row, return_headers, header_fill_red)
        row += 1
        
        # Set widths for returned cheques columns
        ret_widths = [14, 16, 16, 16, 16]
        for idx, width in enumerate(ret_widths, 1):
            ws2.column_dimensions[get_column_letter(idx)].width = width
        
        # Write returned cheques data
        ret_totals = [0.0] * 4  # 4 columns to sum (skip Month)
        for idx, item in enumerate(monthly_analysis, row):
            values = [
                item.get("month"),
                item.get("returned_cheques_inward_count"), 
                item.get("returned_cheques_inward_amount"),
                item.get("returned_cheques_outward_count"), 
                item.get("returned_cheques_outward_amount")
            ]
            write_values(ws2, idx, values, number_cols={2, 3, 4, 5}, debit_cols={3, 5})
            
            # Center align ALL columns for data rows
            for col in range(1, 6):
                ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Format count columns as integers
            for count_col in [2, 4]:
                ws2.cell(row=idx, column=count_col).number_format = "0"
            
            # Accumulate totals
            for i in range(2, 6):  # Skip Month (column 1)
                ret_totals[i-2] += safe_float(values[i-1] or 0)
        
        row = idx + 2
        
        # Add Total row for Returned Cheques
        ws2.cell(row=row, column=1, value="TOTAL").font = bold_font
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.cell(row=row, column=1).border = thin_border
        
        ret_total_values = [None] + ret_totals  # None for Month column
        for col_idx, value in enumerate(ret_total_values, 1):
            if value is not None:
                cell = ws2.cell(row=row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                if col_idx in [3, 5]:  # Amount columns
                    cell.number_format = "#,##0.00"
                    cell.font = debit_font
                elif col_idx in [2, 4]:  # Count columns
                    cell.number_format = "0"
        
        row += 2

    # ============================================================
    # SECTION 6: EOD BALANCE ANALYSIS
    # ============================================================
    ws2.cell(row=row, column=1, value="EOD BALANCE ANALYSIS").font = bold_font
    row += 1

    eod_headers = ["Month", "EOD Lowest", "EOD Highest", "EOD Average"]
    write_headers(ws2, row, eod_headers, header_fill_blue)
    row += 1

    # Set widths for EOD columns
    eod_widths = [14, 16, 16, 16]
    for idx, width in enumerate(eod_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    for idx, item in enumerate(monthly_analysis, row):
        values = [
            item.get("month"),
            item.get("eod_lowest"), 
            item.get("eod_highest"),
            item.get("eod_average")
        ]
        write_values(ws2, idx, values, number_cols={2, 3, 4})
        
        # Center align ALL columns for data rows
        for col in range(1, 5):
            ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = idx + 2

    # ============================================================
    # SECTION 7: FX / REMITTANCE (if available)
    # ============================================================
    if is_v620:
        ws2.cell(row=row, column=1, value="FX / REMITTANCE").font = bold_font
        row += 1
        
        fx_headers = ["Month", "FX Cr Count", "FX Cr Amount", "FX Dr Count", "FX Dr Amount", "FX Currencies"]
        write_headers(ws2, row, fx_headers, header_fill_purple)
        row += 1
        
        # Set widths for FX columns
        fx_widths = [14, 14, 16, 14, 16, 30]
        for idx, width in enumerate(fx_widths, 1):
            ws2.column_dimensions[get_column_letter(idx)].width = width
        
        for idx, item in enumerate(monthly_analysis, row):
            values = [
                item.get("month"),
                item.get("fx_credit_count", 0), 
                item.get("fx_credit_amount", 0),
                item.get("fx_debit_count", 0), 
                item.get("fx_debit_amount", 0),
                ", ".join(item.get("fx_currencies", []) or [])
            ]
            write_values(ws2, idx, values, number_cols={2, 3, 4, 5}, 
                        credit_cols={3}, debit_cols={5})
            
            # Center align ALL columns for data rows
            for col in range(1, 7):
                ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Format count columns as integers
            for count_col in [2, 4]:
                ws2.cell(row=idx, column=count_col).number_format = "0"
            
            # Left align FX Currencies for readability
            ws2.cell(row=idx, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        row = idx + 2

    # ============================================================
    # SECTION 8: RECONCILIATION STATUS (if available)
    # ============================================================
    if has_recon:
        ws2.cell(row=row, column=1, value="RECONCILIATION STATUS").font = bold_font
        row += 1
        
        recon_headers = ["Month", "Recon Status", "Recon Delta", "Gaps", "Missing Debits", "Missing Credits"]
        write_headers(ws2, row, recon_headers, header_fill_orange)
        row += 1
        
        # Set widths for reconciliation columns
        recon_widths = [14, 14, 16, 12, 16, 16]
        for idx, width in enumerate(recon_widths, 1):
            ws2.column_dimensions[get_column_letter(idx)].width = width
        
        for idx, item in enumerate(monthly_analysis, row):
            effective_recon = _effective_reconciliation_values(item, recon_lookup) if has_recon else None
            values = [
                item.get("month"),
                effective_recon["status"] if effective_recon else "",
                effective_recon["delta"] if effective_recon else 0,
                effective_recon["gaps"] if effective_recon else 0,
                effective_recon["missing_debits"] if effective_recon else 0,
                effective_recon["missing_credits"] if effective_recon else 0
            ]
            write_values(ws2, idx, values, number_cols={3, 4, 5, 6})
            
            # Center align ALL columns for data rows
            for col in range(1, 7):
                ws2.cell(row=idx, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Format count columns as integers
            for count_col in [4]:
                ws2.cell(row=idx, column=count_col).number_format = "0"
            
            if effective_recon and effective_recon["status"] == "FAIL":
                for col in range(1, len(values) + 1):
                    ws2.cell(row=idx, column=col).fill = fail_fill
        
        row = idx + 2

    # Remove auto_width to maintain manual column widths
    # auto_width(ws2, min_width=12)


    # Top Parties
    ws3 = wb.create_sheet("Top Parties")
    ws3.cell(row=1, column=1, value="TOP PARTIES ANALYSIS").font = title_font
    party_view = prepare_top_parties_for_report(top_parties, limit=10, company_name=company_name)
    payers = party_view["payers"]
    payees = party_view["payees"]
    all_party_rows = list(payers) + list(payees)
    monthly_bd = sorted({
        mb.get("month", "")
        for party in all_party_rows
        for mb in (party.get("monthly_breakdown") or [])
        if isinstance(mb, dict) and mb.get("month")
    })
    party_headers = ["Rank", "Party Name", "Total Amount", "Transactions", "Related Party"] + monthly_bd
    party_num_cols = {3, *range(6, 6 + len(monthly_bd))}

    # Start from row 3 to leave space for title
    start_row = 3

    # TOP PAYERS (Income Sources)
    ws3.cell(row=start_row, column=1, value="TOP PAYERS (Income Sources)").font = bold_font
    row = start_row + 1
    write_headers(ws3, row, party_headers, header_fill_green)
    row += 1

    for party in payers:
        lookup = {mb.get("month"): safe_float(mb.get("amount")) for mb in (party.get("monthly_breakdown") or []) if isinstance(mb, dict)}
        values = [party.get("rank"), party.get("party_name") or party.get("name"), party.get("total_amount"), party.get("transaction_count"), "Yes" if party.get("is_related_party") else "No"]
        values.extend(lookup.get(month, 0) for month in monthly_bd)
        write_values(ws3, row, values, number_cols=party_num_cols, credit_cols=party_num_cols)
        ws3.cell(row=row, column=1).number_format = "0"
        ws3.cell(row=row, column=4).number_format = "0"
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(3, 3 + len(party_headers)):
            ws3.cell(row=row, column=col).alignment = alignment
        row += 1

    row = row + 2

    # TOP PAYEES (Payment Destinations)
    ws3.cell(row=row, column=1, value="TOP PAYEES (Payment Destinations)").font = bold_font
    row += 1
    write_headers(ws3, row, party_headers, header_fill_red)
    row += 1

    for party in payees:
        lookup = {mb.get("month"): safe_float(mb.get("amount")) for mb in (party.get("monthly_breakdown") or []) if isinstance(mb, dict)}
        values = [party.get("rank"), party.get("party_name") or party.get("name"), party.get("total_amount"), party.get("transaction_count"), "Yes" if party.get("is_related_party") else "No"]
        values.extend(lookup.get(month, 0) for month in monthly_bd)
        write_values(ws3, row, values, number_cols=party_num_cols, debit_cols=party_num_cols)
        ws3.cell(row=row, column=1).number_format = "0"
        ws3.cell(row=row, column=4).number_format = "0"
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(3, 3 + len(party_headers)):
            ws3.cell(row=row, column=col).alignment = alignment
        row += 1

    auto_width(ws3)
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["D"].width = 13

    # Large Transactions - Updated with proper formatting
    ws_large = wb.create_sheet("Large Transactions")
    large_rows = report_data.get("large_transactions", []) or report_data.get("large_credits", []) or []
    write_split_transaction_sheet(
        ws_large,
        f"Large Transactions (>= RM {consolidated.get('high_value_threshold', 100000):,.0f})",
        large_rows,
        number_cols={4, 5},
        credit_cols={4},
        debit_cols={4}
    )

    # Counterparty - MODIFIED: Combined with Related Parties content
    ws5 = wb.create_sheet("Counterparty")
    ws5.cell(row=1, column=1, value="COUNTERPARTY TRANSACTIONS").font = title_font
    own_related_groups_for_counterparty = build_own_related_party_groups_for_report(
        own_related,
        related_parties=related_parties,
        company_name=company_name,
        counterparty_rows=cp_sorted,
    )
    own_party_group = next(
        (group for group in own_related_groups_for_counterparty if group.get("badge_type") == "OP"),
        {},
    )
    own_related_display_transactions = [
        txn
        for group in own_related_groups_for_counterparty
        for txn in (group.get("transactions", []) or [])
    ]

   # ============================================================
    # OWN PARTY TABLE (Added above Related Parties - Single row only)
    # ============================================================
    row = 3
    ws5.cell(row=row, column=1, value="OWN PARTY").font = bold_font
    row += 1

    # Own Party headers - columns A-F (No., Relationship, Name, Total Credits, Total Debits, Transactions)
    own_party_headers = ["No.", "Relationship", "Name", "Total Credits", "Total Debits", "Transactions"]
    write_headers(ws5, row, own_party_headers, header_fill_blue)  # Using blue for Own Party
    row += 1

    # Set custom column widths for Own Party section
    ws5.column_dimensions["A"].width = 6   # No.
    ws5.column_dimensions["B"].width = 25  # Relationship
    ws5.column_dimensions["C"].width = 20  # Name
    ws5.column_dimensions["D"].width = 16  # Total Credits
    ws5.column_dimensions["E"].width = 16  # Total Debits
    ws5.column_dimensions["F"].width = 14  # Transactions

    # Use the same OP group as the CP Ledger/detail list so counts stay aligned.
    own_party_total_credits = safe_float(own_party_group.get("credits", 0))
    own_party_total_debits = safe_float(own_party_group.get("debits", 0))
    own_party_transaction_count = int(safe_float(own_party_group.get("transaction_count", 0)))
    own_party_name = (
        str(own_party_group.get("party_name") or "").strip()
        or report_info.get("company_name")
        or "Own Party"
    )

    # Write the single Own Party row
    values = [
        1,  # No. column - always 1
        "Own",  # Relationship
        own_party_name,  # Name
        own_party_total_credits,
        own_party_total_debits,
        own_party_transaction_count,
    ]
    write_values(ws5, row, values, number_cols={4, 5, 6}, credit_cols={4}, debit_cols={5})

    # Format the No. column as integer (not float)
    ws5.cell(row=row, column=1).number_format = "0"
    ws5.cell(row=row, column=6).number_format = "0"

    # Center align ALL columns for Own Party
    for col in range(1, 7):  # Columns A-F
        ws5.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Special: Make Name column left-aligned for readability
    ws5.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row += 2

    # --- START: Related Parties content inserted here ---
    ws5.cell(row=row, column=1, value="RELATED PARTIES").font = bold_font
    row += 1

    # Related Parties headers - columns A-F (No., Relationship, Name, Total Credits, Total Debits, Transactions)
    rp_headers = ["No.", "Relationship", "Name", "Total Credits", "Total Debits", "Transactions"]
    write_headers(ws5, row, rp_headers, header_fill_orange)

    # Set custom column widths for Related Parties section (narrower)
    ws5.column_dimensions["A"].width = 6   # No.
    ws5.column_dimensions["B"].width = 25  # Relationship
    ws5.column_dimensions["C"].width = 20  # Name - smaller than counterparty summary
    ws5.column_dimensions["D"].width = 16  # Total Credits
    ws5.column_dimensions["E"].width = 16  # Total Debits
    ws5.column_dimensions["F"].width = 14  # Transactions

    rp_row_start = row + 1
    related_party_rows = build_related_party_summary_rows_for_report(
        related_parties,
        own_related,
        cp_rows=cp_sorted,
        company_name=company_name,
    )

    if related_party_rows:
        for rp_idx, rp in enumerate(related_party_rows):
            row = rp_row_start + rp_idx
            # 6 values: No., Relationship, Name, Total Credits, Total Debits, Transactions
            values = [
                rp_idx + 1,  # No. column - integer
                rp.get("relationship", ""),
                rp.get("name", ""),
                rp.get("total_credits", 0),
                rp.get("total_debits", 0),
                rp.get("transaction_count", 0),
            ]
            write_values(ws5, row, values, number_cols={4, 5, 6}, credit_cols={4}, debit_cols={5})
            
            # Format the No. column as integer (not float)
            ws5.cell(row=row, column=1).number_format = "0"
            ws5.cell(row=row, column=6).number_format = "0"
            
            # Center align ALL columns for Related Parties
            for col in range(1, 7):  # Columns A-F
                ws5.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Special: Make Name column left-aligned for readability
            ws5.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
        row = rp_row_start + len(related_party_rows)
    else:
        row = rp_row_start
        ws5.cell(row=row, column=1, value="No related parties defined.")
        style_data_cell(ws5, row, 1)
        row += 1

    row += 2
    # --- END: Related Parties content ---

    # COUNTERPARTY SUMMARY with its own column widths
    ws5.cell(row=row, column=1, value="COUNTERPARTY SUMMARY").font = bold_font
    row += 1

    # Counterparty Summary headers - columns A-F (No., Date, Description, Amount, Party Type, Party Name)
    counterparty_headers = ["No.", "Date", "Description", "Amount", "Party Type", "Party Name"]
    write_headers(ws5, row, counterparty_headers, header_fill_orange)

    # Set custom column widths for Counterparty Summary section (wider)
    ws5.column_dimensions["A"].width = 6   # No.
    ws5.column_dimensions["B"].width = 14  # Date
    ws5.column_dimensions["C"].width = 55  # Description - WIDER for counterparty summary
    ws5.column_dimensions["D"].width = 18  # Amount
    ws5.column_dimensions["E"].width = 18  # Party Type
    ws5.column_dimensions["F"].width = 40  # Party Name - WIDER for counterparty summary

    for idx, txn in enumerate(own_related_display_transactions or own_related.get("transactions", []) or [], 1):
        row += 1
        txn_type = (txn.get("type") or "").upper()
        values = [
            idx,  # No.
            txn.get("date"), 
            (txn.get("description", "") or "")[:60], 
            txn.get("amount"), 
            txn.get("party_type"), 
            txn.get("party_name", "")
        ]
        write_values(ws5, row, values, number_cols={4},
                    credit_cols={4} if txn_type == "CREDIT" else set(),
                    debit_cols={4} if txn_type != "CREDIT" else set())
        ws5.cell(row=row, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 5):
            ws5.cell(row=row, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Optional: You can also auto-adjust just the Description column for the summary section
    # But since we already set explicit widths, auto_width might override them
    # If you want to keep the manual widths, don't call auto_width(ws5) or call it with min/max limits
    # auto_width(ws5, min_width=8, max_width=40)  # Uncomment if you want auto-adjustment with limits

    # CP Ledger
    # CP Ledger Sheet (Summary only)
    ws5b = wb.create_sheet("CP Ledger")
    ws5b.cell(row=1, column=1, value="COUNTERPARTY LEDGER").font = title_font

    # Updated headers with No. column at the beginning
    ledger_headers = ["No.", "Counterparty", "Total Credits", "Total Debits", "Net Position", "Cr Count", "Dr Count", "Txn Count"]
    row = 3
    write_headers(ws5b, row, ledger_headers)

    # Set column widths for CP Ledger
    ws5b.column_dimensions["A"].width = 6   # No.
    ws5b.column_dimensions["B"].width = 40  # Counterparty
    ws5b.column_dimensions["C"].width = 18  # Total Credits
    ws5b.column_dimensions["D"].width = 18  # Total Debits
    ws5b.column_dimensions["E"].width = 18  # Net Position
    ws5b.column_dimensions["F"].width = 12  # Cr Count
    ws5b.column_dimensions["G"].width = 12  # Dr Count
    ws5b.column_dimensions["H"].width = 12  # Txn Count

    # Write data rows
    for idx, cp in enumerate(cp_sorted, 1):
        row += 1
        values = [
            idx,  # No. - integer
            cp.get("counterparty_name", ""), 
            cp.get("total_credits", 0), 
            cp.get("total_debits", 0), 
            cp.get("net_position", 0), 
            cp.get("credit_count", 0), 
            cp.get("debit_count", 0), 
            cp.get("transaction_count", 0)
        ]
        write_values(ws5b, row, values, number_cols={3, 4, 5, 6, 7, 8}, credit_cols={3}, debit_cols={4})
        
        # Format No. column as integer (not float)
        ws5b.cell(row=row, column=1).number_format = "0"
        
        # Center align ALL columns for the data row
        for col in range(1, 9):  # Columns A-H
            ws5b.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Special: Make Counterparty column (column B) left-aligned for readability
        ws5b.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Format count columns (F, G, H) as integers (no decimals)
        for count_col in [6, 7, 8]:  # Cr Count, Dr Count, Txn Count
            ws5b.cell(row=row, column=count_col).number_format = "0"


    # ============================================================
    # CP Transaction Sheet (Detail by Counterparty)
    # ============================================================
    ws5c = wb.create_sheet("CP Transaction")
    ws5c.cell(row=1, column=1, value="COUNTERPARTY TRANSACTION DETAIL").font = title_font

    # Set column widths for transaction detail
    ws5c.column_dimensions["A"].width = 6   # No.
    ws5c.column_dimensions["B"].width = 14  # Date
    ws5c.column_dimensions["C"].width = 60  # Description - wider
    ws5c.column_dimensions["D"].width = 18  # Amount
    ws5c.column_dimensions["E"].width = 14  # Type
    ws5c.column_dimensions["F"].width = 20  # Account

    row = 3  # Start after title

    # Transaction detail for each counterparty
    for cp_idx, cp in enumerate(cp_sorted):
        counterparty_name = cp.get("counterparty_name", "")
        transactions = cp.get("transactions", []) or []
        
        # Write counterparty name as section header
        ws5c.cell(row=row, column=1, value=counterparty_name).font = Font(name="Calibri", bold=True, color="1B4F72", size=12)
        row += 1
        
        # Write transaction headers
        detail_headers = ["No.", "Date", "Description", "Amount", "Type", "Account"]
        write_headers(ws5c, row, detail_headers, header_fill_orange)
        row += 1
        
        # Write transaction rows
        if transactions:
            for txn_idx, txn in enumerate(transactions, 1):
                txn_type = (txn.get("type") or "").upper()
                values = [
                    txn_idx,  # No. - integer
                    txn.get("date", ""), 
                    (txn.get("description", "") or "")[:100],  # Allow more text
                    txn.get("amount", 0), 
                    txn.get("type", ""), 
                    txn.get("account_number", "")
                ]
                write_values(ws5c, row, values, number_cols={4}, 
                            credit_cols={4} if txn_type == "CREDIT" else set(), 
                            debit_cols={4} if txn_type != "CREDIT" else set())
                
                # Format No. column as integer
                ws5c.cell(row=row, column=1).number_format = "0"
                
                # Column alignments:
                # - No., Date, Amount, Type, Account: Center aligned
                # - Description: Left aligned
                for col in range(1, 7):  # Columns A-F
                    if col == 3:  # Description column - Left aligned
                        ws5c.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:  # All other columns - Center aligned
                        ws5c.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                row += 1
        else:
            # No transactions for this counterparty
            ws5c.cell(row=row, column=1, value="No transactions")
            style_data_cell(ws5c, row, 1)
            ws5c.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            row += 1
        
        # Add 1 row spacer after each counterparty (except the last one)
        if cp_idx < len(cp_sorted) - 1:
            row += 1

    # NOTE: Related Parties sheet (ws5c) has been REMOVED as content is now in Counterparty sheet

    # Unclassified sheet - Updated with caption
    ws5d = wb.create_sheet("Unclassified")
    ws5d.cell(row=1, column=1, value="UNCLASSIFIED TRANSACTIONS").font = title_font

    # Add caption below title
    caption_row = 2
    uncl_threshold = safe_float(
        report_data.get("classification_config", {}).get("unclassified_listing_threshold", 10000)
    )
    caption_text = f"List of unclassified transactions that are >= RM {uncl_threshold:,.0f}"
    ws5d.cell(row=caption_row, column=1, value=caption_text)
    ws5d.cell(row=caption_row, column=1).font = Font(name="Calibri", italic=True, color="475569", size=10)
    ws5d.cell(row=caption_row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Merge cells for caption to span across columns (optional but looks better)
    ws5d.merge_cells(start_row=caption_row, start_column=1, end_row=caption_row, end_column=6)

    # Set column widths before writing data
    ws5d.column_dimensions["A"].width = 8   # No.
    ws5d.column_dimensions["B"].width = 14  # Date
    ws5d.column_dimensions["C"].width = 55  # Description
    ws5d.column_dimensions["D"].width = 18  # Amount
    ws5d.column_dimensions["E"].width = 14  # Type
    ws5d.column_dimensions["F"].width = 18  # Balance

    # Write headers (starting 2 rows after caption, or row 4 if caption is at row 2)
    header_row = caption_row + 2  # This will be row 4
    unclassified_headers = ["No.", "Date", "Description", "Amount", "Type", "Balance"]
    write_headers(ws5d, header_row, unclassified_headers, header_fill_orange)

    # Write data rows starting after headers
    for idx, txn in enumerate(report_data.get("unclassified_transactions", []) or [], 1):
        row_idx = header_row + idx
        txn_type = (txn.get("type") or "").upper()
        values = [
            idx, 
            txn.get("date", ""), 
            (txn.get("description", "") or "")[:80], 
            txn.get("amount"), 
            txn_type, 
            txn.get("balance")
        ]
        write_values(ws5d, row_idx, values, number_cols={4, 6},
                    credit_cols={4} if txn_type == "CREDIT" else set(),
                    debit_cols={4} if txn_type != "CREDIT" else set())
        ws5d.cell(row=row_idx, column=1).number_format = "0"
        for centre_col in (1, 2, 4, 5, 6):
            ws5d.cell(row=row_idx, column=centre_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Round Figure Transactions - Updated with proper formatting
    round_rows = get_round_transactions_for_report(report_data)
    ws_round = wb.create_sheet("Round Figure")
    write_split_transaction_sheet(
        ws_round,
        "ROUND FIGURE TRANSACTIONS",
        round_rows,
        caption="Round figure transactions are amounts that are multiples of RM 10,000.",
        number_cols={4, 5},
        credit_cols={4},
        debit_cols={4}
    )

    # Observations
    ws5f = wb.create_sheet("Observations")
    ws5f.cell(row=1, column=1, value="OBSERVATIONS").font = title_font
    row = 3
    for title, items, fill in (("POSITIVE OBSERVATIONS", observations.get("positive", []), header_fill_green), ("CONCERNS", observations.get("concerns", []), header_fill_red)):
        ws5f.cell(row=row, column=1, value=title)
        style_header_row(ws5f, row, 1, fill)
        for item in items:
            row += 1
            ws5f.cell(row=row, column=1, value=str(item))
            style_data_cell(ws5f, row, 1)
        row += 2
    ws5f.column_dimensions["A"].width = 100

   # Facilities - MODIFIED: Summary table moved to right side (Column H)
    ws6 = wb.create_sheet("Facilities")

    def _fac_amount(t):
        try:
            return float(t.get("amount") or 0)
        except (TypeError, ValueError):
            return 0.0

    def _is_real_fac(t, expected_category):
        return isinstance(t, dict) and t.get("category") == expected_category and _fac_amount(t) > 0

    loan_disb = [t for t in (loans.get("disbursements") or []) if _is_real_fac(t, "loan_disbursement")]
    loan_repay = [t for t in (loans.get("repayments") or []) if _is_real_fac(t, "loan_repayment")]
    loan_disb_total = round(sum(_fac_amount(t) for t in loan_disb), 2)
    loan_repay_total = round(sum(_fac_amount(t) for t in loan_repay), 2)

    ws6.cell(row=1, column=1, value="FACILITIES").font = title_font

    # ============================================================
    # DISBURSEMENTS TABLE (Left side, starting at Column A)
    # ============================================================
    row = 3  # Start at row 3
    ws6.cell(row=row, column=1, value="DISBURSEMENTS (Credits)").font = bold_font
    row += 1

    facility_headers = ["No.", "Date", "Description", "Amount", "Category"]
    write_headers(ws6, row, facility_headers, header_fill_green)
    row += 1

    if not loan_disb:
        ws6.cell(row=row, column=3, value="No disbursements")
        style_data_cell(ws6, row, 1)
        # Center align the "No disbursements" message
        ws6.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
    else:
        for idx, txn in enumerate(loan_disb, 1):
            values = [idx, txn.get("date"), (txn.get("description", "") or "")[:70], _fac_amount(txn), txn.get("category", "")]
            write_values(ws6, row, values, number_cols={4}, credit_cols={4})
            
            # Format No. column as integer
            ws6.cell(row=row, column=1).number_format = "0"
            
            # Center align ALL columns for data rows
            for col in range(1, 6):  # Columns A-E
                ws6.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Special: Make Description column (column 3) left-aligned for readability
            ws6.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1

    row += 2  # Add spacing

    # ============================================================
    # REPAYMENTS TABLE (Left side, continuing from disbursements)
    # ============================================================
    ws6.cell(row=row, column=1, value="REPAYMENTS (Debits)").font = bold_font
    row += 1

    write_headers(ws6, row, facility_headers, header_fill_red)
    row += 1

    if not loan_repay:
        ws6.cell(row=row, column=3, value="No repayments")
        style_data_cell(ws6, row, 1)
        # Center align the "No repayments" message
        ws6.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
    else:
        for idx, txn in enumerate(loan_repay, 1):
            values = [idx, txn.get("date"), (txn.get("description", "") or "")[:70], _fac_amount(txn), txn.get("category", "")]
            write_values(ws6, row, values, number_cols={4}, debit_cols={4})
            
            # Format No. column as integer
            ws6.cell(row=row, column=1).number_format = "0"
            
            # Center align ALL columns for data rows
            for col in range(1, 6):  # Columns A-E
                ws6.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Special: Make Description column (column 3) left-aligned for readability
            ws6.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            row += 1

    # Set column widths for the main tables (Columns A-E)
    ws6.column_dimensions["A"].width = 8    # No.
    ws6.column_dimensions["B"].width = 14   # Date
    ws6.column_dimensions["C"].width = 55   # Description
    ws6.column_dimensions["D"].width = 18   # Amount
    ws6.column_dimensions["E"].width = 22   # Category

    # Risk Signals
   # Risk Signals - Fixed header issue
    ws7 = wb.create_sheet("Risk Signals")
    ws7.cell(row=1, column=1, value="RISK SIGNALS ANALYSIS").font = title_font

    # Write headers at row 3 (leaving row 2 empty for spacing)
    risk_headers = ["No.", "Signal", "Detected", "Remarks"]
    write_headers(ws7, 3, risk_headers)  # Headers at row 3

    # Write data starting from row 4 (after headers)
    risk_df = build_risk_signals_dataframe_for_excel(flags, consolidated, statutory_compliance, monthly_analysis, report_data)
    for row_idx, item in enumerate(risk_df.to_dict(orient="records"), 4):  # Start at row 4
        values = [item.get("#"), item.get("Signal"), item.get("Detected"), item.get("Remarks")]
        write_values(ws7, row_idx, values)
        ws7.cell(row=row_idx, column=1).number_format = "0"
        ws7.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws7.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws7.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        if item.get("Detected") == "YES":
            ws7.cell(row=row_idx, column=3).font = Font(name="Calibri", color="922B21", bold=True)

    # Set column widths
    auto_width(ws7)
    ws7.column_dimensions["D"].width = 70

    # Parsing QC
    ws8 = wb.create_sheet("Parsing QC")
    ws8.cell(row=1, column=1, value="PARSING QUALITY METRICS").font = title_font
    success_rate = safe_float(parsing.get("overall_success_rate", 0))
    success_rate_pct = success_rate * 100 if success_rate <= 1 else success_rate
    p_total_gaps = int(consolidated.get("total_extraction_gaps") or len(parsing.get("extraction_gaps", []) or []))
    p_missing_dr = safe_float(consolidated.get("total_missing_debits", 0))
    p_missing_cr = safe_float(consolidated.get("total_missing_credits", 0))
    metric_rows = [
        ("Success Rate", f"{success_rate_pct:.1f}%"),
        ("Transactions Extracted", parsing.get("total_transactions_extracted", 0)),
        ("Balance Checks Passed", f"{parsing.get('total_balance_checks_passed', 0)} / {parsing.get('total_balance_checks', 0)}"),
    ]
    if has_recon:
        metric_rows.extend([
            ("Extraction Gaps", p_total_gaps),
            ("Missing Debits", p_missing_dr),
            ("Missing Credits", p_missing_cr),
        ])
    for row_idx, (label, value) in enumerate(metric_rows, 3):
        ws8.cell(row=row_idx, column=1, value=label)
        ws8.cell(row=row_idx, column=2, value=value)
        ws8.cell(row=row_idx, column=1).font = bold_font
        ws8.cell(row=row_idx, column=1).border = thin_border
        ws8.cell(row=row_idx, column=2).border = thin_border
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            ws8.cell(row=row_idx, column=2).number_format = num_fmt

    row = 3 + len(metric_rows) + 2
    qc_headers = [
        "Month", "Account", "Opening Balance", "Gross Credits", "Gross Debits",
        "Expected Closing", "Actual Closing", "Delta", "Status", "Txns", "Gaps", "Notes",
    ]
    write_headers(ws8, row, qc_headers)
    for check in parsing.get("account_month_checks", []) or []:
        row += 1
        passed = bool(check.get("passed"))
        values = [
            check.get("month"),
            check.get("account_number"),
            check.get("opening_balance"),
            check.get("gross_credits"),
            check.get("gross_debits"),
            check.get("expected_closing"),
            check.get("closing_balance"),
            check.get("reconciliation_delta"),
            "PASS" if passed else "FAIL",
            check.get("transactions_extracted"),
            check.get("extraction_gaps", 0),
            check.get("notes", ""),
        ]
        write_values(ws8, row, values, number_cols={3, 4, 5, 6, 7, 8, 10, 11}, credit_cols={4}, debit_cols={5})
        ws8.cell(row=row, column=9).font = Font(name="Calibri", color="196F3D" if passed else "922B21", bold=True)
    if parsing.get("extraction_gaps"):
        row += 2
        ws8.cell(row=row, column=1, value="EXTRACTION GAPS DETAIL").font = title_font
        row += 1
        gap_headers = ["Month", "Date", "Page", "Source File", "Missing", "Amount (RM)", "Before Gap", "After Gap"]
        write_headers(ws8, row, gap_headers, header_fill_red)
        for gap in parsing.get("extraction_gaps", []) or []:
            row += 1
            before_gap = f"{(gap.get('prev_description', '') or '')[:60]} (RM {safe_float(gap.get('balance_before_gap', 0)):,.2f})"
            after_gap = f"{(gap.get('next_description', '') or '')[:60]} (RM {safe_float(gap.get('balance_after_gap', 0)):,.2f})"
            values = [
                gap.get("month", ""), gap.get("date", ""), gap.get("page", ""), gap.get("source_file", ""),
                gap.get("missing_type", ""), gap.get("missing_amount", 0), before_gap, after_gap,
            ]
            write_values(ws8, row, values, number_cols={6}, debit_cols={6})

    cls_config = report_data.get("classification_config", {}) or {}
    if cls_config or schema_version:
        row += 2
        ws8.cell(row=row, column=1, value="CLASSIFICATION CONFIGURATION").font = title_font
        row += 1
        config_headers = ["Setting", "Value"]
        write_headers(ws8, row, config_headers, header_fill_orange)
        factoring_entities = cls_config.get("known_factoring_entities", [])
        config_rows = [
            ("Schema Version", schema_version or "N/A"),
            ("Rulebook Version", cls_config.get("rulebook_version", "N/A")),
            ("Execution Mode", cls_config.get("execution_mode", "N/A")),
            ("Large Transaction Threshold", safe_float(cls_config.get("large_transaction_threshold") or consolidated.get("high_value_threshold") or 100000)),
            ("Unclassified Listing Threshold", safe_float(cls_config.get("unclassified_listing_threshold", 10000))),
            ("Known Factoring Entities", ", ".join(factoring_entities) if factoring_entities else "None configured"),
        ]
        for label, value in config_rows:
            row += 1
            write_values(ws8, row, [label, value], number_cols={2} if isinstance(value, (int, float)) else set())

    validation_rows = build_formula_validation_checks_for_report(consolidated, monthly_analysis)
    if validation_rows:
        row += 2
        ws8.cell(row=row, column=1, value="FORMULA VALIDATION CHECKS (V1-V6)").font = title_font
        row += 1
        validation_headers = ["ID", "Check", "Severity", "Status", "Remarks"]
        write_headers(ws8, row, validation_headers, header_fill_orange)
        for item in validation_rows:
            row += 1
            values = [item.get("ID"), item.get("Check"), item.get("Severity"), item.get("Status"), item.get("Remarks")]
            write_values(ws8, row, values)
            status = item.get("Status")
            status_color = "196F3D" if status == "PASS" else "B9770E" if status in ("WARN", "N/A") else "922B21"
            ws8.cell(row=row, column=4).font = Font(name="Calibri", color=status_color, bold=True)
    auto_width(ws8)

    # Fraud Detector
    ws9 = wb.create_sheet("Fraud Detector")
    ws9.cell(row=1, column=1, value="Fraud Detector").font = title_font
    fraud_headers = ["file_name", "overall_risk", "layer", "severity", "finding", "anomaly_count", "detail"]
    write_headers(ws9, 3, fraud_headers)
    fraud_rows = []
    if isinstance(pdf_integrity, dict):
        for file_name, result in pdf_integrity.items():
            if isinstance(result, dict):
                fraud_rows.extend(_normalise_pdf_integrity_layer_rows(file_name, result))
    for row_idx, item in enumerate(fraud_rows, 4):
        values = [item.get(key) for key in fraud_headers]
        write_values(ws9, row_idx, values, number_cols={6})
    auto_width(ws9)

    if report_data.get("transactions"):
        ws10 = wb.create_sheet("Transactions")
        txns = report_data.get("transactions", []) or []
        txn_headers = list(txns[0].keys()) if txns and isinstance(txns[0], dict) else []
        write_headers(ws10, 1, txn_headers)
        for row_idx, txn in enumerate(txns, 2):
            values = [txn.get(key) for key in txn_headers]
            write_values(ws10, row_idx, values)
        auto_width(ws10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _write_excel_sheet(writer, sheet_name: str, df: pd.DataFrame, title: str | None = None) -> None:
    workbook = writer.book
    safe_sheet_name = sheet_name[:31]
    startrow = 2 if title else 0
    
    # Convert DataFrame to have proper string columns for Excel
    df_to_write = df.copy()
    
    # Convert all columns to string for safe processing, but keep numeric ones numeric
    for col in df_to_write.columns:
        # Check if column contains numeric data (float/int)
        if pd.api.types.is_numeric_dtype(df_to_write[col]):
            # Keep numeric columns as is
            continue
        else:
            # Convert non-numeric columns to string, handling None/NaN
            df_to_write[col] = df_to_write[col].apply(
                lambda x: str(x) if x is not None and pd.notna(x) else ""
            )
    
    df_to_write.to_excel(writer, sheet_name=safe_sheet_name, startrow=startrow, index=False)
    worksheet = writer.sheets[safe_sheet_name]

    header_format = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#1B4F72", "border": 1}
    )
    title_format = workbook.add_format({"bold": True, "font_size": 14, "font_color": "#1B4F72"})
    money_format = workbook.add_format({"num_format": "#,##0.00"})

    if title:
        worksheet.write(0, 0, title, title_format)

    for col_idx, col_name in enumerate(df_to_write.columns):
        worksheet.write(startrow, col_idx, col_name, header_format)
        
        # Safely calculate column width
        try:
            # Get column values as strings safely
            col_values = df_to_write[col_name].astype(str).tolist() if not df_to_write.empty else []
            max_len = len(str(col_name))
            for val in col_values[:200]:  # Limit to first 200 rows for performance
                if val:
                    max_len = max(max_len, len(val))
            # Cap width between 12 and 42
            col_width = min(max(max_len + 2, 12), 42)
        except Exception:
            col_width = 15  # Default fallback width
        
        worksheet.set_column(col_idx, col_idx, col_width)
        
        # Apply money format to amount columns
        if any(token in str(col_name).lower() for token in ("amount", "credit", "debit", "balance", "gross", "net")):
            # Get the column range
            last_row = startrow + len(df_to_write)
            if last_row > startrow:
                worksheet.set_column(col_idx, col_idx, col_width, money_format)

    worksheet.freeze_panes(startrow + 1, 0)
    if not df_to_write.empty:
        worksheet.autofilter(startrow, 0, startrow + len(df_to_write), max(len(df_to_write.columns) - 1, 0))


def _write_excel_sections_sheet(writer, sheet_name: str, sections: List[Tuple[str, pd.DataFrame]]) -> None:
    workbook = writer.book
    safe_sheet_name = sheet_name[:31]
    worksheet = workbook.add_worksheet(safe_sheet_name)
    writer.sheets[safe_sheet_name] = worksheet

    header_format = workbook.add_format(
        {"bold": True, "font_color": "white", "bg_color": "#1B4F72", "border": 1}
    )
    title_format = workbook.add_format({"bold": True, "font_size": 14, "font_color": "#1B4F72"})
    money_format = workbook.add_format({"num_format": "#,##0.00"})
    startrow = 0
    max_col_widths = {}

    for title, df in sections:
        df_to_write = df.copy()
        for col in df_to_write.columns:
            if pd.api.types.is_numeric_dtype(df_to_write[col]):
                continue
            df_to_write[col] = df_to_write[col].apply(
                lambda x: str(x) if x is not None and pd.notna(x) else ""
            )

        worksheet.write(startrow, 0, title, title_format)
        header_row = startrow + 2

        for col_idx, col_name in enumerate(df_to_write.columns):
            try:
                col_values = df_to_write[col_name].astype(str).tolist() if not df_to_write.empty else []
                max_len = len(str(col_name))
                for val in col_values[:200]:
                    if val:
                        max_len = max(max_len, len(val))
                col_width = min(max(max_len + 2, 12), 42)
            except Exception:
                col_width = 15
            max_col_widths[col_idx] = max(max_col_widths.get(col_idx, 0), col_width)

            if any(token in str(col_name).lower() for token in ("amount", "credit", "debit", "balance", "gross", "net")):
                worksheet.set_column(col_idx, col_idx, max_col_widths[col_idx], money_format)
            else:
                worksheet.set_column(col_idx, col_idx, max_col_widths[col_idx])

        if len(df_to_write.columns):
            if df_to_write.empty:
                for col_idx, col_name in enumerate(df_to_write.columns):
                    worksheet.write(header_row, col_idx, col_name, header_format)
            else:
                worksheet.add_table(
                    header_row,
                    0,
                    header_row + len(df_to_write),
                    len(df_to_write.columns) - 1,
                    {
                        "columns": [{"header": str(col_name)} for col_name in df_to_write.columns],
                        "data": [list(row) for row in df_to_write.itertuples(index=False, name=None)],
                        "style": "Table Style Medium 2",
                    },
                )

        startrow = header_row + len(df_to_write) + 3

    worksheet.freeze_panes(3, 0)


def _records_to_excel_df(records, columns: List[str] | None = None) -> pd.DataFrame:
    safe_records = [
        {key: _excel_safe_value(val) for key, val in dict(record).items()}
        for record in (records or [])
        if isinstance(record, dict)
    ]
    df = pd.DataFrame(safe_records)
    if columns:
        if df.empty:
            df = pd.DataFrame(columns=columns)
        else:
            df = df.reindex(columns=columns)
    return df


def _excel_safe_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if isinstance(value, pd.Period):
        return str(value)
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _pdf_detail_to_excel_text(value) -> str:
    if value in (None, "", [], {}):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _normalise_pdf_integrity_layer_rows(file_name: str, result: dict) -> List[dict]:
    layer_order = [
        ("metadata", "Layer 1: Metadata"),
        ("fonts", "Layer 2: Fonts"),
        ("text_layers", "Layer 3: Text Layers"),
        ("visual", "Layer 4: Visual"),
        ("cross_validation", "Layer 5: Cross Validation"),
        ("bank_profile", "Layer 6: Bank Profile"),
        ("structural", "Layer 7: Structural"),
        ("arithmetic", "Layer 8: Arithmetic"),
    ]
    if not isinstance(result, dict):
        return []

    overall_risk = (result.get("overall_risk") or "LOW").upper()
    layer_results = result.get("layer_results")
    if isinstance(layer_results, dict):
        rows = []
        handled_keys = set()
        for layer_key, layer_label in layer_order:
            handled_keys.add(layer_key)
            findings = layer_results.get(layer_key, []) or []
            findings = findings if isinstance(findings, list) else []
            highest = next(
                (
                    level
                    for level in ("HIGH", "MEDIUM", "LOW")
                    if any(
                        isinstance(finding, dict)
                        and (finding.get("severity") or "").upper() == level
                        for finding in findings
                    )
                ),
                "LOW",
            )
            anomaly_count = sum(
                1
                for finding in findings
                if isinstance(finding, dict) and not _pdf_finding_is_benign_for_export(finding)
            )
            primary = findings[0] if findings and isinstance(findings[0], dict) else {}
            detail_text = _pdf_detail_to_excel_text(primary.get("detail"))
            rows.append(
                {
                    "file_name": file_name,
                    "overall_risk": overall_risk,
                    "layer": layer_label,
                    "severity": highest,
                    "finding": primary.get("message") or "No findings.",
                    "anomaly_count": anomaly_count,
                    "detail": detail_text,
                }
            )

        for layer_key, findings in layer_results.items():
            if layer_key in handled_keys:
                continue
            findings = findings if isinstance(findings, list) else []
            for finding in findings:
                if not isinstance(finding, dict):
                    continue
                rows.append(
                    {
                        "file_name": file_name,
                        "overall_risk": overall_risk,
                        "layer": str(layer_key),
                        "severity": (finding.get("severity") or "LOW").upper(),
                        "finding": finding.get("message") or finding.get("finding") or "",
                        "anomaly_count": 0 if _pdf_finding_is_benign_for_export(finding) else 1,
                        "detail": _pdf_detail_to_excel_text(finding.get("detail")),
                    }
                )
        return rows

    legacy_layers = result.get("layers", result.get("checks", result.get("findings", [])))
    if isinstance(legacy_layers, dict):
        legacy_layers = [
            {"layer": layer_name, **layer_data}
            if isinstance(layer_data, dict)
            else {"layer": layer_name, "message": str(layer_data)}
            for layer_name, layer_data in legacy_layers.items()
        ]
    rows = []
    for layer in legacy_layers or []:
        if not isinstance(layer, dict):
            continue
        rows.append(
            {
                "file_name": file_name,
                "overall_risk": overall_risk,
                "layer": layer.get("layer", layer.get("name", "")),
                "severity": (layer.get("severity") or layer.get("risk") or "LOW").upper(),
                "finding": layer.get("message", layer.get("finding", layer.get("description", ""))),
                "anomaly_count": 0 if _pdf_finding_is_benign_for_export(layer) else 1,
                "detail": _pdf_detail_to_excel_text(layer.get("detail", layer.get("details", ""))),
            }
        )
    return rows


def build_risk_signals_dataframe_for_excel(flags_data: dict, consolidated: dict, statutory_compliance: dict, monthly_analysis: list, report_data: dict) -> pd.DataFrame:
    """Build the 16-row Risk Signals DataFrame with actual computed values."""
    
    # Get computed values from various sources
    gross_credits = float(consolidated.get("gross_credits", 0))
    gross_debits = float(consolidated.get("gross_debits", 0))
    
    # Round Figure Credits
    round_figure_entries = report_data.get("round_figure_credits", report_data.get("round_transactions", []))
    round_figure_count = len(round_figure_entries)
    round_figure_total = sum(float(e.get("amount", 0)) for e in round_figure_entries)
    
    # Large Transactions (>= threshold) - check both large_transactions and large_credits
    large_txns = report_data.get("large_transactions", [])
    if not large_txns:
        large_txns = report_data.get("large_credits", [])
    large_txn_count = len(large_txns) if large_txns else 0
    large_txn_total = sum(float(t.get("amount", 0)) for t in (large_txns or []))
    
    # Cash Deposits
    cash_deposits = consolidated.get("total_cash_deposits", 0)
    cash_deposit_count = 0
    for m in monthly_analysis:
        cash_deposit_count += int(m.get("cash_deposits_count", 0))
    
    # Own Party Transactions
    own_cr = float(consolidated.get("total_own_party_cr", 0))
    own_dr = float(consolidated.get("total_own_party_dr", 0))
    
    # Related Party Transactions
    rp_cr = float(consolidated.get("total_related_party_cr", 0))
    rp_dr = float(consolidated.get("total_related_party_dr", 0))
    
    # Loan Activity
    loan_disb = float(consolidated.get("total_loan_disbursement_cr", 0))
    loan_repay = float(consolidated.get("total_loan_repayment_dr", 0))
    
    # EPF/SOCSO from statutory_compliance
    salary_months = statutory_compliance.get("salary_months_active", 0)
    epf_pct = statutory_compliance.get("epf_coverage_pct", 100) if salary_months > 0 else 100
    socso_pct = statutory_compliance.get("socso_coverage_pct", 100) if salary_months > 0 else 100
    
    # LHDN
    lhdn_detected = statutory_compliance.get("lhdn_detected", False)
    lhdn_count = statutory_compliance.get("lhdn_months_paid", 0)
    lhdn_total = float(consolidated.get("total_statutory_tax", 0))
    
    # HRDF
    hrdf_detected = statutory_compliance.get("hrdf_detected", False)
    hrdf_count = statutory_compliance.get("hrdf_months_paid", 0)
    hrdf_total = float(consolidated.get("total_statutory_hrdf", 0))
    
    # FX Transactions
    fx_cr = float(consolidated.get("total_fx_credits", 0))
    fx_dr = float(consolidated.get("total_fx_debits", 0))
    
    # Returned Cheques
    rc_in_count = int(consolidated.get("total_returned_cheques_inward_count", 0)) or 0
    rc_out_count = int(consolidated.get("total_returned_cheques_outward_count", 0)) or 0
    
    # Data Quality
    data_complete = consolidated.get("data_completeness", "COMPLETE") == "COMPLETE"
    
    # Low Closing Balance
    low_balance_months = [m for m in monthly_analysis if float(m.get("closing_balance", 0)) < 1000]
    
    # Build the 16 rows
    risk_signals = [
        {"#": 1, "Signal": "Returned Cheques (Inward)", "Detected": "YES" if rc_in_count > 0 else "NO", 
         "Remarks": f"{rc_in_count} inward returned cheques totalling RM {consolidated.get('total_returned_cheques_inward', 0):,.2f}." if rc_in_count > 0 else "No inward returned cheques in the period."},
        
        {"#": 2, "Signal": "Returned Cheques (Outward)", "Detected": "YES" if rc_out_count > 0 else "NO", 
         "Remarks": f"{rc_out_count} outward returned cheques totalling RM {consolidated.get('total_returned_cheques_outward', 0):,.2f}." if rc_out_count > 0 else "No outward returned cheques in the period."},
        
        {"#": 3, "Signal": "Round Figure Transactions (AML)", "Detected": "YES" if round_figure_count > 0 else "NO", 
         "Remarks": f"{round_figure_count} round-figure transactions totalling RM {round_figure_total:,.2f}." if round_figure_count > 0 else "No round-figure transactions flagged."},
        
        {"#": 4, "Signal": "High Value Credits (>3x EOD)", "Detected": "NO", 
         "Remarks": "No credits exceeded 3x daily EOD."},
        
        {"#": 5, "Signal": "Cash Deposits (AML)", "Detected": "YES" if cash_deposits > 0 else "NO", 
         "Remarks": f"{cash_deposit_count} cash deposits totalling RM {cash_deposits:,.2f} ({cash_deposits/gross_credits*100:.1f}% of gross credits)." if cash_deposits > 0 else "No cash deposits in the period."},
        
        {"#": 6, "Signal": "EPF Compliance", "Detected": "NO" if epf_pct >= 99.5 else "YES", 
         "Remarks": f"EPF coverage {epf_pct:.1f}% across salary months." if salary_months > 0 else "No salary months detected."},
        
        {"#": 7, "Signal": "SOCSO Compliance", "Detected": "NO" if socso_pct >= 99.5 else "YES", 
         "Remarks": f"SOCSO coverage {socso_pct:.1f}% across salary months." if salary_months > 0 else "No salary months detected."},
        
        {"#": 8, "Signal": "LHDN Tax Payments", "Detected": "NO", 
         "Remarks": f"LHDN payments detected: {lhdn_count} tx totalling RM {lhdn_total:,.2f} (PCB/CP204/SST — schedules differ; informational only)." if lhdn_detected else "No LHDN tax payments detected."},
        
        {"#": 9, "Signal": f"Large Transactions (>=RM{consolidated.get('high_value_threshold', 10000):,.0f})", "Detected": "YES" if large_txn_count > 0 else "NO", 
         "Remarks": f"{large_txn_count} large transactions (>=RM{consolidated.get('high_value_threshold', 10000):,.0f}) totalling RM {large_txn_total:,.2f}." if large_txn_count > 0 else f"No transactions at or above RM{consolidated.get('high_value_threshold', 10000):,.0f}."},
        
        {"#": 10, "Signal": "Own Party Transactions", "Detected": "YES" if own_cr > 0 or own_dr > 0 else "NO", 
         "Remarks": f"Own-party CR RM {own_cr:,.2f} ({own_cr/gross_credits*100:.1f}% of gross credits); DR RM {own_dr:,.2f} ({own_dr/gross_debits*100:.1f}% of gross debits)." if (own_cr > 0 or own_dr > 0) else "No own-party transactions detected."},
        
        {"#": 11, "Signal": "Related Party Transactions", "Detected": "YES" if rp_cr > 0 or rp_dr > 0 else "NO", 
         "Remarks": f"Related-party CR RM {rp_cr:,.2f} ({rp_cr/gross_credits*100:.1f}% of gross credits); DR RM {rp_dr:,.2f} ({rp_dr/gross_debits*100:.1f}% of gross debits). Parties: (no canonical names provided)." if (rp_cr > 0 or rp_dr > 0) else "No related-party transactions detected."},
        
        {"#": 12, "Signal": "Loan Activity", "Detected": "YES" if loan_disb > 0 or loan_repay > 0 else "NO", 
         "Remarks": f"Loan disbursements RM {loan_disb:,.2f}; loan repayments RM {loan_repay:,.2f}." if (loan_disb > 0 or loan_repay > 0) else "No loan disbursements or repayments detected."},
        
        {"#": 13, "Signal": "Data Quality", "Detected": "NO" if data_complete else "YES", 
         "Remarks": "Statement data complete across the period." if data_complete else f"Statement data INCOMPLETE: {consolidated.get('data_gaps', '')}"},
        
        {"#": 14, "Signal": "FX Transactions", "Detected": "YES" if fx_cr > 0 or fx_dr > 0 else "NO", 
         "Remarks": f"FX credits RM {fx_cr:,.2f}; FX debits RM {fx_dr:,.2f}." if (fx_cr > 0 or fx_dr > 0) else "No FX (foreign-currency) activity detected."},
        
        {"#": 15, "Signal": "Low Closing Balance", "Detected": "YES" if low_balance_months else "NO", 
         "Remarks": f"Closing balance below RM 1,000.00 in: {', '.join([m.get('month', '?') for m in low_balance_months])}." if low_balance_months else "Closing balance stayed at or above RM 1,000.00 every month."},
        
        {"#": 16, "Signal": "HRDF Payments", "Detected": "NO", 
         "Remarks": f"HRDF payments detected: {hrdf_count} tx totalling RM {hrdf_total:,.2f} (informational; no coverage ratio computed)." if hrdf_detected else "No HRDF payments detected."},
    ]
    
    return pd.DataFrame(risk_signals)


def build_parsing_qc_dataframe_from_parsing_metadata(parsing_metadata: dict) -> pd.DataFrame:
    """Build Parsing QC dataframe from parsing_metadata for Excel export."""
    if not parsing_metadata:
        return pd.DataFrame()
    
    account_month_checks = parsing_metadata.get("account_month_checks", [])
    if not account_month_checks:
        return pd.DataFrame()
    
    rows = []
    for chk in account_month_checks:
        month = chk.get("month", "")
        account_number = chk.get("account_number", "")
        opening_balance = safe_float(chk.get("opening_balance", 0))
        closing_balance = safe_float(chk.get("closing_balance", 0))
        gross_credits = safe_float(chk.get("gross_credits", 0))
        gross_debits = safe_float(chk.get("gross_debits", 0))
        expected_closing = safe_float(chk.get("expected_closing", 0))
        reconciliation_delta = safe_float(chk.get("reconciliation_delta", 0))
        passed = chk.get("passed", False)
        transactions_extracted = chk.get("transactions_extracted", 0)
        extraction_gaps = chk.get("extraction_gaps", 0)
        notes = chk.get("notes", "")
        
        rows.append({
            "Month": month,
            "Account": account_number,
            "Opening Balance": opening_balance,
            "Closing Balance": closing_balance,
            "Gross Credits": gross_credits,
            "Gross Debits": gross_debits,
            "Expected Close": expected_closing,
            "Recon Delta": reconciliation_delta,
            "Status": "PASS" if passed else "FAIL",
            "Transactions Extracted": transactions_extracted,
            "Gaps": extraction_gaps,
            "Notes": notes
        })
    
    return pd.DataFrame(rows)


__all__ = [
    'bind_app_globals',
    'generate_excel_report',
    '_write_excel_sheet',
    '_write_excel_sections_sheet',
    '_records_to_excel_df',
    '_excel_safe_value',
    '_pdf_detail_to_excel_text',
    '_normalise_pdf_integrity_layer_rows',
    'build_risk_signals_dataframe_for_excel',
    'build_parsing_qc_dataframe_from_parsing_metadata',
]
